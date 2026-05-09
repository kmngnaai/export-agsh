import logging

import os
import re
import json
import threading
import hashlib
import copy
import traceback
import subprocess
import tempfile
import shutil
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Dict, List, Optional, Tuple, Set

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import xlrd
import sys
from pathlib import Path

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception:  # pragma: no cover
    tk = None
    filedialog = None
    messagebox = None
    ttk = None

APP_TITLE = "Sub Detail Collector - Python"
CONFIG_FILE = os.path.join(
    os.path.dirname(__file__),
    "config_detail_sub.json",
)
MAX_WORKERS = 4

SUB_DETAIL_HEADERS = [
    "Folder Created", "Folder Name", "File", "PL Sheet", "INV Sheet",
    "Pair QTY", "Carton QTY", "CBM", "Gross Weight",
    "INV_QtyPairs", "INV_Amount_WO_Label", "INV_LabelAmount", "INV_TotalAmount",
    "Invoice No", "Invoice Date", "Departure", "Destination", "Trade Term",
    "Shipping Date", "Payment Term", "CUS PO#", "Status", "Error",
    "File Path", "File Signature"
]

LOG_HEADERS = ["Run Time", "File", "Step", "Result", "Detail", "File Path"]

INV_DETAIL_HEADERS = [
    "Folder Created", "Folder Name", "File", "INV Sheet",
    "Folder Date", "Folder Invoice No", "Folder Destination",
    "Date", "Invoice No", "Destination", "PO No", "Customer PO#",
    "DESC", "HS CODE", "MODEL NAME / COLOR", "MODEL Y", "COLOR", "WIDTH",
    "QTY (Pairs)", "UNIT PRICE (USD)", "AMOUNT W/OUT LABEL",
    "LABEL COST", "LABEL AMOUNT", "TOTAL AMOUNT",
    "Row Type", "Status", "File Path"
]

PL_DETAIL_HEADERS = [
    "Folder Created", "Folder Name", "File", "PL Sheet",
    "Folder Date", "Folder Invoice No", "Folder Destination",
    "Date", "Invoice No", "Destination", "PO No.", "Customer PO#",
    "MODEL NAME / COLOR", "COLOR", "WIDTH", "HS CODE",
    "QTY (Pairs)", "CARTON QTY", "CBM", "GROSS WEIGHT",
    "Row Type", "Status", "File Path"
]

STATUS_FILLS = {
    "WARNING": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "ERROR": PatternFill(fill_type="solid", fgColor="FFC7CE"),
}
HEADER_FILL_SUB = PatternFill(fill_type="solid", fgColor="DCE6F1")
HEADER_FILL_LOG = PatternFill(fill_type="solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True)

META_KEYS = [
    "invoice_no", "invoice_date", "departure", "destination",
    "trade_term", "shipping_date", "payment_term", "cus_po"
]

COMPARE_KEYS = [
    "invoice_no", "invoice_date", "destination", "trade_term",
    "departure", "shipping_date", "payment_term", "cus_po"
]


@dataclass
class WorkbookData:
    file_path: str
    engine: str
    sheet_names: List[str]
    sheets: Dict[str, List[List[Any]]]


class AppLogger:
    def __init__(self, callback=None):
        self.callback = callback
        self._lock = threading.Lock()

    def log(self, msg: str, level: str = "info"):
        with self._lock:
            if self.callback:
                self.callback(msg, level)
            else:
                print(f"[{level.upper()}] {msg}")


@dataclass
class RepairOptions:
    use_folder_truth: bool = False
    use_manual_values: bool = False
    manual_invoice_no: str = ""
    manual_invoice_date: str = ""
    manual_destination: str = ""

def nz_str(v: Any) -> str:
    if v is None:
        return ""
    try:
        s = str(v).strip()
    except Exception:
        return ""
    return s


_ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

def sanitize_excel_text(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, (int, float, datetime)):
        return v
    s = str(v)
    s = _ILLEGAL_XML_RE.sub(" ", s)
    # Prevent Excel from interpreting text as a formula in logs/details.
    # This is the main cause of "Removed Records: Formula..." when text starts
    # with =, +, -, @ or repeated === markers.
    if s and s[0] in ("=", "+", "-", "@"):
        s = "'" + s
    return s


def normalize_path_text(s: str) -> str:
    t = nz_str(s).replace("/", "\\")
    while "\\\\" in t:
        t = t.replace("\\\\", "\\")
    return t.lower()


def normalize_simple_text(s: str) -> str:
    s = nz_str(s).replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", s).strip()


def normalize_header(s: str) -> str:
    s = nz_str(s).upper().strip()
    replacements = {
        "\r": " ", "\n": " ", "'": "", '"': "", ".": "", ":": "",
        "/": "", "\\": "", "(": "", ")": "", "-": "", "_": "",
        "&": "AND",
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    s = re.sub(r"\s+", " ", s)
    return "".join(ch for ch in s if ch.isalnum())


def normalize_meta_label(s: str) -> str:
    s = nz_str(s).upper().strip()
    for old, new in [("\r", " "), ("\n", " "), (".", ""), (":", ""), ("#", ""), ("/", " ")]:
        s = s.replace(old, new)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_invoice_no(s: str) -> str:
    s = nz_str(s).upper().strip()
    return "".join(ch for ch in s if ch.isalnum() or ch == ".")


def normalize_text_compare(s: str) -> str:
    s = nz_str(s).upper().strip().replace("\r", " ").replace("\n", " ")
    s = s.replace(" ,", ",").replace(", ", ", ")
    s = s.replace("VIETNAM", "VIET NAM")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def try_parse_loose_date(s: str) -> Optional[str]:
    t = nz_str(s).strip()
    if not t:
        return None

    # First: let datetime handle common timestamp strings like:
    # 2026-04-01 00:00:00 / 2026-04-01 / 01/04/2026 00:00:00
    direct_formats = (
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
        "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y",
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d",
        "%d/%m/%y", "%m/%d/%y",
    )
    for fmt in direct_formats:
        try:
            return datetime.strptime(t, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass

    # Fallback: normalize separators and strip time tail if present
    t2 = t.upper().replace(".", "/").replace("-", "/")
    m = re.match(r"^(\d{1,4}/\d{1,2}/\d{1,4})", t2)
    if m:
        t2 = m.group(1)

    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d/%m/%y", "%m/%d/%y"):
        try:
            return datetime.strptime(t2, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return None


def excel_serial_to_datetime(value: Any) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)) and value > 0:
        try:
            return openpyxl.utils.datetime.from_excel(value)
        except Exception:
            pass
    return None


def normalize_date_compare(v: Any) -> str:
    dt = excel_serial_to_datetime(v)
    if dt is not None:
        return dt.strftime("%Y-%m-%d")
    s = nz_str(v)
    if not s:
        return ""
    loose = try_parse_loose_date(s)
    return loose if loose else s


def format_date_display(v: Any) -> str:
    norm = normalize_date_compare(v)
    if not norm:
        return ""
    try:
        return datetime.strptime(norm, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return nz_str(v)


def normalize_number_text(s: Any) -> str:
    t = nz_str(s)
    if not t:
        return ""
    t = t.replace(",", "").replace(" ", "")
    try:
        return str(Decimal(t).normalize()) if "." in t else str(Decimal(t))
    except (InvalidOperation, ValueError):
        return t


def normalize_compare_by_field(v: Any, field_name: str) -> str:
    s = nz_str(v)
    if not s and not isinstance(v, (int, float, datetime)):
        return ""
    field_name = field_name.lower()
    if field_name == "invoice_no":
        return normalize_invoice_no(v)
    if field_name in ("invoice_date", "shipping_date"):
        return normalize_date_compare(v)
    if field_name in ("pair_qty", "inv_total_amount"):
        return normalize_number_text(v)
    return normalize_text_compare(v)


def format_date_for_display(v: Any) -> str:
    s = nz_str(v)
    if not s and not isinstance(v, (int, float, datetime)):
        return ""
    iso_text = normalize_date_compare(v)
    try:
        if iso_text:
            return datetime.strptime(iso_text, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        pass
    try:
        dt = excel_serial_to_datetime(v)
        if dt is not None:
            return dt.strftime("%d/%m/%Y")
    except Exception:
        pass
    return s



def coerce_excel_date_value(v: Any) -> Any:
    """Return a real datetime for Excel date cells when possible."""
    if isinstance(v, datetime):
        return v
    dt = excel_serial_to_datetime(v)
    if dt is not None:
        return dt
    iso_text = normalize_date_compare(v)
    if not iso_text:
        return v
    try:
        return datetime.strptime(iso_text, "%Y-%m-%d")
    except Exception:
        return v


def _v74_excel_serial_text(v: Any) -> str:
    dt = _v74_to_excel_datetime(v)
    if isinstance(dt, datetime):
        try:
            serial = openpyxl.utils.datetime.to_excel(dt)
            return str(int(round(serial)))
        except Exception:
            return ""
    return ""


def _v74_make_sdatewinv(date_val: Any, s_invoice: Any) -> str:
    seq = nz_str(s_invoice).strip()
    if not seq:
        return ""
    serial = _v74_excel_serial_text(date_val)
    return f"{serial}{seq}" if serial else seq


def _v74_coerce_hscode(v: Any) -> Any:
    if v is None or v == "":
        return ""
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if float(v).is_integer() else float(v)
    s = nz_str(v).strip().replace(',', '')
    if not s:
        return ""
    if re.fullmatch(r"[+-]?\d+", s):
        return int(s)
    if re.fullmatch(r"[+-]?\d+\.0+", s):
        return int(float(s))
    if re.fullmatch(r"[+-]?\d+\.\d+", s):
        f = float(s)
        return int(f) if f.is_integer() else f
    return s


def get_meta_key_by_alias(norm_label: str) -> str:
    aliases = {
        "INVOICE NO": "invoice_no",
        "INVOICE NO ": "invoice_no",
        "INVOICE NUMBER": "invoice_no",
        "DATE": "invoice_date",
        "INVOICE DATE": "invoice_date",
        "DEPARTURE": "departure",
        "DESTINATION": "destination",
        "TRADE TERM": "trade_term",
        "TRADE TERMS": "trade_term",
        "TERMS OF DELIVERY": "trade_term",
        "SHIPPING DATE": "shipping_date",
        "SHIPPINGDATE": "shipping_date",
        "PAYMENT TERM": "payment_term",
        "PAYMENT TERMS": "payment_term",
        "CUS PO": "cus_po",
        "CUS PO NO": "cus_po",
        "CUS PO NO ": "cus_po",
        "CUS PO ": "cus_po",
        "CUS PONO": "cus_po",
    }
    return aliases.get(norm_label, "")


def is_pl_pair_qty_header(txt: str) -> bool:
    return ("QTY" in txt and "PAIR" in txt) or ("QTY" in txt and "PRS" in txt)


def is_pl_carton_header(txt: str) -> bool:
    return ("CARTON" in txt and "QTY" in txt) or ("PACKAGE" in txt and "CTN" in txt)


def is_pl_cbm_header(txt: str) -> bool:
    return txt == "CBM" or ("MEASUREMENT" in txt and "CBM" in txt)


def is_pl_gross_header(txt: str) -> bool:
    return ("GROSS" in txt and "WEIGHT" in txt) or ("GW" in txt and "KG" in txt)


def is_inv_pair_qty_header(txt: str) -> bool:
    return ("QTY" in txt and "PAIR" in txt) or ("QTY" in txt and "PRS" in txt)


def is_inv_amount_without_label_header(txt: str) -> bool:
    return ("AMOUNT" in txt and "WOUT" in txt and "LABEL" in txt) or (
        "AMOUNT" in txt and "WITHOUT" in txt and "LABEL" in txt
    )


def is_inv_label_amount_header(txt: str) -> bool:
    txt = txt.upper()
    return (
        "TOTALLABELCOST" in txt or "LABELAMOUNT" in txt or
        ("LABEL" in txt and "AMOUNT" in txt) or
        ("LABEL" in txt and "COST" in txt and "WITHOUT" not in txt)
    )


def is_inv_total_amount_header(txt: str) -> bool:
    return txt in ("TOTALAMOUNT", "AMOUNTUSD")


def match_pl_detail_key(txt: str) -> str:
    t = normalize_header(txt)
    if not t:
        return ""
    if t.startswith("DATE"):
        return "date"
    if "INVOICENO" in t or "INVOICENUMBER" in t:
        return "invoice_no"
    if "DESTINATION" in t:
        return "destination"
    if t.startswith("PONO"):
        return "po_no"
    if "CUSTOMERPO" in t or "CUSPO" in t:
        return "customer_po"
    if "MODELNAMECOLOR" in t:
        return "model_name_color"
    if t == "COLOR":
        return "color"
    if "WIDTH" in t:
        return "width"
    if "HSCODE" in t:
        return "hs_code"
    if is_pl_pair_qty_header(t):
        return "pair_qty"
    if is_pl_carton_header(t):
        return "carton_qty"
    if is_pl_cbm_header(t):
        return "cbm"
    if is_pl_gross_header(t):
        return "gross_weight"
    return ""


def match_inv_detail_key(txt: str) -> str:
    t = normalize_header(txt)
    if not t:
        return ""
    if t.startswith("DATE"):
        return "date"
    if "INVOICENO" in t or "INVOICENUMBER" in t:
        return "invoice_no"
    if "DESTINATION" in t:
        return "destination"
    if t.startswith("PONO"):
        return "po_no"
    if "CUSTOMERPO" in t or "CUSPO" in t:
        return "customer_po"
    if t == "DESC":
        return "desc"
    if "HSCODE" in t:
        return "hs_code"
    if "MODELNAMECOLOR" in t:
        return "model_name_color"
    if t == "MODELY":
        return "model_y"
    if t == "COLOR":
        return "color"
    if "WIDTH" in t:
        return "width"
    if is_inv_pair_qty_header(t):
        return "qty_pairs"
    if "UNITPRICE" in t and "USD" in t:
        return "unit_price_usd"
    if is_inv_amount_without_label_header(t):
        return "amount_wo_label"
    if "LABELCOST" in t and "TOTAL" not in t:
        return "label_cost"
    if is_inv_label_amount_header(t):
        return "label_amount"
    if is_inv_total_amount_header(t):
        return "total_amount"
    return ""


def append_error(base_text: str, add_text: str) -> str:
    add_text = nz_str(add_text)
    if not add_text:
        return base_text
    return f"{base_text}; {add_text}" if nz_str(base_text) else add_text


def get_file_name_from_path(file_path: str) -> str:
    return os.path.basename(file_path)


def get_folder_name_from_path(file_path: str) -> str:
    return os.path.basename(os.path.dirname(file_path))


def get_parent_folder_created(file_path: str) -> str:
    try:
        folder = os.path.dirname(file_path)
        return datetime.fromtimestamp(os.path.getctime(folder)).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""

def parse_folder_truth(folder_name: str) -> Dict[str, str]:
    """
    Parse folder name as source of truth when possible.
    Example: 20260407-INV000000479333-UK-HD427XB-20260406
    -> invoice_date=2026-04-07, invoice_no=INV000000479333, destination=UK
    """
    folder_name = nz_str(folder_name).strip()
    result = {"invoice_no": "", "invoice_date": "", "destination": ""}
    if not folder_name:
        return result

    parts = [p.strip() for p in folder_name.split("-") if nz_str(p).strip()]
    if not parts:
        return result

    m_date = re.match(r"^(20\d{6})$", parts[0])
    if m_date:
        try:
            result["invoice_date"] = datetime.strptime(m_date.group(1), "%Y%m%d").strftime("%Y-%m-%d")
        except Exception:
            pass

    inv_idx = -1
    for i, part in enumerate(parts):
        if re.match(r"^INV[0-9A-Z]+$", part.upper()):
            result["invoice_no"] = part.upper()
            inv_idx = i
            break

    if inv_idx >= 0 and inv_idx + 1 < len(parts):
        dest = parts[inv_idx + 1].upper().strip()
        if re.match(r"^[A-Z]{2,}$", dest):
            result["destination"] = dest

    return result


def compare_with_folder_truth(folder_name: str, invoice_no: Any, invoice_date: Any, destination: Any) -> str:
    truth = parse_folder_truth(folder_name)
    msg = ""

    if truth["invoice_no"]:
        file_inv = normalize_invoice_no(invoice_no)
        folder_inv = normalize_invoice_no(truth["invoice_no"])
        if file_inv and file_inv != folder_inv:
            msg = append_error(msg, f"invoice_no mismatch - FOLDER: {truth['invoice_no']} | FILE: {nz_str(invoice_no)}")
        elif not file_inv:
            msg = append_error(msg, f"invoice_no missing FILE - FOLDER: {truth['invoice_no']}")

    if truth["invoice_date"]:
        file_date = normalize_date_compare(invoice_date)
        folder_date = truth["invoice_date"]
        if file_date and file_date != folder_date:
            msg = append_error(
                msg,
                f"invoice_date mismatch - FOLDER: {datetime.strptime(folder_date, '%Y-%m-%d').strftime('%d/%m/%Y')} | FILE: {format_date_display(invoice_date)}"
            )
        elif not file_date:
            msg = append_error(msg, f"invoice_date missing FILE - FOLDER: {datetime.strptime(folder_date, '%Y-%m-%d').strftime('%d/%m/%Y')}")

    if truth["destination"]:
        file_dest = normalize_text_compare(destination)
        folder_dest = normalize_text_compare(truth["destination"])
        if file_dest and folder_dest not in file_dest and file_dest != folder_dest:
            msg = append_error(msg, f"destination mismatch - FOLDER: {truth['destination']} | FILE: {nz_str(destination)}")
        elif not file_dest:
            msg = append_error(msg, f"destination missing FILE - FOLDER: {truth['destination']}")

    return msg


def detect_file_type(file_path: str) -> Tuple[str, Optional[str]]:
    ext = os.path.splitext(file_path)[1].lower()
    with open(file_path, "rb") as f:
        header = f.read(8)
    if ext == ".xls" and header.startswith(b"\xd0\xcf\x11\xe0"):
        return "xlrd", None
    if ext in (".xlsx", ".xlsm") and header.startswith(b"PK"):
        return "openpyxl", None
    return "", f"File khong hop le: {file_path}"


def cell_to_value(cell: Any) -> Any:
    return cell.value if hasattr(cell, "value") else cell


def trim_matrix(matrix: List[List[Any]]) -> List[List[Any]]:
    max_row = 0
    max_col = 0
    for r_idx, row in enumerate(matrix, start=1):
        for c_idx, value in enumerate(row, start=1):
            if nz_str(value) != "":
                if r_idx > max_row:
                    max_row = r_idx
                if c_idx > max_col:
                    max_col = c_idx
    if max_row == 0:
        return [[""]]
    trimmed = [row[:max_col] + [""] * max(0, max_col - len(row)) for row in matrix[:max_row]]
    return trimmed


def load_workbook_data(file_path: str) -> WorkbookData:
    engine, err = detect_file_type(file_path)
    if not engine:
        raise ValueError(err)

    if engine == "openpyxl":
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheets = {}
            for name in wb.sheetnames:
                ws = wb[name]
                matrix = [[cell_to_value(c) for c in row] for row in ws.iter_rows(values_only=False)]
                sheets[name] = trim_matrix(matrix)
            return WorkbookData(file_path=file_path, engine=engine, sheet_names=list(wb.sheetnames), sheets=sheets)
        finally:
            wb.close()

    book = xlrd.open_workbook(file_path, on_demand=True)
    try:
        sheets = {}
        for name in book.sheet_names():
            sh = book.sheet_by_name(name)
            matrix = []
            for r in range(sh.nrows):
                row = []
                for c in range(sh.ncols):
                    val = sh.cell_value(r, c)
                    if sh.cell_type(r, c) == xlrd.XL_CELL_DATE:
                        try:
                            dt_tuple = xlrd.xldate_as_tuple(val, book.datemode)
                            val = datetime(*dt_tuple)
                        except Exception:
                            pass
                    row.append(val)
                matrix.append(row)
            sheets[name] = trim_matrix(matrix)
        return WorkbookData(file_path=file_path, engine=engine, sheet_names=book.sheet_names(), sheets=sheets)
    finally:
        book.release_resources()


def get_cell(sheet_data: List[List[Any]], r: int, c: int) -> Any:
    if r < 1 or c < 1:
        return ""
    if r > len(sheet_data):
        return ""
    row = sheet_data[r - 1]
    if c > len(row):
        return ""
    return row[c - 1]


def last_used_row(sheet_data: List[List[Any]]) -> int:
    return max(1, len(sheet_data))


def last_used_col_on_row(sheet_data: List[List[Any]], row_num: int) -> int:
    if row_num < 1 or row_num > len(sheet_data):
        return 0
    row = sheet_data[row_num - 1]
    for idx in range(len(row), 0, -1):
        if nz_str(row[idx - 1]) != "":
            return idx
    return 0


def score_pl_sheet(sheet_data: List[List[Any]]) -> int:
    best = 0
    max_rows = min(40, len(sheet_data))
    for r in range(1, max_rows + 1):
        last_col = last_used_col_on_row(sheet_data, r)
        score = 0
        for c in range(1, min(last_col, 60) + 1):
            txt = normalize_header(get_cell(sheet_data, r, c))
            if is_pl_pair_qty_header(txt):
                score += 1
            if is_pl_carton_header(txt):
                score += 1
            if is_pl_cbm_header(txt):
                score += 1
            if is_pl_gross_header(txt):
                score += 1
        best = max(best, score)
    return best


def score_inv_sheet(sheet_data: List[List[Any]]) -> int:
    best = 0
    max_rows = min(40, len(sheet_data))
    for r in range(1, max_rows + 1):
        last_col = last_used_col_on_row(sheet_data, r)
        score = 0
        for c in range(1, min(last_col, 80) + 1):
            txt = normalize_header(get_cell(sheet_data, r, c))
            if is_inv_pair_qty_header(txt):
                score += 1
            if is_inv_amount_without_label_header(txt):
                score += 1
            if is_inv_label_amount_header(txt):
                score += 1
            if is_inv_total_amount_header(txt):
                score += 1
        best = max(best, score)
    return best


def find_best_pl_sheet(wb_data: WorkbookData) -> Optional[str]:
    best_name = None
    best_score = -1
    for name in wb_data.sheet_names:
        nm = name.strip().upper()
        if nm == "PL" or "PL" in nm:
            sc = score_pl_sheet(wb_data.sheets[name])
            if sc > best_score:
                best_score = sc
                best_name = name
    if best_name is None:
        for name in wb_data.sheet_names:
            sc = score_pl_sheet(wb_data.sheets[name])
            if sc > best_score:
                best_score = sc
                best_name = name
    return best_name if best_score >= 2 else None


def find_best_inv_sheet(wb_data: WorkbookData) -> Optional[str]:
    best_name = None
    best_score = -1
    for name in wb_data.sheet_names:
        nm = name.strip().upper()
        if nm in ("INV", "INVOICE") or "INV" in nm or "INVOICE" in nm:
            sc = score_inv_sheet(wb_data.sheets[name])
            if sc > best_score:
                best_score = sc
                best_name = name
    if best_name is None:
        for name in wb_data.sheet_names:
            sc = score_inv_sheet(wb_data.sheets[name])
            if sc > best_score:
                best_score = sc
                best_name = name
    return best_name if best_score >= 2 else None


def detect_pl_header(sheet_data: List[List[Any]]) -> Tuple[Dict[str, int], int]:
    best_map: Dict[str, int] = {}
    best_score = -1
    hdr_row = 0
    for r in range(1, min(40, len(sheet_data)) + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        for c in range(1, min(last_col, 80) + 1):
            txt = normalize_header(get_cell(sheet_data, r, c))
            if is_pl_pair_qty_header(txt) and "pair_qty" not in tmp:
                tmp["pair_qty"] = c
            if is_pl_carton_header(txt) and "carton_qty" not in tmp:
                tmp["carton_qty"] = c
            if is_pl_cbm_header(txt) and "cbm" not in tmp:
                tmp["cbm"] = c
            if is_pl_gross_header(txt) and "gross_weight" not in tmp:
                tmp["gross_weight"] = c
        score = len(tmp)
        if score > best_score:
            best_score = score
            hdr_row = r
            best_map = tmp
    if best_score < 2:
        raise ValueError("Khong tim thay header PL hop le.")
    return best_map, hdr_row


def detect_inv_header(sheet_data: List[List[Any]]) -> Tuple[Dict[str, int], int]:
    best_map: Dict[str, int] = {}
    best_score = -1
    hdr_row = 0
    for r in range(1, min(40, len(sheet_data)) + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        for c in range(1, min(last_col, 120) + 1):
            txt = normalize_header(get_cell(sheet_data, r, c))
            if is_inv_pair_qty_header(txt) and "inv_qty_pairs" not in tmp:
                tmp["inv_qty_pairs"] = c
            if is_inv_amount_without_label_header(txt) and "inv_amount_wo_label" not in tmp:
                tmp["inv_amount_wo_label"] = c
            if is_inv_label_amount_header(txt):
                if "inv_label_amount" not in tmp:
                    tmp["inv_label_amount"] = c
                elif "TOTALLABELCOST" in txt or "LABELAMOUNT" in txt:
                    tmp["inv_label_amount"] = c
            if is_inv_total_amount_header(txt) and "inv_total_amount" not in tmp:
                tmp["inv_total_amount"] = c
        score = len(tmp)
        if score > best_score:
            best_score = score
            hdr_row = r
            best_map = tmp
    if best_score < 2:
        raise ValueError("Khong tim thay header INV hop le.")
    return best_map, hdr_row


def find_total_row(sheet_data: List[List[Any]], start_row: int) -> int:
    last_row = last_used_row(sheet_data)
    for r in range(start_row, last_row + 1):
        for c in range(1, 6):
            txt = normalize_simple_text(get_cell(sheet_data, r, c)).replace(":", "")
            if "TOTAL" in txt.upper():
                return r
    return 0


def get_mapped_cell_value(sheet_data: List[List[Any]], row_num: int, col_map: Dict[str, int], key: str) -> Any:
    if key in col_map:
        return get_cell(sheet_data, row_num, int(col_map[key]))
    return ""


def extract_pl_total(sheet_data: List[List[Any]]) -> Dict[str, Any]:
    col_map, hdr_row = detect_pl_header(sheet_data)
    total_row = find_total_row(sheet_data, hdr_row + 1)
    if total_row == 0:
        raise ValueError("Khong tim thay dong TOTAL trong PL.")
    return {
        "pair_qty": get_mapped_cell_value(sheet_data, total_row, col_map, "pair_qty"),
        "carton_qty": get_mapped_cell_value(sheet_data, total_row, col_map, "carton_qty"),
        "cbm": get_mapped_cell_value(sheet_data, total_row, col_map, "cbm"),
        "gross_weight": get_mapped_cell_value(sheet_data, total_row, col_map, "gross_weight"),
    }


def extract_inv_total(sheet_data: List[List[Any]]) -> Dict[str, Any]:
    col_map, hdr_row = detect_inv_header(sheet_data)
    total_row = find_total_row(sheet_data, hdr_row + 1)
    if total_row == 0:
        raise ValueError("Khong tim thay dong TOTAL trong INV.")
    return {
        "inv_qty_pairs": get_mapped_cell_value(sheet_data, total_row, col_map, "inv_qty_pairs"),
        "inv_amount_wo_label": get_mapped_cell_value(sheet_data, total_row, col_map, "inv_amount_wo_label"),
        "inv_label_amount": get_mapped_cell_value(sheet_data, total_row, col_map, "inv_label_amount"),
        "inv_total_amount": get_mapped_cell_value(sheet_data, total_row, col_map, "inv_total_amount"),
    }


def is_usable_meta_value(s: Any, key_name: str) -> bool:
    t = nz_str(s).strip()
    if not t:
        return False
    if normalize_meta_label(t) in {
        "INVOICE NO", "INVOICE NUMBER", "DATE", "INVOICE DATE", "DEPARTURE", "DESTINATION",
        "TRADE TERM", "TRADE TERMS", "TERMS OF DELIVERY", "SHIPPING DATE", "SHIPPINGDATE",
        "PAYMENT TERM", "PAYMENT TERMS", "CUS PO", "CUS PO NO", "CUS PONO"
    }:
        return False
    key_name = key_name.lower()
    if key_name == "invoice_no":
        return normalize_invoice_no(t) != "" and "INVOICE" not in t.upper()
    if key_name in ("invoice_date", "shipping_date"):
        return normalize_date_compare(t) != "" and ("DATE" not in t.upper() or bool(try_parse_loose_date(t)))
    return True


def get_meta_value_flexible(sheet_data: List[List[Any]], r: int, c: int, key_name: str) -> str:
    candidates = [
        get_cell(sheet_data, r, c + 1), get_cell(sheet_data, r, c + 2), get_cell(sheet_data, r, c + 3),
        get_cell(sheet_data, r + 1, c), get_cell(sheet_data, r + 2, c), get_cell(sheet_data, r + 3, c),
        get_cell(sheet_data, r + 1, c + 1), get_cell(sheet_data, r + 1, c + 2),
        get_cell(sheet_data, r + 2, c + 1), get_cell(sheet_data, r + 2, c + 2),
        get_cell(sheet_data, r + 3, c + 1), get_cell(sheet_data, r + 3, c + 2),
    ]
    for candidate in candidates:
        if is_usable_meta_value(candidate, key_name):
            return nz_str(candidate)
    return ""


def extract_sheet_meta(sheet_data: List[List[Any]]) -> Dict[str, str]:
    meta = {k: "" for k in META_KEYS}
    max_row = min(60, last_used_row(sheet_data))
    for r in range(1, max_row + 1):
        max_col = min(40, last_used_col_on_row(sheet_data, r))
        if max_col < 1:
            max_col = 1
        for c in range(1, max_col + 1):
            raw = nz_str(get_cell(sheet_data, r, c))
            if not raw:
                continue
            norm = normalize_meta_label(raw)
            key_name = get_meta_key_by_alias(norm)
            if key_name:
                val_text = get_meta_value_flexible(sheet_data, r, c, key_name)
                if not meta[key_name] and val_text:
                    meta[key_name] = val_text
    return meta


def compare_pl_inv_meta(pl_meta: Dict[str, str], inv_meta: Dict[str, str]) -> str:
    msg = ""
    for key in COMPARE_KEYS:
        pl_raw = nz_str(pl_meta.get(key, ""))
        inv_raw = nz_str(inv_meta.get(key, ""))
        pl_norm = normalize_compare_by_field(pl_raw, key)
        inv_norm = normalize_compare_by_field(inv_raw, key)
        if pl_norm and inv_norm:
            if pl_norm != inv_norm:
                msg = append_error(msg, f"{key} mismatch - PL: {pl_raw} | INV: {inv_raw}")
        elif pl_norm and not inv_norm:
            msg = append_error(msg, f"{key} missing INV - PL: {pl_raw}")
        elif not pl_norm and inv_norm:
            msg = append_error(msg, f"{key} missing PL - INV: {inv_raw}")
    return msg


def merge_meta_value(pl_meta: Dict[str, str], inv_meta: Dict[str, str], key_name: str) -> str:
    return nz_str(pl_meta.get(key_name, "")) or nz_str(inv_meta.get(key_name, ""))


def build_file_signature(arr: List[Any]) -> str:
    inv_no = normalize_compare_by_field(arr[13], "invoice_no")
    inv_date = normalize_compare_by_field(arr[14], "invoice_date")
    dest = normalize_compare_by_field(arr[16], "destination")
    trade_term = normalize_compare_by_field(arr[17], "trade_term")
    pair_qty = normalize_compare_by_field(arr[5], "pair_qty")
    inv_total = normalize_compare_by_field(arr[12], "inv_total_amount")
    return "|".join([inv_no, inv_date, dest, trade_term, pair_qty, inv_total])



def detect_detail_header_row(sheet_data: List[List[Any]], kind: str) -> Tuple[Dict[str, int], int]:
    best_map: Dict[str, int] = {}
    best_score = -1
    hdr_row = 0
    max_rows = min(40, len(sheet_data))
    for r in range(1, max_rows + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        for c in range(1, min(last_col, 120) + 1):
            raw = get_cell(sheet_data, r, c)
            key = match_pl_detail_key(raw) if kind == "PL" else match_inv_detail_key(raw)
            if key and key not in tmp:
                tmp[key] = c
        score = len(tmp)
        if score > best_score:
            best_score = score
            best_map = tmp
            hdr_row = r
    min_score = 6 if kind == "PL" else 8
    if best_score < min_score:
        raise ValueError(f"Khong tim thay header chi tiet {kind} hop le.")
    return best_map, hdr_row


def classify_row_type(sheet_data: List[List[Any]], row_num: int, po_col: int) -> str:
    po_text = normalize_simple_text(get_cell(sheet_data, row_num, po_col)).upper().replace(":", "")
    if "TOTAL" in po_text:
        return "TOTAL"
    if "HANDLING CHARGE" in po_text:
        return "HANDLING CHARGE"
    return "DETAIL"


def row_has_useful_data(sheet_data: List[List[Any]], row_num: int, cols: List[int]) -> bool:
    for c in cols:
        if c > 0 and nz_str(get_cell(sheet_data, row_num, c)) != "":
            return True
    return False


def extract_pl_detail_rows(file_path: str, folder_name: str, folder_created: str, sheet_name: str, sheet_data: List[List[Any]], status_text: str) -> List[List[Any]]:
    truth = parse_folder_truth(folder_name)
    col_map, hdr_row = detect_detail_header_row(sheet_data, "PL")
    po_col = col_map.get("po_no", 0)
    keep_cols = [col_map.get(k, 0) for k in ["date","invoice_no","destination","po_no","customer_po","model_name_color","color","width","hs_code","pair_qty","carton_qty","cbm","gross_weight"]]
    rows = []
    carry = {"date": "", "invoice_no": "", "destination": ""}
    for r in range(hdr_row + 1, last_used_row(sheet_data) + 1):
        if not row_has_useful_data(sheet_data, r, keep_cols):
            continue
        row_type = classify_row_type(sheet_data, r, po_col) if po_col else "DETAIL"
        date_v = get_mapped_cell_value(sheet_data, r, col_map, "date")
        inv_v = get_mapped_cell_value(sheet_data, r, col_map, "invoice_no")
        dest_v = get_mapped_cell_value(sheet_data, r, col_map, "destination")
        if nz_str(date_v): carry["date"] = format_date_for_display(date_v)
        if nz_str(inv_v): carry["invoice_no"] = nz_str(inv_v)
        if nz_str(dest_v): carry["destination"] = nz_str(dest_v)

        po_v = get_mapped_cell_value(sheet_data, r, col_map, "po_no")
        if row_type == "DETAIL" and not (nz_str(po_v) or nz_str(get_mapped_cell_value(sheet_data, r, col_map, "pair_qty"))):
            continue

        rows.append([
            folder_created, folder_name, get_file_name_from_path(file_path), sheet_name,
            datetime.strptime(truth["invoice_date"], "%Y-%m-%d").strftime("%d/%m/%Y") if truth.get("invoice_date") else "",
            truth.get("invoice_no", ""), truth.get("destination", ""),
            carry["date"] or format_date_for_display(date_v), carry["invoice_no"] or nz_str(inv_v), carry["destination"] or nz_str(dest_v),
            po_v, get_mapped_cell_value(sheet_data, r, col_map, "customer_po"),
            get_mapped_cell_value(sheet_data, r, col_map, "model_name_color"),
            get_mapped_cell_value(sheet_data, r, col_map, "color"),
            get_mapped_cell_value(sheet_data, r, col_map, "width"),
            get_mapped_cell_value(sheet_data, r, col_map, "hs_code"),
            get_mapped_cell_value(sheet_data, r, col_map, "pair_qty"),
            get_mapped_cell_value(sheet_data, r, col_map, "carton_qty"),
            get_mapped_cell_value(sheet_data, r, col_map, "cbm"),
            get_mapped_cell_value(sheet_data, r, col_map, "gross_weight"),
            row_type, status_text, file_path
        ])
    return rows


def extract_inv_detail_rows(file_path: str, folder_name: str, folder_created: str, sheet_name: str, sheet_data: List[List[Any]], status_text: str) -> List[List[Any]]:
    truth = parse_folder_truth(folder_name)
    col_map, hdr_row = detect_detail_header_row(sheet_data, "INV")
    po_col = col_map.get("po_no", 0)
    keep_cols = [col_map.get(k, 0) for k in ["date","invoice_no","destination","po_no","customer_po","desc","hs_code","model_name_color","model_y","color","width","qty_pairs","unit_price_usd","amount_wo_label","label_cost","label_amount","total_amount"]]
    rows = []
    carry = {"date": "", "invoice_no": "", "destination": ""}
    for r in range(hdr_row + 1, last_used_row(sheet_data) + 1):
        if not row_has_useful_data(sheet_data, r, keep_cols):
            continue
        row_type = classify_row_type(sheet_data, r, po_col) if po_col else "DETAIL"
        date_v = get_mapped_cell_value(sheet_data, r, col_map, "date")
        inv_v = get_mapped_cell_value(sheet_data, r, col_map, "invoice_no")
        dest_v = get_mapped_cell_value(sheet_data, r, col_map, "destination")
        if nz_str(date_v): carry["date"] = format_date_for_display(date_v)
        if nz_str(inv_v): carry["invoice_no"] = nz_str(inv_v)
        if nz_str(dest_v): carry["destination"] = nz_str(dest_v)

        po_v = get_mapped_cell_value(sheet_data, r, col_map, "po_no")
        if row_type == "DETAIL" and not (nz_str(po_v) or nz_str(get_mapped_cell_value(sheet_data, r, col_map, "qty_pairs")) or nz_str(get_mapped_cell_value(sheet_data, r, col_map, "label_amount"))):
            continue

        rows.append([
            folder_created, folder_name, get_file_name_from_path(file_path), sheet_name,
            datetime.strptime(truth["invoice_date"], "%Y-%m-%d").strftime("%d/%m/%Y") if truth.get("invoice_date") else "",
            truth.get("invoice_no", ""), truth.get("destination", ""),
            carry["date"] or format_date_for_display(date_v), carry["invoice_no"] or nz_str(inv_v), carry["destination"] or nz_str(dest_v),
            po_v, get_mapped_cell_value(sheet_data, r, col_map, "customer_po"),
            get_mapped_cell_value(sheet_data, r, col_map, "desc"),
            get_mapped_cell_value(sheet_data, r, col_map, "hs_code"),
            get_mapped_cell_value(sheet_data, r, col_map, "model_name_color"),
            get_mapped_cell_value(sheet_data, r, col_map, "model_y"),
            get_mapped_cell_value(sheet_data, r, col_map, "color"),
            get_mapped_cell_value(sheet_data, r, col_map, "width"),
            get_mapped_cell_value(sheet_data, r, col_map, "qty_pairs"),
            get_mapped_cell_value(sheet_data, r, col_map, "unit_price_usd"),
            get_mapped_cell_value(sheet_data, r, col_map, "amount_wo_label"),
            get_mapped_cell_value(sheet_data, r, col_map, "label_cost"),
            get_mapped_cell_value(sheet_data, r, col_map, "label_amount"),
            get_mapped_cell_value(sheet_data, r, col_map, "total_amount"),
            row_type, status_text, file_path
        ])
    return rows


def process_one_sub_file_bundle(file_path: str) -> Tuple[List[Any], List[List[Any]], List[List[Any]]]:
    arr = process_one_sub_file(file_path)
    inv_rows: List[List[Any]] = []
    pl_rows: List[List[Any]] = []
    try:
        wb_data = load_workbook_data(file_path)
        pl_name = nz_str(arr[3]) or find_best_pl_sheet(wb_data) or ""
        inv_name = nz_str(arr[4]) or find_best_inv_sheet(wb_data) or ""
        if pl_name and pl_name in wb_data.sheets:
            pl_rows = extract_pl_detail_rows(file_path, get_folder_name_from_path(file_path), get_parent_folder_created(file_path), pl_name, wb_data.sheets[pl_name], nz_str(arr[21]).upper())
        if inv_name and inv_name in wb_data.sheets:
            inv_rows = extract_inv_detail_rows(file_path, get_folder_name_from_path(file_path), get_parent_folder_created(file_path), inv_name, wb_data.sheets[inv_name], nz_str(arr[21]).upper())
    except Exception:
        pass
    return arr, inv_rows, pl_rows


def process_one_sub_file(file_path: str) -> List[Any]:
    arr = [""] * 25
    arr[0] = get_parent_folder_created(file_path)
    arr[1] = get_folder_name_from_path(file_path)
    arr[2] = get_file_name_from_path(file_path)
    arr[23] = file_path
    try:
        wb_data = load_workbook_data(file_path)
        pl_name = find_best_pl_sheet(wb_data)
        inv_name = find_best_inv_sheet(wb_data)
        if not pl_name:
            raise ValueError("Khong tim thay sheet PL hop le.")
        if not inv_name:
            raise ValueError("Khong tim thay sheet INV/INVOICE hop le.")

        pl_data = extract_pl_total(wb_data.sheets[pl_name])
        inv_data = extract_inv_total(wb_data.sheets[inv_name])
        pl_meta = extract_sheet_meta(wb_data.sheets[pl_name])
        inv_meta = extract_sheet_meta(wb_data.sheets[inv_name])
        compare_msg = compare_pl_inv_meta(pl_meta, inv_meta)

        arr[3] = pl_name
        arr[4] = inv_name
        arr[5] = pl_data["pair_qty"]
        arr[6] = pl_data["carton_qty"]
        arr[7] = pl_data["cbm"]
        arr[8] = pl_data["gross_weight"]
        arr[9] = inv_data["inv_qty_pairs"]
        arr[10] = inv_data["inv_amount_wo_label"]
        arr[11] = inv_data["inv_label_amount"]
        arr[12] = inv_data["inv_total_amount"]
        arr[13] = merge_meta_value(pl_meta, inv_meta, "invoice_no")
        arr[14] = format_date_for_display(merge_meta_value(pl_meta, inv_meta, "invoice_date"))
        arr[15] = merge_meta_value(pl_meta, inv_meta, "departure")
        arr[16] = merge_meta_value(pl_meta, inv_meta, "destination")
        arr[17] = merge_meta_value(pl_meta, inv_meta, "trade_term")
        arr[18] = format_date_for_display(merge_meta_value(pl_meta, inv_meta, "shipping_date"))
        arr[19] = merge_meta_value(pl_meta, inv_meta, "payment_term")
        arr[20] = merge_meta_value(pl_meta, inv_meta, "cus_po")

        folder_msg = compare_with_folder_truth(arr[1], arr[13], arr[14], arr[16])
        final_msg = compare_msg
        if folder_msg:
            final_msg = append_error(final_msg, folder_msg)

        arr[21] = "WARNING" if final_msg else "OK"
        arr[22] = final_msg
        arr[23] = file_path
        arr[24] = build_file_signature(arr)
        return arr
    except Exception as exc:
        arr[21] = "ERROR"
        arr[22] = str(exc)
        arr[24] = ""
        return arr


def ensure_sheet(wb: Workbook, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(title=sheet_name)


def init_sub_detail_header(ws):
    if nz_str(ws.cell(1, 1).value) == "":
        for idx, header in enumerate(SUB_DETAIL_HEADERS, start=1):
            ws.cell(1, idx).value = sanitize_excel_text(header)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL_SUB


def init_log_header(ws):
    if nz_str(ws.cell(1, 1).value) == "":
        for idx, header in enumerate(LOG_HEADERS, start=1):
            ws.cell(1, idx).value = sanitize_excel_text(header)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL_LOG



def init_inv_detail_header(ws):
    if nz_str(ws.cell(1, 1).value) == "":
        for idx, header in enumerate(INV_DETAIL_HEADERS, start=1):
            ws.cell(1, idx).value = sanitize_excel_text(header)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"


def init_pl_detail_header(ws):
    if nz_str(ws.cell(1, 1).value) == "":
        for idx, header in enumerate(PL_DETAIL_HEADERS, start=1):
            ws.cell(1, idx).value = sanitize_excel_text(header)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"


def find_header_col(ws, header_name: str) -> int:
    last_col = ws.max_column
    for c in range(1, last_col + 1):
        if nz_str(ws.cell(1, c).value) == header_name:
            return c
    return 0


def load_processed_keys(ws) -> Tuple[set, set]:
    dict_path = set()
    dict_sig = set()
    path_col = find_header_col(ws, "File Path")
    sig_col = find_header_col(ws, "File Signature")
    if path_col == 0 and sig_col == 0:
        return dict_path, dict_sig
    for r in range(2, ws.max_row + 1):
        if path_col > 0:
            p = nz_str(ws.cell(r, path_col).value).upper()
            if p:
                dict_path.add(p)
        if sig_col > 0:
            s = nz_str(ws.cell(r, sig_col).value).upper()
            if s:
                dict_sig.add(s)
    return dict_path, dict_sig


def append_result_rows(ws, rows: List[List[Any]]):
    if not rows:
        return
    start_row = ws.max_row + 1 if nz_str(ws.cell(ws.max_row, 1).value) else max(2, ws.max_row)
    if start_row < 2:
        start_row = 2
    for idx, arr in enumerate(rows, start=start_row):
        for c, value in enumerate(arr, start=1):
            cell = ws.cell(idx, c)
            cell.value = sanitize_excel_text(value)
            if c in (15, 19):
                cell.number_format = "dd/mm/yyyy"
        status = nz_str(arr[21]).upper()
        if status in STATUS_FILLS:
            fill = STATUS_FILLS[status]
            for c in range(1, 26):
                ws.cell(idx, c).fill = fill


def append_log_rows(ws, rows: List[List[Any]]):
    if not rows:
        return
    start_row = ws.max_row + 1 if nz_str(ws.cell(ws.max_row, 1).value) else max(2, ws.max_row)
    if start_row < 2:
        start_row = 2
    for idx, row in enumerate(rows, start=start_row):
        for c, value in enumerate(row, start=1):
            ws.cell(idx, c).value = sanitize_excel_text(value)


def get_existing_subdetail_rows(ws) -> List[List[Any]]:
    rows = []
    max_col = len(SUB_DETAIL_HEADERS)
    for r in range(2, ws.max_row + 1):
        arr = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if not any(nz_str(v) for v in arr):
            continue
        rows.append(arr)
    return rows


def keep_only_ok_rows(ws) -> Tuple[List[List[Any]], int]:
    status_col = find_header_col(ws, "Status")
    if status_col == 0:
        return [], 0
    kept = []
    rerun_count = 0
    for arr in get_existing_subdetail_rows(ws):
        status = nz_str(arr[status_col - 1]).upper()
        if status == "OK":
            kept.append(arr)
        elif status in ("WARNING", "ERROR"):
            rerun_count += 1
    ws.delete_rows(2, max(0, ws.max_row - 1))
    append_result_rows(ws, kept)
    return kept, rerun_count


def build_processed_sets_from_rows(rows: List[List[Any]]) -> Tuple[set, set]:
    dict_path = set()
    dict_sig = set()
    for arr in rows:
        if len(arr) >= 24:
            p = nz_str(arr[23]).upper()
            if p:
                dict_path.add(p)
        if len(arr) >= 25:
            s = nz_str(arr[24]).upper()
            if s:
                dict_sig.add(s)
    return dict_path, dict_sig


def safe_save_workbook_atomic(wb, output_path: str):
    output_path = os.path.abspath(output_path)
    folder = os.path.dirname(output_path) or "."
    os.makedirs(folder, exist_ok=True)
    fd, temp_path = tempfile.mkstemp(prefix="tmp_sub_detail_", suffix=".xlsx", dir=folder)
    os.close(fd)
    backup_path = output_path + ".bak"
    try:
        wb.save(temp_path)
        if os.path.exists(output_path):
            try:
                shutil.copy2(output_path, backup_path)
            except Exception:
                pass
        try:
            os.replace(temp_path, output_path)
        except PermissionError as exc:
            alt_path = os.path.join(
                folder,
                f"{os.path.splitext(os.path.basename(output_path))[0]}__CLOSE_FILE_AND_RETRY__{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            try:
                shutil.copy2(temp_path, alt_path)
            except Exception:
                alt_path = ""
            detail = (
                f"Khong the ghi de file output vi file dang mo trong Excel hoac dang bi khoa:\n"
                f"{output_path}\n\n"
                f"Hay dong file output roi chay lai."
            )
            if alt_path:
                detail += (
                    f"\n\nDa luu tam sang file khac de tranh mat du lieu:\n"
                    f"{alt_path}"
                )
            raise PermissionError(detail) from exc
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except Exception:
                pass

def autofit_useful(ws, max_col: int):
    for col_idx in range(1, max_col + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            val = nz_str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)
    ws.row_dimensions[1].height = 20



def get_repair_truth(folder_name: str, options: RepairOptions) -> Dict[str, str]:
    truth = {"invoice_no": "", "invoice_date": "", "destination": ""}
    if options.use_folder_truth:
        truth.update(parse_folder_truth(folder_name))
    if options.use_manual_values:
        if nz_str(options.manual_invoice_no):
            truth["invoice_no"] = nz_str(options.manual_invoice_no).strip().upper()
        if nz_str(options.manual_invoice_date):
            manual_date = normalize_date_compare(options.manual_invoice_date)
            truth["invoice_date"] = manual_date or truth["invoice_date"]
        if nz_str(options.manual_destination):
            truth["destination"] = nz_str(options.manual_destination).strip().upper()
    return truth


def find_meta_cells_openpyxl(ws) -> Dict[str, tuple]:
    found = {}
    max_row = min(ws.max_row, 60)
    max_col = min(ws.max_column, 40)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            raw = nz_str(ws.cell(r, c).value)
            if not raw:
                continue
            key_name = get_meta_key_by_alias(normalize_meta_label(raw))
            if key_name and key_name not in found:
                for offset in (1, 2):
                    if c + offset <= ws.max_column + 2:
                        found[key_name] = (r, c + offset)
                        break
    return found


def set_meta_value_to_ws(ws, key_name: str, value: Any):
    cells = find_meta_cells_openpyxl(ws)
    if key_name not in cells:
        return False, f"Khong tim thay nhan {key_name}"
    r, c = cells[key_name]
    if key_name.endswith("date"):
        value = format_date_for_display(value)
    ws.cell(r, c).value = value
    return True, ""


def repair_excel_metadata(file_path: str, arr: List[Any], options: RepairOptions) -> Tuple[bool, str]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return False, "Tu dong sua chi ho tro .xlsx/.xlsm; .xls chi danh dau de xu ly tay"

    truth = get_repair_truth(arr[1], options)
    if not any(truth.values()):
        return False, "Khong co du lieu chuan de sua"

    wb = openpyxl.load_workbook(file_path)
    try:
        target_sheets = []
        pl_name = nz_str(arr[3])
        inv_name = nz_str(arr[4])
        if pl_name in wb.sheetnames:
            target_sheets.append(wb[pl_name])
        if inv_name in wb.sheetnames and inv_name != pl_name:
            target_sheets.append(wb[inv_name])

        if not target_sheets:
            return False, "Khong mo duoc sheet PL/INV de sua"

        changed = []
        errs = []
        for key_name in ("invoice_no", "invoice_date", "destination"):
            desired = truth.get(key_name, "")
            if not desired:
                continue
            for ws in target_sheets:
                ok, err = set_meta_value_to_ws(ws, key_name, desired)
                if ok:
                    changed.append(f"{ws.title}:{key_name}={desired}")
                else:
                    errs.append(f"{ws.title}:{err}")

        if not changed:
            return False, "; ".join(errs) if errs else "Khong co o nao duoc sua"

        wb.save(file_path)
        return True, "; ".join(changed)
    finally:
        wb.close()


def collect_source_files(root_folder: str, output_path: str) -> List[str]:
    output_upper = os.path.abspath(output_path).strip().upper() if output_path else ""
    result = []
    for current_root, dirs, files in os.walk(root_folder):
        dirs.sort()
        for name in sorted(files):
            nm = name.lower()
            ext = os.path.splitext(name)[1].lower()
            full_path = os.path.abspath(os.path.join(current_root, name)).strip().upper()
            if full_path == output_upper:
                continue
            if ext not in (".xls", ".xlsx", ".xlsm"):
                continue
            if nm.startswith("~$"):
                continue
            if nm.startswith("sub") or nm.startswith("vsf") or nm.startswith("sa") or nm.startswith("copy of sa"):
                result.append(os.path.join(current_root, name))
    return result


def make_log_row(file_path: str, step_name: str, result_text: str, detail_text: str) -> List[str]:
    return [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        get_file_name_from_path(file_path) if file_path else "",
        step_name,
        result_text,
        detail_text,
        file_path,
    ]


def open_or_create_output_workbook(full_path: str):
    if not nz_str(full_path):
        raise ValueError("OpenOrCreateOutputWorkbook: duong dan rong")
    if not full_path.lower().endswith(".xlsx"):
        raise ValueError("Output workbook phai la file .xlsx")
    if os.path.exists(full_path):
        try:
            return openpyxl.load_workbook(full_path)
        except Exception:
            bad_path = full_path + ".corrupt_" + datetime.now().strftime("%Y%m%d_%H%M%S")
            shutil.copy2(full_path, bad_path)
    wb = Workbook()
    safe_save_workbook_atomic(wb, full_path)
    return openpyxl.load_workbook(full_path)


class Processor:
    def __init__(self, logger: AppLogger):
        self.logger = logger

    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None) -> Dict[str, int]:
        repair_options = repair_options or RepairOptions()
        wb_out = open_or_create_output_workbook(output_path)
        try:
            ws_out = ensure_sheet(wb_out, "SUB_DETAIL")
            ws_log = ensure_sheet(wb_out, "LOG_SUB_DETAIL")
            ws_inv = ensure_sheet(wb_out, "INV")
            ws_pl = ensure_sheet(wb_out, "PL")
            if "Sheet" in wb_out.sheetnames and wb_out["Sheet"].max_row == 1 and wb_out["Sheet"].max_column == 1 and nz_str(wb_out["Sheet"]["A1"].value) == "":
                try:
                    del wb_out["Sheet"]
                except Exception:
                    pass

            init_sub_detail_header(ws_out)
            init_log_header(ws_log)
            init_inv_detail_header(ws_inv)
            init_pl_detail_header(ws_pl)

            kept_ok_rows, rerun_count = keep_only_ok_rows(ws_out)
            processed_paths, processed_signs = build_processed_sets_from_rows(kept_ok_rows)
            ok_paths = set(processed_paths)
            keep_rows_by_file_paths(ws_inv, len(INV_DETAIL_HEADERS), 27, ok_paths)
            keep_rows_by_file_paths(ws_pl, len(PL_DETAIL_HEADERS), 23, ok_paths)
            src_files = sorted([p for p in (ok_paths if ok_paths is not None else collect_source_files(folder_path, output_path)) if os.path.exists(p)])
            cnt_total = len(src_files)
            cnt_ok = cnt_warn = cnt_err = cnt_skip_path = cnt_skip_sig = 0
            cnt_repaired = 0

            log_rows: List[List[Any]] = []
            result_rows: List[List[Any]] = []
            inv_detail_rows: List[List[Any]] = []
            pl_detail_rows: List[List[Any]] = []
            issue_files: List[Tuple[str, str, str]] = []

            if cnt_total == 0:
                log_rows.append(make_log_row("", "SCAN", "INFO", "Khong tim thay file Excel nao de xu ly."))
                append_log_rows(ws_log, log_rows)
                safe_save_workbook_atomic(wb_out, output_path)
                return {
                    "total": 0, "ok": 0, "warning": 0, "error": 0,
                    "skip_path": 0, "skip_signature": 0, "repaired": 0, "issues": [], "scanned_files": src_files,
                }

            log_rows.append(make_log_row("", "START", "INFO", f"Bat dau quet {cnt_total} file."))
            if rerun_count > 0:
                log_rows.append(make_log_row("", "RERUN", "INFO", f"Se chay lai {rerun_count} file WARNING/ERROR tu lan truoc, dong thoi xu ly file moi phat sinh."))

            def job(path: str):
                arr, inv_rows, pl_rows = process_one_sub_file_bundle(path)
                return path, arr, inv_rows, pl_rows

            futures = {}
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                for fp in src_files:
                    fp_norm = fp.strip().upper()
                    if fp_norm == output_path.strip().upper():
                        log_rows.append(make_log_row(fp, "SKIP", "INFO", "Bo qua output workbook."))
                        continue
                    if fp_norm in processed_paths:
                        cnt_skip_path += 1
                        continue
                    futures[executor.submit(job, fp)] = fp

                completed = 0
                total_futures = max(1, len(futures))
                for future in as_completed(futures):
                    file_path, arr, inv_rows_one, pl_rows_one = future.result()
                    completed += 1
                    sig = nz_str(arr[24])
                    status_text = nz_str(arr[21]).upper()
                    err_text = nz_str(arr[22])

                    if sig and sig.upper() in processed_signs:
                        cnt_skip_sig += 1
                        if progress_callback:
                            progress_callback(completed, total_futures, file_path)
                        continue

                    if status_text in ("WARNING", "ERROR"):
                        issue_files.append((file_path, status_text, err_text))
                        self.logger.log(f"{status_text}: {get_file_name_from_path(file_path)} -> {err_text}", "warn" if status_text == "WARNING" else "error")
                        log_rows.append(make_log_row(file_path, "CHECK", status_text, err_text))

                        if repair_options.use_folder_truth or repair_options.use_manual_values:
                            repaired, detail = repair_excel_metadata(file_path, arr, repair_options)
                            if repaired:
                                cnt_repaired += 1
                                self.logger.log(f"REPAIRED: {get_file_name_from_path(file_path)} -> {detail}", "info")
                                log_rows.append(make_log_row(file_path, "REPAIR", "OK", detail))
                                arr = process_one_sub_file(file_path)
                                status_text = nz_str(arr[21]).upper()
                                err_text = nz_str(arr[22])
                            else:
                                self.logger.log(f"REPAIR-SKIP: {get_file_name_from_path(file_path)} -> {detail}", "warn")
                                log_rows.append(make_log_row(file_path, "REPAIR", "SKIP", detail))

                    if status_text == "OK":
                        cnt_ok += 1
                    elif status_text == "WARNING":
                        cnt_warn += 1
                    else:
                        cnt_err += 1

                    result_rows.append(arr)
                    inv_detail_rows.extend(inv_rows_one)
                    pl_detail_rows.extend(pl_rows_one)
                    if sig:
                        processed_signs.add(sig.upper())
                    processed_paths.add(file_path.strip().upper())

                    if progress_callback:
                        progress_callback(completed, total_futures, file_path)

            if issue_files:
                log_rows.append(make_log_row("", "SUMMARY", "INFO", "TONG HOP FILE WARNING/ERROR"))
                for fp, st, detail in issue_files:
                    log_rows.append(make_log_row(fp, "SUMMARY", st, detail))

            if cnt_skip_path > 0:
                log_rows.append(make_log_row("", "SKIP", "INFO", f"So luong bo qua do da co File Path: {cnt_skip_path}"))
            if cnt_skip_sig > 0:
                log_rows.append(make_log_row("", "SKIP", "INFO", f"So luong bo qua do trung File Signature: {cnt_skip_sig}"))

            log_rows.append(make_log_row("", "DONE", "INFO",
                                         f"Tong={cnt_total}; OK={cnt_ok}; WARNING={cnt_warn}; ERROR={cnt_err}; SkipPath={cnt_skip_path}; SkipSignature={cnt_skip_sig}; Repaired={cnt_repaired}"))

            append_result_rows(ws_out, result_rows)
            append_generic_rows(ws_inv, inv_detail_rows, [5, 8], 26)
            append_generic_rows(ws_pl, pl_detail_rows, [5, 8], 22)
            append_log_rows(ws_log, log_rows)
            autofit_useful(ws_out, len(SUB_DETAIL_HEADERS))
            autofit_useful(ws_inv, len(INV_DETAIL_HEADERS))
            autofit_useful(ws_pl, len(PL_DETAIL_HEADERS))
            autofit_useful(ws_log, len(LOG_HEADERS))
            safe_save_workbook_atomic(wb_out, output_path)

            return {
                "total": cnt_total, "ok": cnt_ok, "warning": cnt_warn, "error": cnt_err,
                "skip_path": cnt_skip_path, "skip_signature": cnt_skip_sig, "repaired": cnt_repaired,
                "issues": issue_files, "scanned_files": src_files,
            }
        finally:
            wb_out.close()



def open_file_with_default_app(path: str):
    path = nz_str(path)
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f"Khong tim thay file: {path}")
    try:
        os.startfile(path)  # type: ignore[attr-defined]
    except AttributeError:
        if sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])


def open_folder_and_select_file(path: str):
    path = os.path.abspath(nz_str(path))
    if not path or not os.path.exists(path):
        raise FileNotFoundError(f"Khong tim thay file: {path}")
    folder = os.path.dirname(path)
    try:
        os.startfile(folder)  # fallback simple open folder on Windows
    except AttributeError:
        if sys.platform == "darwin":
            subprocess.Popen(["open", folder])
        else:
            subprocess.Popen(["xdg-open", folder])

class MainWindow:
    def __init__(self):
        if tk is None:
            raise RuntimeError("Khong co tkinter de tao GUI.")
        self.root = tk.Tk()
        self.root.title(APP_TITLE)
        self.root.geometry("1220x820")
        self.folder_path = ""
        self.output_path = ""
        self.logger = AppLogger(self._append_log)
        self.processor = Processor(self.logger)
        self.fix_by_folder_var = tk.BooleanVar(value=True)
        self.fix_by_manual_var = tk.BooleanVar(value=False)
        self.issue_items = []
        self.issue_items_master = []
        self.filtered_issue_items = []
        self._issue_tree_cleared = False
        self._build_ui()
        self._load_config()

    def _build_ui(self):
        frm = ttk.Frame(self.root, padding=10)
        frm.pack(fill="both", expand=True)

        self.lbl_folder = ttk.Label(frm, text="Thu muc da chon:\n(Chua chon)")
        self.lbl_folder.pack(anchor="w", pady=(0, 6))

        self.lbl_output = ttk.Label(frm, text="File ket qua:\n(Chua chon)")
        self.lbl_output.pack(anchor="w", pady=(0, 10))

        btn_frame = ttk.Frame(frm)
        btn_frame.pack(fill="x", pady=4)
        ttk.Button(btn_frame, text="Chon thu muc chua file Excel", command=self.select_folder).pack(side="left", padx=3)
        ttk.Button(btn_frame, text="Chon file ket qua & Bat dau xu ly", command=self.select_output_file).pack(side="left", padx=3)

        btn2 = ttk.Frame(frm)
        btn2.pack(fill="x", pady=4)
        ttk.Button(btn2, text="Luu mac dinh", command=self.save_config).pack(side="left", padx=3)
        ttk.Button(btn2, text="Reset", command=self.reset_config).pack(side="left", padx=3)
        ttk.Button(btn2, text="Thuc hien lai", command=self.re_run).pack(side="left", padx=3)
        ttk.Button(btn2, text="Mo file output", command=self.open_output_file).pack(side="left", padx=3)

        fix_box = ttk.LabelFrame(frm, text="Tuy chon sua file nguon", padding=8)
        fix_box.pack(fill="x", pady=(8, 4))
        ttk.Checkbutton(fix_box, text="Sua theo chuan ten folder", variable=self.fix_by_folder_var).grid(row=0, column=0, sticky="w", padx=4, pady=2)
        ttk.Checkbutton(fix_box, text="Sua theo thong tin nhap tay", variable=self.fix_by_manual_var).grid(row=0, column=1, sticky="w", padx=4, pady=2)

        ttk.Label(fix_box, text="Invoice No").grid(row=1, column=0, sticky="w", padx=4)
        self.ent_manual_inv = ttk.Entry(fix_box, width=30)
        self.ent_manual_inv.grid(row=1, column=1, sticky="w", padx=4)

        ttk.Label(fix_box, text="Invoice Date (dd/mm/yyyy)").grid(row=2, column=0, sticky="w", padx=4)
        self.ent_manual_date = ttk.Entry(fix_box, width=20)
        self.ent_manual_date.grid(row=2, column=1, sticky="w", padx=4)

        ttk.Label(fix_box, text="Destination").grid(row=3, column=0, sticky="w", padx=4)
        self.ent_manual_dest = ttk.Entry(fix_box, width=20)
        self.ent_manual_dest.grid(row=3, column=1, sticky="w", padx=4)


        issues_box = ttk.LabelFrame(frm, text="Bang tra cuu WARNING / ERROR", padding=8)
        issues_box.pack(fill="both", expand=False, pady=(8, 4))

        search_bar = ttk.Frame(issues_box)
        search_bar.pack(fill="x", pady=(0, 6))
        ttk.Label(search_bar, text="Tim kiem:").pack(side="left")
        self.issue_search_var = tk.StringVar()
        self.issue_search_var.trace_add("write", lambda *args: self.refresh_issue_tree())
        ttk.Entry(search_bar, textvariable=self.issue_search_var, width=50).pack(side="left", padx=6)
        ttk.Label(search_bar, text="(tim theo file, loi, path)").pack(side="left")

        issue_main = ttk.Frame(issues_box)
        issue_main.pack(fill="both", expand=True)

        self.issue_tree = ttk.Treeview(issue_main, columns=("status", "file", "detail"), show="headings", height=9, selectmode="extended")
        self.issue_tree.heading("status", text="Status")
        self.issue_tree.heading("file", text="File")
        self.issue_tree.heading("detail", text="Loi chi tiet")
        self.issue_tree.column("status", width=90, anchor="center")
        self.issue_tree.column("file", width=260, anchor="w")
        self.issue_tree.column("detail", width=520, anchor="w")
        self.issue_tree.pack(side="left", fill="both", expand=True)

        issues_scroll = ttk.Scrollbar(issue_main, orient="vertical", command=self.issue_tree.yview)
        issues_scroll.pack(side="left", fill="y")
        self.issue_tree.configure(yscrollcommand=issues_scroll.set)
        self._enable_issue_tree_multi_copy()

        issue_btn = ttk.Frame(issue_main)
        issue_btn.pack(side="left", fill="y", padx=(8, 0))
        ttk.Button(issue_btn, text="Refresh loi", command=self.refresh_issue_tree).pack(fill="x", pady=2)
        ttk.Button(issue_btn, text="Mo file duoc chon", command=self.open_selected_issue_file).pack(fill="x", pady=2)
        ttk.Button(issue_btn, text="Mo thu muc chua file", command=self.open_selected_issue_folder).pack(fill="x", pady=2)
        ttk.Button(issue_btn, text="Copy duong dan", command=self.copy_selected_issue_path).pack(fill="x", pady=2)
        ttk.Button(issue_btn, text="Copy loi chi tiet", command=self.copy_selected_issue_detail).pack(fill="x", pady=2)
        ttk.Button(issue_btn, text="Copy dong da chon", command=self.copy_selected_issues_tsv).pack(fill="x", pady=2)
        ttk.Button(issue_btn, text="Xoa danh sach", command=self.clear_issue_list).pack(fill="x", pady=2)

        ttk.Label(frm, text="Tien do:").pack(anchor="w", pady=(10, 2))
        self.progress = ttk.Progressbar(frm, orient="horizontal", mode="determinate")
        self.progress.pack(fill="x")

        ttk.Label(frm, text="Log:").pack(anchor="w", pady=(10, 2))
        text_frame = ttk.Frame(frm)
        text_frame.pack(fill="both", expand=True)
        self.log_box = tk.Text(text_frame, wrap="word")
        self.log_box.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=self.log_box.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_box.configure(yscrollcommand=scrollbar.set)
        self.log_box.tag_configure("info", foreground="black")
        self.log_box.tag_configure("warn", foreground="orange")
        self.log_box.tag_configure("error", foreground="red")
        self.log_box.configure(state="disabled")

    def _append_log(self, msg: str, level: str = "info"):
        def _inner():
            self.log_box.configure(state="normal")
            self.log_box.insert("end", msg + "\n", level)
            self.log_box.see("end")
            self.log_box.configure(state="disabled")
        self.root.after(0, _inner)


    def clear_issue_list(self):
        # Chi xoa danh sach dang hien thi tren bang, khong xoa nguon loi goc.
        # Khong duoc goi issue_search_var.set("") truc tiep, vi trace_add se kich hoat
        # refresh_issue_tree() va nap lai du lieu ngay lap tuc.
        self._issue_tree_cleared = True
        self.issue_items = []
        self.filtered_issue_items = []
        for item in self.issue_tree.get_children():
            self.issue_tree.delete(item)

    def load_issue_list(self, items):
        self.issue_items_master = list(items or [])
        self.issue_items = list(self.issue_items_master)
        self.refresh_issue_tree()

    def merge_issue_list(self, items, scanned_files=None):
        new_items = list(items or [])
        scanned_set = {normalize_path_text(x) for x in (scanned_files or []) if nz_str(x)}

        current_master = list(getattr(self, "issue_items_master", []))
        if scanned_set:
            kept = []
            for fp, st, detail in current_master:
                if normalize_path_text(fp) in scanned_set:
                    continue
                kept.append((fp, st, detail))
            current_master = kept

        dedup = {}
        for fp, st, detail in current_master + new_items:
            key = (normalize_path_text(fp), nz_str(st).upper().strip(), normalize_simple_text(detail))
            dedup[key] = (fp, st, detail)
        self.issue_items_master = list(dedup.values())
        self.issue_items = list(self.issue_items_master)
        self.refresh_issue_tree()

    def refresh_issue_tree(self):
        self._issue_tree_cleared = False
        keyword = nz_str(self.issue_search_var.get()).lower() if hasattr(self, "issue_search_var") else ""
        if hasattr(self, "issue_items_master"):
            self.issue_items = list(self.issue_items_master)
        for item in self.issue_tree.get_children():
            self.issue_tree.delete(item)

        self.filtered_issue_items = []
        for fp, st, detail in self.issue_items:
            hay = " | ".join([nz_str(fp), nz_str(st), nz_str(detail), get_file_name_from_path(fp)]).lower()
            if keyword and keyword not in hay:
                continue
            self.filtered_issue_items.append((fp, st, detail))
            self.issue_tree.insert("", "end", values=(st, get_file_name_from_path(fp), detail))

    def _enable_issue_tree_multi_copy(self):
        """Cho phep keo chon nhieu dong va Ctrl+C copy bang WARNING/ERROR."""
        try:
            self.issue_tree.configure(selectmode="extended")
            self.issue_tree.bind("<Control-c>", self.copy_selected_issues_tsv)
            self.issue_tree.bind("<Control-C>", self.copy_selected_issues_tsv)
            self.issue_tree.bind("<Button-1>", self._issue_tree_mouse_down, add="+")
            self.issue_tree.bind("<B1-Motion>", self._issue_tree_mouse_drag_select, add="+")
            self.issue_tree.bind("<ButtonRelease-1>", self._issue_tree_mouse_release, add="+")
        except Exception:
            pass

    def _issue_tree_mouse_down(self, event):
        try:
            row_id = self.issue_tree.identify_row(event.y)
            self._issue_drag_anchor_iid = row_id or None
        except Exception:
            self._issue_drag_anchor_iid = None

    def _issue_tree_mouse_drag_select(self, event):
        """Keo chuot de chon mot vung dong lien tuc trong Treeview."""
        try:
            anchor = getattr(self, "_issue_drag_anchor_iid", None)
            row_id = self.issue_tree.identify_row(event.y)
            if not anchor or not row_id:
                return
            children = list(self.issue_tree.get_children())
            if anchor not in children or row_id not in children:
                return
            a = children.index(anchor)
            b = children.index(row_id)
            lo, hi = (a, b) if a <= b else (b, a)
            selected = children[lo:hi + 1]
            self.issue_tree.selection_set(selected)
            self.issue_tree.focus(row_id)
            self.issue_tree.see(row_id)
            return "break"
        except Exception:
            return

    def _issue_tree_mouse_release(self, event):
        try:
            self._issue_drag_anchor_iid = None
        except Exception:
            pass

    def _selected_issue_indices(self):
        try:
            children = list(self.issue_tree.get_children())
            selected = list(self.issue_tree.selection())
            out = []
            for iid in selected:
                if iid in children:
                    idx = children.index(iid)
                    if 0 <= idx < len(self.filtered_issue_items):
                        out.append(idx)
            return sorted(set(out))
        except Exception:
            return []

    def get_selected_issue(self):
        indices = self._selected_issue_indices()
        if not indices:
            messagebox.showwarning("Chua chon", "Hay chon 1 dong loi/canh bao trong bang.")
            return None
        return self.filtered_issue_items[indices[0]]

    def get_selected_issues(self):
        indices = self._selected_issue_indices()
        return [self.filtered_issue_items[i] for i in indices if 0 <= i < len(self.filtered_issue_items)]

    def open_selected_issue_file(self):
        item = self.get_selected_issue()
        if not item:
            return
        fp, st, detail = item
        try:
            open_file_with_default_app(fp)
        except Exception as exc:
            messagebox.showerror("Khong mo duoc file", str(exc))

    def open_selected_issue_folder(self):
        item = self.get_selected_issue()
        if not item:
            return
        fp, st, detail = item
        try:
            open_folder_and_select_file(fp)
        except Exception as exc:
            messagebox.showerror("Khong mo duoc thu muc", str(exc))

    def copy_selected_issue_path(self):
        items = self.get_selected_issues()
        if not items:
            messagebox.showwarning("Chua chon", "Hay chon it nhat 1 dong trong bang WARNING/ERROR.")
            return
        text = "\n".join(nz_str(fp) for fp, st, detail in items)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self._append_log(f"Da copy {len(items)} duong dan WARNING/ERROR")

    def copy_selected_issue_detail(self):
        items = self.get_selected_issues()
        if not items:
            messagebox.showwarning("Chua chon", "Hay chon it nhat 1 dong trong bang WARNING/ERROR.")
            return
        text = "\n".join(nz_str(detail) for fp, st, detail in items)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self._append_log(f"Da copy {len(items)} noi dung loi/canh bao")

    def copy_selected_issues_tsv(self, event=None):
        """Copy cac dong dang chon thanh TSV, paste thang duoc vao Excel."""
        items = self.get_selected_issues()
        if not items:
            return "break"
        lines = ["Status\tFile\tLoi chi tiet\tFile Path"]
        for fp, st, detail in items:
            vals = [nz_str(st), get_file_name_from_path(fp), nz_str(detail), nz_str(fp)]
            vals = [v.replace("\t", " ").replace("\r", " ").replace("\n", " ") for v in vals]
            lines.append("\t".join(vals))
        text = "\n".join(lines)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self._append_log(f"Da copy {len(items)} dong WARNING/ERROR dang chon vao clipboard")
        return "break"

    def _load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                self.folder_path = cfg.get("folder", "")
                self.output_path = cfg.get("output", "")
                self.fix_by_folder_var.set(bool(cfg.get("fix_by_folder", True)))
                self.fix_by_manual_var.set(bool(cfg.get("fix_by_manual", False)))
            except Exception:
                pass
        self._refresh_labels()

    def _refresh_labels(self):
        self.lbl_folder.config(text=f"Thu muc da chon:\n{self.folder_path or '(Chua chon)'}")
        self.lbl_output.config(text=f"File ket qua:\n{self.output_path or '(Chua chon)'}")

    def select_folder(self):
        folder = filedialog.askdirectory(title="Chon thu muc nguon")
        if folder:
            self.folder_path = folder
            self._refresh_labels()
            self._append_log(f"Da chon thu muc: {folder}")

    def select_output_file(self):
        if not self.folder_path:
            messagebox.showwarning("Thieu thu muc", "Hay chon thu muc truoc.")
            return
        path = filedialog.asksaveasfilename(
            title="Chon hoac tao file ket qua",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=os.path.basename(self.output_path) if self.output_path else "SUB_DETAIL_OUTPUT.xlsx",
        )
        if path:
            if not path.lower().endswith(".xlsx"):
                messagebox.showwarning("Sai dinh dang", "Chi duoc chon file .xlsx")
                return
            self.output_path = path
            self._refresh_labels()
            self.start_process()

    def open_output_file(self):
        if not self.output_path:
            messagebox.showwarning("Chua co file", "Chua co duong dan file output.")
            return
        try:
            open_file_with_default_app(self.output_path)
        except Exception as exc:
            messagebox.showerror("Khong mo duoc file output", str(exc))

    def save_config(self):
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump({"folder": self.folder_path, "output": self.output_path, "fix_by_folder": self.fix_by_folder_var.get(), "fix_by_manual": self.fix_by_manual_var.get()}, f, ensure_ascii=False, indent=2)
        messagebox.showinfo("Da luu", "Da luu cau hinh mac dinh!")

    def reset_config(self):
        if os.path.exists(CONFIG_FILE):
            os.remove(CONFIG_FILE)
        self.folder_path = ""
        self.output_path = ""
        self.ent_manual_inv.delete(0, "end")
        self.ent_manual_date.delete(0, "end")
        self.ent_manual_dest.delete(0, "end")
        self._refresh_labels()
        messagebox.showinfo("Da reset", "Da xoa cau hinh mac dinh!")

    def re_run(self):
        if not self.folder_path or not self.output_path:
            messagebox.showwarning("Thieu thong tin", "Hay chon thu muc va file ket qua truoc.")
            return
        self.start_process()

    def start_process(self):
        self.progress["value"] = 0
        self.refresh_issue_tree()
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")
        self._append_log(f"Bat dau xu ly...\n{self.folder_path}\n{self.output_path}")
        self._append_log("Goi y: voi file .xls, hay dung danh sach WARNING/ERROR de mo file va sua tay nhanh hon.")
        threading.Thread(target=self._run_worker, daemon=True).start()

    def _run_worker(self):
        try:
            def progress_callback(done: int, total: int, file_path: str):
                percent = 0 if total <= 0 else int(done / total * 100)
                def _update():
                    self.progress["value"] = percent
                self.root.after(0, _update)
                self._append_log(f"[{done}/{total}] {file_path}")

            repair_options = RepairOptions(
                use_folder_truth=self.fix_by_folder_var.get(),
                use_manual_values=self.fix_by_manual_var.get(),
                manual_invoice_no=self.ent_manual_inv.get().strip(),
                manual_invoice_date=self.ent_manual_date.get().strip(),
                manual_destination=self.ent_manual_dest.get().strip(),
            )
            summary = self.processor.run(self.folder_path, self.output_path, progress_callback=progress_callback, repair_options=repair_options)
            msg = (
                "Hoan tat.\n"
                f"Tong file quet: {summary['total']}\n"
                f"OK: {summary['ok']}\n"
                f"WARNING: {summary['warning']}\n"
                f"ERROR: {summary['error']}\n"
                f"Skip Path: {summary['skip_path']}\n"
                f"Skip Signature: {summary['skip_signature']}\n"
                f"Repaired: {summary['repaired']}"
            )
            self._append_log(msg)
            self.root.after(0, lambda items=summary.get("issues", []), scanned=summary.get("scanned_files", []): self.merge_issue_list(items, scanned))
            self.root.after(0, lambda: messagebox.showinfo("Hoan tat", msg))
        except Exception as exc:
            detail = f"Loi: {exc}\n\n{traceback.format_exc()}"
            self._append_log(detail, "error")
            self.root.after(0, lambda e=exc: messagebox.showerror("Loi", str(e)))

    def run(self):
        self.root.mainloop()


def main():
    if tk is None:
        raise RuntimeError("Moi truong hien tai khong ho tro tkinter GUI.")
    MainWindow().run()




# ==============================
# V4.2 OVERRIDES - DETAIL REPORT LIKE SAMPLE
# ==============================

INV_DETAIL_HEADERS = [
    "Date:", "Invoice No.:", "DESTINATION:", "PO No", "Customer PO#",
    "DESC", "HS CODE", "MODEL NAME / COLOR", "MODEL Y", "COLOR", "WIDTH",
    "Q'TY (Pairs)", "UNIT PRICE (USD)", "AMOUNT W/OUT LABEL",
    "LABEL COST", "LABEL AMOUNT", "TOTAL AMOUNT",
    "Row Type", "Status", "File", "File Path"
]

PL_DETAIL_HEADERS = [
    "Date:", "Invoice No.:", "DESTINATION:", "PO No.", "Customer PO#",
    "MODEL NAME / COLOR", "COLOR", "WIDTH", "HS CODE",
    "Q'TY (Pairs)", "CARTON Q'TY", "CBM", "GROSS WEIGHT",
    "Row Type", "Status", "File", "File Path"
]

TOTAL_ROW_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
HANDLING_ROW_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
TOTAL_FONT = Font(bold=True, color="FF0000")
HANDLING_FONT = Font(color="00B0F0", bold=True)

def match_inv_detail_key(txt: str) -> str:
    t = normalize_header(txt)
    if not t:
        return ""
    if t.startswith("DATE"):
        return "date"
    if "INVOICENO" in t or "INVOICENUMBER" in t:
        return "invoice_no"
    if "DESTINATION" in t:
        return "destination"
    if t.startswith("PONO"):
        return "po_no"
    if "CUSTOMERPO" in t or "CUSPO" in t:
        return "customer_po"
    if t in ("DESC", "DESCRIPTION"):
        return "desc"
    if "HSCODE" in t:
        return "hs_code"
    if "MODELNAMECOLOR" in t:
        return "model_name_color"
    if t == "MODELY":
        return "model_y"
    if t == "COLOR":
        return "color"
    if "WIDTH" in t:
        return "width"
    if is_inv_pair_qty_header(t):
        return "qty_pairs"
    if "UNITPRICE" in t and "USD" in t:
        return "unit_price_usd"
    if "AMOUNTWOUTLABEL" in t or "AMOUNTWITHOUTLABEL" in t or is_inv_amount_without_label_header(t):
        return "amount_wo_label"
    if ("LABELCOST" in t and "TOTAL" not in t) or t == "LABELCOST":
        return "label_cost"
    if "TOTALLABELCOST" in t or is_inv_label_amount_header(t):
        return "label_amount"
    if is_inv_total_amount_header(t):
        return "total_amount"
    return ""

def init_inv_detail_header(ws):
    if nz_str(ws.cell(1, 1).value) == "":
        for idx, header in enumerate(INV_DETAIL_HEADERS, start=1):
            ws.cell(1, idx).value = sanitize_excel_text(header)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"
    # hide technical columns
    for c in (18, 19, 20, 21):
        ws.column_dimensions[get_column_letter(c)].hidden = True

def init_pl_detail_header(ws):
    if nz_str(ws.cell(1, 1).value) == "":
        for idx, header in enumerate(PL_DETAIL_HEADERS, start=1):
            ws.cell(1, idx).value = sanitize_excel_text(header)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"
    for c in (14, 15, 16, 17):
        ws.column_dimensions[get_column_letter(c)].hidden = True

def append_generic_rows(ws, rows: List[List[Any]], date_cols: List[int], status_col_idx: int):
    if not rows:
        return
    start_row = ws.max_row + 1 if nz_str(ws.cell(ws.max_row, 1).value) else max(2, ws.max_row)
    if start_row < 2:
        start_row = 2
    rowtype_idx = 18 if ws.title == "INV" else 14
    for idx, arr in enumerate(rows, start=start_row):
        for c, value in enumerate(arr, start=1):
            cell = ws.cell(idx, c)
            cell.value = sanitize_excel_text(value)
            if c in date_cols:
                cell.number_format = "dd-mmm"
            # simple number formats for visible numeric columns
            if ws.title == "INV":
                if c in (12,):
                    cell.number_format = '#,##0'
                elif c in (13, 14, 16, 17):
                    cell.number_format = '#,##0.00'
                elif c in (15,):
                    cell.number_format = '0.0000'
            else:
                if c in (10, 11):
                    cell.number_format = '#,##0'
                elif c in (12, 13):
                    cell.number_format = '#,##0.000'
        status = nz_str(arr[status_col_idx - 1]).upper()
        row_type = nz_str(arr[rowtype_idx - 1]).upper()
        if row_type == "TOTAL":
            for c in range(1, len(arr) + 1):
                ws.cell(idx, c).fill = TOTAL_ROW_FILL
                ws.cell(idx, c).font = TOTAL_FONT
        elif row_type == "HANDLING CHARGE":
            for c in range(1, len(arr) + 1):
                ws.cell(idx, c).font = HANDLING_FONT
        elif status in STATUS_FILLS:
            fill = STATUS_FILLS[status]
            for c in range(1, len(arr) + 1):
                ws.cell(idx, c).fill = fill

def keep_rows_by_file_paths(ws, header_len: int, file_path_col_idx: int, valid_paths: set) -> int:
    kept = []
    removed = 0
    for arr in get_existing_rows(ws, header_len):
        path = nz_str(arr[file_path_col_idx - 1]).upper()
        if path and path in valid_paths:
            kept.append(arr)
        else:
            removed += 1
    ws.delete_rows(2, max(0, ws.max_row - 1))
    if ws.title == "INV":
        append_generic_rows(ws, kept, [1], 19)
    elif ws.title == "PL":
        append_generic_rows(ws, kept, [1], 15)
    return removed

def find_detail_range(sheet_data: List[List[Any]], kind: str) -> Tuple[Dict[str, int], int, int]:
    col_map, hdr_row = detect_detail_header_row(sheet_data, kind)
    total_row = 0
    po_col = col_map.get("po_no", 0)
    for r in range(hdr_row + 1, last_used_row(sheet_data) + 1):
        txt1 = normalize_simple_text(get_cell(sheet_data, r, po_col)).upper() if po_col else ""
        txt2 = normalize_simple_text(get_cell(sheet_data, r, 1)).upper()
        probe = (txt1 + " " + txt2).replace(":", "")
        if "TOTAL" in probe:
            total_row = r
            break
    if total_row == 0:
        total_row = last_used_row(sheet_data)
    return col_map, hdr_row, total_row

def get_value_or_fallback(sheet_data: List[List[Any]], row_num: int, col_map: Dict[str, int], key: str) -> Any:
    v = get_mapped_cell_value(sheet_data, row_num, col_map, key)
    if nz_str(v):
        return v
    model_col = col_map.get("model_name_color", 0)
    if model_col <= 0:
        return v
    if key == "model_y":
        return get_cell(sheet_data, row_num, model_col + 1)
    if key == "color":
        return get_cell(sheet_data, row_num, model_col + 2)
    if key == "width":
        return get_cell(sheet_data, row_num, model_col + 3)
    return v

def classify_row_type_by_map(sheet_data: List[List[Any]], row_num: int, col_map: Dict[str, int]) -> str:
    probes = []
    for key in ("po_no", "customer_po", "desc", "model_name_color"):
        c = col_map.get(key, 0)
        if c > 0:
            probes.append(normalize_simple_text(get_cell(sheet_data, row_num, c)).upper().replace(":", ""))
    probes.append(normalize_simple_text(get_cell(sheet_data, row_num, 1)).upper().replace(":", ""))
    full = " | ".join([p for p in probes if p])
    if "TOTAL" in full:
        return "TOTAL"
    if "HANDLING CHARGE" in full:
        return "HANDLING CHARGE"
    return "DETAIL"

def build_report_meta(pl_meta: Dict[str, str], inv_meta: Dict[str, str]) -> Dict[str, str]:
    def pick(key: str) -> str:
        return nz_str(inv_meta.get(key, "")) or nz_str(pl_meta.get(key, ""))
    return {
        "invoice_no": pick("invoice_no"),
        "invoice_date": pick("invoice_date"),
        "destination": pick("destination"),
    }

def extract_pl_detail_rows(file_path: str, folder_name: str, folder_created: str, sheet_name: str,
                           sheet_data: List[List[Any]], status_text: str, report_meta: Optional[Dict[str, str]] = None) -> List[List[Any]]:
    if report_meta is None:
        report_meta = {}
    col_map, hdr_row, total_row = find_detail_range(sheet_data, "PL")
    rows = []
    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))

    for r in range(hdr_row + 1, total_row + 1):
        row_type = classify_row_type_by_map(sheet_data, r, col_map)
        pair_qty = get_mapped_cell_value(sheet_data, r, col_map, "pair_qty")
        po_v = get_mapped_cell_value(sheet_data, r, col_map, "po_no")

        if row_type == "DETAIL" and not (nz_str(po_v) or nz_str(pair_qty)):
            continue

        model_name = get_mapped_cell_value(sheet_data, r, col_map, "model_name_color")
        color_v = get_value_or_fallback(sheet_data, r, col_map, "color")
        width_v = get_value_or_fallback(sheet_data, r, col_map, "width")

        if row_type == "TOTAL":
            po_v = "TOTAL"

        rows.append([
            rep_date, rep_inv, rep_dest,
            po_v,
            get_mapped_cell_value(sheet_data, r, col_map, "customer_po"),
            model_name,
            color_v,
            width_v,
            get_mapped_cell_value(sheet_data, r, col_map, "hs_code"),
            pair_qty,
            get_mapped_cell_value(sheet_data, r, col_map, "carton_qty"),
            get_mapped_cell_value(sheet_data, r, col_map, "cbm"),
            get_mapped_cell_value(sheet_data, r, col_map, "gross_weight"),
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows

def extract_inv_detail_rows(file_path: str, folder_name: str, folder_created: str, sheet_name: str,
                            sheet_data: List[List[Any]], status_text: str, report_meta: Optional[Dict[str, str]] = None) -> List[List[Any]]:
    if report_meta is None:
        report_meta = {}
    col_map, hdr_row, total_row = find_detail_range(sheet_data, "INV")
    rows = []
    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))

    for r in range(hdr_row + 1, total_row + 1):
        row_type = classify_row_type_by_map(sheet_data, r, col_map)
        qty_v = get_mapped_cell_value(sheet_data, r, col_map, "qty_pairs")
        po_v = get_mapped_cell_value(sheet_data, r, col_map, "po_no")

        if row_type == "DETAIL" and not (nz_str(po_v) or nz_str(qty_v)):
            continue

        model_name = get_mapped_cell_value(sheet_data, r, col_map, "model_name_color")
        model_y = get_value_or_fallback(sheet_data, r, col_map, "model_y")
        color_v = get_value_or_fallback(sheet_data, r, col_map, "color")
        width_v = get_value_or_fallback(sheet_data, r, col_map, "width")

        if row_type == "TOTAL":
            po_v = "TOTAL:"
        elif row_type == "HANDLING CHARGE":
            po_v = "HANDLING CHARGE"

        rows.append([
            rep_date, rep_inv, rep_dest,
            po_v,
            get_mapped_cell_value(sheet_data, r, col_map, "customer_po"),
            get_mapped_cell_value(sheet_data, r, col_map, "desc"),
            get_mapped_cell_value(sheet_data, r, col_map, "hs_code"),
            model_name,
            model_y,
            color_v,
            width_v,
            qty_v,
            get_mapped_cell_value(sheet_data, r, col_map, "unit_price_usd"),
            get_mapped_cell_value(sheet_data, r, col_map, "amount_wo_label"),
            get_mapped_cell_value(sheet_data, r, col_map, "label_cost"),
            get_mapped_cell_value(sheet_data, r, col_map, "label_amount"),
            get_mapped_cell_value(sheet_data, r, col_map, "total_amount"),
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows

def process_one_sub_file_bundle(file_path: str) -> Tuple[List[Any], List[List[Any]], List[List[Any]]]:
    arr = process_one_sub_file(file_path)
    inv_rows: List[List[Any]] = []
    pl_rows: List[List[Any]] = []
    try:
        wb_data = load_workbook_data(file_path)
        pl_name = nz_str(arr[3]) or find_best_pl_sheet(wb_data) or ""
        inv_name = nz_str(arr[4]) or find_best_inv_sheet(wb_data) or ""

        pl_meta = extract_sheet_meta(wb_data.sheets[pl_name]) if pl_name and pl_name in wb_data.sheets else {k: "" for k in META_KEYS}
        inv_meta = extract_sheet_meta(wb_data.sheets[inv_name]) if inv_name and inv_name in wb_data.sheets else {k: "" for k in META_KEYS}
        report_meta = build_report_meta(pl_meta, inv_meta)

        # PL detail: try main PL first, then fallback to any sheet with PL-like header
        if pl_name and pl_name in wb_data.sheets:
            try:
                pl_rows = extract_pl_detail_rows(file_path, get_folder_name_from_path(file_path), get_parent_folder_created(file_path),
                                                 pl_name, wb_data.sheets[pl_name], nz_str(arr[21]).upper(), report_meta)
            except Exception:
                pl_rows = []
        if not pl_rows:
            for other_name in wb_data.sheet_names:
                try:
                    pl_rows = extract_pl_detail_rows(file_path, get_folder_name_from_path(file_path), get_parent_folder_created(file_path),
                                                     other_name, wb_data.sheets[other_name], nz_str(arr[21]).upper(), report_meta)
                    if pl_rows:
                        break
                except Exception:
                    pass

        if inv_name and inv_name in wb_data.sheets:
            try:
                inv_rows = extract_inv_detail_rows(file_path, get_folder_name_from_path(file_path), get_parent_folder_created(file_path),
                                                   inv_name, wb_data.sheets[inv_name], nz_str(arr[21]).upper(), report_meta)
            except Exception:
                inv_rows = []
        if not inv_rows:
            for other_name in wb_data.sheet_names:
                try:
                    inv_rows = extract_inv_detail_rows(file_path, get_folder_name_from_path(file_path), get_parent_folder_created(file_path),
                                                       other_name, wb_data.sheets[other_name], nz_str(arr[21]).upper(), report_meta)
                    if inv_rows:
                        break
                except Exception:
                    pass
    except Exception:
        pass
    return arr, inv_rows, pl_rows

# patch Processor.run dependent indexes through keep_rows_by_file_paths override above


# ==============================
# V4.3 OVERRIDES - ROBUST INV/PL DETAIL
# ==============================

TOTAL_ROW_FILL = PatternFill(fill_type="solid", fgColor="FDE9D9")
TOTAL_FONT = Font(bold=True, color="C00000")
HANDLING_ROW_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
HANDLING_FONT = Font(bold=True, color="0070C0")

def _clean_upper(v: Any) -> str:
    return normalize_simple_text(v).upper().replace(":", "").replace("#", "")

def _sheet_name_upper(name: str) -> str:
    return nz_str(name).strip().upper()

def match_pl_detail_key(txt: str) -> str:
    t = normalize_header(txt)
    if not t:
        return ""
    # standard SUB / VSF / SA headers
    if t.startswith("DATE"):
        return "date"
    if "INVOICENO" in t or "INVOICENUMBER" in t:
        return "invoice_no"
    if "DESTINATION" in t:
        return "destination"
    if t in ("PO", "PONO"):
        return "po_no"
    if "CUSTOMERPO" in t or "CUSPO" in t:
        return "customer_po"
    if "MODELNAMECOLOR" in t or "GROUPSUMMODEL" in t or t == "STYLE":
        return "model_name_color"
    if t == "COLOR":
        return "color"
    if "WIDTH" in t or "GENDERSIZE" in t:
        return "width"
    if "HSCODE" in t:
        return "hs_code"
    if is_pl_pair_qty_header(t) or "QTYPRS" in t or "QTYPAIR" in t:
        return "pair_qty"
    if is_pl_carton_header(t) or "PACKAGECTN" in t or "PACKAGE" == t:
        return "carton_qty"
    if is_pl_cbm_header(t) or "MEASUREMENTCBM" in t or t == "MEASUREMENT":
        return "cbm"
    if is_pl_gross_header(t) or t in ("GWKG", "GW", "GWKGS", "GWEIGHT"):
        return "gross_weight"
    return ""

def match_inv_detail_key(txt: str) -> str:
    t = normalize_header(txt)
    if not t:
        return ""
    if t.startswith("DATE"):
        return "date"
    if "INVOICENO" in t or "INVOICENUMBER" in t:
        return "invoice_no"
    if "DESTINATION" in t:
        return "destination"
    if t in ("PO", "PONO"):
        return "po_no"
    if "CUSTOMERPO" in t or "CUSPO" in t:
        return "customer_po"
    if t in ("DESC", "DESCRIPTION", "TYPECATEGORY"):
        return "desc"
    if "HSCODE" in t:
        return "hs_code"
    if "MODELNAMECOLOR" in t or "GROUPSUMMODEL" in t:
        return "model_name_color"
    if t == "STYLE":
        return "model_y"
    if t == "MODELY":
        return "model_y"
    if t == "COLOR":
        return "color"
    if "WIDTH" in t:
        return "width"
    if is_inv_pair_qty_header(t) or "QTYPAIR" in t:
        return "qty_pairs"
    if "UNITPRICE" in t or "UPRICEUSDPAIR" in t:
        return "unit_price_usd"
    if is_inv_amount_without_label_header(t) or t in ("AMOUNTUSD", "AMOUNTWOUTLABELCOST"):
        return "amount_wo_label"
    if ("LABELCOST" in t and "TOTAL" not in t) or t == "LABELCOST":
        return "label_cost"
    if is_inv_label_amount_header(t) or t == "TOTALLABELCOST":
        return "label_amount"
    if is_inv_total_amount_header(t) or t == "TOTALAMOUNT":
        return "total_amount"
    return ""

def init_inv_detail_header(ws):
    for idx, header in enumerate(INV_DETAIL_HEADERS, start=1):
        ws.cell(1, idx).value = sanitize_excel_text(header)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"
    # keep technical columns hidden
    for c in (18, 19, 20, 21):
        ws.column_dimensions[get_column_letter(c)].hidden = True

def init_pl_detail_header(ws):
    for idx, header in enumerate(PL_DETAIL_HEADERS, start=1):
        ws.cell(1, idx).value = sanitize_excel_text(header)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"
    for c in (14, 15, 16, 17):
        ws.column_dimensions[get_column_letter(c)].hidden = True

def append_generic_rows(ws, rows: List[List[Any]], date_cols: List[int], status_col_idx: int):
    if not rows:
        return
    start_row = ws.max_row + 1 if nz_str(ws.cell(ws.max_row, 1).value) else max(2, ws.max_row)
    if start_row < 2:
        start_row = 2
    row_type_col_idx = 18 if ws.title == "INV" else 14
    for idx, arr in enumerate(rows, start=start_row):
        for c, value in enumerate(arr, start=1):
            cell = ws.cell(idx, c)
            cell.value = sanitize_excel_text(value)
            if c in date_cols:
                cell.number_format = "dd-mmm"
            if ws.title == "INV":
                if c == 12:
                    cell.number_format = '#,##0'
                elif c in (13, 14, 16, 17):
                    cell.number_format = '#,##0.00'
                elif c == 15:
                    cell.number_format = '0.0000'
            else:
                if c in (10, 11):
                    cell.number_format = '#,##0'
                elif c in (12, 13):
                    cell.number_format = '#,##0.000'
        row_type = nz_str(arr[row_type_col_idx - 1]).upper()
        status = nz_str(arr[status_col_idx - 1]).upper()
        if row_type == "TOTAL":
            for c in range(1, len(arr) + 1):
                ws.cell(idx, c).fill = TOTAL_ROW_FILL
                ws.cell(idx, c).font = TOTAL_FONT
        elif row_type == "HANDLING CHARGE":
            for c in range(1, len(arr) + 1):
                ws.cell(idx, c).fill = HANDLING_ROW_FILL
                ws.cell(idx, c).font = HANDLING_FONT
        elif status in STATUS_FILLS:
            fill = STATUS_FILLS[status]
            for c in range(1, len(arr) + 1):
                ws.cell(idx, c).fill = fill
def get_existing_rows(ws, header_len):
    rows = []
    max_row = ws.max_row

    for r in range(2, max_row + 1):
        row = [ws.cell(row=r, column=c).value for c in range(1, header_len + 1)]
        if any(v is not None for v in row):
            rows.append(row)

    return rows                

def keep_rows_by_file_paths(ws, header_len: int, file_path_col_idx: int, valid_paths: set) -> int:
    kept = []
    removed = 0
    for arr in get_existing_rows(ws, header_len):
        path = nz_str(arr[file_path_col_idx - 1]).upper()
        if path and path in valid_paths:
            kept.append(arr)
        else:
            removed += 1
    ws.delete_rows(2, max(0, ws.max_row - 1))
    if ws.title == "INV":
        append_generic_rows(ws, kept, [1], 19)
    elif ws.title == "PL":
        append_generic_rows(ws, kept, [1], 15)
    return removed

def get_mapped_with_fallback(sheet_data: List[List[Any]], row_num: int, col_map: Dict[str, int], key: str, layout: str) -> Any:
    v = get_mapped_cell_value(sheet_data, row_num, col_map, key)
    if nz_str(v):
        return v
    model_col = col_map.get("model_name_color", 0)
    if model_col <= 0:
        return v
    # standard sub/vsf layout after MODEL NAME / COLOR
    if layout.startswith("STD"):
        if key == "model_y":
            return get_cell(sheet_data, row_num, model_col + 1)
        if key == "color":
            return get_cell(sheet_data, row_num, model_col + 2)
        if key == "width":
            return get_cell(sheet_data, row_num, model_col + 3)
    # standard PL layout after MODEL NAME / COLOR
    if layout == "STD_PL":
        if key == "color":
            return get_cell(sheet_data, row_num, model_col + 1)
        if key == "width":
            return get_cell(sheet_data, row_num, model_col + 2)
    # SA invoice / PL layout
    if layout in ("SA_INV", "SA_PL"):
        if key == "model_y":
            return get_cell(sheet_data, row_num, model_col + 1)
    return v

def normalize_model_name(v: Any) -> str:
    s = nz_str(v)
    if not s:
        return ""
    # "U2010 v1" -> "U2010"
    parts = s.split()
    if len(parts) >= 1 and re.match(r"^[A-Z0-9]+$", parts[0].upper()):
        return parts[0]
    return s

def is_footer_or_signature_row(sheet_data: List[List[Any]], row_num: int) -> bool:
    probes = []
    maxc = min(12, last_used_col_on_row(sheet_data, row_num))
    for c in range(1, maxc + 1):
        probes.append(_clean_upper(get_cell(sheet_data, row_num, c)))
    full = " | ".join([p for p in probes if p])
    bad_terms = [
        "VERY TRULY YOURS", "GENERAL DIRECTOR", "SIGNED BY", "ACCOUNT NAME",
        "ACCOUNT NO", "BANK NAME", "BANK ADDRESS", "SWIFT CODE"
    ]
    return any(term in full for term in bad_terms)

def detect_inv_detail_header_row_v43(sheet_data: List[List[Any]], sheet_name: str) -> Tuple[Dict[str, int], int, str]:
    best_map = {}
    best_score = -1
    best_row = 0
    best_layout = ""
    for r in range(1, min(50, len(sheet_data)) + 1):
        tmp = {}
        last_col = last_used_col_on_row(sheet_data, r)
        for c in range(1, min(last_col, 80) + 1):
            raw = get_cell(sheet_data, r, c)
            key = match_inv_detail_key(raw)
            if key and key not in tmp:
                tmp[key] = c
        score = len(tmp)
        layout = ""
        row_text = " | ".join(_clean_upper(get_cell(sheet_data, r, c)) for c in range(1, min(last_col, 15) + 1))
        if "GROUP SUM MODEL" in row_text or "UPRICEUSDPAIR" in row_text.replace(" ", ""):
            layout = "SA_INV"
            score += 2
        elif "MODEL NAME COLOR" in row_text:
            layout = "STD_INV"
            score += 1
        if score > best_score:
            best_score = score
            best_map = tmp
            best_row = r
            best_layout = layout or "STD_INV"
    if best_score < 6:
        raise ValueError("Khong tim thay header INV detail hop le.")
    return best_map, best_row, best_layout

def detect_pl_detail_header_row_v43(sheet_data: List[List[Any]], sheet_name: str) -> Tuple[Dict[str, int], int, str]:
    best_map = {}
    best_score = -1
    best_row = 0
    best_layout = ""
    sheet_nm = _sheet_name_upper(sheet_name)
    for r in range(1, min(60, len(sheet_data)) + 1):
        tmp = {}
        last_col = last_used_col_on_row(sheet_data, r)
        for c in range(1, min(last_col, 80) + 1):
            raw = get_cell(sheet_data, r, c)
            key = match_pl_detail_key(raw)
            if key and key not in tmp:
                tmp[key] = c
        score = len(tmp)
        row_text = " | ".join(_clean_upper(get_cell(sheet_data, r, c)) for c in range(1, min(last_col, 16) + 1))
        layout = ""
        if "MODEL NAME COLOR" in row_text:
            layout = "STD_PL"
            score += 2
        elif "GROUP SUM MODEL" in row_text or ("PO" in row_text and "CUSTOMER PO" in row_text and "PACKAGE" in row_text):
            layout = "SA_PL"
            score += 1
        if score > best_score:
            best_score = score
            best_map = tmp
            best_row = r
            best_layout = layout or ("SA_PL" if sheet_nm == "SHEET1" else "STD_PL")
    if best_score < 5:
        raise ValueError("Khong tim thay header PL detail hop le.")
    return best_map, best_row, best_layout

def find_total_row_v43(sheet_data: List[List[Any]], hdr_row: int, col_map: Dict[str, int], layout: str) -> int:
    po_col = col_map.get("po_no", 0)
    qty_col = col_map.get("qty_pairs", 0) or col_map.get("pair_qty", 0)
    # first explicit TOTAL row
    for r in range(hdr_row + 1, last_used_row(sheet_data) + 1):
        if is_footer_or_signature_row(sheet_data, r):
            return max(hdr_row + 1, r - 1)
        probes = []
        for c in [po_col, 1, 2]:
            if c > 0:
                probes.append(_clean_upper(get_cell(sheet_data, r, c)))
        full = " | ".join([p for p in probes if p])
        if "TOTAL" in full:
            return r
    # SA sheet1 may have final numeric totals without TOTAL text
    if layout == "SA_PL":
        for r in range(last_used_row(sheet_data), hdr_row, -1):
            qty_v = get_cell(sheet_data, r, qty_col) if qty_col else ""
            if isinstance(qty_v, (int, float)) and float(qty_v) > 0:
                po_v = nz_str(get_cell(sheet_data, r, po_col)) if po_col else ""
                if not po_v:
                    return r
    return last_used_row(sheet_data)

def classify_row_type_v43(sheet_data: List[List[Any]], row_num: int, col_map: Dict[str, int], layout: str) -> str:
    probes = []
    for key in ("po_no", "customer_po", "desc", "model_name_color"):
        c = col_map.get(key, 0)
        if c > 0:
            probes.append(get_cell(sheet_data, row_num, c))
    probes.append(get_cell(sheet_data, row_num, 1))
    full = " | ".join([_clean_upper(p) for p in probes if nz_str(p)])
    if "TOTAL" in full:
        return "TOTAL"
    fee_type = _v76_detect_fee_type(*probes)
    if fee_type:
        return fee_type
    return "DETAIL"

def build_report_meta(pl_meta: Dict[str, str], inv_meta: Dict[str, str]) -> Dict[str, str]:
    def pick(key: str) -> str:
        return nz_str(inv_meta.get(key, "")) or nz_str(pl_meta.get(key, ""))
    return {
        "invoice_no": pick("invoice_no"),
        "invoice_date": pick("invoice_date"),
        "destination": pick("destination"),
    }

def collect_best_meta(wb_data: WorkbookData, pl_name: str, inv_name: str) -> Tuple[Dict[str, str], Dict[str, str]]:
    blank = {k: "" for k in META_KEYS}
    inv_meta = blank.copy()
    pl_meta = blank.copy()

    candidate_inv = []
    candidate_pl = []
    if inv_name and inv_name in wb_data.sheets:
        candidate_inv.append(inv_name)
    if pl_name and pl_name in wb_data.sheets:
        candidate_pl.append(pl_name)
    for nm in wb_data.sheet_names:
        up = _sheet_name_upper(nm)
        if up in ("INV", "INVOICE") or "INVOICE" in up or "INV" == up:
            if nm not in candidate_inv:
                candidate_inv.append(nm)
        if up.startswith("PL") or up == "SHEET1":
            if nm not in candidate_pl:
                candidate_pl.append(nm)

    for nm in candidate_inv:
        try:
            meta = extract_sheet_meta(wb_data.sheets[nm])
            for k in META_KEYS:
                if not inv_meta[k] and nz_str(meta.get(k, "")):
                    inv_meta[k] = nz_str(meta.get(k, ""))
        except Exception:
            pass
    for nm in candidate_pl:
        try:
            meta = extract_sheet_meta(wb_data.sheets[nm])
            for k in META_KEYS:
                if not pl_meta[k] and nz_str(meta.get(k, "")):
                    pl_meta[k] = nz_str(meta.get(k, ""))
        except Exception:
            pass
    return pl_meta, inv_meta

def extract_inv_detail_rows(file_path: str, folder_name: str, folder_created: str, sheet_name: str,
                            sheet_data: List[List[Any]], status_text: str, report_meta: Optional[Dict[str, str]] = None) -> List[List[Any]]:
    report_meta = report_meta or {}
    col_map, hdr_row, layout = detect_inv_detail_header_row_v43(sheet_data, sheet_name)
    total_row = find_total_row_v43(sheet_data, hdr_row, col_map, layout)
    rows = []

    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))

    for r in range(hdr_row + 1, total_row + 1):
        if is_footer_or_signature_row(sheet_data, r):
            break
        row_type = classify_row_type_v43(sheet_data, r, col_map, layout)
        po_v = get_mapped_cell_value(sheet_data, r, col_map, "po_no")
        qty_v = get_mapped_cell_value(sheet_data, r, col_map, "qty_pairs")
        label_amt = get_mapped_cell_value(sheet_data, r, col_map, "label_amount")
        total_amt = get_mapped_cell_value(sheet_data, r, col_map, "total_amount")
        amount_wo = get_mapped_cell_value(sheet_data, r, col_map, "amount_wo_label")

        if row_type == "DETAIL" and not (nz_str(po_v) or nz_str(qty_v) or nz_str(label_amt) or nz_str(total_amt) or nz_str(amount_wo)):
            continue

        model_name = get_mapped_cell_value(sheet_data, r, col_map, "model_name_color")
        model_y = get_mapped_with_fallback(sheet_data, r, col_map, "model_y", layout)
        color_v = get_mapped_with_fallback(sheet_data, r, col_map, "color", layout)
        width_v = get_mapped_with_fallback(sheet_data, r, col_map, "width", layout)
        desc_v = get_mapped_cell_value(sheet_data, r, col_map, "desc")
        hs_v = get_mapped_cell_value(sheet_data, r, col_map, "hs_code")

        if layout == "SA_INV":
            model_name = normalize_model_name(model_name)
            # SA invoice has no separate color; keep blank
            if not nz_str(desc_v):
                desc_v = get_mapped_cell_value(sheet_data, r, col_map, "desc")
        if row_type == "TOTAL":
            po_v = "TOTAL:"
        elif row_type == "HANDLING CHARGE":
            po_v = "HANDLING CHARGE"

        rows.append([
            rep_date, rep_inv, rep_dest,
            po_v,
            get_mapped_cell_value(sheet_data, r, col_map, "customer_po"),
            desc_v,
            hs_v,
            model_name,
            model_y,
            color_v,
            width_v,
            qty_v,
            get_mapped_cell_value(sheet_data, r, col_map, "unit_price_usd"),
            amount_wo,
            get_mapped_cell_value(sheet_data, r, col_map, "label_cost"),
            label_amt,
            total_amt,
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows

def extract_pl_detail_rows(file_path: str, folder_name: str, folder_created: str, sheet_name: str,
                           sheet_data: List[List[Any]], status_text: str, report_meta: Optional[Dict[str, str]] = None) -> List[List[Any]]:
    report_meta = report_meta or {}
    col_map, hdr_row, layout = detect_pl_detail_header_row_v43(sheet_data, sheet_name)
    total_row = find_total_row_v43(sheet_data, hdr_row, col_map, layout)
    rows = []

    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))

    for r in range(hdr_row + 1, total_row + 1):
        if is_footer_or_signature_row(sheet_data, r):
            break
        row_type = classify_row_type_v43(sheet_data, r, col_map, layout)
        po_v = get_mapped_cell_value(sheet_data, r, col_map, "po_no")
        qty_v = get_mapped_cell_value(sheet_data, r, col_map, "pair_qty")
        carton_v = get_mapped_cell_value(sheet_data, r, col_map, "carton_qty")
        cbm_v = get_mapped_cell_value(sheet_data, r, col_map, "cbm")
        gw_v = get_mapped_cell_value(sheet_data, r, col_map, "gross_weight")

        if row_type == "DETAIL" and not (nz_str(po_v) or nz_str(qty_v) or nz_str(carton_v) or nz_str(cbm_v) or nz_str(gw_v)):
            continue

        model_name = get_mapped_cell_value(sheet_data, r, col_map, "model_name_color")
        color_v = get_mapped_with_fallback(sheet_data, r, col_map, "color", layout)
        width_v = get_mapped_with_fallback(sheet_data, r, col_map, "width", layout)
        hs_v = get_mapped_cell_value(sheet_data, r, col_map, "hs_code")

        if layout == "SA_PL":
            model_name = normalize_model_name(model_name)
        if row_type == "TOTAL":
            po_v = "TOTAL"

        rows.append([
            rep_date, rep_inv, rep_dest,
            po_v,
            get_mapped_cell_value(sheet_data, r, col_map, "customer_po"),
            model_name,
            color_v,
            width_v,
            hs_v,
            qty_v,
            carton_v,
            cbm_v,
            gw_v,
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows

def process_one_sub_file_bundle(file_path: str) -> Tuple[List[Any], List[List[Any]], List[List[Any]]]:
    arr = process_one_sub_file(file_path)
    inv_rows: List[List[Any]] = []
    pl_rows: List[List[Any]] = []
    try:
        wb_data = load_workbook_data(file_path)
        pl_name = nz_str(arr[3]) or find_best_pl_sheet(wb_data) or ""
        inv_name = nz_str(arr[4]) or find_best_inv_sheet(wb_data) or ""
        pl_meta, inv_meta = collect_best_meta(wb_data, pl_name, inv_name)
        report_meta = build_report_meta(pl_meta, inv_meta)

        # INV: prefer INV/INVOICE sheets
        inv_candidates = []
        if inv_name and inv_name in wb_data.sheets:
            inv_candidates.append(inv_name)
        for nm in wb_data.sheet_names:
            up = _sheet_name_upper(nm)
            if up in ("INV", "INVOICE") or "INVOICE" in up or up == "INV":
                if nm not in inv_candidates:
                    inv_candidates.append(nm)
        for nm in inv_candidates:
            try:
                inv_rows = extract_inv_detail_rows(file_path, get_folder_name_from_path(file_path), get_parent_folder_created(file_path),
                                                   nm, wb_data.sheets[nm], nz_str(arr[21]).upper(), report_meta)
                if inv_rows:
                    break
            except Exception:
                pass

        # PL: prefer PL sheet, then any PL-like/Sheet1
        pl_candidates = []
        if pl_name and pl_name in wb_data.sheets:
            pl_candidates.append(pl_name)
        for nm in wb_data.sheet_names:
            up = _sheet_name_upper(nm)
            if up.startswith("PL") or up == "SHEET1" or "PACKING" in up:
                if nm not in pl_candidates:
                    pl_candidates.append(nm)
        for nm in pl_candidates:
            try:
                pl_rows = extract_pl_detail_rows(file_path, get_folder_name_from_path(file_path), get_parent_folder_created(file_path),
                                                 nm, wb_data.sheets[nm], nz_str(arr[21]).upper(), report_meta)
                if pl_rows:
                    break
            except Exception:
                pass
    except Exception:
        pass
    return arr, inv_rows, pl_rows

def _processor_run_v43(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None) -> Dict[str, int]:
    repair_options = repair_options or RepairOptions()
    wb_out = open_or_create_output_workbook(output_path)
    try:
        ws_out = ensure_sheet(wb_out, "SUB_DETAIL")
        ws_log = ensure_sheet(wb_out, "LOG_SUB_DETAIL")
        ws_inv = ensure_sheet(wb_out, "INV")
        ws_pl = ensure_sheet(wb_out, "PL")
        if "Sheet" in wb_out.sheetnames and wb_out["Sheet"].max_row == 1 and wb_out["Sheet"].max_column == 1 and nz_str(wb_out["Sheet"]["A1"].value) == "":
            try:
                del wb_out["Sheet"]
            except Exception:
                pass

        init_sub_detail_header(ws_out)
        init_log_header(ws_log)
        init_inv_detail_header(ws_inv)
        init_pl_detail_header(ws_pl)

        kept_ok_rows, rerun_count = keep_only_ok_rows(ws_out)
        processed_paths, processed_signs = build_processed_sets_from_rows(kept_ok_rows)
        ok_paths = set(processed_paths)
        keep_rows_by_file_paths(ws_inv, len(INV_DETAIL_HEADERS), 21, ok_paths)
        keep_rows_by_file_paths(ws_pl, len(PL_DETAIL_HEADERS), 17, ok_paths)

        src_files = collect_source_files(folder_path, output_path)
        cnt_total = len(src_files)
        cnt_ok = cnt_warn = cnt_err = cnt_skip_path = cnt_skip_sig = 0
        cnt_repaired = 0

        log_rows: List[List[Any]] = []
        result_rows: List[List[Any]] = []
        inv_detail_rows: List[List[Any]] = []
        pl_detail_rows: List[List[Any]] = []
        issue_files: List[Tuple[str, str, str]] = []

        if cnt_total == 0:
            log_rows.append(make_log_row("", "SCAN", "INFO", "Khong tim thay file Excel nao de xu ly."))
            append_log_rows(ws_log, log_rows)
            safe_save_workbook_atomic(wb_out, output_path)
            return {"total": 0, "ok": 0, "warning": 0, "error": 0, "skip_path": 0, "skip_signature": 0, "repaired": 0, "issues": [], "scanned_files": []}

        log_rows.append(make_log_row("", "START", "INFO", f"Bat dau quet {cnt_total} file."))
        if rerun_count > 0:
            log_rows.append(make_log_row("", "RERUN", "INFO", f"Se chay lai {rerun_count} file WARNING/ERROR tu lan truoc, dong thoi xu ly file moi phat sinh."))

        def job(path: str):
            arr, invr, plr = process_one_sub_file_bundle(path)
            return path, arr, invr, plr

        futures = {}
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            for fp in src_files:
                fp_norm = fp.strip().upper()
                if fp_norm == output_path.strip().upper():
                    continue
                if fp_norm in processed_paths:
                    cnt_skip_path += 1
                    continue
                futures[executor.submit(job, fp)] = fp

            completed = 0
            total_futures = max(1, len(futures))
            for future in as_completed(futures):
                file_path, arr, inv_rows_one, pl_rows_one = future.result()
                completed += 1
                sig = nz_str(arr[24])
                status_text = nz_str(arr[21]).upper()
                err_text = nz_str(arr[22])

                if sig and sig.upper() in processed_signs:
                    cnt_skip_sig += 1
                    if progress_callback:
                        progress_callback(completed, total_futures, file_path)
                    continue

                if status_text in ("WARNING", "ERROR"):
                    issue_files.append((file_path, status_text, err_text))
                    self.logger.log(f"{status_text}: {get_file_name_from_path(file_path)} -> {err_text}", "warn" if status_text == "WARNING" else "error")
                    log_rows.append(make_log_row(file_path, "CHECK", status_text, err_text))

                if status_text == "OK":
                    cnt_ok += 1
                elif status_text == "WARNING":
                    cnt_warn += 1
                else:
                    cnt_err += 1

                result_rows.append(arr)
                inv_detail_rows.extend(inv_rows_one)
                pl_detail_rows.extend(pl_rows_one)

                if sig:
                    processed_signs.add(sig.upper())
                processed_paths.add(file_path.strip().upper())

                if progress_callback:
                    progress_callback(completed, total_futures, file_path)

        if issue_files:
            log_rows.append(make_log_row("", "SUMMARY", "INFO", "TONG HOP FILE WARNING/ERROR"))
            for fp, st, detail in issue_files:
                log_rows.append(make_log_row(fp, "SUMMARY", st, detail))
        if cnt_skip_path > 0:
            log_rows.append(make_log_row("", "SKIP", "INFO", f"So luong bo qua do da co File Path: {cnt_skip_path}"))
        if cnt_skip_sig > 0:
            log_rows.append(make_log_row("", "SKIP", "INFO", f"So luong bo qua do trung File Signature: {cnt_skip_sig}"))

        log_rows.append(make_log_row("", "DONE", "INFO",
                                     f"Tong={cnt_total}; OK={cnt_ok}; WARNING={cnt_warn}; ERROR={cnt_err}; SkipPath={cnt_skip_path}; SkipSignature={cnt_skip_sig}; Repaired={cnt_repaired}"))

        append_result_rows(ws_out, result_rows)
        append_log_rows(ws_log, log_rows)
        append_generic_rows(ws_inv, inv_detail_rows, [1], 19)
        append_generic_rows(ws_pl, pl_detail_rows, [1], 15)

        autofit_useful(ws_out, len(SUB_DETAIL_HEADERS))
        autofit_useful(ws_log, len(LOG_HEADERS))
        autofit_useful(ws_inv, len(INV_DETAIL_HEADERS))
        autofit_useful(ws_pl, len(PL_DETAIL_HEADERS))
        safe_save_workbook_atomic(wb_out, output_path)

        return {
            "total": cnt_total, "ok": cnt_ok, "warning": cnt_warn, "error": cnt_err,
            "skip_path": cnt_skip_path, "skip_signature": cnt_skip_sig, "repaired": cnt_repaired,
            "issues": issue_files,
        }
    finally:
        wb_out.close()



# ==============================
# V4.4 OVERRIDES - CLEAN LOGIC / REBUILD INV-PL FROM SCRATCH
# ==============================

INV_REPORT_HEADERS = [
    "Date", "Invoice No", "Destination", "PO No", "Customer PO#",
    "DESC", "HS CODE", "MODEL NAME / COLOR", "MODEL Y", "COLOR", "WIDTH",
    "QTY (Pairs)", "UNIT PRICE (USD)", "AMOUNT W/OUT LABEL",
    "LABEL COST", "LABEL AMOUNT", "TOTAL AMOUNT",
    "Row Type", "Status", "File", "File Path"
]

PL_REPORT_HEADERS = [
    "Date", "Invoice No", "Destination", "PO No.", "Customer PO#",
    "MODEL NAME / COLOR", "COLOR", "WIDTH", "HS CODE",
    "QTY (Pairs)", "CARTON QTY", "CBM", "GROSS WEIGHT",
    "Row Type", "Status", "File", "File Path"
]

TOTAL_ROW_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
TOTAL_FONT = Font(bold=True)
HANDLING_ROW_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HANDLING_FONT = Font(bold=True, color="0070C0")

FOOTER_TERMS = [
    "VERY TRULY YOURS", "GENERAL DIRECTOR", "SIGNED BY", "ACCOUNT NAME",
    "ACCOUNT NO", "BANK NAME", "BANK ADDRESS", "SWIFT CODE", "TERMS OF PAYMENT",
    "IEC CODE", "PAN NO", "GST NO"
]

def reset_sheet_with_headers(ws, headers):
    ws.delete_rows(1, ws.max_row)
    for idx, h in enumerate(headers, 1):
        ws.cell(1, idx).value = h
        ws.cell(1, idx).font = HEADER_FONT
        ws.cell(1, idx).fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"

def init_inv_detail_header(ws):
    reset_sheet_with_headers(ws, INV_REPORT_HEADERS)
    for c in (18, 19, 20, 21):
        ws.column_dimensions[get_column_letter(c)].hidden = True

def init_pl_detail_header(ws):
    reset_sheet_with_headers(ws, PL_REPORT_HEADERS)
    for c in (14, 15, 16, 17):
        ws.column_dimensions[get_column_letter(c)].hidden = True

def append_inv_rows(ws, rows):
    append_generic_rows(ws, rows, [1], 19)

def append_pl_rows(ws, rows):
    append_generic_rows(ws, rows, [1], 15)

def is_footer_row_v44(sheet_data, r):
    txt = " | ".join(_clean_upper(get_cell(sheet_data, r, c)) for c in range(1, min(last_used_col_on_row(sheet_data, r), 15)+1))
    return any(term in txt for term in FOOTER_TERMS)

def row_nonempty_values(sheet_data, r, cols):
    vals = []
    for c in cols:
        if c > 0:
            vals.append(get_cell(sheet_data, r, c))
    return vals

def looks_like_blank_or_noise(vals):
    non = [nz_str(v) for v in vals if nz_str(v)]
    if not non:
        return True
    return False

def detect_std_inv_layout(sheet_data):
    for r in range(1, min(40, len(sheet_data))+1):
        row = [_clean_upper(get_cell(sheet_data, r, c)) for c in range(1, min(20, last_used_col_on_row(sheet_data, r))+1)]
        joined = " | ".join(row)
        if "PO NO" in joined and "CUSTOMER PO" in joined and "AMOUNT W/OUT LABEL" in joined:
            return {
                "po":1, "cust":2, "desc":3, "hs":4, "model":5, "model_y":6, "color":7,
                "width":8, "qty":9, "unit":10, "amount":11, "label_cost":12, "label_amt":13, "total":14
            }, r
    return None, 0

def detect_std_pl_layout(sheet_data):
    for r in range(1, min(40, len(sheet_data))+1):
        row = [_clean_upper(get_cell(sheet_data, r, c)) for c in range(1, min(16, last_used_col_on_row(sheet_data, r))+1)]
        joined = " | ".join(row)
        if "PO NO" in joined and "CUSTOMER PO" in joined and "CARTON QTY" in joined:
            return {
                "po":1, "cust":2, "model":3, "color":4, "width":5, "hs":6,
                "qty":7, "carton":8, "cbm":9, "gross":10, "style":12
            }, r
    return None, 0

def detect_sa_inv_layout(sheet_data):
    for r in range(1, min(45, len(sheet_data))+1):
        row = [_clean_upper(get_cell(sheet_data, r, c)) for c in range(1, min(15, last_used_col_on_row(sheet_data, r))+1)]
        joined = " | ".join(row)
        if "PO" in joined and "CUSTOMER PO" in joined and "GROUP SUM MODEL" in joined and "U.PRICE" in joined:
            return {
                "po":1, "cust":2, "width":3, "group_model":4, "mid":5, "style":6, "desc":8, "hs":9,
                "qty":10, "unit":11, "amount":12
            }, r
    return None, 0

def detect_sa_pl_layout(sheet_data):
    for r in range(1, min(20, len(sheet_data))+1):
        row = [_clean_upper(get_cell(sheet_data, r, c)) for c in range(1, min(15, last_used_col_on_row(sheet_data, r))+1)]
        joined = " | ".join(row)
        if "PO" in joined and "CUSTOMER PO" in joined and "PACKAGE" in joined and "MEASUREMENT" in joined:
            return {
                "po":2, "cust":3, "width":4, "style":5, "desc":6, "qty":7, "carton":8, "cbm":9, "gross":11
            }, r
    return None, 0

def norm_meta_date(v):
    return format_date_for_display(v) if nz_str(v) else ""

def get_best_meta_for_file(wb_data, pl_name="", inv_name=""):
    pl_meta, inv_meta = collect_best_meta(wb_data, pl_name, inv_name)
    invoice_no = nz_str(inv_meta.get("invoice_no")) or nz_str(pl_meta.get("invoice_no"))
    invoice_date = norm_meta_date(nz_str(inv_meta.get("invoice_date")) or nz_str(pl_meta.get("invoice_date")))
    destination = nz_str(inv_meta.get("destination")) or nz_str(pl_meta.get("destination"))
    return invoice_no, invoice_date, destination

def parse_standard_inv(sheet_data, meta_date, meta_inv, meta_dest, file_name, file_path, status):
    cmap, hdr = detect_std_inv_layout(sheet_data)
    rows = []
    if not cmap:
        return rows
    for r in range(hdr+1, last_used_row(sheet_data)+1):
        if is_footer_row_v44(sheet_data, r):
            break
        po = nz_str(get_cell(sheet_data, r, cmap["po"]))
        qty = get_cell(sheet_data, r, cmap["qty"])
        amount = get_cell(sheet_data, r, cmap["amount"])
        label_amt = get_cell(sheet_data, r, cmap["label_amt"])
        total_amt = get_cell(sheet_data, r, cmap["total"])
        rowtxt = " ".join(_clean_upper(get_cell(sheet_data, r, c)) for c in range(1, min(5, last_used_col_on_row(sheet_data, r))+1))
        if "TOTAL" in rowtxt:
            rows.append([meta_date, meta_inv, meta_dest, "TOTAL:", "", "", "", "", "", "", "", qty, "", amount, "", label_amt, total_amt, "TOTAL", status, file_name, file_path])
            break
        fee_type = _v76_detect_fee_type(
            *(get_cell(sheet_data, r, c) for c in range(1, min(8, last_used_col_on_row(sheet_data, r)) + 1))
        )
        if fee_type:
            rows.append([meta_date, meta_inv, meta_dest, fee_type, "", "", "", "", "", "", "", "", "", "", "", label_amt, total_amt, fee_type, status, file_name, file_path])
            continue
        if not (po or nz_str(qty) or nz_str(total_amt) or nz_str(label_amt)):
            continue
        rows.append([
            meta_date, meta_inv, meta_dest,
            po, nz_str(get_cell(sheet_data, r, cmap["cust"])),
            nz_str(get_cell(sheet_data, r, cmap["desc"])),
            nz_str(get_cell(sheet_data, r, cmap["hs"])),
            nz_str(get_cell(sheet_data, r, cmap["model"])),
            nz_str(get_cell(sheet_data, r, cmap["model_y"])),
            nz_str(get_cell(sheet_data, r, cmap["color"])),
            nz_str(get_cell(sheet_data, r, cmap["width"])),
            qty,
            get_cell(sheet_data, r, cmap["unit"]),
            amount,
            get_cell(sheet_data, r, cmap["label_cost"]),
            label_amt,
            total_amt,
            "DETAIL", status, file_name, file_path
        ])
    return rows

def parse_sa_inv(sheet_data, meta_date, meta_inv, meta_dest, file_name, file_path, status):
    cmap, hdr = detect_sa_inv_layout(sheet_data)
    rows = []
    if not cmap:
        return rows
    for r in range(hdr+1, last_used_row(sheet_data)+1):
        if is_footer_row_v44(sheet_data, r):
            break
        rowtxt = " ".join(_clean_upper(get_cell(sheet_data, r, c)) for c in range(1, min(6, last_used_col_on_row(sheet_data, r))+1))
        po = nz_str(get_cell(sheet_data, r, cmap["po"]))
        qty = get_cell(sheet_data, r, cmap["qty"])
        amt = get_cell(sheet_data, r, cmap["amount"])
        if "TOTAL" in rowtxt:
            rows.append([meta_date, meta_inv, meta_dest, "TOTAL:", "", "", "", "", "", "", "", qty, "", amt, "", "", amt, "TOTAL", status, file_name, file_path])
            break
        if not (po or nz_str(qty) or nz_str(amt)):
            continue
        model = normalize_model_name(get_cell(sheet_data, r, cmap["group_model"]))
        model_y = nz_str(get_cell(sheet_data, r, cmap["style"]))
        rows.append([
            meta_date, meta_inv, meta_dest,
            po, nz_str(get_cell(sheet_data, r, cmap["cust"])),
            nz_str(get_cell(sheet_data, r, cmap["desc"])),
            nz_str(get_cell(sheet_data, r, cmap["hs"])),
            model,
            model_y,
            "",  # color not available in SA invoice layout
            nz_str(get_cell(sheet_data, r, cmap["width"])),
            qty,
            get_cell(sheet_data, r, cmap["unit"]),
            amt,
            "", "", amt,
            "DETAIL", status, file_name, file_path
        ])
    return rows

def parse_standard_pl(sheet_data, meta_date, meta_inv, meta_dest, file_name, file_path, status):
    cmap, hdr = detect_std_pl_layout(sheet_data)
    rows = []
    if not cmap:
        return rows
    for r in range(hdr+1, last_used_row(sheet_data)+1):
        if is_footer_row_v44(sheet_data, r):
            break
        po = nz_str(get_cell(sheet_data, r, cmap["po"]))
        qty = get_cell(sheet_data, r, cmap["qty"])
        rowtxt = " ".join(_clean_upper(get_cell(sheet_data, r, c)) for c in range(1, min(4, last_used_col_on_row(sheet_data, r))+1))
        if "TOTAL" in rowtxt:
            rows.append([meta_date, meta_inv, meta_dest, "TOTAL", "", "", "", "", "", qty, get_cell(sheet_data, r, cmap["carton"]), get_cell(sheet_data, r, cmap["cbm"]), get_cell(sheet_data, r, cmap["gross"]), "TOTAL", status, file_name, file_path])
            break
        if not (po or nz_str(qty)):
            continue
        rows.append([
            meta_date, meta_inv, meta_dest,
            po, nz_str(get_cell(sheet_data, r, cmap["cust"])),
            nz_str(get_cell(sheet_data, r, cmap["model"])),
            nz_str(get_cell(sheet_data, r, cmap["color"])),
            nz_str(get_cell(sheet_data, r, cmap["width"])),
            nz_str(get_cell(sheet_data, r, cmap["hs"])),
            qty,
            get_cell(sheet_data, r, cmap["carton"]),
            get_cell(sheet_data, r, cmap["cbm"]),
            get_cell(sheet_data, r, cmap["gross"]),
            "DETAIL", status, file_name, file_path
        ])
    return rows

def parse_sa_pl(sheet_data, meta_date, meta_inv, meta_dest, file_name, file_path, status):
    cmap, hdr = detect_sa_pl_layout(sheet_data)
    rows = []
    if not cmap:
        return rows
    for r in range(hdr+1, last_used_row(sheet_data)+1):
        if is_footer_row_v44(sheet_data, r):
            break
        po = nz_str(get_cell(sheet_data, r, cmap["po"]))
        qty = get_cell(sheet_data, r, cmap["qty"])
        if not (po or nz_str(qty)):
            continue
        # total row in SA PL has blank PO and numeric totals
        if not po and nz_str(qty):
            rows.append([meta_date, meta_inv, meta_dest, "TOTAL", "", "", "", "", "", qty, get_cell(sheet_data, r, cmap["carton"]), get_cell(sheet_data, r, cmap["cbm"]), get_cell(sheet_data, r, cmap["gross"]), "TOTAL", status, file_name, file_path])
            break
        rows.append([
            meta_date, meta_inv, meta_dest,
            po, nz_str(get_cell(sheet_data, r, cmap["cust"])),
            nz_str(get_cell(sheet_data, r, cmap["style"])),
            "",  # color not available
            nz_str(get_cell(sheet_data, r, cmap["width"])),
            "",  # hs code not available in SA Sheet1 PL
            qty,
            get_cell(sheet_data, r, cmap["carton"]),
            get_cell(sheet_data, r, cmap["cbm"]),
            get_cell(sheet_data, r, cmap["gross"]),
            "DETAIL", status, file_name, file_path
        ])
    return rows

def extract_detail_rows_for_file_v44(file_path, status):
    wb_data = load_workbook_data(file_path)
    pl_name = find_best_pl_sheet(wb_data) or ""
    inv_name = find_best_inv_sheet(wb_data) or ""
    meta_inv, meta_date, meta_dest = get_best_meta_for_file(wb_data, pl_name, inv_name)
    file_name = get_file_name_from_path(file_path)

    inv_rows = []
    # prefer INV/INVOICE sheet names
    for nm in wb_data.sheet_names:
        up = _sheet_name_upper(nm)
        if up in ("INV", "INVOICE") or "INV" == up or "INVOICE" in up:
            inv_rows = parse_standard_inv(wb_data.sheets[nm], meta_date, meta_inv, meta_dest, file_name, file_path, status)
            if not inv_rows:
                inv_rows = parse_sa_inv(wb_data.sheets[nm], meta_date, meta_inv, meta_dest, file_name, file_path, status)
            if inv_rows:
                break

    pl_rows = []
    # standard PL sheet first
    for nm in wb_data.sheet_names:
        up = _sheet_name_upper(nm)
        if up.startswith("PL"):
            pl_rows = parse_standard_pl(wb_data.sheets[nm], meta_date, meta_inv, meta_dest, file_name, file_path, status)
            if pl_rows:
                break
    if not pl_rows:
        # SA detail often lives in Sheet1
        for nm in wb_data.sheet_names:
            up = _sheet_name_upper(nm)
            if up == "SHEET1":
                pl_rows = parse_sa_pl(wb_data.sheets[nm], meta_date, meta_inv, meta_dest, file_name, file_path, status)
                if pl_rows:
                    break
    return inv_rows, pl_rows

def process_one_sub_file_bundle(file_path: str):
    arr = process_one_sub_file(file_path)
    try:
        inv_rows, pl_rows = extract_detail_rows_for_file_v44(file_path, nz_str(arr[21]).upper())
    except Exception:
        inv_rows, pl_rows = [], []
    return arr, inv_rows, pl_rows

def _processor_run_v44(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None) -> Dict[str, int]:
    repair_options = repair_options or RepairOptions()
    wb_out = open_or_create_output_workbook(output_path)
    try:
        ws_out = ensure_sheet(wb_out, "SUB_DETAIL")
        ws_log = ensure_sheet(wb_out, "LOG_SUB_DETAIL")
        ws_inv = ensure_sheet(wb_out, "INV")
        ws_pl = ensure_sheet(wb_out, "PL")
        if "Sheet" in wb_out.sheetnames and wb_out["Sheet"].max_row == 1 and wb_out["Sheet"].max_column == 1 and nz_str(wb_out["Sheet"]["A1"].value) == "":
            try:
                del wb_out["Sheet"]
            except Exception:
                pass

        init_sub_detail_header(ws_out)
        init_log_header(ws_log)
        init_inv_detail_header(ws_inv)
        init_pl_detail_header(ws_pl)

        kept_ok_rows, rerun_count = keep_only_ok_rows(ws_out)
        kept_ok_actual_paths = [nz_str(arr[23]) for arr in kept_ok_rows if len(arr) >= 24 and nz_str(arr[23])]
        processed_paths, processed_signs = build_processed_sets_from_rows(kept_ok_rows)
        src_files = collect_source_files(folder_path, output_path)
        cnt_total = len(src_files)
        cnt_ok = cnt_warn = cnt_err = cnt_skip_path = cnt_skip_sig = 0

        log_rows=[]; result_rows=[]; issue_files=[]
        inv_detail_rows=[]; pl_detail_rows=[]

        # rebuild INV/PL completely every run from kept OK + processed current files
        reset_sheet_with_headers(ws_inv, INV_REPORT_HEADERS)
        reset_sheet_with_headers(ws_pl, PL_REPORT_HEADERS)

        if cnt_total == 0:
            log_rows.append(make_log_row("", "SCAN", "INFO", "Khong tim thay file Excel nao de xu ly."))
            append_log_rows(ws_log, log_rows)
            safe_save_workbook_atomic(wb_out, output_path)
            return {"total":0,"ok":0,"warning":0,"error":0,"skip_path":0,"skip_signature":0,"repaired":0,"issues":[]}

        log_rows.append(make_log_row("", "START", "INFO", f"Bat dau quet {cnt_total} file."))
        if rerun_count > 0:
            log_rows.append(make_log_row("", "RERUN", "INFO", f"Se chay lai {rerun_count} file WARNING/ERROR tu lan truoc, dong thoi xu ly file moi phat sinh."))

        def job(fp):
            return fp, process_one_sub_file_bundle(fp)

        futures={}
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            for fp in src_files:
                fp_norm = fp.strip().upper()
                if fp_norm == output_path.strip().upper():
                    continue
                if fp_norm in processed_paths:
                    cnt_skip_path += 1
                    continue
                futures[ex.submit(job, fp)] = fp

            completed=0
            total_futures=max(1,len(futures))
            for fut in as_completed(futures):
                fp, bundle = fut.result()
                arr, invr, plr = bundle
                completed += 1
                sig = nz_str(arr[24]).upper()
                status = nz_str(arr[21]).upper()
                err = nz_str(arr[22])

                if sig and sig in processed_signs:
                    cnt_skip_sig += 1
                    if progress_callback: progress_callback(completed, total_futures, fp)
                    continue

                result_rows.append(arr)
                inv_detail_rows.extend(invr)
                pl_detail_rows.extend(plr)

                if status == "OK":
                    cnt_ok += 1
                elif status == "WARNING":
                    cnt_warn += 1
                else:
                    cnt_err += 1
                    if status:
                        issue_files.append((fp, status, err))
                if status == "WARNING":
                    issue_files.append((fp, status, err))
                    log_rows.append(make_log_row(fp, "CHECK", status, err))
                elif status == "ERROR":
                    log_rows.append(make_log_row(fp, "CHECK", status, err))

                processed_paths.add(fp.strip().upper())
                if sig:
                    processed_signs.add(sig)
                if progress_callback: progress_callback(completed, total_futures, fp)

        # rebuild details for kept OK rows too
        for fp in kept_ok_actual_paths:
            if os.path.exists(fp):
                try:
                    invr, plr = extract_detail_rows_for_file_v44(fp, "OK")
                    inv_detail_rows.extend(invr)
                    pl_detail_rows.extend(plr)
                except Exception:
                    pass

        if issue_files:
            log_rows.append(make_log_row("", "SUMMARY", "INFO", "TONG HOP FILE WARNING/ERROR"))
            for fp, st, detail in issue_files:
                log_rows.append(make_log_row(fp, "SUMMARY", st, detail))
        if cnt_skip_path:
            log_rows.append(make_log_row("", "SKIP", "INFO", f"So luong bo qua do da co File Path: {cnt_skip_path}"))
        if cnt_skip_sig:
            log_rows.append(make_log_row("", "SKIP", "INFO", f"So luong bo qua do trung File Signature: {cnt_skip_sig}"))
        log_rows.append(make_log_row("", "DONE", "INFO", f"Tong={cnt_total}; OK={cnt_ok}; WARNING={cnt_warn}; ERROR={cnt_err}; SkipPath={cnt_skip_path}; SkipSignature={cnt_skip_sig}"))

        append_result_rows(ws_out, result_rows)
        append_log_rows(ws_log, log_rows)
        append_inv_rows(ws_inv, inv_detail_rows)
        append_pl_rows(ws_pl, pl_detail_rows)
        autofit_useful(ws_out, len(SUB_DETAIL_HEADERS))
        autofit_useful(ws_log, len(LOG_HEADERS))
        autofit_useful(ws_inv, len(INV_REPORT_HEADERS))
        autofit_useful(ws_pl, len(PL_REPORT_HEADERS))
        safe_save_workbook_atomic(wb_out, output_path)
        return {"total":cnt_total,"ok":cnt_ok,"warning":cnt_warn,"error":cnt_err,"skip_path":cnt_skip_path,"skip_signature":cnt_skip_sig,"repaired":0,"issues":issue_files,"scanned_files":src_files}
    finally:
        wb_out.close()



# =========================================================
# V6 CLEAN-ROOM DETAIL ENGINE
# Keep SUB_DETAIL + LOG_SUB_DETAIL core stable
# Rebuild detail sheets from scratch every run:
#   INV, PL           -> SUB / VSF
#   INV Others, PL Others -> SA / Copy of SA
# =========================================================

CORE_RUN_V6 = Processor.run

V6_INV_HEADERS = [
    "Date", "Invoice No", "Destination", "PO No", "Customer PO#",
    "DESC", "HS CODE", "MODEL NAME / COLOR", "MODEL Y", "COLOR", "WIDTH",
    "Q'TY (Pairs)", "UNIT PRICE (USD)", "AMOUNT W/OUT LABEL",
    "LABEL COST", "LABEL AMOUNT", "TOTAL AMOUNT",
    "Row Type", "Status", "File", "File Path"
]

V6_PL_HEADERS = [
    "Date", "Invoice No", "Destination", "PO No", "Customer PO#",
    "MODEL NAME / COLOR", "COLOR", "WIDTH", "HS CODE",
    "Q'TY (Pairs)", "CARTON Q'TY", "CBM", "GROSS WEIGHT",
    "Row Type", "Status", "File", "File Path"
]

V76_PL_HEADERS = [
    "Date", "Invoice No", "Destination", "PO No", "Customer PO#",
    "MODEL NAME / COLOR", "COLOR", "WIDTH", "HS CODE",
    "Q'TY (Pairs)", "CARTON Q'TY", "CBM", "GROSS WEIGHT",
    "S. Invoice#", "S.DateWInv#", "Row Type", "Status", "File", "File Path"
]

V6_TOTAL_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
V6_TOTAL_FONT = Font(bold=True)
V6_HANDLING_FILL = PatternFill(fill_type="solid", fgColor="DDEBF7")
V6_HANDLING_FONT = Font(bold=True)

def _v6_is_sa_family(file_path: str) -> bool:
    nm = get_file_name_from_path(file_path).strip().upper()
    return nm.startswith("SA") or nm.startswith("COPY OF SA")

def _v6_uc(v: Any) -> str:
    return normalize_simple_text(v).upper().replace(":", "").replace("#", "")

def _v6_init_sheet(ws, headers: List[str]):
    ws.delete_rows(1, ws.max_row)
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i).value = sanitize_excel_text(h)
        ws.cell(1, i).font = HEADER_FONT
        ws.cell(1, i).fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"

def _v6_append_rows(ws, rows: List[List[Any]], date_cols: List[int], row_type_idx: int, status_idx: int):
    if not rows:
        return
    start_row = 2
    for r_idx, arr in enumerate(rows, start=start_row):
        for c, val in enumerate(arr, start=1):
            cell = ws.cell(r_idx, c)
            cell.value = sanitize_excel_text(val)
            if c in date_cols:
                cell.number_format = "dd-mmm"
            if ws.title.startswith("INV"):
                if c == 12:
                    cell.number_format = '#,##0'
                elif c in (13, 14, 16, 17):
                    cell.number_format = '#,##0.00'
                elif c == 15:
                    cell.number_format = '0.0000'
            else:
                if c in (10, 11):
                    cell.number_format = '#,##0'
                elif c in (12, 13):
                    cell.number_format = '#,##0.000'
        row_type = nz_str(arr[row_type_idx - 1]).upper()
        status = nz_str(arr[status_idx - 1]).upper()
        if row_type == "TOTAL":
            for c in range(1, len(arr) + 1):
                ws.cell(r_idx, c).fill = V6_TOTAL_FILL
                ws.cell(r_idx, c).font = V6_TOTAL_FONT
        elif row_type == "HANDLING CHARGE":
            for c in range(1, len(arr) + 1):
                ws.cell(r_idx, c).fill = V6_HANDLING_FILL
                ws.cell(r_idx, c).font = V6_HANDLING_FONT
        elif status in STATUS_FILLS:
            fill = STATUS_FILLS[status]
            for c in range(1, len(arr) + 1):
                ws.cell(r_idx, c).fill = fill

def _v6_row_has_any(sheet_data: List[List[Any]], r: int, cols: List[int]) -> bool:
    for c in cols:
        if c > 0 and nz_str(get_cell(sheet_data, r, c)):
            return True
    return False

def _v6_is_footer_row(sheet_data: List[List[Any]], r: int) -> bool:
    probes = []
    maxc = min(20, last_used_col_on_row(sheet_data, r))
    for c in range(1, maxc + 1):
        probes.append(_v6_uc(get_cell(sheet_data, r, c)))
    full = " | ".join([p for p in probes if p])
    bad_terms = [
        "VERY TRULY YOURS", "GENERAL DIRECTOR", "SIGNED BY",
        "ACCOUNT NAME", "ACCOUNT NO", "BANK NAME", "BANK ADDRESS",
        "SWIFT CODE"
    ]
    return any(x in full for x in bad_terms)

def _v6_find_total_row(sheet_data: List[List[Any]], start_row: int, probe_cols: List[int]) -> int:
    for r in range(start_row, last_used_row(sheet_data) + 1):
        if _v6_is_footer_row(sheet_data, r):
            return max(start_row, r - 1)
        probe = " | ".join(_v6_uc(get_cell(sheet_data, r, c)) for c in probe_cols if c > 0)
        if "TOTAL" in probe:
            return r
    return last_used_row(sheet_data)

def _v6_collect_meta(wb_data: WorkbookData, inv_sheet_name: str, pl_sheet_name: str) -> Dict[str, str]:
    blank = {k: "" for k in META_KEYS}
    inv_meta = blank.copy()
    pl_meta = blank.copy()

    inv_candidates = []
    pl_candidates = []

    if inv_sheet_name and inv_sheet_name in wb_data.sheets:
        inv_candidates.append(inv_sheet_name)
    if pl_sheet_name and pl_sheet_name in wb_data.sheets:
        pl_candidates.append(pl_sheet_name)

    for nm in wb_data.sheet_names:
        up = nm.strip().upper()
        if up in ("INV", "INVOICE") or "INVOICE" in up:
            if nm not in inv_candidates:
                inv_candidates.append(nm)
        if up.startswith("PL") or up == "SHEET1":
            if nm not in pl_candidates:
                pl_candidates.append(nm)

    for nm in inv_candidates:
        meta = extract_sheet_meta(wb_data.sheets[nm])
        for k in META_KEYS:
            if not inv_meta[k] and nz_str(meta.get(k, "")):
                inv_meta[k] = nz_str(meta.get(k, ""))
    for nm in pl_candidates:
        meta = extract_sheet_meta(wb_data.sheets[nm])
        for k in META_KEYS:
            if not pl_meta[k] and nz_str(meta.get(k, "")):
                pl_meta[k] = nz_str(meta.get(k, ""))

    return {
        "invoice_no": nz_str(inv_meta.get("invoice_no", "")) or nz_str(pl_meta.get("invoice_no", "")),
        "invoice_date": nz_str(inv_meta.get("invoice_date", "")) or nz_str(pl_meta.get("invoice_date", "")),
        "destination": nz_str(inv_meta.get("destination", "")) or nz_str(pl_meta.get("destination", "")),
    }

# ---------- SUB / VSF parsers ----------

def _v6_detect_sub_vsf_inv_header(sheet_data: List[List[Any]]) -> Tuple[Dict[str, int], int]:
    best_map: Dict[str, int] = {}
    best_row = 0
    best_score = -1
    for r in range(1, min(60, len(sheet_data)) + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        for c in range(1, min(last_col, 40) + 1):
            t = normalize_header(get_cell(sheet_data, r, c))
            if t in ("PONO", "PO") and "po_no" not in tmp:
                tmp["po_no"] = c
            elif ("CUSTOMERPO" in t or "CUSPO" in t) and "customer_po" not in tmp:
                tmp["customer_po"] = c
            elif t in ("DESC", "DESCRIPTION") and "desc" not in tmp:
                tmp["desc"] = c
            elif "HSCODE" in t and "hs_code" not in tmp:
                tmp["hs_code"] = c
            elif "MODELNAMECOLOR" in t and "model_name" not in tmp:
                tmp["model_name"] = c
            elif t == "WIDTH" and "width" not in tmp:
                tmp["width"] = c
            elif ("QTY" in t and "PAIR" in t) and "qty_pairs" not in tmp:
                tmp["qty_pairs"] = c
            elif "UNITPRICE" in t and "USD" in t and "unit_price" not in tmp:
                tmp["unit_price"] = c
            elif ("AMOUNTWOUTLABEL" in t or "AMOUNTWITHOUTLABEL" in t or "AMOUNTWOUTLABELCOST" in t) and "amount_wo" not in tmp:
                tmp["amount_wo"] = c
            elif t == "LABELCOST" and "label_cost" not in tmp:
                tmp["label_cost"] = c
            elif ("TOTALLABELCOST" in t or "LABELAMOUNT" in t) and "label_amount" not in tmp:
                tmp["label_amount"] = c
            elif "TOTALAMOUNT" in t and "total_amount" not in tmp:
                tmp["total_amount"] = c
        score = len(tmp)
        if score > best_score:
            best_map = tmp
            best_row = r
            best_score = score
    if best_score < 8:
        raise ValueError("Khong tim thay header INV cho SUB/VSF")
    return best_map, best_row

def _v6_detect_sub_vsf_pl_header(sheet_data: List[List[Any]]) -> Tuple[Dict[str, int], int]:
    best_map: Dict[str, int] = {}
    best_row = 0
    best_score = -1
    for r in range(1, min(60, len(sheet_data)) + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        for c in range(1, min(last_col, 25) + 1):
            t = normalize_header(get_cell(sheet_data, r, c))
            if t in ("PONO", "PO") and "po_no" not in tmp:
                tmp["po_no"] = c
            elif ("CUSTOMERPO" in t or "CUSPO" in t) and "customer_po" not in tmp:
                tmp["customer_po"] = c
            elif "MODELNAMECOLOR" in t and "model_name" not in tmp:
                tmp["model_name"] = c
            elif t == "WIDTH" and "width" not in tmp:
                tmp["width"] = c
            elif "HSCODE" in t and "hs_code" not in tmp:
                tmp["hs_code"] = c
            elif ("QTY" in t and "PAIR" in t) and "qty_pairs" not in tmp:
                tmp["qty_pairs"] = c
            elif ("CARTON" in t and "QTY" in t) and "carton_qty" not in tmp:
                tmp["carton_qty"] = c
            elif t == "CBM" and "cbm" not in tmp:
                tmp["cbm"] = c
            elif "GROSSWEIGHT" in t and "gross_weight" not in tmp:
                tmp["gross_weight"] = c
        score = len(tmp)
        if score > best_score:
            best_map = tmp
            best_row = r
            best_score = score
    if best_score < 7:
        raise ValueError("Khong tim thay header PL cho SUB/VSF")
    return best_map, best_row

def _v6_extract_sub_vsf_inv(file_path: str, wb_data: WorkbookData, report_meta: Dict[str, str], status_text: str) -> List[List[Any]]:
    if "INV" not in wb_data.sheets:
        return []
    sheet = wb_data.sheets["INV"]
    col_map, hdr_row = _v6_detect_sub_vsf_inv_header(sheet)
    total_row = _v6_find_total_row(sheet, hdr_row + 1, [col_map.get("po_no", 0), 1])
    keep_cols = list(col_map.values())

    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))
    rows: List[List[Any]] = []

    for r in range(hdr_row + 1, total_row + 1):
        if _v6_is_footer_row(sheet, r):
            break
        if not _v6_row_has_any(sheet, r, keep_cols):
            continue

        po_raw = get_cell(sheet, r, col_map.get("po_no", 0))
        fee_type = _v76_detect_fee_type(
            po_raw,
            get_cell(sheet, r, col_map.get("customer_po", 0)),
            get_cell(sheet, r, col_map.get("desc", 0)),
            get_cell(sheet, r, col_map.get("model_name", 0)),
            get_cell(sheet, r, 1),
        )
        row_type = "DETAIL"
        if fee_type:
            row_type = fee_type
        elif "TOTAL" in _v6_uc(po_raw):
            row_type = "TOTAL"

        po_no = get_mapped_cell_value(sheet, r, col_map, "po_no")
        customer_po = get_mapped_cell_value(sheet, r, col_map, "customer_po")
        desc = get_mapped_cell_value(sheet, r, col_map, "desc")
        hs_code = get_mapped_cell_value(sheet, r, col_map, "hs_code")
        model_name = get_mapped_cell_value(sheet, r, col_map, "model_name")
        model_col = col_map.get("model_name", 0)
        model_y = get_cell(sheet, r, model_col + 1) if model_col > 0 else ""
        color = get_cell(sheet, r, model_col + 2) if model_col > 0 else ""
        width = get_mapped_cell_value(sheet, r, col_map, "width")
        qty_pairs = get_mapped_cell_value(sheet, r, col_map, "qty_pairs")
        unit_price = get_mapped_cell_value(sheet, r, col_map, "unit_price")
        amount_wo = get_mapped_cell_value(sheet, r, col_map, "amount_wo")
        label_cost = get_mapped_cell_value(sheet, r, col_map, "label_cost")
        label_amount = get_mapped_cell_value(sheet, r, col_map, "label_amount")
        total_amount = get_mapped_cell_value(sheet, r, col_map, "total_amount")

        if row_type == "DETAIL" and not (nz_str(po_no) or nz_str(qty_pairs) or nz_str(total_amount)):
            continue
        if row_type in ("HANDLING CHARGE", "UPCHARGE", "DEDUCT"):
            po_no = row_type
        elif row_type == "TOTAL":
            po_no = "TOTAL:"

        rows.append([
            rep_date, rep_inv, rep_dest,
            po_no, customer_po, desc, hs_code, model_name, model_y, color, width,
            qty_pairs, unit_price, amount_wo, label_cost, label_amount, total_amount,
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows

def _v6_extract_sub_vsf_pl(file_path: str, wb_data: WorkbookData, report_meta: Dict[str, str], status_text: str) -> List[List[Any]]:
    pl_sheet_name = ""
    for nm in wb_data.sheet_names:
        if nm.strip().upper().startswith("PL"):
            pl_sheet_name = nm
            break
    if not pl_sheet_name:
        return []
    sheet = wb_data.sheets[pl_sheet_name]
    col_map, hdr_row = _v6_detect_sub_vsf_pl_header(sheet)
    total_row = _v6_find_total_row(sheet, hdr_row + 1, [col_map.get("po_no", 0), 1])
    keep_cols = list(col_map.values())

    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))
    rows: List[List[Any]] = []

    for r in range(hdr_row + 1, total_row + 1):
        if _v6_is_footer_row(sheet, r):
            break
        if not _v6_row_has_any(sheet, r, keep_cols):
            continue

        po_raw = _v6_uc(get_cell(sheet, r, col_map.get("po_no", 0)))
        row_type = "TOTAL" if "TOTAL" in po_raw else "DETAIL"

        po_no = get_mapped_cell_value(sheet, r, col_map, "po_no")
        customer_po = get_mapped_cell_value(sheet, r, col_map, "customer_po")
        model_name = get_mapped_cell_value(sheet, r, col_map, "model_name")
        model_col = col_map.get("model_name", 0)
        color = get_cell(sheet, r, model_col + 1) if model_col > 0 else ""
        width = get_mapped_cell_value(sheet, r, col_map, "width")
        hs_code = get_mapped_cell_value(sheet, r, col_map, "hs_code")
        qty_pairs = get_mapped_cell_value(sheet, r, col_map, "qty_pairs")
        carton_qty = get_mapped_cell_value(sheet, r, col_map, "carton_qty")
        cbm = get_mapped_cell_value(sheet, r, col_map, "cbm")
        gross_weight = get_mapped_cell_value(sheet, r, col_map, "gross_weight")

        if row_type == "DETAIL" and not (nz_str(po_no) or nz_str(qty_pairs) or nz_str(cbm)):
            continue
        if row_type == "TOTAL":
            po_no = "TOTAL"

        rows.append([
            rep_date, rep_inv, rep_dest,
            po_no, customer_po, model_name, color, width, hs_code,
            qty_pairs, carton_qty, cbm, gross_weight,
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows

# ---------- SA parsers ----------



def _v6_detect_sa_pl_header(sheet_data: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    best_row = 0
    best_map: Dict[str, int] = {}
    best_score = -1
    for r in range(1, min(80, len(sheet_data)) + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        for c in range(1, min(last_col, 30) + 1):
            t = normalize_header(get_cell(sheet_data, r, c))
            if t in ("PO", "PO#", "PONO") and "po_no" not in tmp:
                tmp["po_no"] = c
            elif ("CUSTOMERPO" in t or "CUSPO" in t) and "customer_po" not in tmp:
                tmp["customer_po"] = c
            elif ("GROUPSUMMODEL" in t or "GROUPMODEL" in t or "MODELNAMECOLOR" in t) and "model_name" not in tmp:
                tmp["model_name"] = c
            elif t == "STYLE" and "style" not in tmp:
                tmp["style"] = c
            elif "WIDTH" in t and "width" not in tmp:
                tmp["width"] = c
            elif "HSCODE" in t and "hs_code" not in tmp:
                tmp["hs_code"] = c
            elif (("QTY" in t and "PRS" in t) or ("QTY" in t and "PAIR" in t) or t == "QTY") and "qty_pairs" not in tmp:
                tmp["qty_pairs"] = c
            elif (("PACKAGE" in t or "CARTON" in t) and ("CTN" in t or "QTY" in t)) and "carton_qty" not in tmp:
                tmp["carton_qty"] = c
            elif (("MEASUREMENT" in t and "CBM" in t) or t == "CBM") and "cbm" not in tmp:
                tmp["cbm"] = c
            elif (("GROSS" in t and "WEIGHT" in t) or "GW" in t or "GWKG" in t) and "gross_weight" not in tmp:
                tmp["gross_weight"] = c
        score = len(tmp)
        row_text = " | ".join(_v6_uc(get_cell(sheet_data, r, c)) for c in range(1, min(last_col, 20) + 1))
        if "GROUP SUM/MODEL" in row_text or "GROUP SUM MODEL" in row_text:
            score += 1
        if "PACKAGE" in row_text and "MEASUREMENT" in row_text:
            score += 1
        if score > best_score:
            best_score = score
            best_row = r
            best_map = tmp
    if best_score < 7:
        raise ValueError("Khong tim thay header PL cho SA")
    return best_row, best_map
def _v6_extract_sa_inv(file_path: str, wb_data: WorkbookData, report_meta: Dict[str, str], status_text: str) -> List[List[Any]]:
    if "INVOICE" not in wb_data.sheets:
        return []
    sheet = wb_data.sheets["INVOICE"]
    hdr_row = 0
    for r in range(1, min(80, len(sheet)) + 1):
        row_text = " | ".join(_v6_uc(get_cell(sheet, r, c)) for c in range(1, min(20, last_used_col_on_row(sheet, r)) + 1))
        if "PO" in row_text and "CUSTOMER PO" in row_text and "GROUP SUM MODEL" in row_text and "U PRICE" in row_text:
            hdr_row = r
            break
    if hdr_row == 0:
        return []

    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))
    total_row = _v6_find_total_row(sheet, hdr_row + 1, [1])

    rows: List[List[Any]] = []
    for r in range(hdr_row + 1, total_row + 1):
        if _v6_is_footer_row(sheet, r):
            break
        po_txt = _v6_uc(get_cell(sheet, r, 1))
        qty_v = get_cell(sheet, r, 10)
        amt_v = get_cell(sheet, r, 12)
        if not po_txt and not nz_str(qty_v) and not nz_str(amt_v):
            continue
        row_type = "TOTAL" if "TOTAL" in po_txt else "DETAIL"
        po_no = get_cell(sheet, r, 1)
        customer_po = get_cell(sheet, r, 2)
        width = get_cell(sheet, r, 3)
        model_name = normalize_model_name(get_cell(sheet, r, 4))
        model_y = get_cell(sheet, r, 6)   # STYLE
        desc = get_cell(sheet, r, 8)      # TYPE/CATEGORY
        hs_code = get_cell(sheet, r, 9)
        qty_pairs = qty_v
        unit_price = get_cell(sheet, r, 11)
        amount_wo = amt_v

        if row_type == "TOTAL":
            po_no = "TOTAL:"
        rows.append([
            rep_date, rep_inv, rep_dest,
            po_no, customer_po, desc, hs_code, model_name, model_y, "", width,
            qty_pairs, unit_price, amount_wo, "", "", amount_wo,
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows

def _v6_extract_sa_pl(file_path: str, wb_data: WorkbookData, report_meta: Dict[str, str], status_text: str) -> List[List[Any]]:
    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))

    # Prefer real PL sheet for SA files
    pl_sheet_name = ""
    for nm in wb_data.sheet_names:
        if nm.strip().upper().startswith("PL"):
            pl_sheet_name = nm
            break
    if pl_sheet_name:
        try:
            sheet = wb_data.sheets[pl_sheet_name]
            hdr_row, col_map = _v6_detect_sa_pl_header(sheet)
            rows: List[List[Any]] = []
            for r in range(hdr_row + 1, last_used_row(sheet) + 1):
                if _v6_is_footer_row(sheet, r):
                    break
                po_txt = _v6_uc(get_cell(sheet, r, col_map.get("po_no", 0)))
                qty_v = get_mapped_cell_value(sheet, r, col_map, "qty_pairs")
                cbm_v = get_mapped_cell_value(sheet, r, col_map, "cbm")
                gw_v = get_mapped_cell_value(sheet, r, col_map, "gross_weight")
                if not po_txt and not nz_str(qty_v) and not nz_str(cbm_v) and not nz_str(gw_v):
                    continue
                row_type = "TOTAL" if "TOTAL" in po_txt else "DETAIL"
                po_no = get_mapped_cell_value(sheet, r, col_map, "po_no")
                customer_po = get_mapped_cell_value(sheet, r, col_map, "customer_po")
                model_name = normalize_model_name(get_mapped_cell_value(sheet, r, col_map, "model_name"))
                color = get_mapped_cell_value(sheet, r, col_map, "style")
                width = get_mapped_cell_value(sheet, r, col_map, "width")
                hs_code = get_mapped_cell_value(sheet, r, col_map, "hs_code")
                qty_pairs = qty_v
                carton_qty = get_mapped_cell_value(sheet, r, col_map, "carton_qty")
                cbm = cbm_v
                gross_weight = gw_v
                if row_type == "DETAIL" and not (nz_str(po_no) or nz_str(qty_pairs) or nz_str(cbm)):
                    continue
                if row_type == "TOTAL":
                    po_no = "TOTAL"
                rows.append([
                    rep_date, rep_inv, rep_dest,
                    po_no, customer_po, model_name, color, width, hs_code,
                    qty_pairs, carton_qty, cbm, gross_weight,
                    row_type, status_text, get_file_name_from_path(file_path), file_path
                ])
            if rows:
                return rows
        except Exception:
            pass

    # Fallback: legacy SA layout in Sheet1
    if "Sheet1" not in wb_data.sheets:
        return []
    sheet = wb_data.sheets["Sheet1"]
    hdr_row = 0
    for r in range(1, min(20, len(sheet)) + 1):
        row_text = " | ".join(_v6_uc(get_cell(sheet, r, c)) for c in range(1, min(20, last_used_col_on_row(sheet, r)) + 1))
        if "PO" in row_text and "CUSTOMER PO" in row_text and "GENDER SIZE" in row_text and "PACKAGE" in row_text and "MEASUREMENT" in row_text:
            hdr_row = r
            break
    if hdr_row == 0:
        return []

    rows: List[List[Any]] = []
    for r in range(hdr_row + 1, last_used_row(sheet) + 1):
        if _v6_is_footer_row(sheet, r):
            break
        po_txt = _v6_uc(get_cell(sheet, r, 2))
        qty_v = get_cell(sheet, r, 7)
        cbm_v = get_cell(sheet, r, 9)
        gw_v = get_cell(sheet, r, 11)
        if not po_txt and not nz_str(qty_v) and not nz_str(cbm_v) and not nz_str(gw_v):
            continue
        row_type = "TOTAL" if "TOTAL" in po_txt else "DETAIL"
        po_no = get_cell(sheet, r, 2)
        customer_po = get_cell(sheet, r, 3)
        model_name = normalize_model_name(get_cell(sheet, r, 5))
        color = get_cell(sheet, r, 6)
        width = get_cell(sheet, r, 4)
        hs_code = get_cell(sheet, r, 8)
        qty_pairs = qty_v
        carton_qty = get_cell(sheet, r, 8)
        cbm = cbm_v
        gross_weight = gw_v
        if row_type == "DETAIL" and not (nz_str(po_no) or nz_str(qty_pairs) or nz_str(cbm)):
            continue
        if row_type == "TOTAL":
            po_no = "TOTAL"
        rows.append([
            rep_date, rep_inv, rep_dest,
            po_no, customer_po, model_name, color, width, hs_code,
            qty_pairs, carton_qty, cbm, gross_weight,
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows

def _v6_build_detail_bundle(file_path: str) -> Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]:
    wb_data = load_workbook_data(file_path)

    inv_sheet_name = "INVOICE" if _v6_is_sa_family(file_path) and "INVOICE" in wb_data.sheets else ("INV" if "INV" in wb_data.sheets else "")
    pl_sheet_name = ""
    if _v6_is_sa_family(file_path):
        for nm in wb_data.sheet_names:
            if nm.strip().upper().startswith("PL"):
                pl_sheet_name = nm
                break
        if not pl_sheet_name and "Sheet1" in wb_data.sheets:
            pl_sheet_name = "Sheet1"
    else:
        for nm in wb_data.sheet_names:
            if nm.strip().upper().startswith("PL"):
                pl_sheet_name = nm
                break

    report_meta = _v6_collect_meta(wb_data, inv_sheet_name, pl_sheet_name)
    arr = process_one_sub_file(file_path)
    status_text = nz_str(arr[21]).upper()

    inv_rows: List[List[Any]] = []
    pl_rows: List[List[Any]] = []
    inv_other_rows: List[List[Any]] = []
    pl_other_rows: List[List[Any]] = []

    if _v6_is_sa_family(file_path):
        inv_other_rows = _v6_extract_sa_inv(file_path, wb_data, report_meta, status_text)
        # SA family is a separate flow:
        # - PL: prefer real PL sheet with SA header
        # - fallback to Sheet1 only inside _v6_extract_sa_pl if real PL cannot be parsed
        pl_other_rows = _v6_extract_sa_pl(file_path, wb_data, report_meta, status_text)
    else:
        inv_rows = _v6_extract_sub_vsf_inv(file_path, wb_data, report_meta, status_text)
        pl_rows = _v6_extract_sub_vsf_pl(file_path, wb_data, report_meta, status_text)

    return inv_rows, pl_rows, inv_other_rows, pl_other_rows

def _v6_rebuild_detail_sheets(self, folder_path: str, output_path: str):
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    try:
        for nm in ["INV", "PL", "INV Others", "PL Others"]:
            if nm not in wb.sheetnames:
                wb.create_sheet(nm)

        ws_inv = wb["INV"]
        ws_pl = wb["PL"]
        ws_inv_oth = wb["INV Others"]
        ws_pl_oth = wb["PL Others"]

        _v6_init_sheet(ws_inv, V6_INV_HEADERS)
        _v6_init_sheet(ws_pl, V6_PL_HEADERS)
        _v6_init_sheet(ws_inv_oth, V6_INV_HEADERS)
        _v6_init_sheet(ws_pl_oth, V6_PL_HEADERS)

        src_files = collect_source_files(folder_path, output_path)
        inv_rows_all: List[List[Any]] = []
        pl_rows_all: List[List[Any]] = []
        inv_other_rows_all: List[List[Any]] = []
        pl_other_rows_all: List[List[Any]] = []

        for fp in src_files:
            if fp.strip().upper() == output_path.strip().upper():
                continue
            try:
                inv_rows, pl_rows, inv_other_rows, pl_other_rows = _v6_build_detail_bundle(fp)
                inv_rows_all.extend(inv_rows)
                pl_rows_all.extend(pl_rows)
                inv_other_rows_all.extend(inv_other_rows)
                pl_other_rows_all.extend(pl_other_rows)
            except Exception as exc:
                errors.append((fp, str(exc)))

        _v6_append_rows(ws_inv, inv_rows_all, [1], 18, 19)
        _v6_append_rows(ws_pl, pl_rows_all, [1], 14, 15)
        _v6_append_rows(ws_inv_oth, inv_other_rows_all, [1], 18, 19)
        _v6_append_rows(ws_pl_oth, pl_other_rows_all, [1], 14, 15)

        autofit_useful(ws_inv, len(V6_INV_HEADERS))
        autofit_useful(ws_pl, len(V6_PL_HEADERS))
        autofit_useful(ws_inv_oth, len(V6_INV_HEADERS))
        autofit_useful(ws_pl_oth, len(V6_PL_HEADERS))

        # append explicit rebuild log
        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_REBUILD", "INFO",
                            f"Rebuilt detail tabs: INV={len(inv_rows_all)}, PL={len(pl_rows_all)}, INV Others={len(inv_other_rows_all)}, PL Others={len(pl_other_rows_all)}")])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, "DETAIL_REBUILD", "ERROR", err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

def _v6_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
    _v6_rebuild_detail_sheets(self, folder_path, output_path)
    return summary


# =========================================================
# V7.2 PATCH
# - based on V6 clean room BOOTFIX source
# - SA -> INVOICE -> INV Others parser rieng
# - chi to TOTAL, KHONG to HANDLING CHARGE
# - single entrypoint at end only
# =========================================================

V72_TOTAL_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
V72_TOTAL_FONT = Font(bold=True)

def _v72_init_sheet(ws, headers: List[str]):
    ws.delete_rows(1, ws.max_row)
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i).value = sanitize_excel_text(h)
        ws.cell(1, i).font = HEADER_FONT
        ws.cell(1, i).fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"

def _v72_append_rows(ws, rows: List[List[Any]], date_cols: List[int], row_type_idx: int, status_idx: int):
    if not rows:
        return
    for r_idx, arr in enumerate(rows, start=2):
        for c, val in enumerate(arr, start=1):
            cell = ws.cell(r_idx, c)
            cell.value = sanitize_excel_text(val)
            if c in date_cols:
                cell.number_format = "dd-mmm"
            if ws.title.startswith("INV"):
                if c == 12:
                    cell.number_format = '#,##0'
                elif c in (13, 14, 16, 17):
                    cell.number_format = '#,##0.00'
                elif c == 15:
                    cell.number_format = '0.0000'
            else:
                if c in (10, 11):
                    cell.number_format = '#,##0'
                elif c in (12, 13):
                    cell.number_format = '#,##0.000'
        row_type = nz_str(arr[row_type_idx - 1]).upper()
        status = nz_str(arr[status_idx - 1]).upper()
        if row_type == "TOTAL":
            for c in range(1, len(arr) + 1):
                ws.cell(r_idx, c).fill = V72_TOTAL_FILL
                ws.cell(r_idx, c).font = V72_TOTAL_FONT
        elif status in STATUS_FILLS:
            fill = STATUS_FILLS[status]
            for c in range(1, len(arr) + 1):
                ws.cell(r_idx, c).fill = fill

def _v72_detect_sa_invoice_header(sheet_data: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    best_row = 0
    best_map: Dict[str, int] = {}
    best_score = -1
    for r in range(1, min(80, len(sheet_data)) + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        for c in range(1, min(last_col, 25) + 1):
            t = normalize_header(get_cell(sheet_data, r, c))
            if t in ("PO", "PONO") and "po_no" not in tmp:
                tmp["po_no"] = c
            elif "CUSTOMERPO" in t or "CUSPO" in t:
                tmp["customer_po"] = c
            elif "TYPECATEGORY" in t or t in ("TYPE", "CATEGORY"):
                tmp["desc"] = c
            elif "HSCODE" in t:
                tmp["hs_code"] = c
            elif "GROUPSUMMODEL" in t or "GROUPMODEL" in t or "GROUPSUM" in t:
                tmp["model_name"] = c
            elif "WIDTH" in t:
                tmp["width"] = c
            elif ("QTY" in t and "PAIR" in t) or t == "QTY":
                tmp["qty_pairs"] = c
            elif "UPRICE" in t or ("UNITPRICE" in t and "USD" in t):
                tmp["unit_price"] = c
            elif "AMOUNT" in t and "USD" in t:
                tmp["amount_usd"] = c
        score = len(tmp)
        row_text = " | ".join(_v6_uc(get_cell(sheet_data, r, c)) for c in range(1, min(last_col, 20) + 1))
        if "GROUP SUM MODEL" in row_text or "GROUP SUM/MODEL" in row_text:
            score += 2
        if "TYPE CATEGORY" in row_text or "TYPE/ CATEGORY" in row_text:
            score += 1
        if "U PRICE" in row_text or "U.PRICE" in row_text or "UPRICE" in row_text:
            score += 1
        if score > best_score:
            best_score = score
            best_row = r
            best_map = tmp
    if best_score < 7:
        raise ValueError("Khong tim thay header SA-INVOICE hop le.")
    return best_row, best_map

def _v72_extract_sa_inv(file_path: str, wb_data: WorkbookData, report_meta: Dict[str, str], status_text: str) -> List[List[Any]]:
    if "INVOICE" not in wb_data.sheets:
        return []
    sheet = wb_data.sheets["INVOICE"]
    hdr_row, col_map = _v72_detect_sa_invoice_header(sheet)
    total_row = _v6_find_total_row(sheet, hdr_row + 1, [col_map.get("po_no", 0), 1])

    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))

    rows: List[List[Any]] = []
    for r in range(hdr_row + 1, total_row + 1):
        if _v6_is_footer_row(sheet, r):
            break
        po_raw = _v6_uc(get_cell(sheet, r, col_map.get("po_no", 0)))
        qty_v = get_mapped_cell_value(sheet, r, col_map, "qty_pairs")
        amt_v = get_mapped_cell_value(sheet, r, col_map, "amount_usd")
        if not po_raw and not nz_str(qty_v) and not nz_str(amt_v):
            continue
        row_type = "TOTAL" if "TOTAL" in po_raw else "DETAIL"
        po_no = get_mapped_cell_value(sheet, r, col_map, "po_no")
        customer_po = get_mapped_cell_value(sheet, r, col_map, "customer_po")
        desc = get_mapped_cell_value(sheet, r, col_map, "desc")
        hs_code = get_mapped_cell_value(sheet, r, col_map, "hs_code")
        model_name = normalize_model_name(get_mapped_cell_value(sheet, r, col_map, "model_name"))
        width = get_mapped_cell_value(sheet, r, col_map, "width")
        qty_pairs = qty_v
        unit_price = get_mapped_cell_value(sheet, r, col_map, "unit_price")
        amount_usd = amt_v
        if row_type == "TOTAL":
            po_no = "TOTAL:"
        rows.append([
            rep_date, rep_inv, rep_dest,
            po_no, customer_po, desc, hs_code, model_name, "", "", width,
            qty_pairs, unit_price, amount_usd, "", "", amount_usd,
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows

def _v72_build_detail_bundle(file_path: str) -> Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]:
    wb_data = load_workbook_data(file_path)
    inv_sheet_name = "INVOICE" if _v6_is_sa_family(file_path) and "INVOICE" in wb_data.sheets else ("INV" if "INV" in wb_data.sheets else "")
    pl_sheet_name = ""
    if _v6_is_sa_family(file_path):
        for nm in wb_data.sheet_names:
            if nm.strip().upper().startswith("PL"):
                pl_sheet_name = nm
                break
        if not pl_sheet_name and "Sheet1" in wb_data.sheets:
            pl_sheet_name = "Sheet1"
    else:
        for nm in wb_data.sheet_names:
            if nm.strip().upper().startswith("PL"):
                pl_sheet_name = nm
                break
    report_meta = _v6_collect_meta(wb_data, inv_sheet_name, pl_sheet_name)
    arr = process_one_sub_file(file_path)
    status_text = nz_str(arr[21]).upper()
    inv_rows: List[List[Any]] = []
    pl_rows: List[List[Any]] = []
    inv_other_rows: List[List[Any]] = []
    pl_other_rows: List[List[Any]] = []
    if _v6_is_sa_family(file_path):
        inv_other_rows = _v72_extract_sa_inv(file_path, wb_data, report_meta, status_text)
        # SA family is a separate flow:
        # - PL: prefer real PL sheet with SA header
        # - fallback to Sheet1 only inside _v6_extract_sa_pl if real PL cannot be parsed
        pl_other_rows = _v6_extract_sa_pl(file_path, wb_data, report_meta, status_text)
    else:
        inv_rows = _v6_extract_sub_vsf_inv(file_path, wb_data, report_meta, status_text)
        pl_rows = _v6_extract_sub_vsf_pl(file_path, wb_data, report_meta, status_text)
    return inv_rows, pl_rows, inv_other_rows, pl_other_rows

def _v72_rebuild_detail_sheets(self, folder_path: str, output_path: str):
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    try:
        for nm in ["INV", "PL", "INV Others", "PL Others"]:
            if nm not in wb.sheetnames:
                wb.create_sheet(nm)
        ws_inv = wb["INV"]
        ws_pl = wb["PL"]
        ws_inv_oth = wb["INV Others"]
        ws_pl_oth = wb["PL Others"]
        _v72_init_sheet(ws_inv, V6_INV_HEADERS)
        _v72_init_sheet(ws_pl, V6_PL_HEADERS)
        _v72_init_sheet(ws_inv_oth, V6_INV_HEADERS)
        _v72_init_sheet(ws_pl_oth, V6_PL_HEADERS)
        src_files = collect_source_files(folder_path, output_path)
        inv_rows_all: List[List[Any]] = []
        pl_rows_all: List[List[Any]] = []
        inv_other_rows_all: List[List[Any]] = []
        pl_other_rows_all: List[List[Any]] = []
        for fp in src_files:
            if fp.strip().upper() == output_path.strip().upper():
                continue
            try:
                inv_rows, pl_rows, inv_other_rows, pl_other_rows = _v72_build_detail_bundle(fp)
                inv_rows_all.extend(inv_rows)
                pl_rows_all.extend(pl_rows)
                inv_other_rows_all.extend(inv_other_rows)
                pl_other_rows_all.extend(pl_other_rows)
            except Exception as exc:
                errors.append((fp, str(exc)))
        _v72_append_rows(ws_inv, inv_rows_all, [1], 18, 19)
        _v72_append_rows(ws_pl, pl_rows_all, [1], 14, 15)
        _v72_append_rows(ws_inv_oth, inv_other_rows_all, [1], 18, 19)
        _v72_append_rows(ws_pl_oth, pl_other_rows_all, [1], 14, 15)
        autofit_useful(ws_inv, len(V6_INV_HEADERS))
        autofit_useful(ws_pl, len(V6_PL_HEADERS))
        autofit_useful(ws_inv_oth, len(V6_INV_HEADERS))
        autofit_useful(ws_pl_oth, len(V6_PL_HEADERS))
        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_REBUILD_V7_2", "INFO",
                            f"Rebuilt detail tabs: INV={len(inv_rows_all)}, PL={len(pl_rows_all)}, INV Others={len(inv_other_rows_all)}, PL Others={len(pl_other_rows_all)}")])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, "DETAIL_REBUILD_V7_2", "ERROR", err)])
        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

def _v72_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
    _v72_rebuild_detail_sheets(self, folder_path, output_path)
    return summary




# =========================================================
# V7.3 PATCH
# - Add INV columns:
#   S. Invoice#
#   TOTAL W HANDLING UPCHARGE
# - Logic:
#   * Main detail row = TOTAL AMOUNT + following HANDLING/UPCHARGE rows
#   * Handling/Upcharge rows => blank
#   * Total row => keep TOTAL AMOUNT
# =========================================================

V73_INV_HEADERS = [
    "Date", "Invoice No", "Destination", "PO No", "Customer PO#",
    "DESC", "HS CODE", "MODEL NAME / COLOR", "MODEL Y", "COLOR", "WIDTH",
    "Q'TY (Pairs)", "UNIT PRICE (USD)", "AMOUNT W/OUT LABEL",
    "LABEL COST", "LABEL AMOUNT", "TOTAL AMOUNT",
    "S. Invoice#", "S.DateWInv#", "S.MODEL", "S.MODEL Y", "S.COLOR", "S.WIDTH",
    "S.Q'TY (Pairs)", "S.UNIT PRICE (USD)", "TOTAL W HANDLING UPCHARGE",
    "Row Type", "Status", "File", "File Path"
]

def _v73_parse_num(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = nz_str(v).replace(",", "").replace("$", "").strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None

def _v73_norm_tag(v: Any) -> str:
    return "".join(ch for ch in nz_str(v).upper() if ch.isalnum())

def _v76_fee_text(*vals: Any) -> str:
    return " ".join(nz_str(v) for v in vals if nz_str(v)).upper()

def _v76_detect_fee_type(*vals: Any) -> str:
    raw = _v76_fee_text(*vals)
    norm = _v73_norm_tag(raw)
    if not raw and not norm:
        return ""

    if ("UPCHARGE" in norm) or ("UPCHARG" in norm) or ("UPCHAR" in norm):
        return "UPCHARGE"

    if ("DEDUCT" in norm) or ("DEDUC" in norm) or ("DDUCT" in norm):
        return "DEDUCT"

    if (("HANDLINGCHARGE" in norm) or ("HANDLINGCHAGRE" in norm) or
        ("HANDLINGCHARG" in norm) or ("HNDLINGCHARGE" in norm) or
        (("HANDL" in norm or "HNDL" in norm) and ("CHARG" in norm or "CHAGR" in norm))):
        return "HANDLING CHARGE"

    return ""

def _v73_is_handling(po_no: Any) -> bool:
    return _v76_detect_fee_type(po_no) == "HANDLING CHARGE"

def _v73_is_upcharge(po_no: Any) -> bool:
    return _v76_detect_fee_type(po_no) == "UPCHARGE"

def _v73_is_charge_row(po_no: Any) -> bool:
    return _v76_detect_fee_type(po_no) in ("HANDLING CHARGE", "UPCHARGE", "DEDUCT")

def _v73_is_total_row(row: List[Any]) -> bool:
    # BEFORE inserting 2 new columns, INV row layout is:
    # 0..16 data, 17 Row Type, 18 Status, 19 File, 20 File Path
    return nz_str(row[17]).upper() == "TOTAL" if len(row) > 17 else False

def _v73_is_main_detail_row(row: List[Any]) -> bool:
    if len(row) < 18:
        return False
    po_no = row[3]
    row_type = nz_str(row[17]).upper()
    if row_type == "TOTAL":
        return False
    fee_type = _v76_detect_fee_type(po_no, row[4] if len(row) > 4 else "", row[5] if len(row) > 5 else "")
    if fee_type or row_type in ("HANDLING CHARGE", "UPCHARGE", "DEDUCT"):
        return False
    return True


def _v73_expand_inv_others_rows(rows: List[List[Any]]) -> List[List[Any]]:
    out: List[List[Any]] = []
    for row in rows:
        row = list(row)
        # Base INV row layout before V7.3 extras:
        # 0..16 data, 17 Row Type, 18 Status, 19 File, 20 File Path
        # Expand to:
        # 0..16 data, 17 S.Invoice#, 18 TOTAL W HANDLING UPCHARGE, 19 Row Type, 20 Status, 21 File, 22 File Path
        if len(row) >= 21:
            row.insert(17, "")
            row.insert(18, "")
        out.append(row)
    return out

def _v73_apply_invoice_extras(rows: List[List[Any]]) -> List[List[Any]]:
    if not rows:
        return rows

    counters: Dict[Tuple[str, str], int] = {}
    out: List[List[Any]] = []

    n = len(rows)
    for i, row in enumerate(rows):
        row = list(row)
        invoice_no = nz_str(row[1]).strip()
        inv_date = nz_str(row[0]).strip()
        po_no = row[3]
        total_amount = _v73_parse_num(row[16])

        s_invoice = ""
        total_w = ""

        if _v73_is_total_row(row):
            # TOTAL row: do not number. Keep TOTAL W same as TOTAL AMOUNT.
            s_invoice = ""
            total_w = row[16]
        elif _v73_is_main_detail_row(row):
            key = (invoice_no, inv_date)
            counters[key] = counters.get(key, 0) + 1
            if invoice_no:
                s_invoice = f"{invoice_no}-{counters[key]}"
            base = total_amount if total_amount is not None else 0.0
            addon = 0.0
            j = i + 1
            while j < n:
                next_row = rows[j]
                next_invoice = nz_str(next_row[1]).strip()
                next_date = nz_str(next_row[0]).strip()
                next_po = next_row[3]

                if (next_invoice, next_date) != key:
                    break
                if _v73_is_total_row(next_row):
                    break
                if _v73_is_main_detail_row(next_row):
                    break

                if _v73_is_charge_row(next_po):
                    amt = _v73_parse_num(next_row[16])
                    if amt is not None:
                        addon += amt
                j += 1
            total_w = round(base + addon, 2)
        else:
            # HANDLING / UPCHARGE rows stay blank
            total_w = ""

        row.insert(17, s_invoice)
        row.insert(18, total_w)
        out.append(row)

    return out

def _v73_init_sheet(ws, headers: List[str]):
    ws.delete_rows(1, ws.max_row)
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i).value = sanitize_excel_text(h)
        ws.cell(1, i).font = HEADER_FONT
        ws.cell(1, i).fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"

def _v73_append_rows(ws, rows: List[List[Any]], date_cols: List[int], row_type_idx: int, status_idx: int):
    if not rows:
        return
    for r_idx, arr in enumerate(rows, start=2):
        for c, val in enumerate(arr, start=1):
            cell = ws.cell(r_idx, c)
            if c in date_cols:
                cell.value = coerce_excel_date_value(val)
                cell.number_format = "dd/mm/yyyy"
            else:
                cell.value = sanitize_excel_text(val)
            if ws.title.startswith("INV"):
                if c in (12, 23):
                    cell.number_format = '#,##0'
                elif c in (13, 14, 16, 17, 25):
                    cell.number_format = '#,##0.00'
                elif c == 24:
                    cell.number_format = '#,##0.000000'
                elif c == 15:
                    cell.number_format = '0.0000'
            else:
                if c in (10, 11):
                    cell.number_format = '#,##0'
                elif c in (12, 13):
                    cell.number_format = '#,##0.000'
        row_type = nz_str(arr[row_type_idx - 1]).upper()
        status = nz_str(arr[status_idx - 1]).upper()
        if row_type == "TOTAL":
            for c in range(1, len(arr) + 1):
                ws.cell(r_idx, c).fill = V72_TOTAL_FILL
                ws.cell(r_idx, c).font = V72_TOTAL_FONT
        elif status in STATUS_FILLS:
            fill = STATUS_FILLS[status]
            for c in range(1, len(arr) + 1):
                ws.cell(r_idx, c).fill = fill



# V7.3.5 - unhide all columns in detail sheets
def _v735_unhide_all_columns(ws):
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].hidden = False


def _v73_rebuild_detail_sheets(self, folder_path: str, output_path: str):
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    try:
        for nm in ["INV", "PL", "INV Others", "PL Others"]:
            if nm not in wb.sheetnames:
                wb.create_sheet(nm)

        ws_inv = wb["INV"]
        ws_pl = wb["PL"]
        ws_inv_oth = wb["INV Others"]
        ws_pl_oth = wb["PL Others"]

        _v73_init_sheet(ws_inv, V73_INV_HEADERS)
        _v73_init_sheet(ws_pl, V6_PL_HEADERS)
        _v73_init_sheet(ws_inv_oth, V73_INV_HEADERS)
        _v73_init_sheet(ws_pl_oth, V6_PL_HEADERS)

        src_files = collect_source_files(folder_path, output_path)

        inv_rows_all: List[List[Any]] = []
        pl_rows_all: List[List[Any]] = []
        inv_other_rows_all: List[List[Any]] = []
        pl_other_rows_all: List[List[Any]] = []

        for fp in src_files:
            if fp.strip().upper() == output_path.strip().upper():
                continue
            try:
                inv_rows, pl_rows, inv_other_rows, pl_other_rows = _v72_build_detail_bundle(fp)
                inv_rows_all.extend(inv_rows)
                pl_rows_all.extend(pl_rows)
                inv_other_rows_all.extend(inv_other_rows)
                pl_other_rows_all.extend(pl_other_rows)
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_rows_all = _v73_apply_invoice_extras(inv_rows_all)
        inv_other_rows_all = _v73_apply_invoice_extras(inv_other_rows_all)

        _v73_append_rows(ws_inv, inv_rows_all, [1], 20, 21)
        _v73_append_rows(ws_pl, pl_rows_all, [1], 14, 15)
        _v73_append_rows(ws_inv_oth, inv_other_rows_all, [1], 20, 21)
        _v72_append_rows(ws_pl_oth, pl_other_rows_all, [1], 14, 15)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V6_PL_HEADERS))
        autofit_useful(ws_inv_oth, len(V73_INV_HEADERS))
        autofit_useful(ws_pl_oth, len(V6_PL_HEADERS))

        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)
        _v735_unhide_all_columns(ws_inv_oth)
        _v735_unhide_all_columns(ws_pl_oth)

        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_REBUILD_V7_3_5", "INFO",
                            f"Rebuilt detail tabs: INV={len(inv_rows_all)}, PL={len(pl_rows_all)}, INV Others={len(inv_other_rows_all)}, PL Others={len(pl_other_rows_all)}")])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, "DETAIL_REBUILD_V7_3_5", "ERROR", err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

def _v73_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
    _v73_rebuild_detail_sheets(self, folder_path, output_path)
    return summary









# =========================================================
# V7.4 PATCH - APPEND ONLY, KEEP ORIGINAL EXTRACTION LOGIC
# Fixes:
# - Do NOT rebuild/delete detail sheets
# - Preserve original header <-> row layout strictly
# - INV / INV Others must be transformed to V73 layout BEFORE append
# - Append starts at true last used row
# - De-duplicate existing rows already on sheet
# - S. Invoice# resets by (Invoice No, Date) and uses format -1, -2, ...
# - HANDLING CHARGE / UPCHARGE keep S. Invoice# and TOTAL W HANDLING UPCHARGE blank
# - Only TOTAL rows are highlighted
# =========================================================

def _v74_last_used_row(ws) -> int:
    for r in range(ws.max_row, 0, -1):
        for c in range(1, ws.max_column + 1):
            if nz_str(ws.cell(r, c).value) != "":
                return r
    return 1


def _v74_get_header_map(ws) -> Dict[str, int]:
    mp: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        mp[nz_str(ws.cell(1, c).value)] = c
    return mp


def _v74_ensure_headers(ws, headers: List[str]):
    need_write = False
    for i, h in enumerate(headers, start=1):
        if nz_str(ws.cell(1, i).value) != h:
            need_write = True
            break
    if need_write or ws.max_row < 1:
        for i, h in enumerate(headers, start=1):
            ws.cell(1, i).value = sanitize_excel_text(h)
            ws.cell(1, i).font = HEADER_FONT
            ws.cell(1, i).fill = HEADER_FILL_SUB
    ws.freeze_panes = "A2"


def _v74_norm_cell_for_key(v: Any, header_name: str) -> str:
    h = nz_str(header_name).strip().upper()
    if h in ("DATE", "INVOICE DATE", "SHIPPING DATE"):
        return normalize_date_compare(v)
    if h in ("Q'TY (PAIRS)", "QTY (PAIRS)", "UNIT PRICE (USD)", "AMOUNT W/OUT LABEL", "LABEL COST", "LABEL AMOUNT", "TOTAL AMOUNT", "TOTAL W HANDLING UPCHARGE", "CARTON Q'TY", "CBM", "GROSS WEIGHT"):
        return normalize_number_text(v)
    if h == "INVOICE NO":
        return normalize_invoice_no(v)
    if h == "FILE PATH":
        return normalize_path_text(v)
    return normalize_text_compare(v)


def _v74_row_key_from_arr(arr: List[Any], headers: List[str], exclude_headers: Optional[set] = None) -> Tuple[str, ...]:
    exclude_headers = exclude_headers or set()
    key = []
    for i, h in enumerate(headers):
        if h in exclude_headers:
            continue
        val = arr[i] if i < len(arr) else ""
        key.append(_v74_norm_cell_for_key(val, h))
    return tuple(key)


def _v74_existing_keys(ws, headers: List[str], exclude_headers: Optional[set] = None) -> set:
    exclude_headers = exclude_headers or set()
    keys = set()
    last_row = _v74_last_used_row(ws)
    for r in range(2, last_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        if all(nz_str(v) == "" for v in vals):
            continue
        keys.add(_v74_row_key_from_arr(vals, headers, exclude_headers))
    return keys



def _v74_to_excel_datetime(val: Any) -> Any:
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val
    dt = excel_serial_to_datetime(val)
    if dt is not None:
        return dt
    s = nz_str(val).strip()
    if not s:
        return ""
    for fmt in (
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
        "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M", "%d-%m-%Y",
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d",
        "%d/%m/%y", "%m/%d/%y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    loose = try_parse_loose_date(s)
    if loose:
        try:
            return datetime.strptime(loose, "%Y-%m-%d")
        except Exception:
            pass
    return sanitize_excel_text(val)

def _v74_append_rows_exact(ws, rows: List[List[Any]], headers: List[str], date_cols: List[int], row_type_idx: int):
    if not rows:
        return 0
    start_row = _v74_last_used_row(ws) + 1
    added = 0
    date_col_set = set(date_cols or [])
    for offset, arr in enumerate(rows):
        arr = _v77_align_row_to_headers(arr, headers)
        if len(arr) != len(headers):
            raise ValueError(f"Sheet {ws.title}: row len {len(arr)} != header len {len(headers)}")
        r_idx = start_row + offset
        for c, val in enumerate(arr, start=1):
            header_name = headers[c - 1] if c - 1 < len(headers) else ""
            cell = ws.cell(r_idx, c)
            if c in date_col_set:
                cell.value = _v74_to_excel_datetime(val)
                if cell.value != "":
                    cell.number_format = "dd/mm/yyyy"
            elif header_name == "HS CODE":
                cell.value = _v74_coerce_hscode(val)
            else:
                cell.value = sanitize_excel_text(val)

            if header_name in ("Q'TY (Pairs)", "S.Q'TY (Pairs)", "CARTON Q'TY"):
                cell.number_format = '#,##0'
            elif header_name in ("CBM", "GROSS WEIGHT"):
                cell.number_format = '#,##0.000'
            elif header_name in ("UNIT PRICE (USD)", "AMOUNT W/OUT LABEL", "LABEL AMOUNT", "TOTAL AMOUNT", "TOTAL W HANDLING UPCHARGE"):
                cell.number_format = '#,##0.00'
            elif header_name == "S.UNIT PRICE (USD)":
                cell.number_format = '#,##0.000000'
            elif header_name == "LABEL COST":
                cell.number_format = '0.0000'
        row_type = nz_str(arr[row_type_idx - 1]).upper()
        if row_type == "TOTAL":
            for c in range(1, len(arr) + 1):
                ws.cell(r_idx, c).fill = V72_TOTAL_FILL
                ws.cell(r_idx, c).font = V72_TOTAL_FONT
        added += 1
    return added

def _v74_seq_seed_from_sheet(ws) -> Dict[Tuple[str, str], int]:
    seeds: Dict[Tuple[str, str], int] = {}
    if ws.max_row < 2:
        return seeds
    hdr = _v74_get_header_map(ws)
    c_date = hdr.get("Date", 1)
    c_inv = hdr.get("Invoice No", 2)
    c_seq = hdr.get("S. Invoice#", 18)
    c_rowtype = hdr.get("Row Type", 20)
    last_row = _v74_last_used_row(ws)
    for r in range(2, last_row + 1):
        row_type = nz_str(ws.cell(r, c_rowtype).value).upper()
        seq = nz_str(ws.cell(r, c_seq).value).strip()
        if row_type == "TOTAL" or row_type in ("HANDLING CHARGE", "UPCHARGE", "DEDUCT") or not seq:
            continue

        # Accept both legacy forms:
        #   -1
        #   -2
        # and prefixed forms:
        #   INV000000123456-1
        #   INV000000123456-26
        m = re.search(r'-(\d+)$', seq)
        if not m:
            continue
        try:
            n = int(m.group(1))
        except Exception:
            continue

        key = (normalize_invoice_no(ws.cell(r, c_inv).value), normalize_date_compare(ws.cell(r, c_date).value))
        if key != ("", ""):
            seeds[key] = max(seeds.get(key, 0), n)
    return seeds


def _v73_apply_invoice_extras(rows: List[List[Any]], existing_seeds: Optional[Dict[Tuple[str, str], int]] = None) -> List[List[Any]]:
    if not rows:
        return rows

    counters: Dict[Tuple[str, str], int] = {}
    existing_seeds = existing_seeds or {}
    out: List[List[Any]] = []
    n = len(rows)

    for i, row in enumerate(rows):
        row = list(row)
        if len(row) != 21:
            out.append(row)
            continue

        invoice_no = normalize_invoice_no(row[1])
        inv_date = normalize_date_compare(row[0])
        po_no = row[3]
        qty_pairs = _v73_parse_num(row[11])
        total_amount = _v73_parse_num(row[16])
        s_invoice = ""
        s_datewinv = ""
        s_model = ""
        s_model_y = ""
        s_color = ""
        s_width = ""
        s_qty_pairs: Any = ""
        s_unit_price: Any = ""
        total_w: Any = ""

        if _v73_is_total_row(row):
            s_qty_pairs = row[11]
            total_w = row[16]
        elif _v73_is_main_detail_row(row):
            key = (invoice_no, inv_date)
            if key not in counters:
                counters[key] = int(existing_seeds.get(key, 0))
            counters[key] += 1

            if invoice_no:
                s_invoice = f"{invoice_no}-{counters[key]}"
            else:
                s_invoice = f"-{counters[key]}"

            s_datewinv = _v74_make_sdatewinv(row[0], s_invoice)
            s_model = row[7]
            s_model_y = row[8]
            s_color = row[9]
            s_width = row[10]
            s_qty_pairs = row[11]

            base = total_amount if total_amount is not None else 0.0
            addon = 0.0
            j = i + 1
            while j < n:
                next_row = rows[j]
                if len(next_row) < 21:
                    break
                next_key = (normalize_invoice_no(next_row[1]), normalize_date_compare(next_row[0]))
                next_po = next_row[3]
                if next_key != key:
                    break
                if _v73_is_total_row(next_row):
                    break
                if _v73_is_main_detail_row(next_row):
                    break
                if _v73_is_charge_row(next_po):
                    amt = _v73_parse_num(next_row[16])
                    if amt is not None:
                        addon += amt
                j += 1
            total_w = round(base + addon, 2)
            if qty_pairs not in (None, 0):
                try:
                    s_unit_price = round(float(total_w) / float(qty_pairs), 6)
                except Exception:
                    s_unit_price = ""
        else:
            total_w = ""

        expanded = list(row[:17]) + [
            s_invoice, s_datewinv, s_model, s_model_y, s_color, s_width,
            s_qty_pairs, s_unit_price, total_w,
            row[17], row[18], row[19], row[20]
        ]
        out.append(expanded)

    return out

def _v76_apply_pl_sequence(rows: List[List[Any]], existing_seeds: Optional[Dict[Tuple[str, str], int]] = None) -> List[List[Any]]:
    if not rows:
        return rows

    counters: Dict[Tuple[str, str], int] = {}
    existing_seeds = existing_seeds or {}
    out: List[List[Any]] = []

    for row in rows:
        row = list(row)
        if len(row) != 17:
            out.append(row)
            continue

        invoice_no = normalize_invoice_no(row[1])
        inv_date = normalize_date_compare(row[0])
        row_type = nz_str(row[13]).upper()
        fee_type = _v76_detect_fee_type(row[3], row[4], row[5])
        s_invoice = ""
        s_datewinv = ""

        if row_type == "DETAIL" and not fee_type:
            key = (invoice_no, inv_date)
            if key not in counters:
                counters[key] = int(existing_seeds.get(key, 0))
            counters[key] += 1
            if invoice_no:
                s_invoice = f"{invoice_no}-{counters[key]}"
            else:
                s_invoice = f"-{counters[key]}"
            s_datewinv = _v74_make_sdatewinv(row[0], s_invoice)
        elif fee_type and row_type == "DETAIL":
            row_type = fee_type
            row[13] = fee_type
            row[3] = fee_type

        row.insert(13, s_invoice)
        row.insert(14, s_datewinv)
        out.append(row)

    return out

def _v77_align_row_to_headers(arr: List[Any], headers: List[str]) -> List[Any]:
    arr = list(arr)
    if len(arr) == len(headers):
        return arr

    if headers == V73_INV_HEADERS:
        base_old = [
            "Date", "Invoice No", "Destination", "PO No", "Customer PO#",
            "DESC", "HS CODE", "MODEL NAME / COLOR", "MODEL Y", "COLOR", "WIDTH",
            "Q'TY (Pairs)", "UNIT PRICE (USD)", "AMOUNT W/OUT LABEL",
            "LABEL COST", "LABEL AMOUNT", "TOTAL AMOUNT",
            "S. Invoice#", "S.MODEL", "S.MODEL Y", "S.COLOR", "S.WIDTH",
            "S.Q'TY (Pairs)", "S.UNIT PRICE (USD)", "TOTAL W HANDLING UPCHARGE",
            "Row Type", "Status", "File", "File Path"
        ]
        if len(arr) == len(base_old):
            old_map = {h: i for i, h in enumerate(base_old)}
            return [arr[old_map[h]] if h in old_map else "" for h in headers]

    if headers == V76_PL_HEADERS:
        base_old = [
            "Date", "Invoice No", "Destination", "PO No", "Customer PO#",
            "MODEL NAME / COLOR", "COLOR", "WIDTH", "HS CODE",
            "Q'TY (Pairs)", "CARTON Q'TY", "CBM", "GROSS WEIGHT",
            "S. Invoice#", "Row Type", "Status", "File", "File Path"
        ]
        if len(arr) == len(base_old):
            old_map = {h: i for i, h in enumerate(base_old)}
            return [arr[old_map[h]] if h in old_map else "" for h in headers]

    if len(arr) < len(headers):
        return arr + [""] * (len(headers) - len(arr))
    return arr[:len(headers)]


def _v74_filter_new_rows(ws, rows: List[List[Any]], headers: List[str], exclude_headers: Optional[set] = None) -> List[List[Any]]:
    if not rows:
        return []
    seen = _v74_existing_keys(ws, headers, exclude_headers)
    new_rows: List[List[Any]] = []
    for arr in rows:
        arr = _v77_align_row_to_headers(arr, headers)
        if len(arr) != len(headers):
            raise ValueError(f"Sheet {ws.title}: row len {len(arr)} != header len {len(headers)} before dedupe")
        k = _v74_row_key_from_arr(arr, headers, exclude_headers)
        if k in seen:
            continue
        seen.add(k)
        new_rows.append(arr)
    return new_rows


def _v74_rebuild_detail_sheets_append(self, folder_path: str, output_path: str):
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    try:
        for nm in ["INV", "PL", "INV Others", "PL Others"]:
            if nm not in wb.sheetnames:
                wb.create_sheet(nm)

        ws_inv = wb["INV"]
        ws_pl = wb["PL"]
        ws_inv_oth = wb["INV Others"]
        ws_pl_oth = wb["PL Others"]

        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V6_PL_HEADERS)
        _v74_ensure_headers(ws_inv_oth, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl_oth, V6_PL_HEADERS)

        src_files = collect_source_files(folder_path, output_path)

        inv_rows_all: List[List[Any]] = []
        pl_rows_all: List[List[Any]] = []
        inv_other_rows_all: List[List[Any]] = []
        pl_other_rows_all: List[List[Any]] = []

        for fp in src_files:
            if fp.strip().upper() == output_path.strip().upper():
                continue
            try:
                inv_rows, pl_rows, inv_other_rows, pl_other_rows = _v72_build_detail_bundle(fp)
                inv_rows_all.extend(inv_rows)
                pl_rows_all.extend(pl_rows)
                inv_other_rows_all.extend(inv_other_rows)
                pl_other_rows_all.extend(pl_other_rows)
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_seed = _v74_seq_seed_from_sheet(ws_inv)
        inv_other_seed = _v74_seq_seed_from_sheet(ws_inv_oth)

        inv_rows_all = _v73_apply_invoice_extras(inv_rows_all, inv_seed)
        inv_other_rows_all = _v73_apply_invoice_extras(inv_other_rows_all, inv_other_seed)

        inv_rows_all = _v74_filter_new_rows(ws_inv, inv_rows_all, V73_INV_HEADERS, {"S. Invoice#", "TOTAL W HANDLING UPCHARGE"})
        pl_rows_all = _v74_filter_new_rows(ws_pl, pl_rows_all, V6_PL_HEADERS)
        inv_other_rows_all = _v74_filter_new_rows(ws_inv_oth, inv_other_rows_all, V73_INV_HEADERS, {"S. Invoice#", "TOTAL W HANDLING UPCHARGE"})
        pl_other_rows_all = _v74_filter_new_rows(ws_pl_oth, pl_other_rows_all, V6_PL_HEADERS)

        add_inv = _v74_append_rows_exact(ws_inv, inv_rows_all, V73_INV_HEADERS, [1], 20)
        add_pl = _v74_append_rows_exact(ws_pl, pl_rows_all, V6_PL_HEADERS, [1], 14)
        add_inv_oth = _v74_append_rows_exact(ws_inv_oth, inv_other_rows_all, V73_INV_HEADERS, [1], 20)
        add_pl_oth = _v74_append_rows_exact(ws_pl_oth, pl_other_rows_all, V6_PL_HEADERS, [1], 14)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V6_PL_HEADERS))
        autofit_useful(ws_inv_oth, len(V73_INV_HEADERS))
        autofit_useful(ws_pl_oth, len(V6_PL_HEADERS))
        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)
        _v735_unhide_all_columns(ws_inv_oth)
        _v735_unhide_all_columns(ws_pl_oth)

        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_APPEND_V7_4", "INFO",
                            f"Appended detail tabs: INV={add_inv}, PL={add_pl}, INV Others={add_inv_oth}, PL Others={add_pl_oth}")])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, "DETAIL_APPEND_V7_4", "ERROR", err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()


def _v74_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
    _v74_rebuild_detail_sheets_append(self, folder_path, output_path)
    return summary



# =========================================================
# V7.5 PATCH
# - True append for 4 detail sheets based on NEW OK files only
# - Protect existing detail sheets from CORE_RUN_V6 legacy cleanup
# =========================================================

def _v75_norm_abs_upper(path_text: Any) -> str:
    p = nz_str(path_text).strip()
    if not p:
        return ""
    try:
        return os.path.abspath(p).strip().upper()
    except Exception:
        return p.strip().upper()


def _v75_collect_ok_paths_from_subdetail(ws) -> set:
    result = set()
    hdr = _v74_get_header_map(ws)
    c_status = hdr.get("Status", 22)
    c_path = hdr.get("File Path", 24)
    last_row = _v74_last_used_row(ws)
    for r in range(2, last_row + 1):
        status = nz_str(ws.cell(r, c_status).value).upper()
        if status != "OK":
            continue
        p = _v75_norm_abs_upper(ws.cell(r, c_path).value)
        if p:
            result.add(p)
    return result


def _v75_snapshot_sheet_rows(ws, headers: List[str]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    last_row = _v74_last_used_row(ws)
    for r in range(2, last_row + 1):
        arr = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        if all(nz_str(v) == "" for v in arr):
            continue
        rows.append(arr)
    return rows


def _v75_snapshot_detail_tabs(output_path: str) -> Tuple[Dict[str, List[List[Any]]], set]:
    snapshots = {
        "INV": [],
        "PL": [],
    }
    old_ok_paths = set()
    if not output_path or not os.path.exists(output_path):
        return snapshots, old_ok_paths
    wb = open_or_create_output_workbook(output_path)
    try:
        if "SUB_DETAIL" in wb.sheetnames:
            old_ok_paths = _v75_collect_ok_paths_from_subdetail(wb["SUB_DETAIL"])
        if "INV" in wb.sheetnames:
            snapshots["INV"] = _v75_snapshot_sheet_rows(wb["INV"], V73_INV_HEADERS)
        if "PL" in wb.sheetnames:
            snapshots["PL"] = _v75_snapshot_sheet_rows(wb["PL"], V76_PL_HEADERS)
    finally:
        wb.close()
    return snapshots, old_ok_paths


def _v75_restore_detail_tabs(wb, snapshots: Dict[str, List[List[Any]]]):
    for nm, headers, date_cols, row_type_idx in [
        ("INV", V73_INV_HEADERS, [1], 27),
        ("PL", V76_PL_HEADERS, [1], 16),
    ]:
        ws = wb[nm] if nm in wb.sheetnames else wb.create_sheet(nm)
        _v74_ensure_headers(ws, headers)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        rows = [list(r) for r in snapshots.get(nm, []) if len(r) == len(headers)]
        if rows:
            _v74_append_rows_exact(ws, rows, headers, date_cols, row_type_idx)
    for nm in ["INV Others", "PL Others"]:
        if nm in wb.sheetnames:
            del wb[nm]


def _v75_append_new_detail_rows(output_path: str, new_ok_paths: set):
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    try:
        for nm in ["INV", "PL"]:
            if nm not in wb.sheetnames:
                wb.create_sheet(nm)
        for nm in ["INV Others", "PL Others"]:
            if nm in wb.sheetnames:
                del wb[nm]

        ws_inv = wb["INV"]
        ws_pl = wb["PL"]

        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        inv_rows_all: List[List[Any]] = []
        pl_rows_all: List[List[Any]] = []

        file_list = sorted([p for p in new_ok_paths if os.path.exists(p)])
        for fp in file_list:
            try:
                inv_rows, pl_rows, inv_other_rows, pl_other_rows = _v72_build_detail_bundle(fp)
                inv_rows_all.extend(inv_rows)
                inv_rows_all.extend(inv_other_rows)
                pl_rows_all.extend(pl_rows)
                pl_rows_all.extend(pl_other_rows)
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_seed = _v74_seq_seed_from_sheet(ws_inv)
        pl_seed = _v74_seq_seed_from_sheet(ws_pl)

        inv_rows_all = _v73_apply_invoice_extras(inv_rows_all, inv_seed)
        pl_rows_all = _v76_apply_pl_sequence(pl_rows_all, pl_seed)

        inv_rows_all = _v74_filter_new_rows(ws_inv, inv_rows_all, V73_INV_HEADERS, {"S. Invoice#", "S.DateWInv#", "S.MODEL", "S.MODEL Y", "S.COLOR", "S.WIDTH", "S.Q'TY (Pairs)", "S.UNIT PRICE (USD)", "TOTAL W HANDLING UPCHARGE"})
        pl_rows_all = _v74_filter_new_rows(ws_pl, pl_rows_all, V76_PL_HEADERS, {"S. Invoice#", "S.DateWInv#"})

        add_inv = _v74_append_rows_exact(ws_inv, inv_rows_all, V73_INV_HEADERS, [1], 27)
        add_pl = _v74_append_rows_exact(ws_pl, pl_rows_all, V76_PL_HEADERS, [1], 16)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V76_PL_HEADERS))
        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)

        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_APPEND_V7_5", "INFO",
                            f"New OK files={len(file_list)} | Appended detail tabs: INV={add_inv}, PL={add_pl}")])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, "DETAIL_APPEND_V7_5", "ERROR", err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()


def _v75_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    snapshots, old_ok_paths = _v75_snapshot_detail_tabs(output_path)

    summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)

    wb = open_or_create_output_workbook(output_path)
    try:
        for nm in ["INV", "PL"]:
            if nm not in wb.sheetnames:
                wb.create_sheet(nm)
        _v75_restore_detail_tabs(wb, snapshots)
        current_ok_paths = _v75_collect_ok_paths_from_subdetail(wb["SUB_DETAIL"]) if "SUB_DETAIL" in wb.sheetnames else set()
        new_ok_paths = current_ok_paths - old_ok_paths
        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_PREP_V7_5", "INFO",
                            f"Old OK={len(old_ok_paths)} | Current OK={len(current_ok_paths)} | New OK={len(new_ok_paths)} | Merge mode=2 sheets")])
        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

    _v75_append_new_detail_rows(output_path, new_ok_paths)
    return summary




# =========================================================
# V7.6 FINAL PATCH
# Final goal:
# - Keep SUB_DETAIL + LOG_SUB_DETAIL logic from CORE_RUN_V6
# - Do NOT rebuild detail sheets
# - Restore 4 detail sheets after CORE_RUN_V6 touches them
# - Re-scan current source folder and append only rows not already present
# - Works for: INV, PL, INV Others, PL Others
# =========================================================

def _v74_seq_seed_from_rows(rows: List[List[Any]], headers: List[str], seq_header: str = "S. Invoice#") -> Dict[Tuple[str, str], int]:
    seeds: Dict[Tuple[str, str], int] = {}
    try:
        idx_date = headers.index("Date")
        idx_inv = headers.index("Invoice No")
        idx_seq = headers.index(seq_header)
    except ValueError:
        return seeds
    for arr in rows:
        if len(arr) <= max(idx_date, idx_inv, idx_seq):
            continue
        inv = normalize_invoice_no(arr[idx_inv])
        dt = normalize_date_compare(arr[idx_date])
        if not inv or not dt:
            continue
        seq_text = nz_str(arr[idx_seq]).strip()
        m = re.search(r'-(\d+)$', seq_text)
        if not m:
            continue
        n = int(m.group(1))
        key = (inv, dt)
        if n > seeds.get(key, 0):
            seeds[key] = n
    return seeds


def _v76_append_all_detail_rows(folder_path: str, output_path: str, ok_paths: Optional[set] = None):
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    try:
        for nm in ["INV", "PL"]:
            if nm not in wb.sheetnames:
                wb.create_sheet(nm)
        for nm in ["INV Others", "PL Others"]:
            if nm in wb.sheetnames:
                del wb[nm]

        ws_inv = wb["INV"]
        ws_pl = wb["PL"]

        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        src_files = collect_source_files(folder_path, output_path)

        inv_rows_all: List[List[Any]] = []
        pl_rows_all: List[List[Any]] = []
        inv_other_rows_all: List[List[Any]] = []
        pl_other_rows_all: List[List[Any]] = []

        for fp in src_files:
            if fp.strip().upper() == output_path.strip().upper():
                continue
            try:
                inv_rows, pl_rows, inv_other_rows, pl_other_rows = _v72_build_detail_bundle(fp)
                inv_rows_all.extend(inv_rows)
                pl_rows_all.extend(pl_rows)
                inv_other_rows_all.extend(inv_other_rows)
                pl_other_rows_all.extend(pl_other_rows)
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_seed = _v74_seq_seed_from_sheet(ws_inv)
        pl_seed = _v74_seq_seed_from_sheet(ws_pl)

        inv_rows_main = _v73_apply_invoice_extras(inv_rows_all, inv_seed)
        inv_seed_after_main = _v74_seq_seed_from_rows(inv_rows_main, V73_INV_HEADERS, "S. Invoice#")
        inv_rows_other = _v73_apply_invoice_extras(inv_other_rows_all, inv_seed_after_main)
        inv_rows_all = inv_rows_main + inv_rows_other

        pl_rows_main = _v76_apply_pl_sequence(pl_rows_all, pl_seed)
        pl_seed_after_main = _v74_seq_seed_from_rows(pl_rows_main, V76_PL_HEADERS, "S. Invoice#")
        pl_rows_other = _v76_apply_pl_sequence(pl_other_rows_all, pl_seed_after_main)
        pl_rows_all = pl_rows_main + pl_rows_other

        inv_rows_new = _v74_filter_new_rows(ws_inv, inv_rows_all, V73_INV_HEADERS, {"S. Invoice#", "S.DateWInv#", "S.MODEL", "S.MODEL Y", "S.COLOR", "S.WIDTH", "S.Q'TY (Pairs)", "S.UNIT PRICE (USD)", "TOTAL W HANDLING UPCHARGE"})
        pl_rows_new = _v74_filter_new_rows(ws_pl, pl_rows_all, V76_PL_HEADERS, {"S. Invoice#", "S.DateWInv#"})

        add_inv = _v74_append_rows_exact(ws_inv, inv_rows_new, V73_INV_HEADERS, [1], 27)
        add_pl = _v74_append_rows_exact(ws_pl, pl_rows_new, V76_PL_HEADERS, [1], 16)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V76_PL_HEADERS))
        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)

        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_APPEND_V7_6", "INFO",
                            f"Source files={len(src_files)} | Appended detail tabs: INV={add_inv}, PL={add_pl} | Merge mode=2 sheets")])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, "DETAIL_APPEND_V7_6", "ERROR", err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()


def _v76_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    snapshots, old_ok_paths = _v75_snapshot_detail_tabs(output_path)

    summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)

    wb = open_or_create_output_workbook(output_path)
    try:
        for nm in ["INV", "PL"]:
            if nm not in wb.sheetnames:
                wb.create_sheet(nm)
        _v75_restore_detail_tabs(wb, snapshots)
        current_ok_paths = _v75_collect_ok_paths_from_subdetail(wb["SUB_DETAIL"]) if "SUB_DETAIL" in wb.sheetnames else set()
        new_ok_paths = current_ok_paths - old_ok_paths
        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_PREP_V7_6", "INFO",
                            f"Old OK={len(old_ok_paths)} | Current OK={len(current_ok_paths)} | New OK={len(new_ok_paths)} | Restore detail tabs done | Merge mode=2 sheets")])
        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

    _v76_append_all_detail_rows(folder_path, output_path, new_ok_paths)
    return summary





# =========================================================
# M3 PATCH
# - 2 real sheets only: INV and PL
# - Rebuild detail tabs from CURRENT OK files after CORE_RUN_V6
# - Remove wrong/duplicate old rows, then write correct result
# =========================================================

def _m3_rebuild_merged_detail_rows(output_path: str, ok_paths: set):
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    try:
        for nm in ["INV", "PL"]:
            if nm not in wb.sheetnames:
                wb.create_sheet(nm)
        for nm in ["INV Others", "PL Others"]:
            if nm in wb.sheetnames:
                del wb[nm]

        ws_inv = wb["INV"]
        ws_pl = wb["PL"]
        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        if ws_inv.max_row > 1:
            ws_inv.delete_rows(2, ws_inv.max_row - 1)
        if ws_pl.max_row > 1:
            ws_pl.delete_rows(2, ws_pl.max_row - 1)

        file_list = sorted([p for p in ok_paths if os.path.exists(p)])

        inv_rows_main: List[List[Any]] = []
        inv_rows_other: List[List[Any]] = []
        pl_rows_main: List[List[Any]] = []
        pl_rows_other: List[List[Any]] = []

        for fp in file_list:
            try:
                inv_rows, pl_rows, inv_other_rows, pl_other_rows = _v72_build_detail_bundle(fp)
                inv_rows_main.extend(inv_rows)
                inv_rows_other.extend(inv_other_rows)
                pl_rows_main.extend(pl_rows)
                pl_rows_other.extend(pl_other_rows)
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_seed: Dict[Tuple[str, str], int] = {}
        pl_seed: Dict[Tuple[str, str], int] = {}

        inv_rows_main = _v73_apply_invoice_extras(inv_rows_main, inv_seed)
        inv_seed_after_main = _v74_seq_seed_from_rows(inv_rows_main, V73_INV_HEADERS, "S. Invoice#")
        inv_rows_other = _v73_apply_invoice_extras(inv_rows_other, inv_seed_after_main)
        inv_rows_all = inv_rows_main + inv_rows_other

        pl_rows_main = _v76_apply_pl_sequence(pl_rows_main, pl_seed)
        pl_seed_after_main = _v74_seq_seed_from_rows(pl_rows_main, V76_PL_HEADERS, "S. Invoice#")
        pl_rows_other = _v76_apply_pl_sequence(pl_rows_other, pl_seed_after_main)
        pl_rows_all = pl_rows_main + pl_rows_other

        inv_rows_all = _v74_filter_new_rows(ws_inv, inv_rows_all, V73_INV_HEADERS, {"S. Invoice#", "S.DateWInv#", "S.MODEL", "S.MODEL Y", "S.COLOR", "S.WIDTH", "S.Q'TY (Pairs)", "S.UNIT PRICE (USD)", "TOTAL W HANDLING UPCHARGE"})
        pl_rows_all = _v74_filter_new_rows(ws_pl, pl_rows_all, V76_PL_HEADERS, {"S. Invoice#", "S.DateWInv#"})

        add_inv = _v74_append_rows_exact(ws_inv, inv_rows_all, V73_INV_HEADERS, [1], 27)
        add_pl = _v74_append_rows_exact(ws_pl, pl_rows_all, V76_PL_HEADERS, [1], 16)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V76_PL_HEADERS))
        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)

        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_REBUILD_M3", "INFO",
                            f"Current OK files={len(file_list)} | Rebuilt detail tabs: INV={add_inv}, PL={add_pl} | Merge mode=2 sheets")])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, "DETAIL_REBUILD_M3", "ERROR", err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()


def _m3_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)

    wb = open_or_create_output_workbook(output_path)
    try:
        current_ok_paths = _v75_collect_ok_paths_from_subdetail(wb["SUB_DETAIL"]) if "SUB_DETAIL" in wb.sheetnames else set()
        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_PREP_M3", "INFO",
                            f"Current OK={len(current_ok_paths)} | Start rebuild detail tabs | Merge mode=2 sheets")])
        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

    _m3_rebuild_merged_detail_rows(output_path, current_ok_paths)
    return summary



# __main__ moved to end so late patches are active during normal script execution

# =========================================================
# U77 PATCH - SA PL PRIORITY + 2-SHEET MERGE KEEPING SUB_DETAIL
# User requirements implemented:
# - Keep SUB_DETAIL / LOG_SUB_DETAIL flow untouched
# - Keep 2 output detail sheets only: INV, PL
# - SA uses separate parser logic, but still merges into INV / PL
# - Prefer real PL sheet for SA
# - SA PL column mapping:
#   PO# -> PO No
#   CUSTOMER PO# -> Customer PO#
#   GROUP SUM/ MODEL -> MODEL NAME / COLOR
#   STYLE -> COLOR
#   WIDTH -> WIDTH
#   HS CODE -> HS CODE
#   QTY (PRS) -> Q'TY (Pairs)
#   PACKAGE (CTN) -> CARTON Q'TY
#   MEASUREMENT (CBM) -> CBM
#   G.W. (KG) -> GROSS WEIGHT
# - SUB / VSF stays common logic
# =========================================================

def _u77_sheet_by_prefix(wb_data: WorkbookData, prefixes: List[str]) -> str:
    pfx = tuple(p.upper() for p in prefixes)
    for nm in wb_data.sheet_names:
        up = nm.strip().upper()
        if up.startswith(pfx):
            return nm
    return ""


def _u77_detect_sa_pl_header(sheet_data: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    best_row = 0
    best_map: Dict[str, int] = {}
    best_score = -1

    for r in range(1, min(120, len(sheet_data)) + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        if last_col <= 0:
            continue
        row_text = " | ".join(_v6_uc(get_cell(sheet_data, r, c)) for c in range(1, min(last_col, 40) + 1))

        for c in range(1, min(last_col, 40) + 1):
            t = normalize_header(get_cell(sheet_data, r, c))
            if not t:
                continue
            if t in ("PO", "PO#", "PONO", "PONO.") and "po_no" not in tmp:
                tmp["po_no"] = c
            elif ("CUSTOMERPO" in t or "CUSPO" in t) and "customer_po" not in tmp:
                tmp["customer_po"] = c
            elif (("GROUPSUMMODEL" in t) or ("MODELNAMECOLOR" in t) or ("GROUPMODEL" in t)) and "model_name" not in tmp:
                tmp["model_name"] = c
            elif t == "STYLE" and "style" not in tmp:
                tmp["style"] = c
            elif "WIDTH" in t and "width" not in tmp:
                tmp["width"] = c
            elif "HSCODE" in t and "hs_code" not in tmp:
                tmp["hs_code"] = c
            elif (("QTY" in t and "PRS" in t) or ("QTY" in t and "PAIR" in t) or t in ("QTY", "QTYPRS", "QTYPAIRS")) and "qty_pairs" not in tmp:
                tmp["qty_pairs"] = c
            elif ((("PACKAGE" in t) or ("CARTON" in t)) and (("CTN" in t) or ("QTY" in t))) and "carton_qty" not in tmp:
                tmp["carton_qty"] = c
            elif ((("MEASUREMENT" in t) and ("CBM" in t)) or t == "CBM") and "cbm" not in tmp:
                tmp["cbm"] = c
            elif ((("GROSS" in t) and ("WEIGHT" in t)) or ("GWKG" in t) or (t in ("GW", "GWKG", "GROSSWEIGHTKG"))) and "gross_weight" not in tmp:
                tmp["gross_weight"] = c

        score = len(tmp)
        if "GROUP SUM/MODEL" in row_text or "GROUP SUM MODEL" in row_text:
            score += 2
        if "PACKAGE" in row_text and "MEASUREMENT" in row_text:
            score += 1
        if "G.W" in row_text or "G W" in row_text or "GROSS WEIGHT" in row_text:
            score += 1
        if "STYLE" in row_text:
            score += 1

        if score > best_score:
            best_score = score
            best_row = r
            best_map = tmp

    if best_score < 7:
        raise ValueError("Khong tim thay header PL SA hop le")
    return best_row, best_map


def _u77_extract_sa_pl(file_path: str, wb_data: WorkbookData, report_meta: Dict[str, str], status_text: str) -> List[List[Any]]:
    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))

    # Strong priority: real PL sheet first
    candidate_names: List[str] = []
    for nm in wb_data.sheet_names:
        up = nm.strip().upper()
        if up.startswith("PL"):
            candidate_names.append(nm)
    if "Sheet1" in wb_data.sheets:
        candidate_names.append("Sheet1")

    for sheet_name in candidate_names:
        try:
            sheet = wb_data.sheets[sheet_name]
            hdr_row, col_map = _u77_detect_sa_pl_header(sheet)
            rows: List[List[Any]] = []
            for r in range(hdr_row + 1, last_used_row(sheet) + 1):
                if _v6_is_footer_row(sheet, r):
                    break

                po_txt = _v6_uc(get_cell(sheet, r, col_map.get("po_no", 0)))
                qty_v = get_mapped_cell_value(sheet, r, col_map, "qty_pairs")
                ctn_v = get_mapped_cell_value(sheet, r, col_map, "carton_qty")
                cbm_v = get_mapped_cell_value(sheet, r, col_map, "cbm")
                gw_v = get_mapped_cell_value(sheet, r, col_map, "gross_weight")
                model_v = get_mapped_cell_value(sheet, r, col_map, "model_name")
                style_v = get_mapped_cell_value(sheet, r, col_map, "style")

                if not po_txt and not nz_str(qty_v) and not nz_str(ctn_v) and not nz_str(cbm_v) and not nz_str(gw_v) and not nz_str(model_v) and not nz_str(style_v):
                    continue

                row_type = "TOTAL" if "TOTAL" in po_txt else "DETAIL"
                po_no = get_mapped_cell_value(sheet, r, col_map, "po_no")
                customer_po = get_mapped_cell_value(sheet, r, col_map, "customer_po")
                model_name = normalize_model_name(model_v)
                color = style_v  # STYLE -> COLOR
                width = get_mapped_cell_value(sheet, r, col_map, "width")
                hs_code = get_mapped_cell_value(sheet, r, col_map, "hs_code")
                qty_pairs = qty_v
                carton_qty = ctn_v
                cbm = cbm_v
                gross_weight = gw_v

                if row_type == "DETAIL" and not (nz_str(po_no) or nz_str(qty_pairs) or nz_str(carton_qty) or nz_str(cbm) or nz_str(gross_weight) or nz_str(model_name)):
                    continue
                if row_type == "TOTAL":
                    po_no = "TOTAL"

                rows.append([
                    rep_date, rep_inv, rep_dest,
                    po_no, customer_po, model_name, color, width, hs_code,
                    qty_pairs, carton_qty, cbm, gross_weight,
                    row_type, status_text, get_file_name_from_path(file_path), file_path
                ])
            if rows:
                return rows
        except Exception:
            continue
    return []


def _u77_build_detail_bundle(file_path: str) -> Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]:
    wb_data = load_workbook_data(file_path)

    if _v6_is_sa_family(file_path):
        inv_sheet_name = "INVOICE" if "INVOICE" in wb_data.sheets else _u77_sheet_by_prefix(wb_data, ["INV"])
        pl_sheet_name = _u77_sheet_by_prefix(wb_data, ["PL"])
        if not pl_sheet_name and "Sheet1" in wb_data.sheets:
            pl_sheet_name = "Sheet1"
    else:
        inv_sheet_name = _u77_sheet_by_prefix(wb_data, ["INV", "INVOICE"])
        pl_sheet_name = _u77_sheet_by_prefix(wb_data, ["PL"])

    report_meta = _v6_collect_meta(wb_data, inv_sheet_name, pl_sheet_name)
    arr = process_one_sub_file(file_path)
    status_text = nz_str(arr[21]).upper()

    # Keep 2-sheet output architecture. SA is parsed separately but merged later.
    inv_rows: List[List[Any]] = []
    pl_rows: List[List[Any]] = []
    inv_other_rows: List[List[Any]] = []
    pl_other_rows: List[List[Any]] = []

    if _v6_is_sa_family(file_path):
        inv_other_rows = _v72_extract_sa_inv(file_path, wb_data, report_meta, status_text)
        pl_other_rows = _u77_extract_sa_pl(file_path, wb_data, report_meta, status_text)
    else:
        inv_rows = _v6_extract_sub_vsf_inv(file_path, wb_data, report_meta, status_text)
        pl_rows = _v6_extract_sub_vsf_pl(file_path, wb_data, report_meta, status_text)

    return inv_rows, pl_rows, inv_other_rows, pl_other_rows


# Force the latest merged 2-sheet rebuild to use patched bundle above.
_v72_build_detail_bundle = _u77_build_detail_bundle



# =========================================================
# U78 PATCH - SA exact fix from sample output
# - Make late patches effective in normal script run
# - Keep SUB_DETAIL / LOG_SUB_DETAIL flow unchanged
# - Keep SUB / VSF logic unchanged
# - Keep merged detail output only: INV, PL
# - Fix SA parser to match provided sample files/output
# =========================================================

def _u78_detect_sa_invoice_header(sheet_data: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    best_row = 0
    best_map: Dict[str, int] = {}
    best_score = -1
    for r in range(1, min(120, len(sheet_data)) + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        if last_col <= 0:
            continue
        row_text = " | ".join(_v6_uc(get_cell(sheet_data, r, c)) for c in range(1, min(last_col, 40) + 1))
        for c in range(1, min(last_col, 60) + 1):
            t = normalize_header(get_cell(sheet_data, r, c))
            if not t:
                continue
            if t in ("PO", "PO#", "PONO", "PONO.") and "po_no" not in tmp:
                tmp["po_no"] = c
            elif ("CUSTOMERPO" in t or "CUSPO" in t) and "customer_po" not in tmp:
                tmp["customer_po"] = c
            elif ("REFERENCE" in t) and "reference" not in tmp:
                tmp["reference"] = c
            elif (("GENDERSIZE" in t) or ("WIDTH" in t)) and "width" not in tmp:
                tmp["width"] = c
            elif (("GROUPSUMMODEL" in t) or ("GROUPMODEL" in t) or ("MODELNAMECOLOR" in t)) and "model_name" not in tmp:
                tmp["model_name"] = c
            elif (t == "STYLE") and "style" not in tmp:
                tmp["style"] = c
            elif (("TYPE" in t and "CATEGORY" in t) or t in ("TYPECATEGORY", "TYPE", "CATEGORY")) and "desc" not in tmp:
                tmp["desc"] = c
            elif ("HSCODE" in t) and "hs_code" not in tmp:
                tmp["hs_code"] = c
            elif ((("QTY" in t) and ("PRS" in t or "PAIR" in t)) or t in ("QTY", "QTYPRS", "QTYPAIRS")) and "qty_pairs" not in tmp:
                tmp["qty_pairs"] = c
            elif (("UPRICE" in t) or ("UNITPRICE" in t and "USD" in t)) and "unit_price" not in tmp:
                tmp["unit_price"] = c
            elif ("AMOUNT" in t and "USD" in t) and "amount_usd" not in tmp:
                tmp["amount_usd"] = c
        score = len(tmp)
        if "GROUP SUM/MODEL" in row_text or "GROUP SUM MODEL" in row_text:
            score += 2
        if "TYPE/ CATEGORY" in row_text or "TYPE CATEGORY" in row_text:
            score += 1
        if "U.PRICE" in row_text or "U PRICE" in row_text or "UPRICE" in row_text:
            score += 1
        if "GENDER SIZE" in row_text:
            score += 1
        if "STYLE" in row_text:
            score += 1
        if score > best_score:
            best_score = score
            best_row = r
            best_map = tmp
    if best_score < 8:
        raise ValueError("Khong tim thay header SA-INVOICE hop le.")
    return best_row, best_map


def _u78_extract_sa_inv(file_path: str, wb_data: WorkbookData, report_meta: Dict[str, str], status_text: str) -> List[List[Any]]:
    sheet_name = "INVOICE" if "INVOICE" in wb_data.sheets else _u77_sheet_by_prefix(wb_data, ["INV"])
    if not sheet_name or sheet_name not in wb_data.sheets:
        return []
    sheet = wb_data.sheets[sheet_name]
    hdr_row, col_map = _u78_detect_sa_invoice_header(sheet)
    total_row = _v6_find_total_row(sheet, hdr_row + 1, [col_map.get("po_no", 0), 1])

    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))

    rows: List[List[Any]] = []
    for r in range(hdr_row + 1, total_row + 1):
        if _v6_is_footer_row(sheet, r):
            break
        po_raw = _v6_uc(get_cell(sheet, r, col_map.get("po_no", 0)))
        qty_v = get_mapped_cell_value(sheet, r, col_map, "qty_pairs")
        amt_v = get_mapped_cell_value(sheet, r, col_map, "amount_usd")
        if not po_raw and not nz_str(qty_v) and not nz_str(amt_v):
            continue

        row_type = "TOTAL" if "TOTAL" in po_raw else "DETAIL"
        po_no = get_mapped_cell_value(sheet, r, col_map, "po_no")
        customer_po = get_mapped_cell_value(sheet, r, col_map, "customer_po")
        desc = get_mapped_cell_value(sheet, r, col_map, "desc")
        hs_code = get_mapped_cell_value(sheet, r, col_map, "hs_code")
        model_name = get_mapped_cell_value(sheet, r, col_map, "model_name")
        model_y = get_mapped_cell_value(sheet, r, col_map, "style")
        width = get_mapped_cell_value(sheet, r, col_map, "width")
        qty_pairs = qty_v
        unit_price = get_mapped_cell_value(sheet, r, col_map, "unit_price")
        amount_usd = amt_v

        if row_type == "TOTAL":
            po_no = "TOTAL:"

        rows.append([
            rep_date, rep_inv, rep_dest,
            po_no, customer_po, desc, hs_code, model_name, model_y, "", width,
            qty_pairs, unit_price, amount_usd, "", "", amount_usd,
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows


def _u78_detect_sa_pl_header(sheet_data: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    best_row = 0
    best_map: Dict[str, int] = {}
    best_score = -1
    for r in range(1, min(120, len(sheet_data)) + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        if last_col <= 0:
            continue
        row_text = " | ".join(_v6_uc(get_cell(sheet_data, r, c)) for c in range(1, min(last_col, 60) + 1))
        for c in range(1, min(last_col, 80) + 1):
            t = normalize_header(get_cell(sheet_data, r, c))
            if not t:
                continue
            if t in ("PO", "PO#", "PONO", "PONO.") and "po_no" not in tmp:
                tmp["po_no"] = c
            elif ("CUSTOMERPO" in t or "CUSPO" in t) and "customer_po_raw" not in tmp:
                tmp["customer_po_raw"] = c
            elif ("REFERENCE" in t) and "reference" not in tmp:
                tmp["reference"] = c
            elif (("GENDERSIZE" in t) or ("WIDTH" in t)) and "width" not in tmp:
                tmp["width"] = c
            elif (("GROUPSUMMODEL" in t) or ("MODELNAMECOLOR" in t) or ("GROUPMODEL" in t)) and "model_name" not in tmp:
                tmp["model_name"] = c
            elif t == "STYLE" and "style" not in tmp:
                tmp["style"] = c
            elif (("TYPE" in t and "CATEGORY" in t) or t in ("TYPECATEGORY", "TYPE", "CATEGORY")) and "desc" not in tmp:
                tmp["desc"] = c
            elif "HSCODE" in t and "hs_code" not in tmp:
                tmp["hs_code"] = c
            elif ((("QTY" in t) and ("PRS" in t or "PAIR" in t)) or t in ("QTY", "QTYPRS", "QTYPAIRS")) and "qty_pairs" not in tmp:
                tmp["qty_pairs"] = c
            elif ((("PACKAGE" in t) or ("CARTON" in t)) and (("CTN" in t) or ("QTY" in t))) and "carton_qty" not in tmp:
                tmp["carton_qty"] = c
            elif ((("MEASUREMENT" in t) and ("CBM" in t)) or t == "CBM") and "cbm" not in tmp:
                tmp["cbm"] = c
            elif ((("GROSS" in t) and ("WEIGHT" in t)) or ("GWKG" in t) or (t in ("GW", "GWKG", "GROSSWEIGHTKG"))) and "gross_weight" not in tmp:
                tmp["gross_weight"] = c
        score = len(tmp)
        if "GROUP SUM/MODEL" in row_text or "GROUP SUM MODEL" in row_text:
            score += 2
        if "PACKAGE" in row_text and "MEASUREMENT" in row_text:
            score += 1
        if "G.W" in row_text or "G W" in row_text or "GROSS WEIGHT" in row_text:
            score += 1
        if "STYLE" in row_text:
            score += 1
        if "GENDER SIZE" in row_text:
            score += 1
        if score > best_score:
            best_score = score
            best_row = r
            best_map = tmp
    if best_score < 8:
        raise ValueError("Khong tim thay header PL SA hop le")
    return best_row, best_map


def _u78_extract_sa_pl(file_path: str, wb_data: WorkbookData, report_meta: Dict[str, str], status_text: str) -> List[List[Any]]:
    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))

    candidate_names: List[str] = []
    for nm in wb_data.sheet_names:
        if nm.strip().upper().startswith("PL"):
            candidate_names.append(nm)
    if "Sheet1" in wb_data.sheets:
        candidate_names.append("Sheet1")

    for sheet_name in candidate_names:
        try:
            sheet = wb_data.sheets[sheet_name]
            hdr_row, col_map = _u78_detect_sa_pl_header(sheet)
            rows: List[List[Any]] = []
            for r in range(hdr_row + 1, last_used_row(sheet) + 1):
                if _v6_is_footer_row(sheet, r):
                    break
                po_txt = _v6_uc(get_cell(sheet, r, col_map.get("po_no", 0)))
                qty_v = get_mapped_cell_value(sheet, r, col_map, "qty_pairs")
                ctn_v = get_mapped_cell_value(sheet, r, col_map, "carton_qty")
                cbm_v = get_mapped_cell_value(sheet, r, col_map, "cbm")
                gw_v = get_mapped_cell_value(sheet, r, col_map, "gross_weight")
                model_v = get_mapped_cell_value(sheet, r, col_map, "model_name")
                style_v = get_mapped_cell_value(sheet, r, col_map, "style")
                desc_v = get_mapped_cell_value(sheet, r, col_map, "desc")
                width_v = get_mapped_cell_value(sheet, r, col_map, "width")
                if not po_txt and not nz_str(qty_v) and not nz_str(ctn_v) and not nz_str(cbm_v) and not nz_str(gw_v) and not nz_str(model_v) and not nz_str(style_v):
                    continue
                row_type = "TOTAL" if "TOTAL" in po_txt else "DETAIL"
                po_no = get_mapped_cell_value(sheet, r, col_map, "po_no")
                # Match sample PL output: this column carries TYPE/CATEGORY (FINISH SOLE)
                customer_po = desc_v
                model_name = model_v
                color = style_v
                width = width_v
                hs_code = get_mapped_cell_value(sheet, r, col_map, "hs_code")
                qty_pairs = qty_v
                carton_qty = ctn_v
                cbm = cbm_v
                gross_weight = gw_v
                if row_type == "DETAIL" and not (nz_str(po_no) or nz_str(qty_pairs) or nz_str(carton_qty) or nz_str(cbm) or nz_str(gross_weight) or nz_str(model_name)):
                    continue
                if row_type == "TOTAL":
                    po_no = ""
                    customer_po = ""
                    model_name = ""
                    color = ""
                    width = ""
                    hs_code = ""
                rows.append([
                    rep_date, rep_inv, rep_dest,
                    po_no, customer_po, model_name, color, width, hs_code,
                    qty_pairs, carton_qty, cbm, gross_weight,
                    row_type, status_text, get_file_name_from_path(file_path), file_path
                ])
            if rows:
                return rows
        except Exception:
            continue
    return []


def _u78_build_detail_bundle(file_path: str) -> Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]:
    wb_data = load_workbook_data(file_path)

    if _v6_is_sa_family(file_path):
        inv_sheet_name = "INVOICE" if "INVOICE" in wb_data.sheets else _u77_sheet_by_prefix(wb_data, ["INV"])
        pl_sheet_name = _u77_sheet_by_prefix(wb_data, ["PL"])
        if not pl_sheet_name and "Sheet1" in wb_data.sheets:
            pl_sheet_name = "Sheet1"
    else:
        inv_sheet_name = _u77_sheet_by_prefix(wb_data, ["INV", "INVOICE"])
        pl_sheet_name = _u77_sheet_by_prefix(wb_data, ["PL"])

    report_meta = _v6_collect_meta(wb_data, inv_sheet_name, pl_sheet_name)
    arr = process_one_sub_file(file_path)
    status_text = nz_str(arr[21]).upper()

    inv_rows: List[List[Any]] = []
    pl_rows: List[List[Any]] = []
    inv_other_rows: List[List[Any]] = []
    pl_other_rows: List[List[Any]] = []

    if _v6_is_sa_family(file_path):
        inv_other_rows = _u78_extract_sa_inv(file_path, wb_data, report_meta, status_text)
        pl_other_rows = _u78_extract_sa_pl(file_path, wb_data, report_meta, status_text)
    else:
        inv_rows = _v6_extract_sub_vsf_inv(file_path, wb_data, report_meta, status_text)
        pl_rows = _v6_extract_sub_vsf_pl(file_path, wb_data, report_meta, status_text)

    return inv_rows, pl_rows, inv_other_rows, pl_other_rows


_v72_build_detail_bundle = _u78_build_detail_bundle


# =========================================================
# U79 PATCH - SA INV amount fix + include WARNING files in merged detail rebuild
# Root causes fixed:
# 1) SA INV headers may use "TOTAL AMOUNT" (khong co USD) or "AMOUNT (USD)".
#    U78 chi bat duoc header co chu USD nen AMOUNT W/OUT LABEL va TOTAL AMOUNT bi rong
#    o cac file nhu SA2.20260317IS / SA2.20260317S.
# 2) Merge-detail M3 chi lay file status=OK tu SUB_DETAIL.
#    Cac file SA hop le nhung dang WARNING (vi thieu invoice_date metadata trong source)
#    se khong duoc dua vao sheet INV/PL. Vi vay SA2.20260317UPP va SA2.20260311
#    "khong ghi duoc chi tiet" du parser van doc ra dong.
# Keep unchanged:
# - SUB_DETAIL / LOG_SUB_DETAIL flow
# - SUB / VSF parsing logic
# - merged output detail only: INV, PL
# =========================================================

def _u79_collect_detail_paths_from_subdetail(ws) -> set:
    result = set()
    hdr = _v74_get_header_map(ws)
    c_status = hdr.get("Status", 22)
    c_path = hdr.get("File Path", 24)
    last_row = _v74_last_used_row(ws)
    for r in range(2, last_row + 1):
        status = nz_str(ws.cell(r, c_status).value).upper()
        if status not in ("OK", "WARNING"):
            continue
        p = _v75_norm_abs_upper(ws.cell(r, c_path).value)
        if p:
            result.add(p)
    return result


def _u79_detect_sa_invoice_header(sheet_data: List[List[Any]]) -> Tuple[int, Dict[str, int]]:
    best_row = 0
    best_map: Dict[str, int] = {}
    best_score = -1
    for r in range(1, min(120, len(sheet_data)) + 1):
        tmp: Dict[str, int] = {}
        last_col = last_used_col_on_row(sheet_data, r)
        if last_col <= 0:
            continue
        row_text = " | ".join(_v6_uc(get_cell(sheet_data, r, c)) for c in range(1, min(last_col, 60) + 1))
        for c in range(1, min(last_col, 80) + 1):
            t = normalize_header(get_cell(sheet_data, r, c))
            if not t:
                continue
            if t in ("PO", "PO#", "PONO", "PONO."):
                tmp.setdefault("po_no", c)
            elif ("CUSTOMERPO" in t or "CUSPO" in t):
                tmp.setdefault("customer_po", c)
            elif "REFERENCE" in t:
                tmp.setdefault("reference", c)
            elif ("GENDERSIZE" in t or "WIDTH" in t):
                tmp.setdefault("width", c)
            elif ("GROUPSUMMODEL" in t or "GROUPMODEL" in t or "MODELNAMECOLOR" in t):
                tmp.setdefault("model_name", c)
            elif t == "STYLE":
                tmp.setdefault("style", c)
            elif (("TYPE" in t and "CATEGORY" in t) or t in ("TYPECATEGORY", "TYPE", "CATEGORY")):
                tmp.setdefault("desc", c)
            elif "HSCODE" in t:
                tmp.setdefault("hs_code", c)
            elif ((("QTY" in t) and ("PRS" in t or "PAIR" in t)) or t in ("QTY", "QTYPRS", "QTYPAIRS")):
                tmp.setdefault("qty_pairs", c)
            elif ("UPRICE" in t or ("UNITPRICE" in t and "USD" in t)):
                tmp.setdefault("unit_price", c)
            elif (
                "TOTALAMOUNT" in t or
                ("AMOUNT" in t and "USD" in t) or
                t in ("AMOUNT", "AMOUNTUSD")
            ):
                tmp.setdefault("amount_usd", c)
        score = len(tmp)
        if "GROUP SUM/MODEL" in row_text or "GROUP SUM MODEL" in row_text:
            score += 2
        if "TYPE/ CATEGORY" in row_text or "TYPE CATEGORY" in row_text:
            score += 1
        if "U.PRICE" in row_text or "U PRICE" in row_text or "UPRICE" in row_text:
            score += 1
        if "GENDER SIZE" in row_text:
            score += 1
        if "STYLE" in row_text:
            score += 1
        if "TOTAL AMOUNT" in row_text or "AMOUNT (USD)" in row_text or "AMOUNT USD" in row_text:
            score += 1
        if score > best_score:
            best_score = score
            best_row = r
            best_map = tmp
    if best_score < 8:
        raise ValueError("Khong tim thay header SA-INVOICE hop le.")
    return best_row, best_map


def _u79_extract_sa_inv(file_path: str, wb_data: WorkbookData, report_meta: Dict[str, str], status_text: str) -> List[List[Any]]:
    sheet_name = "INVOICE" if "INVOICE" in wb_data.sheets else _u77_sheet_by_prefix(wb_data, ["INV"])
    if not sheet_name or sheet_name not in wb_data.sheets:
        return []
    sheet = wb_data.sheets[sheet_name]
    hdr_row, col_map = _u79_detect_sa_invoice_header(sheet)
    total_row = _v6_find_total_row(sheet, hdr_row + 1, [col_map.get("po_no", 0), 1])

    rep_date = format_date_for_display(report_meta.get("invoice_date", ""))
    rep_inv = nz_str(report_meta.get("invoice_no", ""))
    rep_dest = nz_str(report_meta.get("destination", ""))

    rows: List[List[Any]] = []
    for r in range(hdr_row + 1, total_row + 1):
        if _v6_is_footer_row(sheet, r):
            break
        po_raw = _v6_uc(get_cell(sheet, r, col_map.get("po_no", 0)))
        qty_v = get_mapped_cell_value(sheet, r, col_map, "qty_pairs")
        amt_v = get_mapped_cell_value(sheet, r, col_map, "amount_usd")
        unit_price = get_mapped_cell_value(sheet, r, col_map, "unit_price")

        if not po_raw and not nz_str(qty_v) and not nz_str(amt_v):
            continue

        row_type = "TOTAL" if "TOTAL" in po_raw else "DETAIL"
        po_no = get_mapped_cell_value(sheet, r, col_map, "po_no")
        customer_po = get_mapped_cell_value(sheet, r, col_map, "customer_po")
        desc = get_mapped_cell_value(sheet, r, col_map, "desc")
        hs_code = get_mapped_cell_value(sheet, r, col_map, "hs_code")
        model_name = get_mapped_cell_value(sheet, r, col_map, "model_name")
        model_y = get_mapped_cell_value(sheet, r, col_map, "style")
        width = get_mapped_cell_value(sheet, r, col_map, "width")

        amount_wo_label = amt_v
        total_amount = amt_v

        if row_type == "TOTAL":
            po_no = "TOTAL:"

        rows.append([
            rep_date, rep_inv, rep_dest,
            po_no, customer_po, desc, hs_code, model_name, model_y, "", width,
            qty_v, unit_price, amount_wo_label, "", "", total_amount,
            row_type, status_text, get_file_name_from_path(file_path), file_path
        ])
    return rows


def _u79_build_detail_bundle(file_path: str) -> Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]:
    wb_data = load_workbook_data(file_path)

    if _v6_is_sa_family(file_path):
        inv_sheet_name = "INVOICE" if "INVOICE" in wb_data.sheets else _u77_sheet_by_prefix(wb_data, ["INV"])
        pl_sheet_name = _u77_sheet_by_prefix(wb_data, ["PL"])
        if not pl_sheet_name and "Sheet1" in wb_data.sheets:
            pl_sheet_name = "Sheet1"
    else:
        inv_sheet_name = _u77_sheet_by_prefix(wb_data, ["INV", "INVOICE"])
        pl_sheet_name = _u77_sheet_by_prefix(wb_data, ["PL"])

    report_meta = _v6_collect_meta(wb_data, inv_sheet_name, pl_sheet_name)
    arr = process_one_sub_file(file_path)
    status_text = nz_str(arr[21]).upper()

    inv_rows: List[List[Any]] = []
    pl_rows: List[List[Any]] = []
    inv_other_rows: List[List[Any]] = []
    pl_other_rows: List[List[Any]] = []

    if _v6_is_sa_family(file_path):
        inv_other_rows = _u79_extract_sa_inv(file_path, wb_data, report_meta, status_text)
        pl_other_rows = _u78_extract_sa_pl(file_path, wb_data, report_meta, status_text)
    else:
        inv_rows = _v6_extract_sub_vsf_inv(file_path, wb_data, report_meta, status_text)
        pl_rows = _v6_extract_sub_vsf_pl(file_path, wb_data, report_meta, status_text)

    return inv_rows, pl_rows, inv_other_rows, pl_other_rows


_v72_build_detail_bundle = _u79_build_detail_bundle


def _u79_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)

    wb = open_or_create_output_workbook(output_path)
    try:
        detail_paths = _u79_collect_detail_paths_from_subdetail(wb["SUB_DETAIL"]) if "SUB_DETAIL" in wb.sheetnames else set()
        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", "DETAIL_PREP_U79", "INFO",
                            f"Paths for merged detail (OK+WARNING)={len(detail_paths)} | Start rebuild detail tabs | Merge mode=2 sheets")])
        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

    _m3_rebuild_merged_detail_rows(output_path, detail_paths)
    return summary


# =========================================================
# U82 LOGIC-SAFE WRITE PATCH
# Nguyen tac:
# - Lay 00.Detaisublog lam chuan 100%
# - Giu nguyen logic CORE_RUN_V6 va logic rebuild merged detail
# - Chi toi uu co che ghi/save/autofit, khong doi logic du lieu
# =========================================================

def _u82_fast_autofit_useful(ws, max_col: int, sample_head: int = 120, sample_tail: int = 300):
    last_row = ws.max_row
    if last_row <= 1:
        for col_idx in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12
        ws.row_dimensions[1].height = 20
        return

    row_indexes = [1]
    if last_row > 1:
        head_end = min(last_row, 1 + sample_head)
        row_indexes.extend(range(2, head_end + 1))
        tail_start = max(head_end + 1, last_row - sample_tail + 1)
        if tail_start <= last_row:
            row_indexes.extend(range(tail_start, last_row + 1))

    # remove duplicates but keep order
    seen = set()
    compact_rows = []
    for r in row_indexes:
        if r not in seen:
            seen.add(r)
            compact_rows.append(r)

    for col_idx in range(1, max_col + 1):
        max_len = 0
        for r in compact_rows:
            val = nz_str(ws.cell(r, col_idx).value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)
    ws.row_dimensions[1].height = 20


autofit_useful = _u82_fast_autofit_useful


def _u82_rebuild_merged_detail_rows(output_path: str, ok_paths: set, prep_label: str = "DETAIL_PREP_U79"):
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    try:
        for nm in ["INV", "PL"]:
            if nm not in wb.sheetnames:
                wb.create_sheet(nm)
        for nm in ["INV Others", "PL Others"]:
            if nm in wb.sheetnames:
                del wb[nm]

        ws_inv = wb["INV"]
        ws_pl = wb["PL"]
        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        if ws_inv.max_row > 1:
            ws_inv.delete_rows(2, ws_inv.max_row - 1)
        if ws_pl.max_row > 1:
            ws_pl.delete_rows(2, ws_pl.max_row - 1)

        file_list = sorted([p for p in ok_paths if os.path.exists(p)])

        inv_rows_main: List[List[Any]] = []
        inv_rows_other: List[List[Any]] = []
        pl_rows_main: List[List[Any]] = []
        pl_rows_other: List[List[Any]] = []

        for fp in file_list:
            try:
                inv_rows, pl_rows, inv_other_rows, pl_other_rows = _v72_build_detail_bundle(fp)
                inv_rows_main.extend(inv_rows)
                inv_rows_other.extend(inv_other_rows)
                pl_rows_main.extend(pl_rows)
                pl_rows_other.extend(pl_other_rows)
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_seed: Dict[Tuple[str, str], int] = {}
        pl_seed: Dict[Tuple[str, str], int] = {}

        inv_rows_main = _v73_apply_invoice_extras(inv_rows_main, inv_seed)
        inv_seed_after_main = _v74_seq_seed_from_rows(inv_rows_main, V73_INV_HEADERS, "S. Invoice#")
        inv_rows_other = _v73_apply_invoice_extras(inv_rows_other, inv_seed_after_main)
        inv_rows_all = inv_rows_main + inv_rows_other

        pl_rows_main = _v76_apply_pl_sequence(pl_rows_main, pl_seed)
        pl_seed_after_main = _v74_seq_seed_from_rows(pl_rows_main, V76_PL_HEADERS, "S. Invoice#")
        pl_rows_other = _v76_apply_pl_sequence(pl_rows_other, pl_seed_after_main)
        pl_rows_all = pl_rows_main + pl_rows_other

        inv_rows_all = _v74_filter_new_rows(ws_inv, inv_rows_all, V73_INV_HEADERS, {"S. Invoice#", "S.DateWInv#", "S.MODEL", "S.MODEL Y", "S.COLOR", "S.WIDTH", "S.Q'TY (Pairs)", "S.UNIT PRICE (USD)", "TOTAL W HANDLING UPCHARGE"})
        pl_rows_all = _v74_filter_new_rows(ws_pl, pl_rows_all, V76_PL_HEADERS, {"S. Invoice#", "S.DateWInv#"})

        add_inv = _v74_append_rows_exact(ws_inv, inv_rows_all, V73_INV_HEADERS, [1], 27)
        add_pl = _v74_append_rows_exact(ws_pl, pl_rows_all, V76_PL_HEADERS, [1], 16)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V76_PL_HEADERS))
        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)

        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", prep_label, "INFO",
                            f"Paths for merged detail (OK+WARNING)={len(ok_paths)} | Start rebuild detail tabs | Merge mode=2 sheets")])
        append_log_rows(ws_log, [make_log_row("", "DETAIL_REBUILD_M3", "INFO",
                            f"Current OK files={len(file_list)} | Rebuilt detail tabs: INV={add_inv}, PL={add_pl} | Merge mode=2 sheets")])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, "DETAIL_REBUILD_M3", "ERROR", err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()


def _u82_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)

    wb = open_or_create_output_workbook(output_path)
    try:
        detail_paths = _u79_collect_detail_paths_from_subdetail(wb["SUB_DETAIL"]) if "SUB_DETAIL" in wb.sheetnames else set()
    finally:
        wb.close()

    _u82_rebuild_merged_detail_rows(output_path, detail_paths, prep_label="DETAIL_PREP_U79")
    return summary



# =========================================================
# U83 SPEED PATCH
# Keep U82 logic/result, optimize only write/reset path for merged detail
# =========================================================

def _u83_recreate_sheet_keep_index(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        old = wb[sheet_name]
        idx = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name, index=idx)
        return ws
    return wb.create_sheet(title=sheet_name)


def _u83_append_rows_exact_fast(ws, rows: List[List[Any]], headers: List[str], date_cols: List[int], row_type_idx: int):
    if not rows:
        return 0

    date_col_set = set(date_cols or [])
    qty_headers = {"Q'TY (Pairs)", "S.Q'TY (Pairs)", "CARTON Q'TY"}
    dec3_headers = {"CBM", "GROSS WEIGHT"}
    dec2_headers = {"UNIT PRICE (USD)", "AMOUNT W/OUT LABEL", "LABEL AMOUNT", "TOTAL AMOUNT", "TOTAL W HANDLING UPCHARGE"}
    dec6_headers = {"S.UNIT PRICE (USD)"}
    dec4_headers = {"LABEL COST"}

    start_row = _v74_last_used_row(ws) + 1
    total_rows = []

    aligned_rows: List[List[Any]] = []
    for arr in rows:
        arr = _v77_align_row_to_headers(arr, headers)
        if len(arr) != len(headers):
            raise ValueError(f"Sheet {ws.title}: row len {len(arr)} != header len {len(headers)}")
        out = []
        for c, val in enumerate(arr, start=1):
            header_name = headers[c - 1] if c - 1 < len(headers) else ""
            if c in date_col_set:
                out.append(_v74_to_excel_datetime(val))
            elif header_name == "HS CODE":
                out.append(_v74_coerce_hscode(val))
            else:
                out.append(sanitize_excel_text(val))
        aligned_rows.append(out)
        if nz_str(arr[row_type_idx - 1]).upper() == "TOTAL":
            total_rows.append(start_row + len(aligned_rows) - 1)

    for out in aligned_rows:
        ws.append(out)

    end_row = start_row + len(aligned_rows) - 1
    fmt_map = {}
    for c, header_name in enumerate(headers, start=1):
        if c in date_col_set:
            fmt_map[c] = "dd/mm/yyyy"
        elif header_name in qty_headers:
            fmt_map[c] = '#,##0'
        elif header_name in dec3_headers:
            fmt_map[c] = '#,##0.000'
        elif header_name in dec2_headers:
            fmt_map[c] = '#,##0.00'
        elif header_name in dec6_headers:
            fmt_map[c] = '#,##0.000000'
        elif header_name in dec4_headers:
            fmt_map[c] = '0.0000'

    for c, fmt in fmt_map.items():
        for r in range(start_row, end_row + 1):
            ws.cell(r, c).number_format = fmt

    for r_idx in total_rows:
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r_idx, c)
            cell.fill = V72_TOTAL_FILL
            cell.font = V72_TOTAL_FONT

    return len(aligned_rows)


def _u83_rebuild_merged_detail_rows(output_path: str, ok_paths: set, prep_label: str = "DETAIL_PREP_U79"):
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    try:
        # Recreate detail sheets instead of delete_rows on large sheets.
        ws_inv = _u83_recreate_sheet_keep_index(wb, "INV")
        ws_pl = _u83_recreate_sheet_keep_index(wb, "PL")
        for nm in ["INV Others", "PL Others"]:
            if nm in wb.sheetnames:
                del wb[nm]

        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        file_list = sorted([p for p in ok_paths if os.path.exists(p)])

        inv_rows_main: List[List[Any]] = []
        inv_rows_other: List[List[Any]] = []
        pl_rows_main: List[List[Any]] = []
        pl_rows_other: List[List[Any]] = []

        for fp in file_list:
            try:
                inv_rows, pl_rows, inv_other_rows, pl_other_rows = _v72_build_detail_bundle(fp)
                inv_rows_main.extend(inv_rows)
                inv_rows_other.extend(inv_other_rows)
                pl_rows_main.extend(pl_rows)
                pl_rows_other.extend(pl_other_rows)
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_seed: Dict[Tuple[str, str], int] = {}
        pl_seed: Dict[Tuple[str, str], int] = {}

        inv_rows_main = _v73_apply_invoice_extras(inv_rows_main, inv_seed)
        inv_seed_after_main = _v74_seq_seed_from_rows(inv_rows_main, V73_INV_HEADERS, "S. Invoice#")
        inv_rows_other = _v73_apply_invoice_extras(inv_rows_other, inv_seed_after_main)
        inv_rows_all = inv_rows_main + inv_rows_other

        pl_rows_main = _v76_apply_pl_sequence(pl_rows_main, pl_seed)
        pl_seed_after_main = _v74_seq_seed_from_rows(pl_rows_main, V76_PL_HEADERS, "S. Invoice#")
        pl_rows_other = _v76_apply_pl_sequence(pl_rows_other, pl_seed_after_main)
        pl_rows_all = pl_rows_main + pl_rows_other

        # Newly recreated sheets are empty except header, so result is identical to filtering against empty target.
        add_inv = _u83_append_rows_exact_fast(ws_inv, inv_rows_all, V73_INV_HEADERS, [1], 27)
        add_pl = _u83_append_rows_exact_fast(ws_pl, pl_rows_all, V76_PL_HEADERS, [1], 16)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V76_PL_HEADERS))
        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)

        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row("", prep_label, "INFO",
                            f"Paths for merged detail (OK+WARNING)={len(ok_paths)} | Start rebuild detail tabs | Merge mode=2 sheets")])
        append_log_rows(ws_log, [make_log_row("", "DETAIL_REBUILD_M3", "INFO",
                            f"Current OK files={len(file_list)} | Rebuilt detail tabs: INV={add_inv}, PL={add_pl} | Merge mode=2 sheets")])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, "DETAIL_REBUILD_M3", "ERROR", err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()


def _u83_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)

    wb = open_or_create_output_workbook(output_path)
    try:
        detail_paths = _u79_collect_detail_paths_from_subdetail(wb["SUB_DETAIL"]) if "SUB_DETAIL" in wb.sheetnames else set()
    finally:
        wb.close()

    _u83_rebuild_merged_detail_rows(output_path, detail_paths, prep_label="DETAIL_PREP_U79")
    return summary



# =========================================================
# U84 WRITE-FORMAT + WIDTH-CACHE PATCH
# Nguyen tac:
# - Giu nguyen 100% logic ket qua cua U83 / 00.Detaisublog
# - Chi giam chi phi formatting tung o va cache do rong cot
# =========================================================

_U84_WIDTH_CACHE_CTX = {"output_path": "", "cache": {}}


def _u84_width_cache_path(output_path: str) -> str:
    base = os.path.abspath(nz_str(output_path))
    return base + '.widthcache.json' if base else 'detail_sub.widthcache.json'


def _u84_load_width_cache(output_path: str) -> Dict[str, Any]:
    path = _u84_width_cache_path(output_path)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _u84_save_width_cache(output_path: str, cache: Dict[str, Any]):
    path = _u84_width_cache_path(output_path)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _u84_sheet_width_signature(ws, max_col: int, sample_head: int = 120, sample_tail: int = 300) -> str:
    last_row = ws.max_row
    row_indexes = [1]
    if last_row > 1:
        head_end = min(last_row, 1 + sample_head)
        row_indexes.extend(range(2, head_end + 1))
        tail_start = max(head_end + 1, last_row - sample_tail + 1)
        if tail_start <= last_row:
            row_indexes.extend(range(tail_start, last_row + 1))
    seen = set()
    compact_rows = []
    for r in row_indexes:
        if r not in seen:
            seen.add(r)
            compact_rows.append(r)
    parts = [ws.title, str(last_row), str(max_col)]
    for c in range(1, max_col + 1):
        max_len = 0
        non_empty = 0
        for r in compact_rows:
            val = nz_str(ws.cell(r, c).value)
            if val:
                non_empty += 1
                if len(val) > max_len:
                    max_len = len(val)
        parts.append(f"{c}:{non_empty}:{max_len}")
    return '|'.join(parts)


def _u84_autofit_useful_cached(ws, max_col: int, sample_head: int = 120, sample_tail: int = 300):
    output_path = nz_str(_U84_WIDTH_CACHE_CTX.get('output_path', ''))
    cache = _U84_WIDTH_CACHE_CTX.get('cache') or {}
    cache_key = ws.title
    sig = _u84_sheet_width_signature(ws, max_col, sample_head=sample_head, sample_tail=sample_tail)
    item = cache.get(cache_key, {}) if isinstance(cache, dict) else {}

    if item.get('sig') == sig and isinstance(item.get('widths'), dict):
        widths = item.get('widths', {})
        for col_idx in range(1, max_col + 1):
            w = widths.get(str(col_idx))
            if isinstance(w, (int, float)):
                ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.row_dimensions[1].height = 20
        return

    _u82_fast_autofit_useful(ws, max_col, sample_head=sample_head, sample_tail=sample_tail)
    widths = {}
    for col_idx in range(1, max_col + 1):
        try:
            widths[str(col_idx)] = float(ws.column_dimensions[get_column_letter(col_idx)].width)
        except Exception:
            pass
    if isinstance(cache, dict):
        cache[cache_key] = {'sig': sig, 'widths': widths}
    if output_path:
        _u84_save_width_cache(output_path, cache)


autofit_useful = _u84_autofit_useful_cached


def _u84_set_number_formats_sparse(ws, start_row: int, headers: List[str], date_cols: List[int], end_row: int):
    if end_row < start_row:
        return
    date_col_set = set(date_cols or [])
    qty_headers = {"Q'TY (Pairs)", "S.Q'TY (Pairs)", "CARTON Q'TY"}
    dec3_headers = {"CBM", "GROSS WEIGHT"}
    dec2_headers = {"UNIT PRICE (USD)", "AMOUNT W/OUT LABEL", "LABEL AMOUNT", "TOTAL AMOUNT", "TOTAL W HANDLING UPCHARGE"}
    dec6_headers = {"S.UNIT PRICE (USD)"}
    dec4_headers = {"LABEL COST"}

    fmt_map = {}
    for c, header_name in enumerate(headers, start=1):
        if c in date_col_set:
            fmt_map[c] = 'dd/mm/yyyy'
        elif header_name in qty_headers:
            fmt_map[c] = '#,##0'
        elif header_name in dec3_headers:
            fmt_map[c] = '#,##0.000'
        elif header_name in dec2_headers:
            fmt_map[c] = '#,##0.00'
        elif header_name in dec6_headers:
            fmt_map[c] = '#,##0.000000'
        elif header_name in dec4_headers:
            fmt_map[c] = '0.0000'

    # Format only non-empty cells in relevant columns.
    for c, fmt in fmt_map.items():
        for r in range(start_row, end_row + 1):
            cell = ws.cell(r, c)
            if cell.value not in (None, ''):
                cell.number_format = fmt


def _u84_apply_row_fill_fast(ws, row_indexes: List[int], fill_obj, font_obj, col_count: int):
    for r_idx in row_indexes:
        for c in range(1, col_count + 1):
            cell = ws.cell(r_idx, c)
            cell.fill = fill_obj
            if font_obj is not None:
                cell.font = font_obj


def append_result_rows(ws, rows: List[List[Any]]):
    if not rows:
        return
    start_row = ws.max_row + 1 if nz_str(ws.cell(ws.max_row, 1).value) else max(2, ws.max_row)
    if start_row < 2:
        start_row = 2

    warning_rows = []
    error_rows = []
    for arr in rows:
        out = [sanitize_excel_text(v) for v in arr]
        ws.append(out)
        r_idx = ws.max_row
        status = nz_str(arr[21]).upper()
        if status == 'WARNING':
            warning_rows.append(r_idx)
        elif status == 'ERROR':
            error_rows.append(r_idx)

    end_row = ws.max_row
    for c in (15, 19):
        for r in range(start_row, end_row + 1):
            cell = ws.cell(r, c)
            if cell.value not in (None, ''):
                cell.number_format = 'dd/mm/yyyy'

    if warning_rows:
        _u84_apply_row_fill_fast(ws, warning_rows, STATUS_FILLS['WARNING'], None, 25)
    if error_rows:
        _u84_apply_row_fill_fast(ws, error_rows, STATUS_FILLS['ERROR'], None, 25)


def append_log_rows(ws, rows: List[List[Any]]):
    if not rows:
        return
    for row in rows:
        ws.append([sanitize_excel_text(v) for v in row])


def append_generic_rows(ws, rows: List[List[Any]], date_cols: List[int], status_col_idx: int):
    if not rows:
        return
    start_row = ws.max_row + 1 if nz_str(ws.cell(ws.max_row, 1).value) else max(2, ws.max_row)
    if start_row < 2:
        start_row = 2
    row_type_col_idx = 18 if ws.title == 'INV' else 14

    total_rows = []
    handling_rows = []
    warning_rows = []
    error_rows = []

    for arr in rows:
        ws.append([sanitize_excel_text(v) for v in arr])
        r_idx = ws.max_row
        row_type = nz_str(arr[row_type_col_idx - 1]).upper()
        status = nz_str(arr[status_col_idx - 1]).upper()
        if row_type == 'TOTAL':
            total_rows.append(r_idx)
        elif row_type == 'HANDLING CHARGE':
            handling_rows.append(r_idx)
        elif status == 'WARNING':
            warning_rows.append(r_idx)
        elif status == 'ERROR':
            error_rows.append(r_idx)

    end_row = ws.max_row

    if ws.title == 'INV':
        header_names = INV_DETAIL_HEADERS
        local_date_cols = [1]
        fmt_map = {
            12: '#,##0',
            13: '#,##0.00', 14: '#,##0.00', 16: '#,##0.00', 17: '#,##0.00',
        }
    else:
        header_names = PL_DETAIL_HEADERS
        local_date_cols = [1]
        fmt_map = {
            10: '#,##0', 11: '#,##0',
            12: '#,##0.000', 13: '#,##0.000',
        }

    for c in local_date_cols:
        for r in range(start_row, end_row + 1):
            cell = ws.cell(r, c)
            if cell.value not in (None, ''):
                cell.number_format = 'dd-mmm'
    for c, fmt in fmt_map.items():
        for r in range(start_row, end_row + 1):
            cell = ws.cell(r, c)
            if cell.value not in (None, ''):
                cell.number_format = fmt

    col_count = len(rows[0]) if rows else 0
    if total_rows:
        _u84_apply_row_fill_fast(ws, total_rows, TOTAL_ROW_FILL, TOTAL_FONT, col_count)
    if handling_rows:
        _u84_apply_row_fill_fast(ws, handling_rows, HANDLING_ROW_FILL, HANDLING_FONT, col_count)
    if warning_rows:
        _u84_apply_row_fill_fast(ws, warning_rows, STATUS_FILLS['WARNING'], None, col_count)
    if error_rows:
        _u84_apply_row_fill_fast(ws, error_rows, STATUS_FILLS['ERROR'], None, col_count)


def _u84_append_rows_exact_fast(ws, rows: List[List[Any]], headers: List[str], date_cols: List[int], row_type_idx: int):
    if not rows:
        return 0

    start_row = _v74_last_used_row(ws) + 1
    total_rows = []
    aligned_rows: List[List[Any]] = []
    for arr in rows:
        arr = _v77_align_row_to_headers(arr, headers)
        if len(arr) != len(headers):
            raise ValueError(f"Sheet {ws.title}: row len {len(arr)} != header len {len(headers)}")
        out = []
        for c, val in enumerate(arr, start=1):
            header_name = headers[c - 1] if c - 1 < len(headers) else ''
            if c in set(date_cols or []):
                out.append(_v74_to_excel_datetime(val))
            elif header_name == 'HS CODE':
                out.append(_v74_coerce_hscode(val))
            else:
                out.append(sanitize_excel_text(val))
        aligned_rows.append(out)
        if nz_str(arr[row_type_idx - 1]).upper() == 'TOTAL':
            total_rows.append(start_row + len(aligned_rows) - 1)

    for out in aligned_rows:
        ws.append(out)

    end_row = start_row + len(aligned_rows) - 1
    _u84_set_number_formats_sparse(ws, start_row, headers, date_cols, end_row)
    if total_rows:
        _u84_apply_row_fill_fast(ws, total_rows, V72_TOTAL_FILL, V72_TOTAL_FONT, len(headers))
    return len(aligned_rows)


def _u84_rebuild_merged_detail_rows(output_path: str, ok_paths: set, prep_label: str = 'DETAIL_PREP_U79'):
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    try:
        ws_inv = _u83_recreate_sheet_keep_index(wb, 'INV')
        ws_pl = _u83_recreate_sheet_keep_index(wb, 'PL')
        for nm in ['INV Others', 'PL Others']:
            if nm in wb.sheetnames:
                del wb[nm]

        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        file_list = sorted([p for p in ok_paths if os.path.exists(p)])

        inv_rows_main: List[List[Any]] = []
        inv_rows_other: List[List[Any]] = []
        pl_rows_main: List[List[Any]] = []
        pl_rows_other: List[List[Any]] = []

        for fp in file_list:
            try:
                inv_rows, pl_rows, inv_other_rows, pl_other_rows = _v72_build_detail_bundle(fp)
                inv_rows_main.extend(inv_rows)
                inv_rows_other.extend(inv_other_rows)
                pl_rows_main.extend(pl_rows)
                pl_rows_other.extend(pl_other_rows)
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_seed: Dict[Tuple[str, str], int] = {}
        pl_seed: Dict[Tuple[str, str], int] = {}

        inv_rows_main = _v73_apply_invoice_extras(inv_rows_main, inv_seed)
        inv_seed_after_main = _v74_seq_seed_from_rows(inv_rows_main, V73_INV_HEADERS, 'S. Invoice#')
        inv_rows_other = _v73_apply_invoice_extras(inv_rows_other, inv_seed_after_main)
        inv_rows_all = inv_rows_main + inv_rows_other

        pl_rows_main = _v76_apply_pl_sequence(pl_rows_main, pl_seed)
        pl_seed_after_main = _v74_seq_seed_from_rows(pl_rows_main, V76_PL_HEADERS, 'S. Invoice#')
        pl_rows_other = _v76_apply_pl_sequence(pl_rows_other, pl_seed_after_main)
        pl_rows_all = pl_rows_main + pl_rows_other

        add_inv = _u84_append_rows_exact_fast(ws_inv, inv_rows_all, V73_INV_HEADERS, [1], 27)
        add_pl = _u84_append_rows_exact_fast(ws_pl, pl_rows_all, V76_PL_HEADERS, [1], 16)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V76_PL_HEADERS))
        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)

        ws_log = ensure_sheet(wb, 'LOG_SUB_DETAIL')
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row('', prep_label, 'INFO',
                            f'Paths for merged detail (OK+WARNING)={len(ok_paths)} | Start rebuild detail tabs | Merge mode=2 sheets')])
        append_log_rows(ws_log, [make_log_row('', 'DETAIL_REBUILD_M3', 'INFO',
                            f'Current OK files={len(file_list)} | Rebuilt detail tabs: INV={add_inv}, PL={add_pl} | Merge mode=2 sheets')])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, 'DETAIL_REBUILD_M3', 'ERROR', err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()


def _u84_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _U84_WIDTH_CACHE_CTX['output_path'] = output_path
    _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)
    try:
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)

        wb = open_or_create_output_workbook(output_path)
        try:
            detail_paths = _u79_collect_detail_paths_from_subdetail(wb['SUB_DETAIL']) if 'SUB_DETAIL' in wb.sheetnames else set()
        finally:
            wb.close()

        _u84_rebuild_merged_detail_rows(output_path, detail_paths, prep_label='DETAIL_PREP_U79')
        _u84_save_width_cache(output_path, _U84_WIDTH_CACHE_CTX.get('cache') or {})
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}




# =========================================================
# U85 BENCHMARK INTERNAL
# - Khong doi logic ket qua
# - Ghi thoi gian tung pha vao LOG_SUB_DETAIL
# - Theo doi ca time cua autofit va save workbook
# - Ghi ro trang thai file width cache
# =========================================================

import time

_U85_BENCH_CTX = {
    'enabled': False,
    'scan_seconds': 0.0,
    'autofit_seconds': 0.0,
    'save_seconds': 0.0,
    'cache_path': '',
    'cache_saved': False,
    'cache_error': '',
    'save_calls': 0,
    'autofit_calls': 0,
}

_u85_collect_source_files_base = collect_source_files
def collect_source_files(root_folder: str, output_path: str) -> List[str]:
    if not _U85_BENCH_CTX.get('enabled'):
        return _u85_collect_source_files_base(root_folder, output_path)
    t0 = time.perf_counter()
    result = _u85_collect_source_files_base(root_folder, output_path)
    _U85_BENCH_CTX['scan_seconds'] += max(0.0, time.perf_counter() - t0)
    return result

_u85_safe_save_workbook_atomic_base = safe_save_workbook_atomic
def safe_save_workbook_atomic(wb, output_path: str):
    if not _U85_BENCH_CTX.get('enabled'):
        return _u85_safe_save_workbook_atomic_base(wb, output_path)
    t0 = time.perf_counter()
    try:
        return _u85_safe_save_workbook_atomic_base(wb, output_path)
    finally:
        _U85_BENCH_CTX['save_seconds'] += max(0.0, time.perf_counter() - t0)
        _U85_BENCH_CTX['save_calls'] += 1

_u85_autofit_useful_cached_base = autofit_useful
def autofit_useful(ws, max_col: int, sample_head: int = 120, sample_tail: int = 300):
    if not _U85_BENCH_CTX.get('enabled'):
        return _u85_autofit_useful_cached_base(ws, max_col, sample_head=sample_head, sample_tail=sample_tail)
    t0 = time.perf_counter()
    try:
        return _u85_autofit_useful_cached_base(ws, max_col, sample_head=sample_head, sample_tail=sample_tail)
    finally:
        _U85_BENCH_CTX['autofit_seconds'] += max(0.0, time.perf_counter() - t0)
        _U85_BENCH_CTX['autofit_calls'] += 1

_u85_save_width_cache_base = _u84_save_width_cache
def _u84_save_width_cache(output_path: str, cache: Dict[str, Any]):
    path = _u84_width_cache_path(output_path)
    _U85_BENCH_CTX['cache_path'] = path
    try:
        _u85_save_width_cache_base(output_path, cache)
        _U85_BENCH_CTX['cache_saved'] = os.path.exists(path)
        if not _U85_BENCH_CTX['cache_saved']:
            _U85_BENCH_CTX['cache_error'] = 'save_called_but_file_not_found'
    except Exception as exc:
        _U85_BENCH_CTX['cache_saved'] = False
        _U85_BENCH_CTX['cache_error'] = str(exc)
        raise

def _u85_reset_bench():
    _U85_BENCH_CTX.update({
        'enabled': True,
        'scan_seconds': 0.0,
        'autofit_seconds': 0.0,
        'save_seconds': 0.0,
        'cache_path': '',
        'cache_saved': False,
        'cache_error': '',
        'save_calls': 0,
        'autofit_calls': 0,
    })

def _u85_make_bench_rows(folder_path: str, output_path: str, core_seconds: float, rebuild_seconds: float, total_seconds: float) -> List[List[Any]]:
    scan_s = float(_U85_BENCH_CTX.get('scan_seconds') or 0.0)
    autofit_s = float(_U85_BENCH_CTX.get('autofit_seconds') or 0.0)
    save_s = float(_U85_BENCH_CTX.get('save_seconds') or 0.0)
    save_calls = int(_U85_BENCH_CTX.get('save_calls') or 0)
    autofit_calls = int(_U85_BENCH_CTX.get('autofit_calls') or 0)
    cache_path = nz_str(_U85_BENCH_CTX.get('cache_path'))
    cache_saved = bool(_U85_BENCH_CTX.get('cache_saved'))
    cache_error = nz_str(_U85_BENCH_CTX.get('cache_error'))

    phases = [
        ("BENCHMARK", "INFO", f"folder={folder_path}"),
        ("BENCH_TOTAL", "INFO", f"{total_seconds:.3f}s"),
        ("BENCH_SCAN", "INFO", f"{scan_s:.3f}s"),
        ("BENCH_CORE_WRITE", "INFO", f"{core_seconds:.3f}s"),
        ("BENCH_REBUILD_DETAIL", "INFO", f"{rebuild_seconds:.3f}s"),
        ("BENCH_AUTOFIT", "INFO", f"{autofit_s:.3f}s | calls={autofit_calls}"),
        ("BENCH_SAVE_WORKBOOK", "INFO", f"{save_s:.3f}s | calls={save_calls}"),
        ("BENCH_CACHE_PATH", "INFO", cache_path or "(empty)"),
        ("BENCH_CACHE_STATUS", "INFO", "saved" if cache_saved else ("not_saved" if not cache_error else f"error: {cache_error}")),
    ]

    rows = []
    for step_name, result_text, detail_text in phases:
        rows.append(make_log_row(output_path, step_name, result_text, detail_text))

    ranked = sorted([
        ("SCAN", scan_s),
        ("CORE_WRITE", core_seconds),
        ("REBUILD_DETAIL", rebuild_seconds),
        ("AUTOFIT", autofit_s),
        ("SAVE_WORKBOOK", save_s),
    ], key=lambda x: x[1], reverse=True)

    rows.append(make_log_row(output_path, "BENCH_HOTSPOT", "INFO", " > ".join([f"{name}={sec:.3f}s" for name, sec in ranked])))
    return rows

def _u85_append_bench_log_rows(output_path: str, bench_rows: List[List[Any]]):
    wb = open_or_create_output_workbook(output_path)
    try:
        ws_log = ensure_sheet(wb, "LOG_SUB_DETAIL")
        init_log_header(ws_log)
        append_log_rows(ws_log, bench_rows)
        _u85_safe_save_workbook_atomic_base(wb, output_path)
    finally:
        wb.close()

def _u85_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _u84_processor_run_base = _u84_processor_run
    _u85_reset_bench()
    t_all_0 = time.perf_counter()
    try:
        _U84_WIDTH_CACHE_CTX['output_path'] = output_path
        _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

        t0 = time.perf_counter()
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        core_seconds = max(0.0, time.perf_counter() - t0)

        wb = open_or_create_output_workbook(output_path)
        try:
            detail_paths = _u79_collect_detail_paths_from_subdetail(wb['SUB_DETAIL']) if 'SUB_DETAIL' in wb.sheetnames else set()
        finally:
            wb.close()

        t1 = time.perf_counter()
        _u84_rebuild_merged_detail_rows(output_path, detail_paths, prep_label='DETAIL_PREP_U79')
        _u84_save_width_cache(output_path, _U84_WIDTH_CACHE_CTX.get('cache') or {})
        rebuild_seconds = max(0.0, time.perf_counter() - t1)

        total_seconds = max(0.0, time.perf_counter() - t_all_0)
        bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
        _u85_append_bench_log_rows(output_path, bench_rows)
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}
        _U85_BENCH_CTX['enabled'] = False


# =========================
# U86 FIXED ENTRY MARKER
# =========================
_u86_processor_run_base = _u85_processor_run

def _u86_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    self.logger.log('BENCH_ENTER_U86', 'info')
    return _u86_processor_run_base(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)




# =========================================================
# U87 SMART REBUILD - EXACT RESULT MATCH WITH ORIGINAL LOGIC
# Nguyen tac:
# - Ket qua cuoi cung phai 100%% giong full rebuild goc
# - Khong incremental seed / khong patch row truc tiep tren INV/PL
# - Van reconstruct toan bo INV/PL theo sorted active file_list
# - NHUNG bundle chi tiet cua moi file duoc cache rieng, file khong doi thi khong can doc/parsing lai
# =========================================================

import pickle

_U87_DETAIL_CACHE_VERSION = 1


def _u87_detail_cache_path(output_path: str) -> str:
    base = os.path.abspath(nz_str(output_path))
    return base + '.detailcache.pkl' if base else 'detail_sub.detailcache.pkl'


def _u87_load_detail_cache(output_path: str) -> Dict[str, Any]:
    path = _u87_detail_cache_path(output_path)
    try:
        with open(path, 'rb') as f:
            data = pickle.load(f)
        if not isinstance(data, dict):
            return {'version': _U87_DETAIL_CACHE_VERSION, 'files': {}}
        if int(data.get('version', 0) or 0) != _U87_DETAIL_CACHE_VERSION:
            return {'version': _U87_DETAIL_CACHE_VERSION, 'files': {}}
        files = data.get('files')
        if not isinstance(files, dict):
            files = {}
        return {'version': _U87_DETAIL_CACHE_VERSION, 'files': files}
    except Exception:
        return {'version': _U87_DETAIL_CACHE_VERSION, 'files': {}}


def _u87_save_detail_cache(output_path: str, cache: Dict[str, Any]):
    path = _u87_detail_cache_path(output_path)
    tmp = path + '.tmp'
    try:
        with open(tmp, 'wb') as f:
            pickle.dump(cache, f, protocol=pickle.HIGHEST_PROTOCOL)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except Exception:
                pass


def _u87_file_fingerprint(file_path: str) -> Tuple[int, int]:
    try:
        st = os.stat(file_path)
        return int(getattr(st, 'st_mtime_ns', int(st.st_mtime * 1_000_000_000))), int(st.st_size)
    except Exception:
        return (0, 0)


def _u87_norm_path_set(paths: Any) -> set:
    result = set()
    if not paths:
        return result
    for p in paths:
        n = _v75_norm_abs_upper(p)
        if n:
            result.add(n)
    return result


def _u87_get_bundle_from_cache_or_build(file_path: str, cache_files: Dict[str, Any], force_refresh: bool = False):
    key = _v75_norm_abs_upper(file_path)
    fp = _u87_file_fingerprint(file_path)
    item = cache_files.get(key, {}) if isinstance(cache_files, dict) else {}
    if (not force_refresh) and isinstance(item, dict) and tuple(item.get('fingerprint', ())) == fp and 'bundle' in item:
        bundle = item.get('bundle')
        if isinstance(bundle, tuple) and len(bundle) == 4:
            return bundle, 'cache'
    bundle = _v72_build_detail_bundle(file_path)
    if isinstance(cache_files, dict):
        cache_files[key] = {
            'fingerprint': fp,
            'bundle': bundle,
        }
    return bundle, 'rebuild'


def _u87_collect_rows_from_bundle_map(file_list: List[str], bundle_map: Dict[str, Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]]):
    inv_rows_main: List[List[Any]] = []
    inv_rows_other: List[List[Any]] = []
    pl_rows_main: List[List[Any]] = []
    pl_rows_other: List[List[Any]] = []

    for fp in file_list:
        bundle = bundle_map.get(fp)
        if not bundle:
            continue
        inv_rows, pl_rows, inv_other_rows, pl_other_rows = bundle
        inv_rows_main.extend(inv_rows or [])
        inv_rows_other.extend(inv_other_rows or [])
        pl_rows_main.extend(pl_rows or [])
        pl_rows_other.extend(pl_other_rows or [])

    inv_seed: Dict[Tuple[str, str], int] = {}
    pl_seed: Dict[Tuple[str, str], int] = {}

    inv_rows_main = _v73_apply_invoice_extras(inv_rows_main, inv_seed)
    inv_seed_after_main = _v74_seq_seed_from_rows(inv_rows_main, V73_INV_HEADERS, 'S. Invoice#')
    inv_rows_other = _v73_apply_invoice_extras(inv_rows_other, inv_seed_after_main)
    inv_rows_all = inv_rows_main + inv_rows_other

    pl_rows_main = _v76_apply_pl_sequence(pl_rows_main, pl_seed)
    pl_seed_after_main = _v74_seq_seed_from_rows(pl_rows_main, V76_PL_HEADERS, 'S. Invoice#')
    pl_rows_other = _v76_apply_pl_sequence(pl_rows_other, pl_seed_after_main)
    pl_rows_all = pl_rows_main + pl_rows_other
    return inv_rows_all, pl_rows_all


def _u87_rebuild_merged_detail_rows_smart(output_path: str, ok_paths: set, changed_paths: Optional[set] = None, prep_label: str = 'DETAIL_PREP_U79') -> Dict[str, Any]:
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    cache = _u87_load_detail_cache(output_path)
    cache_files: Dict[str, Any] = cache.get('files', {}) if isinstance(cache, dict) else {}
    if not isinstance(cache_files, dict):
        cache_files = {}
    file_list = sorted([p for p in ok_paths if os.path.exists(p)])
    active_set = set(file_list)
    changed_paths = _u87_norm_path_set(changed_paths)
    changed_paths &= active_set
    cache_hit = 0
    cache_miss = 0
    rebuilt_files = 0
    dirty_cache = False
    try:
        # recreate sheets => ket qua cuoi giong full rebuild, nhanh hon delete_rows
        ws_inv = _u83_recreate_sheet_keep_index(wb, 'INV')
        ws_pl = _u83_recreate_sheet_keep_index(wb, 'PL')
        for nm in ['INV Others', 'PL Others']:
            if nm in wb.sheetnames:
                del wb[nm]
        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        # remove stale cache entries not active anymore
        stale_keys = [k for k in list(cache_files.keys()) if k not in active_set]
        for k in stale_keys:
            try:
                del cache_files[k]
                dirty_cache = True
            except Exception:
                pass

        bundle_map: Dict[str, Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]] = {}
        for fp in file_list:
            try:
                force_refresh = (fp in changed_paths) or (fp not in cache_files)
                bundle, source = _u87_get_bundle_from_cache_or_build(fp, cache_files, force_refresh=force_refresh)
                bundle_map[fp] = bundle
                if source == 'cache':
                    cache_hit += 1
                else:
                    cache_miss += 1
                    rebuilt_files += 1
                    dirty_cache = True
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_rows_all, pl_rows_all = _u87_collect_rows_from_bundle_map(file_list, bundle_map)

        add_inv = _u83_append_rows_exact_fast(ws_inv, inv_rows_all, V73_INV_HEADERS, [1], 27)
        add_pl = _u83_append_rows_exact_fast(ws_pl, pl_rows_all, V76_PL_HEADERS, [1], 16)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V76_PL_HEADERS))
        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)

        ws_log = ensure_sheet(wb, 'LOG_SUB_DETAIL')
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row('', prep_label, 'INFO',
                            f'Paths for merged detail (OK+WARNING)={len(ok_paths)} | Start rebuild detail tabs | Merge mode=2 sheets | SmartCache=ON')])
        append_log_rows(ws_log, [make_log_row('', 'DETAIL_REBUILD_U87', 'INFO',
                            f'Current active files={len(file_list)} | Rebuilt detail tabs: INV={add_inv}, PL={add_pl} | cache_hit={cache_hit} | rebuilt={rebuilt_files} | stale_removed={len(stale_keys)}')])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, 'DETAIL_REBUILD_U87', 'ERROR', err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

    cache['version'] = _U87_DETAIL_CACHE_VERSION
    cache['files'] = cache_files
    if dirty_cache:
        _u87_save_detail_cache(output_path, cache)

    return {
        'active_files': len(file_list),
        'cache_hit': cache_hit,
        'cache_miss': cache_miss,
        'rebuilt_files': rebuilt_files,
        'stale_removed': len(stale_keys),
        'errors': len(errors),
        'cache_path': _u87_detail_cache_path(output_path),
    }


def _u87_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    # benchmark-safe + smart rebuild exact logic
    _u85_reset_bench()
    t_all_0 = time.perf_counter()
    try:
        self.logger.log('SMART_REBUILD_ENTER_U87', 'info')
        _U84_WIDTH_CACHE_CTX['output_path'] = output_path
        _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

        t0 = time.perf_counter()
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        core_seconds = max(0.0, time.perf_counter() - t0)

        wb = open_or_create_output_workbook(output_path)
        try:
            detail_paths = _u79_collect_detail_paths_from_subdetail(wb['SUB_DETAIL']) if 'SUB_DETAIL' in wb.sheetnames else set()
        finally:
            wb.close()

        changed_paths = _u87_norm_path_set((summary or {}).get('scanned_files', []))

        t1 = time.perf_counter()
        smart_stats = _u87_rebuild_merged_detail_rows_smart(output_path, detail_paths, changed_paths=changed_paths, prep_label='DETAIL_PREP_U79')
        _u84_save_width_cache(output_path, _U84_WIDTH_CACHE_CTX.get('cache') or {})
        rebuild_seconds = max(0.0, time.perf_counter() - t1)

        total_seconds = max(0.0, time.perf_counter() - t_all_0)
        bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
        bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                        f"active={smart_stats.get('active_files',0)} | cache_hit={smart_stats.get('cache_hit',0)} | rebuilt={smart_stats.get('rebuilt_files',0)} | stale_removed={smart_stats.get('stale_removed',0)} | cache={smart_stats.get('cache_path','')}"))
        _u85_append_bench_log_rows(output_path, bench_rows)
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}
        _U85_BENCH_CTX['enabled'] = False



# [u90] disabled earlier entrypoint


# =========================================================
# U88 SOURCE PARSE CACHE
# Nguyen tac:
# - Giu ket qua 100%% giong logic goc / u87
# - Khong doi output cuoi
# - Giam CORE_WRITE bang cache ket qua parse file nguon theo fingerprint
# - Cache ca 3 lop:
#     + process_one_sub_file(file_path) -> arr 25 cot
#     + process_one_sub_file_bundle(file_path) -> bundle core (arr, inv_rows, pl_rows)
#     + _v72_build_detail_bundle(file_path) -> bundle smart rebuild 4 tap rows
# =========================================================

_U88_SOURCE_CACHE_VERSION = 1
_U88_SOURCE_CACHE_CTX = {
    'output_path': '',
    'cache': None,
}


def _u88_source_cache_path(output_path: str) -> str:
    base = os.path.abspath(nz_str(output_path))
    return base + '.sourcecache.pkl' if base else 'detail_sub.sourcecache.pkl'


def _u88_empty_source_cache() -> Dict[str, Any]:
    return {
        'version': _U88_SOURCE_CACHE_VERSION,
        'files': {},
        'stats': {
            'arr_hit': 0,
            'arr_miss': 0,
            'bundle_hit': 0,
            'bundle_miss': 0,
            'detail_hit': 0,
            'detail_miss': 0,
        }
    }


def _u88_get_source_cache() -> Dict[str, Any]:
    cache = _U88_SOURCE_CACHE_CTX.get('cache')
    if isinstance(cache, dict):
        if 'files' not in cache or not isinstance(cache.get('files'), dict):
            cache['files'] = {}
        if 'stats' not in cache or not isinstance(cache.get('stats'), dict):
            cache['stats'] = {}
        return cache
    cache = _u88_empty_source_cache()
    _U88_SOURCE_CACHE_CTX['cache'] = cache
    return cache


def _u88_load_source_cache(output_path: str) -> Dict[str, Any]:
    path = _u88_source_cache_path(output_path)
    try:
        with open(path, 'rb') as f:
            data = pickle.load(f)
        if not isinstance(data, dict):
            return _u88_empty_source_cache()
        if int(data.get('version', 0) or 0) != _U88_SOURCE_CACHE_VERSION:
            return _u88_empty_source_cache()
        files = data.get('files')
        if not isinstance(files, dict):
            files = {}
        stats = data.get('stats')
        if not isinstance(stats, dict):
            stats = {}
        loaded = {'version': _U88_SOURCE_CACHE_VERSION, 'files': files, 'stats': stats}
        return loaded
    except Exception:
        return _u88_empty_source_cache()


def _u88_save_source_cache(output_path: str, cache: Dict[str, Any]):
    path = _u88_source_cache_path(output_path)
    tmp = path + '.tmp'
    try:
        with open(tmp, 'wb') as f:
            pickle.dump(cache, f, protocol=pickle.HIGHEST_PROTOCOL)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except Exception:
                pass


def _u88_source_key(file_path: str) -> str:
    return _v75_norm_abs_upper(file_path)


def _u88_source_fp(file_path: str) -> Tuple[int, int]:
    try:
        st = os.stat(file_path)
        return int(getattr(st, 'st_mtime_ns', int(st.st_mtime * 1_000_000_000))), int(st.st_size)
    except Exception:
        return (0, 0)


def _u88_source_item(cache: Dict[str, Any], file_path: str) -> Dict[str, Any]:
    files = cache.setdefault('files', {})
    key = _u88_source_key(file_path)
    item = files.get(key)
    if not isinstance(item, dict):
        item = {}
        files[key] = item
    return item


def _u88_stats_add(cache: Dict[str, Any], name: str, inc: int = 1):
    stats = cache.setdefault('stats', {})
    try:
        stats[name] = int(stats.get(name, 0) or 0) + int(inc)
    except Exception:
        pass


_u88_process_one_sub_file_base = process_one_sub_file
_u88_process_one_sub_file_bundle_base = process_one_sub_file_bundle
_u88_build_detail_bundle_base = _v72_build_detail_bundle


def process_one_sub_file(file_path: str) -> List[Any]:
    cache = _u88_get_source_cache()
    item = _u88_source_item(cache, file_path)
    fp = _u88_source_fp(file_path)
    if tuple(item.get('fingerprint', ())) == fp and 'arr25' in item:
        arr = item.get('arr25')
        if isinstance(arr, list) and len(arr) >= 25:
            _u88_stats_add(cache, 'arr_hit')
            return list(arr)
    arr = _u88_process_one_sub_file_base(file_path)
    item['fingerprint'] = fp
    item['arr25'] = list(arr) if isinstance(arr, list) else arr
    _u88_stats_add(cache, 'arr_miss')
    return arr



def process_one_sub_file_bundle(file_path: str):
    cache = _u88_get_source_cache()
    item = _u88_source_item(cache, file_path)
    fp = _u88_source_fp(file_path)
    if tuple(item.get('fingerprint', ())) == fp and 'core_bundle' in item:
        bundle = item.get('core_bundle')
        if isinstance(bundle, tuple) and len(bundle) == 3:
            _u88_stats_add(cache, 'bundle_hit')
            arr, inv_rows, pl_rows = bundle
            return (list(arr), list(inv_rows), list(pl_rows))
    bundle = _u88_process_one_sub_file_bundle_base(file_path)
    try:
        arr, inv_rows, pl_rows = bundle
        item['fingerprint'] = fp
        item['arr25'] = list(arr) if isinstance(arr, list) else arr
        item['core_bundle'] = (
            list(arr) if isinstance(arr, list) else arr,
            list(inv_rows) if isinstance(inv_rows, list) else inv_rows,
            list(pl_rows) if isinstance(pl_rows, list) else pl_rows,
        )
    except Exception:
        pass
    _u88_stats_add(cache, 'bundle_miss')
    return bundle



def _v72_build_detail_bundle(file_path: str):
    cache = _u88_get_source_cache()
    item = _u88_source_item(cache, file_path)
    fp = _u88_source_fp(file_path)
    if tuple(item.get('fingerprint', ())) == fp and 'detail_bundle4' in item:
        bundle = item.get('detail_bundle4')
        if isinstance(bundle, tuple) and len(bundle) == 4:
            _u88_stats_add(cache, 'detail_hit')
            a, b, c, d = bundle
            return (list(a), list(b), list(c), list(d))
    bundle = _u88_build_detail_bundle_base(file_path)
    try:
        a, b, c, d = bundle
        item['fingerprint'] = fp
        item['detail_bundle4'] = (
            list(a) if isinstance(a, list) else a,
            list(b) if isinstance(b, list) else b,
            list(c) if isinstance(c, list) else c,
            list(d) if isinstance(d, list) else d,
        )
    except Exception:
        pass
    _u88_stats_add(cache, 'detail_miss')
    return bundle


# Force late references to use cache-enabled builder
_v72_build_detail_bundle = _v72_build_detail_bundle


_u88_processor_run_base = _u87_processor_run


def _u88_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _U88_SOURCE_CACHE_CTX['output_path'] = output_path
    _U88_SOURCE_CACHE_CTX['cache'] = _u88_load_source_cache(output_path)
    self.logger.log('SOURCE_PARSE_CACHE_ENTER_U88', 'info')
    try:
        summary = _u88_processor_run_base(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        cache = _u88_get_source_cache()
        _u88_save_source_cache(output_path, cache)
        stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
        rows = [make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u88_source_cache_path(output_path)}")]
        _u85_append_bench_log_rows(output_path, rows)
        return summary
    finally:
        _U88_SOURCE_CACHE_CTX['output_path'] = ''
        _U88_SOURCE_CACHE_CTX['cache'] = None



# [u90] disabled earlier entrypoint


# =========================================================
# U90 FINAL ZERO-READ CACHE + FIX CORE SOURCE SCAN
# Nguyen tac:
# - Giu ket qua cuoi 100% giong logic u87/u88
# - Chi toi uu bang cach:
#   1) CORE_RUN_V6 chi scan folder hien tai, khong lap lai toan bo ok_paths cu
#   2) File nguon khong doi => zero-read: khong mo Excel, tra thang cache arr/core/detail
# - Ghi cache rieng de tranh dung file cache cu:
#   *.sourcecache_u90.pkl
# =========================================================

_U90_SOURCE_CACHE_VERSION = 1
_U90_SOURCE_CACHE_CTX = {
    'output_path': '',
    'cache': None,
}

def _u90_source_cache_path(output_path: str) -> str:
    base = os.path.abspath(nz_str(output_path))
    return base + '.sourcecache_u90.pkl' if base else 'detail_sub.sourcecache_u90.pkl'

def _u90_empty_source_cache() -> Dict[str, Any]:
    return {
        'version': _U90_SOURCE_CACHE_VERSION,
        'files': {},
        'stats': {
            'arr_hit': 0,
            'arr_miss': 0,
            'bundle_hit': 0,
            'bundle_miss': 0,
            'detail_hit': 0,
            'detail_miss': 0,
        }
    }

def _u90_get_source_cache() -> Dict[str, Any]:
    cache = _U90_SOURCE_CACHE_CTX.get('cache')
    if isinstance(cache, dict):
        if 'files' not in cache or not isinstance(cache.get('files'), dict):
            cache['files'] = {}
        if 'stats' not in cache or not isinstance(cache.get('stats'), dict):
            cache['stats'] = {}
        return cache
    cache = _u90_empty_source_cache()
    _U90_SOURCE_CACHE_CTX['cache'] = cache
    return cache

def _u90_load_source_cache(output_path: str) -> Dict[str, Any]:
    path = _u90_source_cache_path(output_path)
    try:
        with open(path, 'rb') as f:
            data = pickle.load(f)
        if not isinstance(data, dict):
            return _u90_empty_source_cache()
        if int(data.get('version', 0) or 0) != _U90_SOURCE_CACHE_VERSION:
            return _u90_empty_source_cache()
        files = data.get('files')
        if not isinstance(files, dict):
            files = {}
        loaded = {'version': _U90_SOURCE_CACHE_VERSION, 'files': files, 'stats': _u90_empty_source_cache()['stats']}
        return loaded
    except Exception:
        return _u90_empty_source_cache()

def _u90_save_source_cache(output_path: str, cache: Dict[str, Any]):
    path = _u90_source_cache_path(output_path)
    tmp = path + '.tmp'
    try:
        with open(tmp, 'wb') as f:
            pickle.dump(cache, f, protocol=pickle.HIGHEST_PROTOCOL)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except Exception:
                pass

def _u90_source_key(file_path: str) -> str:
    return _v75_norm_abs_upper(file_path)

def _u90_source_fp(file_path: str) -> Tuple[int, int]:
    try:
        st = os.stat(file_path)
        return int(getattr(st, 'st_mtime_ns', int(st.st_mtime * 1_000_000_000))), int(st.st_size)
    except Exception:
        return (0, 0)

def _u90_source_item(cache: Dict[str, Any], file_path: str) -> Dict[str, Any]:
    files = cache.setdefault('files', {})
    key = _u90_source_key(file_path)
    item = files.get(key)
    if not isinstance(item, dict):
        item = {}
        files[key] = item
    return item

def _u90_stats_add(cache: Dict[str, Any], name: str, inc: int = 1):
    stats = cache.setdefault('stats', {})
    try:
        stats[name] = int(stats.get(name, 0) or 0) + int(inc)
    except Exception:
        pass

# Luu base goc truoc khi override lan nua
_u90_process_one_sub_file_base = _u88_process_one_sub_file_base
_u90_process_one_sub_file_bundle_base = _u88_process_one_sub_file_bundle_base
_u90_build_detail_bundle_base = _u88_build_detail_bundle_base

def process_one_sub_file(file_path: str) -> List[Any]:
    cache = _u90_get_source_cache()
    item = _u90_source_item(cache, file_path)
    fp = _u90_source_fp(file_path)
    if tuple(item.get('fingerprint', ())) == fp and 'arr25' in item:
        arr = item.get('arr25')
        if isinstance(arr, list) and len(arr) >= 25:
            _u90_stats_add(cache, 'arr_hit')
            return list(arr)
    arr = _u90_process_one_sub_file_base(file_path)
    item['fingerprint'] = fp
    item['arr25'] = list(arr) if isinstance(arr, list) else arr
    _u90_stats_add(cache, 'arr_miss')
    return arr

def process_one_sub_file_bundle(file_path: str):
    cache = _u90_get_source_cache()
    item = _u90_source_item(cache, file_path)
    fp = _u90_source_fp(file_path)
    if tuple(item.get('fingerprint', ())) == fp and 'core_bundle' in item:
        bundle = item.get('core_bundle')
        if isinstance(bundle, tuple) and len(bundle) == 3:
            _u90_stats_add(cache, 'bundle_hit')
            arr, inv_rows, pl_rows = bundle
            return (
                list(arr) if isinstance(arr, list) else arr,
                list(inv_rows) if isinstance(inv_rows, list) else inv_rows,
                list(pl_rows) if isinstance(pl_rows, list) else pl_rows,
            )
    bundle = _u90_process_one_sub_file_bundle_base(file_path)
    try:
        arr, inv_rows, pl_rows = bundle
        item['fingerprint'] = fp
        item['arr25'] = list(arr) if isinstance(arr, list) else arr
        item['core_bundle'] = (
            list(arr) if isinstance(arr, list) else arr,
            list(inv_rows) if isinstance(inv_rows, list) else inv_rows,
            list(pl_rows) if isinstance(pl_rows, list) else pl_rows,
        )
    except Exception:
        pass
    _u90_stats_add(cache, 'bundle_miss')
    return bundle

def _v72_build_detail_bundle(file_path: str):
    cache = _u90_get_source_cache()
    item = _u90_source_item(cache, file_path)
    fp = _u90_source_fp(file_path)
    if tuple(item.get('fingerprint', ())) == fp and 'detail_bundle4' in item:
        bundle = item.get('detail_bundle4')
        if isinstance(bundle, tuple) and len(bundle) == 4:
            _u90_stats_add(cache, 'detail_hit')
            a, b, c, d = bundle
            return (
                list(a) if isinstance(a, list) else a,
                list(b) if isinstance(b, list) else b,
                list(c) if isinstance(c, list) else c,
                list(d) if isinstance(d, list) else d,
            )
    bundle = _u90_build_detail_bundle_base(file_path)
    try:
        a, b, c, d = bundle
        item['fingerprint'] = fp
        item['detail_bundle4'] = (
            list(a) if isinstance(a, list) else a,
            list(b) if isinstance(b, list) else b,
            list(c) if isinstance(c, list) else c,
            list(d) if isinstance(d, list) else d,
        )
    except Exception:
        pass
    _u90_stats_add(cache, 'detail_miss')
    return bundle

# Force late references
_v72_build_detail_bundle = _v72_build_detail_bundle


def _u90_core_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None) -> Dict[str, int]:
    repair_options = repair_options or RepairOptions()
    wb_out = open_or_create_output_workbook(output_path)
    try:
        ws_out = ensure_sheet(wb_out, "SUB_DETAIL")
        ws_log = ensure_sheet(wb_out, "LOG_SUB_DETAIL")
        ws_inv = ensure_sheet(wb_out, "INV")
        ws_pl = ensure_sheet(wb_out, "PL")
        if "Sheet" in wb_out.sheetnames and wb_out["Sheet"].max_row == 1 and wb_out["Sheet"].max_column == 1 and nz_str(wb_out["Sheet"]["A1"].value) == "":
            try:
                del wb_out["Sheet"]
            except Exception:
                pass

        init_sub_detail_header(ws_out)
        init_log_header(ws_log)
        init_inv_detail_header(ws_inv)
        init_pl_detail_header(ws_pl)

        kept_ok_rows, rerun_count = keep_only_ok_rows(ws_out)
        processed_paths, processed_signs = build_processed_sets_from_rows(kept_ok_rows)
        ok_paths = set(processed_paths)
        keep_rows_by_file_paths(ws_inv, len(INV_DETAIL_HEADERS), 27, ok_paths)
        keep_rows_by_file_paths(ws_pl, len(PL_DETAIL_HEADERS), 23, ok_paths)

        # FIX LON NHAT CUA U90:
        # Chi scan file trong folder hien tai, khong lap qua toan bo OK paths cu.
        src_files = sorted([p for p in collect_source_files(folder_path, output_path) if os.path.exists(p)])

        cnt_total = len(src_files)
        cnt_ok = cnt_warn = cnt_err = cnt_skip_path = cnt_skip_sig = 0
        cnt_repaired = 0

        log_rows: List[List[Any]] = []
        result_rows: List[List[Any]] = []
        inv_detail_rows: List[List[Any]] = []
        pl_detail_rows: List[List[Any]] = []
        issue_files: List[Tuple[str, str, str]] = []

        if cnt_total == 0:
            log_rows.append(make_log_row("", "SCAN", "INFO", "Khong tim thay file Excel nao de xu ly."))
            append_log_rows(ws_log, log_rows)
            safe_save_workbook_atomic(wb_out, output_path)
            return {
                "total": 0, "ok": 0, "warning": 0, "error": 0,
                "skip_path": 0, "skip_signature": 0, "repaired": 0, "issues": [], "scanned_files": src_files,
            }

        log_rows.append(make_log_row("", "START", "INFO", f"Bat dau quet {cnt_total} file."))
        if rerun_count > 0:
            log_rows.append(make_log_row("", "RERUN", "INFO", f"Se chay lai {rerun_count} file WARNING/ERROR tu lan truoc, dong thoi xu ly file moi phat sinh."))

        def job(path: str):
            arr, inv_rows, pl_rows = process_one_sub_file_bundle(path)
            return path, arr, inv_rows, pl_rows

        futures = {}
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            for fp in src_files:
                fp_norm = fp.strip().upper()
                if fp_norm == output_path.strip().upper():
                    log_rows.append(make_log_row(fp, "SKIP", "INFO", "Bo qua output workbook."))
                    continue
                if fp_norm in processed_paths:
                    cnt_skip_path += 1
                    continue
                futures[executor.submit(job, fp)] = fp

            completed = 0
            total_futures = max(1, len(futures))
            for future in as_completed(futures):
                file_path, arr, inv_rows_one, pl_rows_one = future.result()
                completed += 1
                sig = nz_str(arr[24])
                status_text = nz_str(arr[21]).upper()
                err_text = nz_str(arr[22])

                if sig and sig.upper() in processed_signs:
                    cnt_skip_sig += 1
                    if progress_callback:
                        progress_callback(completed, total_futures, file_path)
                    continue

                if status_text in ("WARNING", "ERROR"):
                    issue_files.append((file_path, status_text, err_text))
                    self.logger.log(f"{status_text}: {get_file_name_from_path(file_path)} -> {err_text}", "warn" if status_text == "WARNING" else "error")
                    log_rows.append(make_log_row(file_path, "CHECK", status_text, err_text))

                    if repair_options.use_folder_truth or repair_options.use_manual_values:
                        repaired, detail = repair_excel_metadata(file_path, arr, repair_options)
                        if repaired:
                            cnt_repaired += 1
                            self.logger.log(f"REPAIRED: {get_file_name_from_path(file_path)} -> {detail}", "info")
                            log_rows.append(make_log_row(file_path, "REPAIR", "OK", detail))
                            arr = process_one_sub_file(file_path)
                            status_text = nz_str(arr[21]).upper()
                            err_text = nz_str(arr[22])
                        else:
                            self.logger.log(f"REPAIR-SKIP: {get_file_name_from_path(file_path)} -> {detail}", "warn")
                            log_rows.append(make_log_row(file_path, "REPAIR", "SKIP", detail))

                if status_text == "OK":
                    cnt_ok += 1
                elif status_text == "WARNING":
                    cnt_warn += 1
                else:
                    cnt_err += 1

                result_rows.append(arr)
                inv_detail_rows.extend(inv_rows_one)
                pl_detail_rows.extend(pl_rows_one)
                if sig:
                    processed_signs.add(sig.upper())
                processed_paths.add(file_path.strip().upper())

                if progress_callback:
                    progress_callback(completed, total_futures, file_path)

        if issue_files:
            log_rows.append(make_log_row("", "SUMMARY", "INFO", "TONG HOP FILE WARNING/ERROR"))
            for fp, st, detail in issue_files:
                log_rows.append(make_log_row(fp, "SUMMARY", st, detail))

        if cnt_skip_path > 0:
            log_rows.append(make_log_row("", "SKIP", "INFO", f"So luong bo qua do da co File Path: {cnt_skip_path}"))
        if cnt_skip_sig > 0:
            log_rows.append(make_log_row("", "SKIP", "INFO", f"So luong bo qua do trung File Signature: {cnt_skip_sig}"))

        log_rows.append(make_log_row("", "DONE", "INFO",
                                     f"Tong={cnt_total}; OK={cnt_ok}; WARNING={cnt_warn}; ERROR={cnt_err}; SkipPath={cnt_skip_path}; SkipSignature={cnt_skip_sig}; Repaired={cnt_repaired}"))

        append_result_rows(ws_out, result_rows)
        append_generic_rows(ws_inv, inv_detail_rows, [5, 8], 26)
        append_generic_rows(ws_pl, pl_detail_rows, [5, 8], 22)
        append_log_rows(ws_log, log_rows)
        autofit_useful(ws_out, len(SUB_DETAIL_HEADERS))
        autofit_useful(ws_inv, len(INV_DETAIL_HEADERS))
        autofit_useful(ws_pl, len(PL_DETAIL_HEADERS))
        autofit_useful(ws_log, len(LOG_HEADERS))
        safe_save_workbook_atomic(wb_out, output_path)

        return {
            "total": cnt_total, "ok": cnt_ok, "warning": cnt_warn, "error": cnt_err,
            "skip_path": cnt_skip_path, "skip_signature": cnt_skip_sig, "repaired": cnt_repaired,
            "issues": issue_files, "scanned_files": src_files,
        }
    finally:
        wb_out.close()


# IMPORTANT: U87 benchmark/smart rebuild se goi CORE_RUN_V6 nay
CORE_RUN_V6 = _u90_core_run


_u90_processor_run_base = _u87_processor_run

def _u90_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _U90_SOURCE_CACHE_CTX['output_path'] = output_path
    _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
    cache = _u90_get_source_cache()
    cache['stats'] = _u90_empty_source_cache()['stats']
    self.logger.log('SOURCE_PARSE_CACHE_ENTER_U90', 'info')
    try:
        summary = _u90_processor_run_base(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        cache = _u90_get_source_cache()
        _u90_save_source_cache(output_path, cache)
        stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
        rows = [make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}")]
        _u85_append_bench_log_rows(output_path, rows)
        return summary
    finally:
        _U90_SOURCE_CACHE_CTX['output_path'] = ''
        _U90_SOURCE_CACHE_CTX['cache'] = None



# =========================================================
# U92 SMART REBUILD TRUE CACHE HIT
# Nguyen tac:
# - Giu ket qua cuoi giong U87/U90
# - Core lean: chi ghi SUB_DETAIL + LOG_SUB_DETAIL
# - Khong ghi INV/PL trong core
# - Smart rebuild chi force refresh cho file THUC SU thay doi trong lan run nay
# - Gop benchmark + source cache stats vao 1 lan append log cuoi de giam 1 lan save
# =========================================================

def _u92_core_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None) -> Dict[str, int]:
    repair_options = repair_options or RepairOptions()
    wb_out = open_or_create_output_workbook(output_path)
    try:
        ws_out = ensure_sheet(wb_out, "SUB_DETAIL")
        ws_log = ensure_sheet(wb_out, "LOG_SUB_DETAIL")
        if "Sheet" in wb_out.sheetnames and wb_out["Sheet"].max_row == 1 and wb_out["Sheet"].max_column == 1 and nz_str(wb_out["Sheet"]["A1"].value) == "":
            try:
                del wb_out["Sheet"]
            except Exception:
                pass

        init_sub_detail_header(ws_out)
        init_log_header(ws_log)

        kept_ok_rows, rerun_count = keep_only_ok_rows(ws_out)
        processed_paths, processed_signs = build_processed_sets_from_rows(kept_ok_rows)

        src_files = sorted([p for p in collect_source_files(folder_path, output_path) if os.path.exists(p)])

        cnt_total = len(src_files)
        cnt_ok = cnt_warn = cnt_err = cnt_skip_path = cnt_skip_sig = 0
        cnt_repaired = 0

        log_rows: List[List[Any]] = []
        result_rows: List[List[Any]] = []
        issue_files: List[Tuple[str, str, str]] = []
        changed_detail_files: List[str] = []

        if cnt_total == 0:
            log_rows.append(make_log_row("", "SCAN", "INFO", "Khong tim thay file Excel nao de xu ly."))
            append_log_rows(ws_log, log_rows)
            autofit_useful(ws_out, len(SUB_DETAIL_HEADERS))
            autofit_useful(ws_log, len(LOG_HEADERS))
            safe_save_workbook_atomic(wb_out, output_path)
            return {
                "total": 0, "ok": 0, "warning": 0, "error": 0,
                "skip_path": 0, "skip_signature": 0, "repaired": 0,
                "issues": [], "scanned_files": src_files, "changed_detail_files": [],
            }

        log_rows.append(make_log_row("", "START", "INFO", f"Bat dau quet {cnt_total} file."))
        if rerun_count > 0:
            log_rows.append(make_log_row("", "RERUN", "INFO", f"Se chay lai {rerun_count} file WARNING/ERROR tu lan truoc, dong thoi xu ly file moi phat sinh."))

        def job(path: str):
            arr = process_one_sub_file(path)
            return path, arr

        futures = {}
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            for fp in src_files:
                fp_norm = fp.strip().upper()
                if fp_norm == output_path.strip().upper():
                    log_rows.append(make_log_row(fp, "SKIP", "INFO", "Bo qua output workbook."))
                    continue
                if fp_norm in processed_paths:
                    cnt_skip_path += 1
                    continue
                futures[executor.submit(job, fp)] = fp

            completed = 0
            total_futures = max(1, len(futures))
            for future in as_completed(futures):
                file_path, arr = future.result()
                completed += 1
                sig = nz_str(arr[24]) if isinstance(arr, list) and len(arr) > 24 else ""
                status_text = nz_str(arr[21]).upper() if isinstance(arr, list) and len(arr) > 21 else "ERROR"
                err_text = nz_str(arr[22]) if isinstance(arr, list) and len(arr) > 22 else ""

                if sig and sig.upper() in processed_signs:
                    cnt_skip_sig += 1
                    if progress_callback:
                        progress_callback(completed, total_futures, file_path)
                    continue

                if status_text in ("WARNING", "ERROR"):
                    issue_files.append((file_path, status_text, err_text))
                    self.logger.log(f"{status_text}: {get_file_name_from_path(file_path)} -> {err_text}", "warn" if status_text == "WARNING" else "error")
                    log_rows.append(make_log_row(file_path, "CHECK", status_text, err_text))

                    if repair_options.use_folder_truth or repair_options.use_manual_values:
                        repaired, detail = repair_excel_metadata(file_path, arr, repair_options)
                        if repaired:
                            cnt_repaired += 1
                            self.logger.log(f"REPAIRED: {get_file_name_from_path(file_path)} -> {detail}", "info")
                            log_rows.append(make_log_row(file_path, "REPAIR", "OK", detail))
                            arr = process_one_sub_file(file_path)
                            status_text = nz_str(arr[21]).upper() if isinstance(arr, list) and len(arr) > 21 else "ERROR"
                            err_text = nz_str(arr[22]) if isinstance(arr, list) and len(arr) > 22 else ""
                        else:
                            self.logger.log(f"REPAIR-SKIP: {get_file_name_from_path(file_path)} -> {detail}", "warn")
                            log_rows.append(make_log_row(file_path, "REPAIR", "SKIP", detail))

                if status_text == "OK":
                    cnt_ok += 1
                elif status_text == "WARNING":
                    cnt_warn += 1
                else:
                    cnt_err += 1

                result_rows.append(arr)
                changed_detail_files.append(file_path)

                if sig:
                    processed_signs.add(sig.upper())
                processed_paths.add(file_path.strip().upper())

                if progress_callback:
                    progress_callback(completed, total_futures, file_path)

        if issue_files:
            log_rows.append(make_log_row("", "SUMMARY", "INFO", "TONG HOP FILE WARNING/ERROR"))
            for fp, st, detail in issue_files:
                log_rows.append(make_log_row(fp, "SUMMARY", st, detail))

        if cnt_skip_path > 0:
            log_rows.append(make_log_row("", "SKIP", "INFO", f"So luong bo qua do da co File Path: {cnt_skip_path}"))
        if cnt_skip_sig > 0:
            log_rows.append(make_log_row("", "SKIP", "INFO", f"So luong bo qua do trung File Signature: {cnt_skip_sig}"))

        log_rows.append(make_log_row("", "DONE", "INFO",
                                     f"Tong={cnt_total}; OK={cnt_ok}; WARNING={cnt_warn}; ERROR={cnt_err}; SkipPath={cnt_skip_path}; SkipSignature={cnt_skip_sig}; Repaired={cnt_repaired}"))

        append_result_rows(ws_out, result_rows)
        append_log_rows(ws_log, log_rows)
        autofit_useful(ws_out, len(SUB_DETAIL_HEADERS))
        autofit_useful(ws_log, len(LOG_HEADERS))
        safe_save_workbook_atomic(wb_out, output_path)

        return {
            "total": cnt_total, "ok": cnt_ok, "warning": cnt_warn, "error": cnt_err,
            "skip_path": cnt_skip_path, "skip_signature": cnt_skip_sig, "repaired": cnt_repaired,
            "issues": issue_files, "scanned_files": src_files, "changed_detail_files": changed_detail_files,
        }
    finally:
        wb_out.close()


# IMPORTANT: U92 thay core cho U87/U90
CORE_RUN_V6 = _u92_core_run


def _u93_collect_existing_active_paths(output_path: str) -> set:
    wb = open_or_create_output_workbook(output_path)
    try:
        if 'SUB_DETAIL' not in wb.sheetnames:
            return set()
        return _u79_collect_detail_paths_from_subdetail(wb['SUB_DETAIL'])
    finally:
        wb.close()


def _u94_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _u85_reset_bench()
    t_all_0 = time.perf_counter()
    _U90_SOURCE_CACHE_CTX['output_path'] = output_path
    _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
    cache = _u90_get_source_cache()
    cache['stats'] = _u90_empty_source_cache()['stats']

    try:
        self.logger.log('SMART_REBUILD_ENTER_U94', 'info')
        self.logger.log('SOURCE_PARSE_CACHE_ENTER_U94', 'info')

        _U84_WIDTH_CACHE_CTX['output_path'] = output_path
        _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

        # IMPORTANT U94:
        # Lay active paths TRUOC KHI core run.
        existing_active_paths = _u93_collect_existing_active_paths(output_path)

        # Scanned files cua lan chay nay - dung de tinh delta dung thoi diem.
        scanned_files = _u87_norm_path_set(collect_source_files(folder_path, output_path))

        t0 = time.perf_counter()
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        core_seconds = max(0.0, time.perf_counter() - t0)

        # Sau core moi doc current active set de rebuild ket qua cuoi.
        current_active_paths = _u93_collect_existing_active_paths(output_path)

        # U94 TRUE DELTA CORRECT TIMING:
        # - KHONG lay current - existing
        # - CHI lay file scanned cua lan nay ma truoc do CHUA co trong active set
        changed_paths = _u87_norm_path_set(p for p in scanned_files if p not in existing_active_paths)

        t1 = time.perf_counter()
        smart_stats = _u87_rebuild_merged_detail_rows_smart(
            output_path,
            current_active_paths,
            changed_paths=changed_paths,
            prep_label='DETAIL_PREP_U79'
        )
        _u84_save_width_cache(output_path, _U84_WIDTH_CACHE_CTX.get('cache') or {})
        rebuild_seconds = max(0.0, time.perf_counter() - t1)

        total_seconds = max(0.0, time.perf_counter() - t_all_0)

        cache = _u90_get_source_cache()
        _u90_save_source_cache(output_path, cache)
        stats = cache.get('stats', {}) if isinstance(cache, dict) else {}

        bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
        bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                        f"active={smart_stats.get('active_files',0)} | scanned={len(scanned_files)} | changed={len(changed_paths)} | cache_hit={smart_stats.get('cache_hit',0)} | rebuilt={smart_stats.get('rebuilt_files',0)} | stale_removed={smart_stats.get('stale_removed',0)} | cache={smart_stats.get('cache_path','')}"))
        bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                        f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
        _u85_append_bench_log_rows(output_path, bench_rows)
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}
        _U85_BENCH_CTX['enabled'] = False
        _U90_SOURCE_CACHE_CTX['output_path'] = ''
        _U90_SOURCE_CACHE_CTX['cache'] = None




# =========================================================
# U95 INCREMENTAL WRITE FOR INV / PL
# Nguyen tac:
# - Giu logic ket qua cuoi giong U94/U87
# - Khong full rewrite neu chi co phan duoi thay doi
# - Ghi theo block-row architecture
# - Moi file co the tao 2 block tren moi sheet: main / other
# - Neu metadata index hu / khong dong bo -> fallback full rewrite
# - Neu co common-prefix lon -> chi rewrite tail
# =========================================================

import hashlib

_U95_BLOCK_INDEX_VERSION = 1
_U95_BLOCK_INDEX_SHEET = "_U95_BLOCK_INDEX"


def _u95_block_id(kind: str, file_path: str) -> str:
    return f"{kind}|{_v75_norm_abs_upper(file_path)}"


def _u95_split_rows_by_lengths(rows: List[List[Any]], lengths: List[int]) -> List[List[List[Any]]]:
    out: List[List[List[Any]]] = []
    pos = 0
    for n in lengths:
        n = int(n or 0)
        if n <= 0:
            out.append([])
            continue
        out.append(rows[pos:pos+n])
        pos += n
    return out


def _u95_hash_rows(rows: List[List[Any]], headers: List[str]) -> str:
    if not rows:
        return "0"
    payload = []
    for arr in rows:
        arr = _v77_align_row_to_headers(arr, headers)
        vals = []
        for v in arr:
            if isinstance(v, datetime):
                vals.append(v.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                vals.append(nz_str(v))
        payload.append(vals)
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return hashlib.md5(raw.encode("utf-8", errors="ignore")).hexdigest()


def _u95_collect_final_blocks(file_list: List[str], bundle_map: Dict[str, Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]]):
    inv_main_by_file: List[Tuple[str, List[List[Any]]]] = []
    inv_other_by_file: List[Tuple[str, List[List[Any]]]] = []
    pl_main_by_file: List[Tuple[str, List[List[Any]]]] = []
    pl_other_by_file: List[Tuple[str, List[List[Any]]]] = []

    inv_main_all: List[List[Any]] = []
    inv_other_all: List[List[Any]] = []
    pl_main_all: List[List[Any]] = []
    pl_other_all: List[List[Any]] = []

    inv_main_lengths: List[int] = []
    inv_other_lengths: List[int] = []
    pl_main_lengths: List[int] = []
    pl_other_lengths: List[int] = []

    for fp in file_list:
        bundle = bundle_map.get(fp)
        if not bundle:
            inv_rows = []
            pl_rows = []
            inv_other_rows = []
            pl_other_rows = []
        else:
            inv_rows, pl_rows, inv_other_rows, pl_other_rows = bundle

        inv_rows = list(inv_rows or [])
        pl_rows = list(pl_rows or [])
        inv_other_rows = list(inv_other_rows or [])
        pl_other_rows = list(pl_other_rows or [])

        inv_main_by_file.append((fp, inv_rows))
        inv_other_by_file.append((fp, inv_other_rows))
        pl_main_by_file.append((fp, pl_rows))
        pl_other_by_file.append((fp, pl_other_rows))

        inv_main_lengths.append(len(inv_rows))
        inv_other_lengths.append(len(inv_other_rows))
        pl_main_lengths.append(len(pl_rows))
        pl_other_lengths.append(len(pl_other_rows))

        inv_main_all.extend(inv_rows)
        inv_other_all.extend(inv_other_rows)
        pl_main_all.extend(pl_rows)
        pl_other_all.extend(pl_other_rows)

    inv_seed: Dict[Tuple[str, str], int] = {}
    pl_seed: Dict[Tuple[str, str], int] = {}

    inv_main_processed = _v73_apply_invoice_extras(inv_main_all, inv_seed)
    inv_seed_after_main = _v74_seq_seed_from_rows(inv_main_processed, V73_INV_HEADERS, "S. Invoice#")
    inv_other_processed = _v73_apply_invoice_extras(inv_other_all, inv_seed_after_main)

    pl_main_processed = _v76_apply_pl_sequence(pl_main_all, pl_seed)
    pl_seed_after_main = _v74_seq_seed_from_rows(pl_main_processed, V76_PL_HEADERS, "S. Invoice#")
    pl_other_processed = _v76_apply_pl_sequence(pl_other_all, pl_seed_after_main)

    inv_main_split = _u95_split_rows_by_lengths(inv_main_processed, inv_main_lengths)
    inv_other_split = _u95_split_rows_by_lengths(inv_other_processed, inv_other_lengths)
    pl_main_split = _u95_split_rows_by_lengths(pl_main_processed, pl_main_lengths)
    pl_other_split = _u95_split_rows_by_lengths(pl_other_processed, pl_other_lengths)

    inv_blocks: List[Tuple[str, str, List[List[Any]]]] = []
    pl_blocks: List[Tuple[str, str, List[List[Any]]]] = []

    for (fp, _), rows in zip(inv_main_by_file, inv_main_split):
        if rows:
            inv_blocks.append(("main", fp, rows))
    for (fp, _), rows in zip(inv_other_by_file, inv_other_split):
        if rows:
            inv_blocks.append(("other", fp, rows))

    for (fp, _), rows in zip(pl_main_by_file, pl_main_split):
        if rows:
            pl_blocks.append(("main", fp, rows))
    for (fp, _), rows in zip(pl_other_by_file, pl_other_split):
        if rows:
            pl_blocks.append(("other", fp, rows))

    return inv_blocks, pl_blocks


def _u95_ensure_block_index_sheet(wb):
    if _U95_BLOCK_INDEX_SHEET in wb.sheetnames:
        ws = wb[_U95_BLOCK_INDEX_SHEET]
    else:
        ws = wb.create_sheet(_U95_BLOCK_INDEX_SHEET)
    ws.sheet_state = "hidden"
    return ws


def _u95_load_block_index(wb) -> Dict[str, List[Dict[str, Any]]]:
    result = {"INV": [], "PL": []}
    if _U95_BLOCK_INDEX_SHEET not in wb.sheetnames:
        return result
    ws = wb[_U95_BLOCK_INDEX_SHEET]
    if ws.max_row < 2:
        return result
    headers = [nz_str(ws.cell(1, c).value) for c in range(1, 9)]
    expected = ["Version", "Sheet", "Order", "BlockID", "FilePath", "Kind", "RowCount", "Hash"]
    if headers[:8] != expected:
        return result

    for r in range(2, ws.max_row + 1):
        sh = nz_str(ws.cell(r, 2).value).upper()
        if sh not in ("INV", "PL"):
            continue
        result[sh].append({
            "sheet": sh,
            "order": int(ws.cell(r, 3).value or 0),
            "block_id": nz_str(ws.cell(r, 4).value),
            "file_path": nz_str(ws.cell(r, 5).value),
            "kind": nz_str(ws.cell(r, 6).value),
            "row_count": int(ws.cell(r, 7).value or 0),
            "hash": nz_str(ws.cell(r, 8).value),
        })

    result["INV"] = sorted(result["INV"], key=lambda x: x["order"])
    result["PL"] = sorted(result["PL"], key=lambda x: x["order"])
    return result


def _u95_save_block_index(wb, inv_meta: List[Dict[str, Any]], pl_meta: List[Dict[str, Any]]):
    ws = _u95_ensure_block_index_sheet(wb)
    ws.delete_rows(1, ws.max_row)
    headers = ["Version", "Sheet", "Order", "BlockID", "FilePath", "Kind", "RowCount", "Hash"]
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h

    row_no = 2
    for sh_name, items in (("INV", inv_meta), ("PL", pl_meta)):
        for order_no, item in enumerate(items, start=1):
            ws.cell(row_no, 1).value = _U95_BLOCK_INDEX_VERSION
            ws.cell(row_no, 2).value = sh_name
            ws.cell(row_no, 3).value = order_no
            ws.cell(row_no, 4).value = item.get("block_id", "")
            ws.cell(row_no, 5).value = item.get("file_path", "")
            ws.cell(row_no, 6).value = item.get("kind", "")
            ws.cell(row_no, 7).value = int(item.get("row_count", 0) or 0)
            ws.cell(row_no, 8).value = item.get("hash", "")
            row_no += 1
    ws.sheet_state = "hidden"


def _u95_build_meta_from_blocks(blocks: List[Tuple[str, str, List[List[Any]]]], headers: List[str]) -> List[Dict[str, Any]]:
    meta: List[Dict[str, Any]] = []
    for kind, fp, rows in blocks:
        rows = list(rows or [])
        if not rows:
            continue
        meta.append({
            "block_id": _u95_block_id(kind, fp),
            "file_path": _v75_norm_abs_upper(fp),
            "kind": kind,
            "row_count": len(rows),
            "hash": _u95_hash_rows(rows, headers),
        })
    return meta


def _u95_common_prefix_len(existing_meta: List[Dict[str, Any]], desired_meta: List[Dict[str, Any]]) -> int:
    n = min(len(existing_meta), len(desired_meta))
    i = 0
    while i < n:
        a = existing_meta[i]
        b = desired_meta[i]
        if (
            nz_str(a.get("block_id")) == nz_str(b.get("block_id")) and
            int(a.get("row_count", 0) or 0) == int(b.get("row_count", 0) or 0) and
            nz_str(a.get("hash")) == nz_str(b.get("hash"))
        ):
            i += 1
            continue
        break
    return i


def _u95_incremental_write_sheet(ws, headers: List[str], blocks: List[Tuple[str, str, List[List[Any]]]], existing_meta: List[Dict[str, Any]], date_cols: List[int], row_type_idx: int) -> Tuple[int, int, int]:
    desired_meta = _u95_build_meta_from_blocks(blocks, headers)
    prefix = _u95_common_prefix_len(existing_meta, desired_meta)

    # ensure headers
    _v74_ensure_headers(ws, headers)

    # count retained rows from common prefix
    retained_rows = 0
    for i in range(prefix):
        retained_rows += int(existing_meta[i].get("row_count", 0) or 0)

    last_used = _v74_last_used_row(ws)
    start_delete = retained_rows + 2  # row 1 is header
    if start_delete <= last_used:
        ws.delete_rows(start_delete, last_used - start_delete + 1)

    tail_blocks = blocks[prefix:]
    added_rows = 0
    for kind, fp, rows in tail_blocks:
        if not rows:
            continue
        added_rows += _u83_append_rows_exact_fast(ws, rows, headers, date_cols, row_type_idx)

    return prefix, retained_rows, added_rows


def _u95_full_rewrite_sheet(ws, headers: List[str], blocks: List[Tuple[str, str, List[List[Any]]]], date_cols: List[int], row_type_idx: int) -> int:
    ws = ws
    ws.delete_rows(2, max(0, ws.max_row - 1))
    total = 0
    for kind, fp, rows in blocks:
        if rows:
            total += _u83_append_rows_exact_fast(ws, rows, headers, date_cols, row_type_idx)
    return total


def _u95_rebuild_merged_detail_rows_incremental(output_path: str, ok_paths: set, changed_paths: Optional[set] = None, prep_label: str = 'DETAIL_PREP_U79') -> Dict[str, Any]:
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    cache = _u87_load_detail_cache(output_path)
    cache_files: Dict[str, Any] = cache.get('files', {}) if isinstance(cache, dict) else {}
    if not isinstance(cache_files, dict):
        cache_files = {}
    file_list = sorted([p for p in ok_paths if os.path.exists(p)])
    active_set = set(file_list)
    changed_paths = _u87_norm_path_set(changed_paths)
    changed_paths &= active_set

    cache_hit = 0
    cache_miss = 0
    rebuilt_files = 0
    dirty_cache = False
    stale_keys = []
    inv_added = 0
    pl_added = 0
    inv_prefix = 0
    pl_prefix = 0

    try:
        ws_inv = ensure_sheet(wb, 'INV')
        ws_pl = ensure_sheet(wb, 'PL')
        for nm in ['INV Others', 'PL Others']:
            if nm in wb.sheetnames:
                del wb[nm]
        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        # remove stale cache entries not active anymore
        stale_keys = [k for k in list(cache_files.keys()) if k not in active_set]
        for k in stale_keys:
            try:
                del cache_files[k]
                dirty_cache = True
            except Exception:
                pass

        bundle_map: Dict[str, Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]] = {}
        for fp in file_list:
            try:
                force_refresh = (fp in changed_paths) or (fp not in cache_files)
                bundle, source = _u87_get_bundle_from_cache_or_build(fp, cache_files, force_refresh=force_refresh)
                bundle_map[fp] = bundle
                if source == 'cache':
                    cache_hit += 1
                else:
                    cache_miss += 1
                    rebuilt_files += 1
                    dirty_cache = True
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_blocks, pl_blocks = _u95_collect_final_blocks(file_list, bundle_map)

        index_data = _u95_load_block_index(wb)
        existing_inv_meta = index_data.get("INV", [])
        existing_pl_meta = index_data.get("PL", [])

        desired_inv_meta = _u95_build_meta_from_blocks(inv_blocks, V73_INV_HEADERS)
        desired_pl_meta = _u95_build_meta_from_blocks(pl_blocks, V76_PL_HEADERS)

        # incremental write by common prefix; fallback full rewrite only if metadata absent and there are rows
        if existing_inv_meta:
            inv_prefix, inv_retained, inv_added = _u95_incremental_write_sheet(ws_inv, V73_INV_HEADERS, inv_blocks, existing_inv_meta, [1], 27)
        else:
            inv_prefix = 0
            inv_added = _u95_full_rewrite_sheet(ws_inv, V73_INV_HEADERS, inv_blocks, [1], 27)

        if existing_pl_meta:
            pl_prefix, pl_retained, pl_added = _u95_incremental_write_sheet(ws_pl, V76_PL_HEADERS, pl_blocks, existing_pl_meta, [1], 16)
        else:
            pl_prefix = 0
            pl_added = _u95_full_rewrite_sheet(ws_pl, V76_PL_HEADERS, pl_blocks, [1], 16)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V76_PL_HEADERS))
        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)

        _u95_save_block_index(wb, desired_inv_meta, desired_pl_meta)

        ws_log = ensure_sheet(wb, 'LOG_SUB_DETAIL')
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row('', prep_label, 'INFO',
                            f'Paths for merged detail (OK+WARNING)={len(ok_paths)} | Start rebuild detail tabs | Merge mode=2 sheets | SmartCache=ON | WriteMode=INCREMENTAL')])
        append_log_rows(ws_log, [make_log_row('', 'DETAIL_REBUILD_U95', 'INFO',
                            f'Current active files={len(file_list)} | INV_added={inv_added} | PL_added={pl_added} | inv_prefix={inv_prefix} | pl_prefix={pl_prefix} | cache_hit={cache_hit} | rebuilt={rebuilt_files} | stale_removed={len(stale_keys)}')])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, 'DETAIL_REBUILD_U95', 'ERROR', err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

    cache['version'] = _U87_DETAIL_CACHE_VERSION
    cache['files'] = cache_files
    if dirty_cache:
        _u87_save_detail_cache(output_path, cache)

    return {
        'active_files': len(file_list),
        'cache_hit': cache_hit,
        'cache_miss': cache_miss,
        'rebuilt_files': rebuilt_files,
        'stale_removed': len(stale_keys),
        'errors': len(errors),
        'cache_path': _u87_detail_cache_path(output_path),
        'inv_prefix': inv_prefix,
        'pl_prefix': pl_prefix,
        'inv_blocks': len(desired_inv_meta),
        'pl_blocks': len(desired_pl_meta),
    }


# U95: monkey patch smart rebuild implementation used by U94 processor
_u87_rebuild_merged_detail_rows_smart = _u95_rebuild_merged_detail_rows_incremental


def _u95_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    return _u94_processor_run(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)




# =========================================================
# U96 TRUE BLOCK SURGERY
# Muc tieu:
# - Giu logic ket qua cuoi giong U94/U95
# - Khong full rewrite toan sheet neu co the tranh duoc
# - Su dung block index co start_row/end_row
# - Giu common prefix + common suffix, chi rewrite middle segment
# - Neu metadata hu/khong hop le -> fallback full rewrite an toan
# =========================================================

_U96_BLOCK_INDEX_VERSION = 1
_U96_BLOCK_INDEX_SHEET = "_U96_BLOCK_INDEX"


def _u96_ensure_block_index_sheet(wb):
    if _U96_BLOCK_INDEX_SHEET in wb.sheetnames:
        ws = wb[_U96_BLOCK_INDEX_SHEET]
    else:
        ws = wb.create_sheet(_U96_BLOCK_INDEX_SHEET)
    ws.sheet_state = "hidden"
    return ws


def _u96_build_meta_from_blocks(blocks: List[Tuple[str, str, List[List[Any]]]], headers: List[str]) -> List[Dict[str, Any]]:
    meta: List[Dict[str, Any]] = []
    row_cursor = 2
    for kind, fp, rows in blocks:
        rows = list(rows or [])
        if not rows:
            continue
        row_count = len(rows)
        meta.append({
            "block_id": _u95_block_id(kind, fp),
            "file_path": _v75_norm_abs_upper(fp),
            "kind": kind,
            "row_count": row_count,
            "hash": _u95_hash_rows(rows, headers),
            "start_row": row_cursor,
            "end_row": row_cursor + row_count - 1,
        })
        row_cursor += row_count
    return meta


def _u96_save_block_index(wb, inv_meta: List[Dict[str, Any]], pl_meta: List[Dict[str, Any]]):
    ws = _u96_ensure_block_index_sheet(wb)
    ws.delete_rows(1, ws.max_row)
    headers = ["Version", "Sheet", "Order", "BlockID", "FilePath", "Kind", "RowCount", "Hash", "StartRow", "EndRow"]
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h

    r = 2
    for sh_name, items in (("INV", inv_meta), ("PL", pl_meta)):
        for order_no, item in enumerate(items, start=1):
            ws.cell(r, 1).value = _U96_BLOCK_INDEX_VERSION
            ws.cell(r, 2).value = sh_name
            ws.cell(r, 3).value = order_no
            ws.cell(r, 4).value = item.get("block_id", "")
            ws.cell(r, 5).value = item.get("file_path", "")
            ws.cell(r, 6).value = item.get("kind", "")
            ws.cell(r, 7).value = int(item.get("row_count", 0) or 0)
            ws.cell(r, 8).value = item.get("hash", "")
            ws.cell(r, 9).value = int(item.get("start_row", 0) or 0)
            ws.cell(r, 10).value = int(item.get("end_row", 0) or 0)
            r += 1
    ws.sheet_state = "hidden"


def _u96_load_block_index(wb) -> Dict[str, List[Dict[str, Any]]]:
    result = {"INV": [], "PL": []}
    if _U96_BLOCK_INDEX_SHEET not in wb.sheetnames:
        return result
    ws = wb[_U96_BLOCK_INDEX_SHEET]
    if ws.max_row < 2:
        return result
    headers = [nz_str(ws.cell(1, c).value) for c in range(1, 11)]
    expected = ["Version", "Sheet", "Order", "BlockID", "FilePath", "Kind", "RowCount", "Hash", "StartRow", "EndRow"]
    if headers[:10] != expected:
        return result

    for r in range(2, ws.max_row + 1):
        sh = nz_str(ws.cell(r, 2).value).upper()
        if sh not in ("INV", "PL"):
            continue
        result[sh].append({
            "sheet": sh,
            "order": int(ws.cell(r, 3).value or 0),
            "block_id": nz_str(ws.cell(r, 4).value),
            "file_path": nz_str(ws.cell(r, 5).value),
            "kind": nz_str(ws.cell(r, 6).value),
            "row_count": int(ws.cell(r, 7).value or 0),
            "hash": nz_str(ws.cell(r, 8).value),
            "start_row": int(ws.cell(r, 9).value or 0),
            "end_row": int(ws.cell(r, 10).value or 0),
        })

    result["INV"] = sorted(result["INV"], key=lambda x: x["order"])
    result["PL"] = sorted(result["PL"], key=lambda x: x["order"])
    return result


def _u96_meta_valid(ws, meta: List[Dict[str, Any]]) -> bool:
    if not meta:
        return False
    prev_end = 1
    for item in meta:
        sr = int(item.get("start_row", 0) or 0)
        er = int(item.get("end_row", 0) or 0)
        rc = int(item.get("row_count", 0) or 0)
        if sr < 2 or er < sr or rc != (er - sr + 1):
            return False
        if sr != prev_end + 1:
            return False
        prev_end = er
    return prev_end <= max(ws.max_row, 1)


def _u96_common_prefix_len(old_meta: List[Dict[str, Any]], new_meta: List[Dict[str, Any]]) -> int:
    n = min(len(old_meta), len(new_meta))
    i = 0
    while i < n:
        a = old_meta[i]
        b = new_meta[i]
        if (
            nz_str(a.get("block_id")) == nz_str(b.get("block_id")) and
            int(a.get("row_count", 0) or 0) == int(b.get("row_count", 0) or 0) and
            nz_str(a.get("hash")) == nz_str(b.get("hash"))
        ):
            i += 1
            continue
        break
    return i


def _u96_common_suffix_len(old_meta: List[Dict[str, Any]], new_meta: List[Dict[str, Any]], prefix_len: int) -> int:
    old_n = len(old_meta)
    new_n = len(new_meta)
    k = 0
    while (old_n - 1 - k) >= prefix_len and (new_n - 1 - k) >= prefix_len:
        a = old_meta[old_n - 1 - k]
        b = new_meta[new_n - 1 - k]
        if (
            nz_str(a.get("block_id")) == nz_str(b.get("block_id")) and
            int(a.get("row_count", 0) or 0) == int(b.get("row_count", 0) or 0) and
            nz_str(a.get("hash")) == nz_str(b.get("hash"))
        ):
            k += 1
            continue
        break
    return k


def _u96_flatten_blocks(blocks: List[Tuple[str, str, List[List[Any]]]]) -> List[List[Any]]:
    out: List[List[Any]] = []
    for kind, fp, rows in blocks:
        if rows:
            out.extend(rows)
    return out


def _u96_write_rows_into_range(ws, start_row: int, rows: List[List[Any]], headers: List[str], date_cols: List[int], row_type_idx: int) -> int:
    rows = list(rows or [])
    if not rows:
        return 0
    for offset, arr in enumerate(rows):
        rr = start_row + offset
        aligned = _v77_align_row_to_headers(arr, headers)
        for c, v in enumerate(aligned, start=1):
            cell = ws.cell(rr, c)
            cell.value = sanitize_excel_text(v)
            if c in date_cols:
                try:
                    cell.number_format = 'dd/mm/yyyy'
                except Exception:
                    pass
        status = ''
        if 1 <= row_type_idx <= len(aligned):
            try:
                status = nz_str(aligned[row_type_idx - 1]).upper()
            except Exception:
                status = ''
        # row_type_idx cua U83 dang dung cho dong TOTAL/HANDLING CHARGE -> dung lai helper cu
        if row_type_idx == 27:
            try:
                status = nz_str(aligned[26]).upper()
            except Exception:
                status = ''
        elif row_type_idx == 16:
            try:
                status = nz_str(aligned[21]).upper()
            except Exception:
                status = ''
        if status in STATUS_FILLS:
            fill = STATUS_FILLS[status]
            for c in range(1, len(headers) + 1):
                ws.cell(rr, c).fill = fill
    return len(rows)


def _u96_full_rewrite_sheet(ws, headers: List[str], blocks: List[Tuple[str, str, List[List[Any]]]], date_cols: List[int], row_type_idx: int) -> int:
    _v74_ensure_headers(ws, headers)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    rows = _u96_flatten_blocks(blocks)
    return _u96_write_rows_into_range(ws, 2, rows, headers, date_cols, row_type_idx)


def _u96_block_surgery_sheet(ws, headers: List[str], blocks: List[Tuple[str, str, List[List[Any]]]], old_meta: List[Dict[str, Any]], date_cols: List[int], row_type_idx: int) -> Dict[str, int]:
    _v74_ensure_headers(ws, headers)
    desired_meta = _u96_build_meta_from_blocks(blocks, headers)

    if not old_meta or not _u96_meta_valid(ws, old_meta):
        added = _u96_full_rewrite_sheet(ws, headers, blocks, date_cols, row_type_idx)
        return {"mode": 0, "prefix": 0, "suffix": 0, "added_rows": added}

    prefix = _u96_common_prefix_len(old_meta, desired_meta)
    suffix = _u96_common_suffix_len(old_meta, desired_meta, prefix)

    # If nothing common, fallback full rewrite for safety and simplicity
    if prefix == 0 and suffix == 0:
        added = _u96_full_rewrite_sheet(ws, headers, blocks, date_cols, row_type_idx)
        return {"mode": 0, "prefix": 0, "suffix": 0, "added_rows": added}

    old_mid_start_idx = prefix
    old_mid_end_idx = len(old_meta) - suffix - 1
    new_mid_start_idx = prefix
    new_mid_end_idx = len(desired_meta) - suffix - 1

    if old_mid_start_idx <= old_mid_end_idx:
        delete_start_row = int(old_meta[old_mid_start_idx]["start_row"])
        delete_end_row = int(old_meta[old_mid_end_idx]["end_row"])
        if delete_end_row >= delete_start_row:
            ws.delete_rows(delete_start_row, delete_end_row - delete_start_row + 1)
    else:
        delete_start_row = int(old_meta[prefix - 1]["end_row"]) + 1 if prefix > 0 else 2

    middle_blocks = []
    if new_mid_start_idx <= new_mid_end_idx:
        middle_blocks = blocks[new_mid_start_idx:new_mid_end_idx + 1]

    middle_rows = _u96_flatten_blocks(middle_blocks)
    if middle_rows:
        ws.insert_rows(delete_start_row, len(middle_rows))
        _u96_write_rows_into_range(ws, delete_start_row, middle_rows, headers, date_cols, row_type_idx)

    return {
        "mode": 1,
        "prefix": prefix,
        "suffix": suffix,
        "added_rows": len(middle_rows),
    }


def _u96_rebuild_merged_detail_rows_block_surgery(output_path: str, ok_paths: set, changed_paths: Optional[set] = None, prep_label: str = 'DETAIL_PREP_U79') -> Dict[str, Any]:
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    cache = _u87_load_detail_cache(output_path)
    cache_files: Dict[str, Any] = cache.get('files', {}) if isinstance(cache, dict) else {}
    if not isinstance(cache_files, dict):
        cache_files = {}
    file_list = sorted([p for p in ok_paths if os.path.exists(p)])
    active_set = set(file_list)
    changed_paths = _u87_norm_path_set(changed_paths)
    changed_paths &= active_set

    cache_hit = 0
    cache_miss = 0
    rebuilt_files = 0
    dirty_cache = False
    stale_keys = []

    try:
        ws_inv = ensure_sheet(wb, 'INV')
        ws_pl = ensure_sheet(wb, 'PL')
        for nm in ['INV Others', 'PL Others']:
            if nm in wb.sheetnames:
                del wb[nm]
        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        stale_keys = [k for k in list(cache_files.keys()) if k not in active_set]
        for k in stale_keys:
            try:
                del cache_files[k]
                dirty_cache = True
            except Exception:
                pass

        bundle_map: Dict[str, Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]] = {}
        for fp in file_list:
            try:
                force_refresh = (fp in changed_paths) or (fp not in cache_files)
                bundle, source = _u87_get_bundle_from_cache_or_build(fp, cache_files, force_refresh=force_refresh)
                bundle_map[fp] = bundle
                if source == 'cache':
                    cache_hit += 1
                else:
                    cache_miss += 1
                    rebuilt_files += 1
                    dirty_cache = True
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_blocks, pl_blocks = _u95_collect_final_blocks(file_list, bundle_map)

        index_data = _u96_load_block_index(wb)
        old_inv_meta = index_data.get("INV", [])
        old_pl_meta = index_data.get("PL", [])

        inv_stats = _u96_block_surgery_sheet(ws_inv, V73_INV_HEADERS, inv_blocks, old_inv_meta, [1], 27)
        pl_stats = _u96_block_surgery_sheet(ws_pl, V76_PL_HEADERS, pl_blocks, old_pl_meta, [1], 16)

        desired_inv_meta = _u96_build_meta_from_blocks(inv_blocks, V73_INV_HEADERS)
        desired_pl_meta = _u96_build_meta_from_blocks(pl_blocks, V76_PL_HEADERS)
        _u96_save_block_index(wb, desired_inv_meta, desired_pl_meta)

        autofit_useful(ws_inv, len(V73_INV_HEADERS))
        autofit_useful(ws_pl, len(V76_PL_HEADERS))
        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)

        ws_log = ensure_sheet(wb, 'LOG_SUB_DETAIL')
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row('', prep_label, 'INFO',
                            f'Paths for merged detail (OK+WARNING)={len(ok_paths)} | Start rebuild detail tabs | Merge mode=2 sheets | SmartCache=ON | WriteMode=BLOCK_SURGERY')])
        append_log_rows(ws_log, [make_log_row('', 'DETAIL_REBUILD_U96', 'INFO',
                            f'Current active files={len(file_list)} | INV_mode={inv_stats.get("mode",0)} | PL_mode={pl_stats.get("mode",0)} | inv_prefix={inv_stats.get("prefix",0)} | inv_suffix={inv_stats.get("suffix",0)} | pl_prefix={pl_stats.get("prefix",0)} | pl_suffix={pl_stats.get("suffix",0)} | inv_mid_added={inv_stats.get("added_rows",0)} | pl_mid_added={pl_stats.get("added_rows",0)} | cache_hit={cache_hit} | rebuilt={rebuilt_files} | stale_removed={len(stale_keys)}')])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, 'DETAIL_REBUILD_U96', 'ERROR', err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

    cache['version'] = _U87_DETAIL_CACHE_VERSION
    cache['files'] = cache_files
    if dirty_cache:
        _u87_save_detail_cache(output_path, cache)

    return {
        'active_files': len(file_list),
        'cache_hit': cache_hit,
        'cache_miss': cache_miss,
        'rebuilt_files': rebuilt_files,
        'stale_removed': len(stale_keys),
        'errors': len(errors),
        'cache_path': _u87_detail_cache_path(output_path),
        'inv_mode': inv_stats.get("mode", 0),
        'pl_mode': pl_stats.get("mode", 0),
        'inv_prefix': inv_stats.get("prefix", 0),
        'inv_suffix': inv_stats.get("suffix", 0),
        'pl_prefix': pl_stats.get("prefix", 0),
        'pl_suffix': pl_stats.get("suffix", 0),
    }


# U96: monkey patch smart rebuild implementation used by U94 processor
_u87_rebuild_merged_detail_rows_smart = _u96_rebuild_merged_detail_rows_block_surgery


def _u96_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    return _u94_processor_run(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)




# =========================================================
# U97 APPEND-ONLY PRODUCTION
# Nguyen tac:
# - 1 file duy nhat
# - Khong rebuild / reorder INV, PL
# - Sheet nguon khong can dep theo thu tu vat ly
# - CHI append them dong moi
# - WARNING / ERROR theo doi rieng trong SUB_DETAIL + LOG_SUB_DETAIL
# - Chi dua file OK vao INV / PL
# - Neu file WARNING/ERROR duoc sua thanh OK o lan sau, se append them du lieu luc do
# - Khong xoa dong cu, khong rewrite dong cu
# =========================================================

def _u97_collect_paths_by_status(ws_sub, statuses: set) -> set:
    out = set()
    if ws_sub is None:
        return out
    path_col = find_header_col(ws_sub, "File Path")
    status_col = find_header_col(ws_sub, "Status")
    if path_col <= 0 or status_col <= 0:
        return out
    for r in range(2, ws_sub.max_row + 1):
        st = nz_str(ws_sub.cell(r, status_col).value).upper()
        if st in statuses:
            p = _v75_norm_abs_upper(ws_sub.cell(r, path_col).value)
            if p:
                out.add(p)
    return out


def _u97_collect_ok_paths_from_output(output_path: str) -> set:
    wb = open_or_create_output_workbook(output_path)
    try:
        if "SUB_DETAIL" not in wb.sheetnames:
            return set()
        return _u97_collect_paths_by_status(wb["SUB_DETAIL"], {"OK"})
    finally:
        wb.close()


def _u97_row_hash(arr: List[Any], headers: List[str]) -> str:
    aligned = _v77_align_row_to_headers(arr, headers)
    vals = []
    for v in aligned:
        if isinstance(v, datetime):
            vals.append(v.strftime("%Y-%m-%d %H:%M:%S"))
        else:
            vals.append(nz_str(v))
    raw = json.dumps(vals, ensure_ascii=False, separators=(",", ":"))
    return hashlib.md5(raw.encode("utf-8", errors="ignore")).hexdigest()


def _u97_collect_existing_row_hashes(ws, headers: List[str]) -> set:
    hashes = set()
    if ws is None or ws.max_row < 2:
        return hashes
    header_len = len(headers)
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, header_len + 1)]
        if not any(nz_str(v) for v in row):
            continue
        hashes.add(_u97_row_hash(row, headers))
    return hashes


def _u97_append_rows_minimal(ws, rows: List[List[Any]], headers: List[str]) -> int:
    if not rows:
        return 0
    count = 0
    for arr in rows:
        aligned = _v77_align_row_to_headers(arr, headers)
        ws.append([sanitize_excel_text(v) for v in aligned])
        count += 1
    return count


def _u97_rebuild_append_only(output_path: str, ok_paths: set, changed_ok_paths: set, prep_label: str = 'DETAIL_PREP_U79') -> Dict[str, Any]:
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    cache = _u87_load_detail_cache(output_path)
    cache_files: Dict[str, Any] = cache.get('files', {}) if isinstance(cache, dict) else {}
    if not isinstance(cache_files, dict):
        cache_files = {}

    file_list = sorted([p for p in ok_paths if os.path.exists(p)])
    changed_ok_paths = _u87_norm_path_set(changed_ok_paths) & set(file_list)

    cache_hit = 0
    rebuilt_files = 0
    dirty_cache = False
    inv_appended = 0
    pl_appended = 0
    inv_skipped_dup = 0
    pl_skipped_dup = 0

    try:
        ws_inv = ensure_sheet(wb, 'INV')
        ws_pl = ensure_sheet(wb, 'PL')
        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        existing_inv_hashes = _u97_collect_existing_row_hashes(ws_inv, V73_INV_HEADERS)
        existing_pl_hashes = _u97_collect_existing_row_hashes(ws_pl, V76_PL_HEADERS)

        inv_rows_to_append: List[List[Any]] = []
        pl_rows_to_append: List[List[Any]] = []

        for fp in sorted(changed_ok_paths):
            try:
                bundle, source = _u87_get_bundle_from_cache_or_build(fp, cache_files, force_refresh=True)
                if source == 'cache':
                    cache_hit += 1
                else:
                    rebuilt_files += 1
                    dirty_cache = True

                inv_rows, pl_rows, inv_other_rows, pl_other_rows = bundle
                inv_all = list(inv_rows or []) + list(inv_other_rows or [])
                pl_all = list(pl_rows or []) + list(pl_other_rows or [])

                inv_seed = {}
                pl_seed = {}
                inv_processed = _v73_apply_invoice_extras(inv_all, inv_seed)
                pl_processed = _v76_apply_pl_sequence(pl_all, pl_seed)

                for arr in inv_processed:
                    h = _u97_row_hash(arr, V73_INV_HEADERS)
                    if h in existing_inv_hashes:
                        inv_skipped_dup += 1
                        continue
                    existing_inv_hashes.add(h)
                    inv_rows_to_append.append(arr)

                for arr in pl_processed:
                    h = _u97_row_hash(arr, V76_PL_HEADERS)
                    if h in existing_pl_hashes:
                        pl_skipped_dup += 1
                        continue
                    existing_pl_hashes.add(h)
                    pl_rows_to_append.append(arr)

            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_appended = _u97_append_rows_minimal(ws_inv, inv_rows_to_append, V73_INV_HEADERS)
        pl_appended = _u97_append_rows_minimal(ws_pl, pl_rows_to_append, V76_PL_HEADERS)

        _v735_unhide_all_columns(ws_inv)
        _v735_unhide_all_columns(ws_pl)

        ws_log = ensure_sheet(wb, 'LOG_SUB_DETAIL')
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row('', prep_label, 'INFO',
                            f'Paths for merged detail (OK only)={len(ok_paths)} | Start append-only detail tabs | Merge mode=1 file | SmartCache=ON | WriteMode=APPEND_ONLY')])
        append_log_rows(ws_log, [make_log_row('', 'DETAIL_REBUILD_U97', 'INFO',
                            f'Current OK files={len(file_list)} | changed_ok={len(changed_ok_paths)} | INV_appended={inv_appended} | PL_appended={pl_appended} | inv_skipped_dup={inv_skipped_dup} | pl_skipped_dup={pl_skipped_dup} | cache_hit={cache_hit} | rebuilt={rebuilt_files}')])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, 'DETAIL_REBUILD_U97', 'ERROR', err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

    cache['version'] = _U87_DETAIL_CACHE_VERSION
    cache['files'] = cache_files
    if dirty_cache:
        _u87_save_detail_cache(output_path, cache)

    return {
        'ok_files': len(file_list),
        'changed_ok': len(changed_ok_paths),
        'inv_appended': inv_appended,
        'pl_appended': pl_appended,
        'inv_skipped_dup': inv_skipped_dup,
        'pl_skipped_dup': pl_skipped_dup,
        'cache_hit': cache_hit,
        'rebuilt_files': rebuilt_files,
        'errors': len(errors),
        'cache_path': _u87_detail_cache_path(output_path),
    }


def _u97_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _u85_reset_bench()
    t_all_0 = time.perf_counter()
    _U90_SOURCE_CACHE_CTX['output_path'] = output_path
    _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
    cache = _u90_get_source_cache()
    cache['stats'] = _u90_empty_source_cache()['stats']

    try:
        self.logger.log('APPEND_ONLY_ENTER_U97', 'info')
        self.logger.log('SOURCE_PARSE_CACHE_ENTER_U97', 'info')

        _U84_WIDTH_CACHE_CTX['output_path'] = output_path
        _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

        existing_ok_paths = _u97_collect_ok_paths_from_output(output_path)
        scanned_files = _u87_norm_path_set(collect_source_files(folder_path, output_path))

        t0 = time.perf_counter()
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        core_seconds = max(0.0, time.perf_counter() - t0)

        current_ok_paths = _u97_collect_ok_paths_from_output(output_path)
        changed_ok_paths = _u87_norm_path_set(p for p in current_ok_paths if p not in existing_ok_paths)

        t1 = time.perf_counter()
        append_stats = _u97_rebuild_append_only(
            output_path,
            current_ok_paths,
            changed_ok_paths=changed_ok_paths,
            prep_label='DETAIL_PREP_U79'
        )
        rebuild_seconds = max(0.0, time.perf_counter() - t1)

        total_seconds = max(0.0, time.perf_counter() - t_all_0)

        cache = _u90_get_source_cache()
        _u90_save_source_cache(output_path, cache)
        stats = cache.get('stats', {}) if isinstance(cache, dict) else {}

        bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
        bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                        f"ok_files={append_stats.get('ok_files',0)} | scanned={len(scanned_files)} | changed_ok={append_stats.get('changed_ok',0)} | inv_appended={append_stats.get('inv_appended',0)} | pl_appended={append_stats.get('pl_appended',0)} | cache_hit={append_stats.get('cache_hit',0)} | rebuilt={append_stats.get('rebuilt_files',0)} | cache={append_stats.get('cache_path','')}"))
        bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                        f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
        _u85_append_bench_log_rows(output_path, bench_rows)
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}
        _U85_BENCH_CTX['enabled'] = False
        _U90_SOURCE_CACHE_CTX['output_path'] = ''
        _U90_SOURCE_CACHE_CTX['cache'] = None




# =========================================================
# U98 FILE UPSERT WARNING -> OK
# Nguyen tac:
# - WARNING van duoc ghi vao INV / PL neu co detail
# - ERROR khong ghi vao INV / PL
# - Moi file chi co 1 block hien hanh tren moi sheet
# - Neu file da ton tai:
#     + Noi dung khong doi -> bo qua
#     + Noi dung doi, row_count giong nhau -> overwrite tai cho
#     + Noi dung doi, row_count khac nhau -> replace block tai vi tri cu
# - Neu file truoc WARNING, nay OK -> sua truc tiep block cu + doi status ngay tren block do
# - Khong append duplicate version
# =========================================================

_U98_BLOCK_INDEX_SHEET = "_U98_FILE_BLOCK_INDEX"


def _u98_ensure_index_sheet(wb):
    if _U98_BLOCK_INDEX_SHEET in wb.sheetnames:
        ws = wb[_U98_BLOCK_INDEX_SHEET]
    else:
        ws = wb.create_sheet(_U98_BLOCK_INDEX_SHEET)
    ws.sheet_state = "hidden"
    return ws


def _u98_norm_path(v) -> str:
    return _v75_norm_abs_upper(v)


def _u98_detail_status(arr_list: List[List[Any]], status_idx: int) -> str:
    for arr in arr_list or []:
        if len(arr) >= status_idx:
            s = nz_str(arr[status_idx - 1]).upper()
            if s:
                return s
    return ""


def _u98_file_hash(rows: List[List[Any]], headers: List[str]) -> str:
    payload = []
    for arr in rows or []:
        aligned = _v77_align_row_to_headers(arr, headers)
        vals = []
        for v in aligned:
            if isinstance(v, datetime):
                vals.append(v.strftime("%Y-%m-%d %H:%M:%S"))
            else:
                vals.append(nz_str(v))
        payload.append(vals)
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    import hashlib
    return hashlib.md5(raw.encode("utf-8", errors="ignore")).hexdigest()


def _u98_collect_supported_paths(ws_sub) -> set:
    out = set()
    if ws_sub is None:
        return out
    path_col = find_header_col(ws_sub, "File Path")
    status_col = find_header_col(ws_sub, "Status")
    if path_col <= 0 or status_col <= 0:
        return out
    for r in range(2, ws_sub.max_row + 1):
        st = nz_str(ws_sub.cell(r, status_col).value).upper()
        if st in ("OK", "WARNING"):
            p = _u98_norm_path(ws_sub.cell(r, path_col).value)
            if p:
                out.add(p)
    return out


def _u98_collect_supported_paths_from_output(output_path: str) -> set:
    wb = open_or_create_output_workbook(output_path)
    try:
        if "SUB_DETAIL" not in wb.sheetnames:
            return set()
        return _u98_collect_supported_paths(wb["SUB_DETAIL"])
    finally:
        wb.close()


def _u98_load_index(wb) -> Dict[str, Dict[str, Dict[str, Any]]]:
    out = {"INV": {}, "PL": {}}
    if _U98_BLOCK_INDEX_SHEET not in wb.sheetnames:
        return out
    ws = wb[_U98_BLOCK_INDEX_SHEET]
    if ws.max_row < 2:
        return out
    headers = [nz_str(ws.cell(1, c).value) for c in range(1, 8)]
    expected = ["Sheet", "FilePath", "StartRow", "EndRow", "RowCount", "Status", "Hash"]
    if headers[:7] != expected:
        return out
    for r in range(2, ws.max_row + 1):
        sh = nz_str(ws.cell(r, 1).value).upper()
        if sh not in ("INV", "PL"):
            continue
        fp = _u98_norm_path(ws.cell(r, 2).value)
        if not fp:
            continue
        out[sh][fp] = {
            "sheet": sh,
            "file_path": fp,
            "start_row": int(ws.cell(r, 3).value or 0),
            "end_row": int(ws.cell(r, 4).value or 0),
            "row_count": int(ws.cell(r, 5).value or 0),
            "status": nz_str(ws.cell(r, 6).value).upper(),
            "hash": nz_str(ws.cell(r, 7).value),
        }
    return out


def _u98_save_index(wb, inv_index: Dict[str, Dict[str, Any]], pl_index: Dict[str, Dict[str, Any]]):
    ws = _u98_ensure_index_sheet(wb)
    ws.delete_rows(1, ws.max_row)
    headers = ["Sheet", "FilePath", "StartRow", "EndRow", "RowCount", "Status", "Hash"]
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h

    row_no = 2
    for sh_name, idx_map in (("INV", inv_index), ("PL", pl_index)):
        items = sorted(idx_map.values(), key=lambda x: int(x.get("start_row", 0) or 0))
        for item in items:
            ws.cell(row_no, 1).value = sh_name
            ws.cell(row_no, 2).value = item.get("file_path", "")
            ws.cell(row_no, 3).value = int(item.get("start_row", 0) or 0)
            ws.cell(row_no, 4).value = int(item.get("end_row", 0) or 0)
            ws.cell(row_no, 5).value = int(item.get("row_count", 0) or 0)
            ws.cell(row_no, 6).value = item.get("status", "")
            ws.cell(row_no, 7).value = item.get("hash", "")
            row_no += 1
    ws.sheet_state = "hidden"


def _u98_shift_index_after(index_map: Dict[str, Dict[str, Any]], from_row: int, delta: int, skip_fp: str = ""):
    if delta == 0:
        return
    for fp, meta in index_map.items():
        if skip_fp and fp == skip_fp:
            continue
        sr = int(meta.get("start_row", 0) or 0)
        er = int(meta.get("end_row", 0) or 0)
        if sr >= from_row:
            meta["start_row"] = sr + delta
            meta["end_row"] = er + delta


def _u98_collect_existing_file_paths_from_sheet(ws, file_path_idx: int) -> set:
    out = set()
    if ws is None or ws.max_row < 2:
        return out
    for r in range(2, ws.max_row + 1):
        if file_path_idx <= ws.max_column:
            p = _u98_norm_path(ws.cell(r, file_path_idx).value)
            if p:
                out.add(p)
    return out


def _u98_append_block(ws, rows: List[List[Any]], headers: List[str], date_cols: List[int], file_path_idx: int, status_idx: int) -> Tuple[int, int]:
    if not rows:
        return 0, 0
    start_row = max(2, ws.max_row + 1)
    count = 0
    for arr in rows:
        aligned = _v77_align_row_to_headers(arr, headers)
        ws.append([sanitize_excel_text(v) for v in aligned])
        rr = start_row + count
        for c in date_cols:
            if c <= len(headers):
                try:
                    ws.cell(rr, c).number_format = 'dd/mm/yyyy'
                except Exception:
                    pass
        count += 1
    return start_row, count


def _u98_overwrite_block_same_size(ws, start_row: int, rows: List[List[Any]], headers: List[str], date_cols: List[int]):
    for i, arr in enumerate(rows):
        rr = start_row + i
        aligned = _v77_align_row_to_headers(arr, headers)
        for c, v in enumerate(aligned, start=1):
            cell = ws.cell(rr, c)
            cell.value = sanitize_excel_text(v)
            if c in date_cols:
                try:
                    cell.number_format = 'dd/mm/yyyy'
                except Exception:
                    pass


def _u98_replace_block_resize(ws, start_row: int, old_count: int, rows: List[List[Any]], headers: List[str], date_cols: List[int]):
    new_count = len(rows)
    if old_count > 0:
        ws.delete_rows(start_row, old_count)
    if new_count > 0:
        ws.insert_rows(start_row, new_count)
        _u98_overwrite_block_same_size(ws, start_row, rows, headers, date_cols)


def _u98_build_rows_for_file(fp: str, cache_files: Dict[str, Any], force_refresh: bool = True):
    bundle, source = _u87_get_bundle_from_cache_or_build(fp, cache_files, force_refresh=force_refresh)
    inv_rows, pl_rows, inv_other_rows, pl_other_rows = bundle

    inv_all = list(inv_rows or []) + list(inv_other_rows or [])
    pl_all = list(pl_rows or []) + list(pl_other_rows or [])

    inv_seed = {}
    pl_seed = {}
    inv_processed = _v73_apply_invoice_extras(inv_all, inv_seed)
    pl_processed = _v76_apply_pl_sequence(pl_all, pl_seed)

    inv_status = _u98_detail_status(inv_processed, 27)
    pl_status = _u98_detail_status(pl_processed, 22)
    inv_hash = _u98_file_hash(inv_processed, V73_INV_HEADERS)
    pl_hash = _u98_file_hash(pl_processed, V76_PL_HEADERS)

    return {
        "source": source,
        "inv_rows": inv_processed,
        "pl_rows": pl_processed,
        "inv_status": inv_status,
        "pl_status": pl_status,
        "inv_hash": inv_hash,
        "pl_hash": pl_hash,
    }


def _u98_upsert_sheet(ws, headers: List[str], index_map: Dict[str, Dict[str, Any]], file_rows_map: Dict[str, Dict[str, Any]], kind: str, date_cols: List[int], file_path_idx: int, status_field: str, hash_field: str, rows_field: str) -> Dict[str, int]:
    appended = 0
    updated_same = 0
    updated_resize = 0
    skipped_same = 0
    existing_physical = _u98_collect_existing_file_paths_from_sheet(ws, file_path_idx)

    # Preserve sheet physical order: existing stay, new append at tail.
    for fp, payload in file_rows_map.items():
        rows = payload.get(rows_field, []) or []
        status_val = nz_str(payload.get(status_field, "")).upper()
        hash_val = nz_str(payload.get(hash_field, ""))

        if not rows:
            continue

        meta = index_map.get(fp)
        if meta is None and fp in existing_physical:
            # physical data exists but index missing => infer by scan and skip risky rewrite
            skipped_same += 1
            continue

        if meta is None:
            start_row, count = _u98_append_block(ws, rows, headers, date_cols, file_path_idx, 0)
            if count > 0:
                index_map[fp] = {
                    "sheet": kind,
                    "file_path": fp,
                    "start_row": start_row,
                    "end_row": start_row + count - 1,
                    "row_count": count,
                    "status": status_val,
                    "hash": hash_val,
                }
                appended += count
            continue

        old_count = int(meta.get("row_count", 0) or 0)
        old_hash = nz_str(meta.get("hash", ""))
        old_status = nz_str(meta.get("status", "")).upper()
        start_row = int(meta.get("start_row", 0) or 0)

        if old_hash == hash_val and old_status == status_val and old_count == len(rows):
            skipped_same += 1
            continue

        if old_count == len(rows):
            _u98_overwrite_block_same_size(ws, start_row, rows, headers, date_cols)
            meta["status"] = status_val
            meta["hash"] = hash_val
            updated_same += len(rows)
        else:
            delta = len(rows) - old_count
            _u98_replace_block_resize(ws, start_row, old_count, rows, headers, date_cols)
            meta["row_count"] = len(rows)
            meta["start_row"] = start_row
            meta["end_row"] = start_row + len(rows) - 1
            meta["status"] = status_val
            meta["hash"] = hash_val
            _u98_shift_index_after(index_map, start_row + old_count, delta, skip_fp=fp)
            updated_resize += len(rows)

        meta["row_count"] = len(rows)
        meta["start_row"] = start_row
        meta["end_row"] = start_row + len(rows) - 1
        meta["status"] = status_val
        meta["hash"] = hash_val

    return {
        "appended": appended,
        "updated_same": updated_same,
        "updated_resize": updated_resize,
        "skipped_same": skipped_same,
    }


def _u98_rebuild_upsert(output_path: str, supported_paths: set, changed_paths: set, prep_label: str = 'DETAIL_PREP_U79') -> Dict[str, Any]:
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    cache = _u87_load_detail_cache(output_path)
    cache_files: Dict[str, Any] = cache.get('files', {}) if isinstance(cache, dict) else {}
    if not isinstance(cache_files, dict):
        cache_files = {}

    file_list = sorted([p for p in supported_paths if os.path.exists(p)])
    changed_paths = _u87_norm_path_set(changed_paths) & set(file_list)

    cache_hit = 0
    rebuilt_files = 0
    dirty_cache = False

    try:
        ws_inv = ensure_sheet(wb, 'INV')
        ws_pl = ensure_sheet(wb, 'PL')
        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        idx_all = _u98_load_index(wb)
        inv_index = idx_all.get("INV", {})
        pl_index = idx_all.get("PL", {})

        file_rows_map: Dict[str, Dict[str, Any]] = {}
        for fp in sorted(changed_paths):
            try:
                payload = _u98_build_rows_for_file(fp, cache_files, force_refresh=True)
                if payload["source"] == "cache":
                    cache_hit += 1
                else:
                    rebuilt_files += 1
                    dirty_cache = True
                file_rows_map[fp] = payload
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_stats = _u98_upsert_sheet(
            ws_inv, V73_INV_HEADERS, inv_index, file_rows_map,
            kind="INV", date_cols=[1], file_path_idx=29,
            status_field="inv_status", hash_field="inv_hash", rows_field="inv_rows"
        )
        pl_stats = _u98_upsert_sheet(
            ws_pl, V76_PL_HEADERS, pl_index, file_rows_map,
            kind="PL", date_cols=[1], file_path_idx=24,
            status_field="pl_status", hash_field="pl_hash", rows_field="pl_rows"
        )

        _u98_save_index(wb, inv_index, pl_index)

        ws_log = ensure_sheet(wb, 'LOG_SUB_DETAIL')
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row('', prep_label, 'INFO',
                            f'Paths for detail (OK+WARNING)={len(supported_paths)} | Start file-upsert detail tabs | Merge mode=1 file | SmartCache=ON | WriteMode=FILE_UPSERT')])
        append_log_rows(ws_log, [make_log_row('', 'DETAIL_REBUILD_U98', 'INFO',
                            f'Current supported files={len(file_list)} | changed={len(changed_paths)} | INV_append={inv_stats.get("appended",0)} | INV_update_same={inv_stats.get("updated_same",0)} | INV_update_resize={inv_stats.get("updated_resize",0)} | PL_append={pl_stats.get("appended",0)} | PL_update_same={pl_stats.get("updated_same",0)} | PL_update_resize={pl_stats.get("updated_resize",0)} | cache_hit={cache_hit} | rebuilt={rebuilt_files}')])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, 'DETAIL_REBUILD_U98', 'ERROR', err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

    cache['version'] = _U87_DETAIL_CACHE_VERSION
    cache['files'] = cache_files
    if dirty_cache:
        _u87_save_detail_cache(output_path, cache)

    return {
        'supported_files': len(file_list),
        'changed': len(changed_paths),
        'inv_append': inv_stats.get("appended", 0),
        'inv_update_same': inv_stats.get("updated_same", 0),
        'inv_update_resize': inv_stats.get("updated_resize", 0),
        'pl_append': pl_stats.get("appended", 0),
        'pl_update_same': pl_stats.get("updated_same", 0),
        'pl_update_resize': pl_stats.get("updated_resize", 0),
        'cache_hit': cache_hit,
        'rebuilt_files': rebuilt_files,
        'errors': len(errors),
        'cache_path': _u87_detail_cache_path(output_path),
    }


def _u98_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _u85_reset_bench()
    t_all_0 = time.perf_counter()
    _U90_SOURCE_CACHE_CTX['output_path'] = output_path
    _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
    cache = _u90_get_source_cache()
    cache['stats'] = _u90_empty_source_cache()['stats']

    try:
        self.logger.log('FILE_UPSERT_ENTER_U98', 'info')
        self.logger.log('SOURCE_PARSE_CACHE_ENTER_U98', 'info')

        _U84_WIDTH_CACHE_CTX['output_path'] = output_path
        _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

        existing_supported = _u98_collect_supported_paths_from_output(output_path)
        scanned_files = _u87_norm_path_set(collect_source_files(folder_path, output_path))

        t0 = time.perf_counter()
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        core_seconds = max(0.0, time.perf_counter() - t0)

        current_supported = _u98_collect_supported_paths_from_output(output_path)
        changed_paths = _u87_norm_path_set(p for p in current_supported if p not in existing_supported)

        t1 = time.perf_counter()
        upsert_stats = _u98_rebuild_upsert(
            output_path,
            current_supported,
            changed_paths=changed_paths,
            prep_label='DETAIL_PREP_U79'
        )
        rebuild_seconds = max(0.0, time.perf_counter() - t1)

        total_seconds = max(0.0, time.perf_counter() - t_all_0)

        cache = _u90_get_source_cache()
        _u90_save_source_cache(output_path, cache)
        stats = cache.get('stats', {}) if isinstance(cache, dict) else {}

        bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
        bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                        f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | changed={upsert_stats.get('changed',0)} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
        bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                        f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
        _u85_append_bench_log_rows(output_path, bench_rows)
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}
        _U85_BENCH_CTX['enabled'] = False
        _U90_SOURCE_CACHE_CTX['output_path'] = ''
        _U90_SOURCE_CACHE_CTX['cache'] = None




# =========================================================
# U99 FIXED FILE UPSERT
# Muc tieu:
# - Sua dung U98
# - Chuan hoa path triet de de index match dung
# - Dung file_path_idx chinh xac cho INV / PL
# - Khong chi xu ly file moi, ma xu ly tat ca file duoc scan trong lan run nay
#   neu file do co support detail (OK/WARNING)
# - WARNING -> OK se update truc tiep block cu
# - Neu hash va row_count khong doi -> skip
# - Neu hash doi:
#     + row_count giong nhau -> overwrite tai cho
#     + row_count khac nhau -> replace block tai vi tri cu
# =========================================================

def _u99_norm_path(v) -> str:
    p = nz_str(v).strip()
    if not p:
        return ""
    try:
        p = os.path.abspath(p)
    except Exception:
        pass
    p = p.replace("\\", "/")
    while "//" in p:
        p = p.replace("//", "/")
    return p.strip().upper()


def _u99_collect_supported_paths(ws_sub) -> set:
    out = set()
    if ws_sub is None:
        return out
    path_col = find_header_col(ws_sub, "File Path")
    status_col = find_header_col(ws_sub, "Status")
    if path_col <= 0 or status_col <= 0:
        return out
    for r in range(2, ws_sub.max_row + 1):
        st = nz_str(ws_sub.cell(r, status_col).value).upper()
        if st in ("OK", "WARNING"):
            p = _u99_norm_path(ws_sub.cell(r, path_col).value)
            if p:
                out.add(p)
    return out


def _u99_collect_supported_paths_from_output(output_path: str) -> set:
    wb = open_or_create_output_workbook(output_path)
    try:
        if "SUB_DETAIL" not in wb.sheetnames:
            return set()
        return _u99_collect_supported_paths(wb["SUB_DETAIL"])
    finally:
        wb.close()


def _u99_load_index(wb) -> Dict[str, Dict[str, Dict[str, Any]]]:
    out = {"INV": {}, "PL": {}}
    if _U98_BLOCK_INDEX_SHEET not in wb.sheetnames:
        return out
    ws = wb[_U98_BLOCK_INDEX_SHEET]
    if ws.max_row < 2:
        return out
    headers = [nz_str(ws.cell(1, c).value) for c in range(1, 8)]
    expected = ["Sheet", "FilePath", "StartRow", "EndRow", "RowCount", "Status", "Hash"]
    if headers[:7] != expected:
        return out
    for r in range(2, ws.max_row + 1):
        sh = nz_str(ws.cell(r, 1).value).upper()
        if sh not in ("INV", "PL"):
            continue
        fp = _u99_norm_path(ws.cell(r, 2).value)
        if not fp:
            continue
        out[sh][fp] = {
            "sheet": sh,
            "file_path": fp,
            "start_row": int(ws.cell(r, 3).value or 0),
            "end_row": int(ws.cell(r, 4).value or 0),
            "row_count": int(ws.cell(r, 5).value or 0),
            "status": nz_str(ws.cell(r, 6).value).upper(),
            "hash": nz_str(ws.cell(r, 7).value),
        }
    return out


def _u99_save_index(wb, inv_index: Dict[str, Dict[str, Any]], pl_index: Dict[str, Dict[str, Any]]):
    ws = _u98_ensure_index_sheet(wb)
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    headers = ["Sheet", "FilePath", "StartRow", "EndRow", "RowCount", "Status", "Hash"]
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c).value = h
    row_no = 2
    for sh_name, idx_map in (("INV", inv_index), ("PL", pl_index)):
        items = sorted(idx_map.values(), key=lambda x: int(x.get("start_row", 0) or 0))
        for item in items:
            ws.cell(row_no, 1).value = sh_name
            ws.cell(row_no, 2).value = item.get("file_path", "")
            ws.cell(row_no, 3).value = int(item.get("start_row", 0) or 0)
            ws.cell(row_no, 4).value = int(item.get("end_row", 0) or 0)
            ws.cell(row_no, 5).value = int(item.get("row_count", 0) or 0)
            ws.cell(row_no, 6).value = item.get("status", "")
            ws.cell(row_no, 7).value = item.get("hash", "")
            row_no += 1
    ws.sheet_state = "hidden"


def _u99_collect_existing_file_paths_from_sheet(ws, file_path_idx: int) -> set:
    out = set()
    if ws is None or ws.max_row < 2:
        return out
    for r in range(2, ws.max_row + 1):
        if file_path_idx <= ws.max_column:
            p = _u99_norm_path(ws.cell(r, file_path_idx).value)
            if p:
                out.add(p)
    return out


def _u99_upsert_sheet(ws, headers: List[str], index_map: Dict[str, Dict[str, Any]], file_rows_map: Dict[str, Dict[str, Any]], kind: str, date_cols: List[int], file_path_idx: int, status_field: str, hash_field: str, rows_field: str) -> Dict[str, int]:
    appended = 0
    updated_same = 0
    updated_resize = 0
    skipped_same = 0
    repaired_index = 0

    existing_physical = _u99_collect_existing_file_paths_from_sheet(ws, file_path_idx)

    for fp in sorted(file_rows_map.keys()):
        payload = file_rows_map.get(fp) or {}
        rows = payload.get(rows_field, []) or []
        status_val = nz_str(payload.get(status_field, "")).upper()
        hash_val = nz_str(payload.get(hash_field, ""))

        if not rows:
            continue

        meta = index_map.get(fp)

        # If physical sheet has path but index missing, rebuild a minimal index entry by scanning contiguous rows of same file_path.
        if meta is None and fp in existing_physical:
            start_row = 0
            end_row = 0
            for r in range(2, ws.max_row + 1):
                p = _u99_norm_path(ws.cell(r, file_path_idx).value)
                if p != fp:
                    continue
                if start_row == 0:
                    start_row = r
                end_row = r
            if start_row > 0 and end_row >= start_row:
                meta = {
                    "sheet": kind,
                    "file_path": fp,
                    "start_row": start_row,
                    "end_row": end_row,
                    "row_count": end_row - start_row + 1,
                    "status": "",
                    "hash": "",
                }
                index_map[fp] = meta
                repaired_index += 1

        if meta is None:
            start_row, count = _u98_append_block(ws, rows, headers, date_cols, file_path_idx, 0)
            if count > 0:
                index_map[fp] = {
                    "sheet": kind,
                    "file_path": fp,
                    "start_row": start_row,
                    "end_row": start_row + count - 1,
                    "row_count": count,
                    "status": status_val,
                    "hash": hash_val,
                }
                appended += count
            continue

        old_count = int(meta.get("row_count", 0) or 0)
        old_hash = nz_str(meta.get("hash", ""))
        old_status = nz_str(meta.get("status", "")).upper()
        start_row = int(meta.get("start_row", 0) or 0)

        if old_hash == hash_val and old_status == status_val and old_count == len(rows):
            skipped_same += 1
            continue

        if old_count == len(rows):
            _u98_overwrite_block_same_size(ws, start_row, rows, headers, date_cols)
            updated_same += len(rows)
        else:
            delta = len(rows) - old_count
            _u98_replace_block_resize(ws, start_row, old_count, rows, headers, date_cols)
            _u98_shift_index_after(index_map, start_row + old_count, delta, skip_fp=fp)
            updated_resize += len(rows)

        meta["row_count"] = len(rows)
        meta["start_row"] = start_row
        meta["end_row"] = start_row + len(rows) - 1
        meta["status"] = status_val
        meta["hash"] = hash_val

    return {
        "appended": appended,
        "updated_same": updated_same,
        "updated_resize": updated_resize,
        "skipped_same": skipped_same,
        "repaired_index": repaired_index,
    }


def _u99_rebuild_upsert(output_path: str, supported_paths: set, changed_paths: set, prep_label: str = 'DETAIL_PREP_U79') -> Dict[str, Any]:
    wb = open_or_create_output_workbook(output_path)
    errors: List[Tuple[str, str]] = []
    cache = _u87_load_detail_cache(output_path)
    cache_files: Dict[str, Any] = cache.get('files', {}) if isinstance(cache, dict) else {}
    if not isinstance(cache_files, dict):
        cache_files = {}

    file_list = sorted([p for p in supported_paths if os.path.exists(p)])
    changed_paths = _u87_norm_path_set(changed_paths) & set(_u87_norm_path_set(file_list))

    cache_hit = 0
    rebuilt_files = 0
    dirty_cache = False

    try:
        ws_inv = ensure_sheet(wb, 'INV')
        ws_pl = ensure_sheet(wb, 'PL')
        _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
        _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

        idx_all = _u99_load_index(wb)
        inv_index = idx_all.get("INV", {})
        pl_index = idx_all.get("PL", {})

        # IMPORTANT U99: process all supported files scanned in current run, not only new files.
        file_rows_map: Dict[str, Dict[str, Any]] = {}
        for fp in sorted(changed_paths):
            try:
                payload = _u98_build_rows_for_file(fp, cache_files, force_refresh=True)
                if payload["source"] == "cache":
                    cache_hit += 1
                else:
                    rebuilt_files += 1
                    dirty_cache = True
                file_rows_map[_u99_norm_path(fp)] = payload
            except Exception as exc:
                errors.append((fp, str(exc)))

        inv_stats = _u99_upsert_sheet(
            ws_inv, V73_INV_HEADERS, inv_index, file_rows_map,
            kind="INV", date_cols=[1], file_path_idx=30,
            status_field="inv_status", hash_field="inv_hash", rows_field="inv_rows"
        )
        pl_stats = _u99_upsert_sheet(
            ws_pl, V76_PL_HEADERS, pl_index, file_rows_map,
            kind="PL", date_cols=[1], file_path_idx=19,
            status_field="pl_status", hash_field="pl_hash", rows_field="pl_rows"
        )

        _u99_save_index(wb, inv_index, pl_index)

        ws_log = ensure_sheet(wb, 'LOG_SUB_DETAIL')
        init_log_header(ws_log)
        append_log_rows(ws_log, [make_log_row('', prep_label, 'INFO',
                            f'Paths for detail (OK+WARNING)={len(supported_paths)} | Start file-upsert detail tabs | Merge mode=1 file | SmartCache=ON | WriteMode=FILE_UPSERT_U99')])
        append_log_rows(ws_log, [make_log_row('', 'DETAIL_REBUILD_U99', 'INFO',
                            f'Current supported files={len(file_list)} | changed={len(changed_paths)} | INV_append={inv_stats.get("appended",0)} | INV_update_same={inv_stats.get("updated_same",0)} | INV_update_resize={inv_stats.get("updated_resize",0)} | INV_index_repaired={inv_stats.get("repaired_index",0)} | PL_append={pl_stats.get("appended",0)} | PL_update_same={pl_stats.get("updated_same",0)} | PL_update_resize={pl_stats.get("updated_resize",0)} | PL_index_repaired={pl_stats.get("repaired_index",0)} | cache_hit={cache_hit} | rebuilt={rebuilt_files}')])
        for fp, err in errors:
            append_log_rows(ws_log, [make_log_row(fp, 'DETAIL_REBUILD_U99', 'ERROR', err)])

        safe_save_workbook_atomic(wb, output_path)
    finally:
        wb.close()

    cache['version'] = _U87_DETAIL_CACHE_VERSION
    cache['files'] = cache_files
    if dirty_cache:
        _u87_save_detail_cache(output_path, cache)

    return {
        'supported_files': len(file_list),
        'changed': len(changed_paths),
        'inv_append': inv_stats.get("appended", 0),
        'inv_update_same': inv_stats.get("updated_same", 0),
        'inv_update_resize': inv_stats.get("updated_resize", 0),
        'inv_index_repaired': inv_stats.get("repaired_index", 0),
        'pl_append': pl_stats.get("appended", 0),
        'pl_update_same': pl_stats.get("updated_same", 0),
        'pl_update_resize': pl_stats.get("updated_resize", 0),
        'pl_index_repaired': pl_stats.get("repaired_index", 0),
        'cache_hit': cache_hit,
        'rebuilt_files': rebuilt_files,
        'errors': len(errors),
        'cache_path': _u87_detail_cache_path(output_path),
    }


def _u99_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _u85_reset_bench()
    t_all_0 = time.perf_counter()
    _U90_SOURCE_CACHE_CTX['output_path'] = output_path
    _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
    cache = _u90_get_source_cache()
    cache['stats'] = _u90_empty_source_cache()['stats']

    try:
        self.logger.log('FILE_UPSERT_ENTER_U99', 'info')
        self.logger.log('SOURCE_PARSE_CACHE_ENTER_U99', 'info')

        _U84_WIDTH_CACHE_CTX['output_path'] = output_path
        _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

        scanned_files = _u87_norm_path_set(_u99_norm_path(p) for p in collect_source_files(folder_path, output_path))

        t0 = time.perf_counter()
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        core_seconds = max(0.0, time.perf_counter() - t0)

        current_supported = _u99_collect_supported_paths_from_output(output_path)

        # IMPORTANT U99:
        # process only files scanned in this run that are currently supported (OK/WARNING),
        # including files that already existed before but need update WARNING->OK or content update.
        changed_paths = _u87_norm_path_set(p for p in scanned_files if p in current_supported)

        t1 = time.perf_counter()
        upsert_stats = _u99_rebuild_upsert(
            output_path,
            current_supported,
            changed_paths=changed_paths,
            prep_label='DETAIL_PREP_U79'
        )
        rebuild_seconds = max(0.0, time.perf_counter() - t1)

        total_seconds = max(0.0, time.perf_counter() - t_all_0)

        cache = _u90_get_source_cache()
        _u90_save_source_cache(output_path, cache)
        stats = cache.get('stats', {}) if isinstance(cache, dict) else {}

        bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
        bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                        f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | changed={upsert_stats.get('changed',0)} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | inv_index_repaired={upsert_stats.get('inv_index_repaired',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | pl_index_repaired={upsert_stats.get('pl_index_repaired',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
        bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                        f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
        _u85_append_bench_log_rows(output_path, bench_rows)
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}
        _U85_BENCH_CTX['enabled'] = False
        _U90_SOURCE_CACHE_CTX['output_path'] = ''
        _U90_SOURCE_CACHE_CTX['cache'] = None




# =========================================================
# U100 BOOTSTRAP FIX FOR EMPTY INV / PL
# Muc tieu:
# - Neu SUB_DETAIL da co du lieu nhung INV/PL dang rong -> phai seed lai toan bo
# - Khong duoc de changed=0 lam bo qua rebuild khi 2 sheet detail chua co data
# - Giu nguyen logic file-upsert cua U99 cho cac lan chay sau
# =========================================================

def _u100_sheet_has_data_rows(ws) -> bool:
    try:
        return ws is not None and ws.max_row >= 2
    except Exception:
        return False


def _u100_need_bootstrap(output_path: str) -> bool:
    wb = open_or_create_output_workbook(output_path)
    try:
        ws_inv = wb["INV"] if "INV" in wb.sheetnames else None
        ws_pl = wb["PL"] if "PL" in wb.sheetnames else None

        inv_has = _u100_sheet_has_data_rows(ws_inv)
        pl_has = _u100_sheet_has_data_rows(ws_pl)

        # Nếu một trong hai sheet chưa có data rows thì coi là chưa bootstrap xong.
        if not inv_has or not pl_has:
            return True

        # Nếu hidden index thiếu hoàn toàn cũng phải bootstrap lại để dựng index chuẩn.
        idx_map = _u98_load_index(wb) if "_U98_FILE_BLOCK_INDEX" in wb.sheetnames else {"INV": {}, "PL": {}}
        inv_idx = idx_map.get("INV", {}) if isinstance(idx_map, dict) else {}
        pl_idx = idx_map.get("PL", {}) if isinstance(idx_map, dict) else {}

        if not inv_idx or not pl_idx:
            return True

        return False
    finally:
        wb.close()


def _u100_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _u85_reset_bench()
    t_all_0 = time.perf_counter()
    _U90_SOURCE_CACHE_CTX['output_path'] = output_path
    _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
    cache = _u90_get_source_cache()
    cache['stats'] = _u90_empty_source_cache()['stats']

    try:
        self.logger.log('FILE_UPSERT_ENTER_U100', 'info')
        self.logger.log('SOURCE_PARSE_CACHE_ENTER_U100', 'info')

        _U84_WIDTH_CACHE_CTX['output_path'] = output_path
        _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

        existing_supported = _u98_collect_supported_paths_from_output(output_path)
        scanned_files = _u87_norm_path_set(collect_source_files(folder_path, output_path))

        t0 = time.perf_counter()
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        core_seconds = max(0.0, time.perf_counter() - t0)

        current_supported = _u98_collect_supported_paths_from_output(output_path)

        bootstrap_needed = _u100_need_bootstrap(output_path)

        if bootstrap_needed:
            changed_paths = set(current_supported)
        else:
            # U99 fixed logic: process files scanned in this run that are currently supported (OK/WARNING),
            # not just "new vs old SUB_DETAIL" by stale workbook state.
            changed_paths = _u87_norm_path_set(p for p in scanned_files if p in current_supported)

        t1 = time.perf_counter()
        upsert_stats = _u98_rebuild_upsert(
            output_path,
            current_supported,
            changed_paths=changed_paths,
            prep_label='DETAIL_PREP_U79'
        )
        rebuild_seconds = max(0.0, time.perf_counter() - t1)

        total_seconds = max(0.0, time.perf_counter() - t_all_0)

        cache = _u90_get_source_cache()
        _u90_save_source_cache(output_path, cache)
        stats = cache.get('stats', {}) if isinstance(cache, dict) else {}

        bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
        bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                        f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | changed={upsert_stats.get('changed',0)} | bootstrap={1 if bootstrap_needed else 0} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
        bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                        f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
        _u85_append_bench_log_rows(output_path, bench_rows)
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}
        _U85_BENCH_CTX['enabled'] = False
        _U90_SOURCE_CACHE_CTX['output_path'] = ''
        _U90_SOURCE_CACHE_CTX['cache'] = None




# =========================================================
# U101 FIX DISPATCH
# Muc tieu:
# - Giu bootstrap fix cua U100
# - NHUNG phai goi dung nhom ham U99 (path normalize + upsert)
# - Khong duoc tiep tuc goi _u98_* vi se append lai va log sai nhan
# =========================================================

def _u101_need_bootstrap(output_path: str) -> bool:
    wb = open_or_create_output_workbook(output_path)
    try:
        ws_inv = wb["INV"] if "INV" in wb.sheetnames else None
        ws_pl = wb["PL"] if "PL" in wb.sheetnames else None

        inv_has = ws_inv is not None and ws_inv.max_row >= 2
        pl_has = ws_pl is not None and ws_pl.max_row >= 2
        if not inv_has or not pl_has:
            return True

        idx_map = _u99_load_index(wb) if "_U98_FILE_BLOCK_INDEX" in wb.sheetnames else {"INV": {}, "PL": {}}
        inv_idx = idx_map.get("INV", {}) if isinstance(idx_map, dict) else {}
        pl_idx = idx_map.get("PL", {}) if isinstance(idx_map, dict) else {}
        return (not inv_idx) or (not pl_idx)
    finally:
        wb.close()


def _u101_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _u85_reset_bench()
    t_all_0 = time.perf_counter()
    _U90_SOURCE_CACHE_CTX['output_path'] = output_path
    _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
    cache = _u90_get_source_cache()
    cache['stats'] = _u90_empty_source_cache()['stats']

    try:
        self.logger.log('FILE_UPSERT_ENTER_U101', 'info')
        self.logger.log('SOURCE_PARSE_CACHE_ENTER_U101', 'info')

        _U84_WIDTH_CACHE_CTX['output_path'] = output_path
        _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

        scanned_files = _u87_norm_path_set(_u99_norm_path(p) for p in collect_source_files(folder_path, output_path))

        t0 = time.perf_counter()
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        core_seconds = max(0.0, time.perf_counter() - t0)

        current_supported = _u99_collect_supported_paths_from_output(output_path)
        bootstrap_needed = _u101_need_bootstrap(output_path)

        if bootstrap_needed:
            changed_paths = set(current_supported)
        else:
            changed_paths = _u87_norm_path_set(p for p in scanned_files if p in current_supported)

        t1 = time.perf_counter()
        upsert_stats = _u99_rebuild_upsert(
            output_path,
            current_supported,
            changed_paths=changed_paths,
            prep_label='DETAIL_PREP_U79'
        )
        rebuild_seconds = max(0.0, time.perf_counter() - t1)

        total_seconds = max(0.0, time.perf_counter() - t_all_0)

        cache = _u90_get_source_cache()
        _u90_save_source_cache(output_path, cache)
        stats = cache.get('stats', {}) if isinstance(cache, dict) else {}

        bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
        bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                        f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | changed={upsert_stats.get('changed',0)} | bootstrap={1 if bootstrap_needed else 0} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | inv_index_repaired={upsert_stats.get('inv_index_repaired',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | pl_index_repaired={upsert_stats.get('pl_index_repaired',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
        bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                        f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
        _u85_append_bench_log_rows(output_path, bench_rows)
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}
        _U85_BENCH_CTX['enabled'] = False
        _U90_SOURCE_CACHE_CTX['output_path'] = ''
        _U90_SOURCE_CACHE_CTX['cache'] = None




# =========================================================
# U102 FIX CURRENT-FOLDER PATH MATCH
# Muc tieu:
# - Khong tinh changed bang intersection set mo ho nua
# - Doc truc tiep SUB_DETAIL sau CORE_RUN
# - Chi lay cac dong co Status in (OK, WARNING) va File Path thuoc scanned_files cua lan nay
# - Giai quyet loi thang 4 / thang 1 khong vao INV/PL du changed=0 sai
# =========================================================

def _u102_collect_supported_scanned_paths(output_path: str, scanned_files: set) -> set:
    scanned_files = _u87_norm_path_set(_u99_norm_path(p) for p in (scanned_files or set()))
    wb = open_or_create_output_workbook(output_path)
    out = set()
    try:
        if "SUB_DETAIL" not in wb.sheetnames:
            return out
        ws = wb["SUB_DETAIL"]
        path_col = find_header_col(ws, "File Path")
        status_col = find_header_col(ws, "Status")
        if path_col <= 0 or status_col <= 0:
            return out

        for r in range(2, ws.max_row + 1):
            st = nz_str(ws.cell(r, status_col).value).upper()
            if st not in ("OK", "WARNING"):
                continue
            fp = _u99_norm_path(ws.cell(r, path_col).value)
            if fp and fp in scanned_files:
                out.add(fp)
        return out
    finally:
        wb.close()


def _u102_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _u85_reset_bench()
    t_all_0 = time.perf_counter()
    _U90_SOURCE_CACHE_CTX['output_path'] = output_path
    _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
    cache = _u90_get_source_cache()
    cache['stats'] = _u90_empty_source_cache()['stats']

    try:
        self.logger.log('FILE_UPSERT_ENTER_U102', 'info')
        self.logger.log('SOURCE_PARSE_CACHE_ENTER_U102', 'info')

        _U84_WIDTH_CACHE_CTX['output_path'] = output_path
        _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

        scanned_files = _u87_norm_path_set(_u99_norm_path(p) for p in collect_source_files(folder_path, output_path))

        t0 = time.perf_counter()
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        core_seconds = max(0.0, time.perf_counter() - t0)

        current_supported = _u99_collect_supported_paths_from_output(output_path)
        bootstrap_needed = _u101_need_bootstrap(output_path)

        if bootstrap_needed:
            changed_paths = set(current_supported)
        else:
            changed_paths = _u102_collect_supported_scanned_paths(output_path, scanned_files)

        t1 = time.perf_counter()
        upsert_stats = _u99_rebuild_upsert(
            output_path,
            current_supported,
            changed_paths=changed_paths,
            prep_label='DETAIL_PREP_U79'
        )
        rebuild_seconds = max(0.0, time.perf_counter() - t1)

        total_seconds = max(0.0, time.perf_counter() - t_all_0)

        cache = _u90_get_source_cache()
        _u90_save_source_cache(output_path, cache)
        stats = cache.get('stats', {}) if isinstance(cache, dict) else {}

        bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
        bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                        f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | changed={upsert_stats.get('changed',0)} | bootstrap={1 if bootstrap_needed else 0} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | inv_index_repaired={upsert_stats.get('inv_index_repaired',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | pl_index_repaired={upsert_stats.get('pl_index_repaired',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
        bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                        f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
        _u85_append_bench_log_rows(output_path, bench_rows)
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}
        _U85_BENCH_CTX['enabled'] = False
        _U90_SOURCE_CACHE_CTX['output_path'] = ''
        _U90_SOURCE_CACHE_CTX['cache'] = None




# =========================================================
# U103 USE CORE SUMMARY PATHS
# Muc tieu:
# - Bo suy doan changed_paths bang cach doc lai workbook / intersection de tranh sai lech
# - Dung truc tiep danh sach file da duoc xu ly trong lan run hien tai do CORE_RUN_V6 tra ve
# - Sau do filter theo current_supported (OK/WARNING) de dua vao INV/PL
# =========================================================

def _u103_extract_processed_paths_from_summary(summary) -> set:
    out = set()
    try:
        paths = (summary or {}).get('changed_detail_files', []) or []
        for p in paths:
            pp = _u99_norm_path(p)
            if pp:
                out.add(pp)
    except Exception:
        pass
    return out


def _u103_processor_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
    _u85_reset_bench()
    t_all_0 = time.perf_counter()
    _U90_SOURCE_CACHE_CTX['output_path'] = output_path
    _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
    cache = _u90_get_source_cache()
    cache['stats'] = _u90_empty_source_cache()['stats']

    try:
        self.logger.log('FILE_UPSERT_ENTER_U103', 'info')
        self.logger.log('SOURCE_PARSE_CACHE_ENTER_U103', 'info')

        _U84_WIDTH_CACHE_CTX['output_path'] = output_path
        _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

        scanned_files = _u87_norm_path_set(_u99_norm_path(p) for p in collect_source_files(folder_path, output_path))

        t0 = time.perf_counter()
        summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
        core_seconds = max(0.0, time.perf_counter() - t0)

        current_supported = _u99_collect_supported_paths_from_output(output_path)
        bootstrap_needed = _u101_need_bootstrap(output_path)

        if bootstrap_needed:
            changed_paths = set(current_supported)
        else:
            processed_paths = _u103_extract_processed_paths_from_summary(summary)
            changed_paths = _u87_norm_path_set(p for p in processed_paths if p in current_supported)

            if not changed_paths and scanned_files:
                changed_paths = _u102_collect_supported_scanned_paths(output_path, scanned_files)

        t1 = time.perf_counter()
        upsert_stats = _u99_rebuild_upsert(
            output_path,
            current_supported,
            changed_paths=changed_paths,
            prep_label='DETAIL_PREP_U79'
        )
        rebuild_seconds = max(0.0, time.perf_counter() - t1)

        total_seconds = max(0.0, time.perf_counter() - t_all_0)

        cache = _u90_get_source_cache()
        _u90_save_source_cache(output_path, cache)
        stats = cache.get('stats', {}) if isinstance(cache, dict) else {}

        bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
        bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                        f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | changed={upsert_stats.get('changed',0)} | bootstrap={1 if bootstrap_needed else 0} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | inv_index_repaired={upsert_stats.get('inv_index_repaired',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | pl_index_repaired={upsert_stats.get('pl_index_repaired',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
        bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                        f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
        _u85_append_bench_log_rows(output_path, bench_rows)
        return summary
    finally:
        _U84_WIDTH_CACHE_CTX['output_path'] = ''
        _U84_WIDTH_CACHE_CTX['cache'] = {}
        _U85_BENCH_CTX['enabled'] = False
        _U90_SOURCE_CACHE_CTX['output_path'] = ''
        _U90_SOURCE_CACHE_CTX['cache'] = None




# =========================================================
# REFACTORED RUNTIME ENTRY
# - Single Processor.run uses latest U103 flow
# - No active monkey patching
# - Structured logging for unexpected exceptions
# =========================================================

_logger_runtime = logging.getLogger("detail_processor")
if not _logger_runtime.handlers:
    _logger_runtime.setLevel(logging.INFO)
    _h = logging.StreamHandler()
    _h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    _logger_runtime.addHandler(_h)


_U130_PARSE_CACHE_VERSION = 1


def _u130_parse_cache_path(output_path: str) -> str:
    base = output_path or 'detail_sub'
    return base + '.sourcecache_u130.pkl' if base else 'detail_sub.sourcecache_u130.pkl'


def _u130_empty_parse_cache() -> Dict[str, Any]:
    return {
        'version': _U130_PARSE_CACHE_VERSION,
        'entries': {},
        'stats': {
            'lookup_total': 0,
            'cache_hit_reuse': 0,
            'cache_miss_parse': 0,
            'cache_bypass': 0,
            'parse_real_total': 0,
            'parse_reused_total': 0,
            'cache_store_total': 0,
        }
    }


def _u130_load_parse_cache(output_path: str) -> Dict[str, Any]:
    path = _u130_parse_cache_path(output_path)
    try:
        with open(path, 'rb') as f:
            data = pickle.load(f)
        if not isinstance(data, dict):
            return _u130_empty_parse_cache()
        if int(data.get('version', 0) or 0) != _U130_PARSE_CACHE_VERSION:
            return _u130_empty_parse_cache()
        entries = data.get('entries')
        stats = data.get('stats')
        if not isinstance(entries, dict):
            entries = {}
        if not isinstance(stats, dict):
            stats = _u130_empty_parse_cache()['stats']
        return {'version': _U130_PARSE_CACHE_VERSION, 'entries': entries, 'stats': stats}
    except Exception:
        return _u130_empty_parse_cache()


def _u130_save_parse_cache(output_path: str, cache: Dict[str, Any]):
    path = _u130_parse_cache_path(output_path)
    tmp = path + '.tmp'
    try:
        with open(tmp, 'wb') as f:
            pickle.dump(cache, f, protocol=pickle.HIGHEST_PROTOCOL)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except Exception:
                pass


def _u130_stats_add(cache: Dict[str, Any], name: str, inc: int = 1):
    stats = cache.setdefault('stats', {})
    try:
        stats[name] = int(stats.get(name, 0) or 0) + int(inc)
    except Exception:
        pass


def _u130_get_stats(cache: Dict[str, Any]) -> Dict[str, Any]:
    stats = dict((cache or {}).get('stats') or {})
    for k in ('lookup_total', 'cache_hit_reuse', 'cache_miss_parse', 'cache_bypass', 'parse_real_total', 'parse_reused_total', 'cache_store_total'):
        stats[k] = int(stats.get(k, 0) or 0)
    return stats


def _u130_repair_signature(repair_options: Optional[RepairOptions]) -> str:
    ro = repair_options or RepairOptions()
    return '|'.join([
        'folder=' + ('1' if bool(getattr(ro, 'use_folder_truth', False)) else '0'),
        'manual=' + ('1' if bool(getattr(ro, 'use_manual_values', False)) else '0'),
        'inv=' + nz_str(getattr(ro, 'manual_invoice_no', '')),
        'date=' + nz_str(getattr(ro, 'manual_invoice_date', '')),
        'dest=' + nz_str(getattr(ro, 'manual_destination', '')),
    ])


def _u130_file_hash(file_path: str, full_limit: int = 2 * 1024 * 1024, chunk: int = 65536) -> str:
    h = hashlib.md5()
    try:
        st = os.stat(file_path)
        size = int(getattr(st, 'st_size', 0) or 0)
        with open(file_path, 'rb') as f:
            if size <= full_limit:
                while True:
                    data = f.read(1024 * 1024)
                    if not data:
                        break
                    h.update(data)
            else:
                head = f.read(chunk)
                h.update(head)
                if size > chunk:
                    f.seek(max(0, size - chunk))
                    tail = f.read(chunk)
                    h.update(tail)
                h.update(str(size).encode('utf-8'))
        return h.hexdigest()
    except Exception:
        return ''


def _u130_build_parse_fp(file_path: str, repair_options: Optional[RepairOptions]) -> str:
    try:
        st = os.stat(file_path)
        mtime_ns = int(getattr(st, 'st_mtime_ns', int(st.st_mtime * 1_000_000_000)))
        size = int(getattr(st, 'st_size', 0) or 0)
    except Exception:
        mtime_ns = 0
        size = 0
    return '|'.join([
        f'mtime_ns={mtime_ns}',
        f'size={size}',
        f'hash={_u130_file_hash(file_path)}',
        f'repair={_u130_repair_signature(repair_options)}',
        f'engine={U107_ENGINE_VERSION}',
    ])


def _u130_parse_cache_key(file_path: str) -> str:
    return _u104_canonical_path(file_path)


def _u130_deepcopy_payload(obj: Any) -> Any:
    try:
        return copy.deepcopy(obj)
    except Exception:
        try:
            return pickle.loads(pickle.dumps(obj, protocol=pickle.HIGHEST_PROTOCOL))
        except Exception:
            return obj


def _u130_lookup_parse_cache(cache: Dict[str, Any], file_path: str, parse_fp: str, force_resync: bool = False) -> Tuple[bool, Optional[Dict[str, Any]], str]:
    _u130_stats_add(cache, 'lookup_total')
    if force_resync:
        _u130_stats_add(cache, 'cache_bypass')
        return False, None, 'FORCE_RESYNC'
    key = _u130_parse_cache_key(file_path)
    entry = (cache.get('entries') or {}).get(key)
    if not isinstance(entry, dict):
        _u130_stats_add(cache, 'cache_miss_parse')
        return False, None, 'CACHE_MISSING'
    if nz_str(entry.get('parse_fp')) != nz_str(parse_fp):
        _u130_stats_add(cache, 'cache_bypass')
        return False, None, 'FP_CHANGED'
    if 'arr' not in entry or 'bundle4' not in entry:
        _u130_stats_add(cache, 'cache_bypass')
        return False, None, 'CACHE_INCOMPLETE'
    _u130_stats_add(cache, 'cache_hit_reuse')
    _u130_stats_add(cache, 'parse_reused_total')
    return True, entry, 'REUSE_CACHE'


def _u130_store_parse_cache(cache: Dict[str, Any], file_path: str, parse_fp: str, arr: List[Any], bundle4: Any):
    entries = cache.setdefault('entries', {})
    key = _u130_parse_cache_key(file_path)
    entries[key] = {
        'parse_fp': parse_fp,
        'arr': _u130_deepcopy_payload(arr),
        'bundle4': _u130_deepcopy_payload(bundle4),
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }
    _u130_stats_add(cache, 'cache_store_total')


def _u130_prune_parse_cache(cache: Dict[str, Any], valid_pathkeys: Set[str]):
    entries = cache.setdefault('entries', {})
    drop = [k for k in list(entries.keys()) if k not in valid_pathkeys]
    for k in drop:
        entries.pop(k, None)


def _u130_append_runtime_bench_log(output_path: str, line: str):
    p = output_path + '.runtime_bench.log'
    os.makedirs(os.path.dirname(p) or '.', exist_ok=True)
    with open(p, 'a', encoding='utf-8') as f:
        f.write(line.rstrip() + "\n")


_U131_DETAIL_INDEX_VERSION = 1


def _u131_detail_index_path(output_path: str) -> str:
    base = output_path or 'detail_sub'
    return base + '.detailindex_u131.pkl' if base else 'detail_sub.detailindex_u131.pkl'


def _u131_empty_detail_index() -> Dict[str, Any]:
    return {'version': _U131_DETAIL_INDEX_VERSION, 'inv': {'rows': {}, 'max_row': 0}, 'pl': {'rows': {}, 'max_row': 0}}


def _u131_load_detail_index(output_path: str) -> Dict[str, Any]:
    path = _u131_detail_index_path(output_path)
    try:
        with open(path, 'rb') as f:
            data = pickle.load(f)
        if not isinstance(data, dict):
            return _u131_empty_detail_index()
        if int(data.get('version', 0) or 0) != _U131_DETAIL_INDEX_VERSION:
            return _u131_empty_detail_index()
        return data
    except Exception:
        return _u131_empty_detail_index()


def _u131_save_detail_index(output_path: str, data: Dict[str, Any]):
    path = _u131_detail_index_path(output_path)
    tmp = path + '.tmp'
    try:
        with open(tmp, 'wb') as f:
            pickle.dump(data, f, protocol=pickle.HIGHEST_PROTOCOL)
        os.replace(tmp, path)
    finally:
        if os.path.exists(tmp):
            try:
                os.remove(tmp)
            except Exception:
                pass


def _u131_build_detail_index_payload(inv_map: Dict[str, List[int]], pl_map: Dict[str, List[int]], ws_inv, ws_pl) -> Dict[str, Any]:
    return {
        'version': _U131_DETAIL_INDEX_VERSION,
        'inv': {'rows': inv_map or {}, 'max_row': int(getattr(ws_inv, 'max_row', 0) or 0)},
        'pl': {'rows': pl_map or {}, 'max_row': int(getattr(ws_pl, 'max_row', 0) or 0)},
    }


def _u131_collect_delete_rows_for_keys(ws, cache_section: Dict[str, Any], affected_keys: Set[str], pathkey_col: int) -> Tuple[List[int], bool]:
    rows_map = (cache_section or {}).get('rows') or {}
    max_row = int((cache_section or {}).get('max_row', 0) or 0)
    if not isinstance(rows_map, dict) or not affected_keys:
        return [], False
    current_max = int(getattr(ws, 'max_row', 0) or 0)
    if max_row and current_max != max_row:
        return [], False
    out: List[int] = []
    for key in affected_keys:
        rownums = rows_map.get(key)
        if not isinstance(rownums, list) or not rownums:
            continue
        first = int(rownums[0])
        last = int(rownums[-1])
        if first < 2 or last > current_max:
            return [], False
        if nz_str(ws.cell(first, pathkey_col).value) != key:
            return [], False
        if nz_str(ws.cell(last, pathkey_col).value) != key:
            return [], False
        out.extend(int(r) for r in rownums if 2 <= int(r) <= current_max)
    return sorted(set(out)), True


def _u133_stage_progress(progress_callback, step: int, total: int, message: str):
    try:
        if progress_callback:
            progress_callback(step, total, message)
    except Exception:
        pass


def _u134_substate_path(output_path: str) -> str:
    base = output_path or 'detail_sub'
    return base + '.substate_u134.json' if base else 'detail_sub.substate_u134.json'


def _u138_issue_signature(status_text: str, error_text: str) -> str:
    status = nz_str(status_text).upper().strip()
    error = ' '.join(nz_str(error_text).upper().split())
    return f"{status}|{error}".strip('|')


def _u138_rows_signature(rows: List[List[Any]]) -> str:
    try:
        payload = []
        for row in list(rows or []):
            norm = []
            for v in list(row or []):
                if v is None:
                    norm.append('')
                elif isinstance(v, float):
                    norm.append(round(v, 6))
                else:
                    norm.append(v)
            payload.append(norm)
        raw = json.dumps(payload, ensure_ascii=False, default=str, separators=(',', ':'))
        return hashlib.md5(raw.encode('utf-8')).hexdigest()
    except Exception:
        return ''


def _u138_normalize_state_entry(raw: Any) -> Dict[str, Any]:
    if isinstance(raw, dict):
        return {
            'status': nz_str(raw.get('status', '')).upper(),
            'error': nz_str(raw.get('error', '')),
            'issue_signature': nz_str(raw.get('issue_signature', '')),
            'inv_rows': int(raw.get('inv_rows', 0) or 0),
            'pl_rows': int(raw.get('pl_rows', 0) or 0),
            'inv_signature': nz_str(raw.get('inv_signature', '')),
            'pl_signature': nz_str(raw.get('pl_signature', '')),
        }
    return {
        'status': nz_str(raw).upper(),
        'error': '',
        'issue_signature': _u138_issue_signature(nz_str(raw), ''),
        'inv_rows': 0,
        'pl_rows': 0,
        'inv_signature': '',
        'pl_signature': '',
    }


def _u138_state_from_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    arr = list((payload or {}).get('arr', []) or [])
    status_text = nz_str((payload or {}).get('status', '') or (arr[21] if len(arr) > 21 else '')).upper()
    error_text = nz_str((payload or {}).get('error', '') or (arr[22] if len(arr) > 22 else ''))
    inv_main, pl_main, inv_other, pl_other = (payload or {}).get('bundle4', ([], [], [], []))
    inv_rows = list(inv_main or []) + list(inv_other or [])
    pl_rows = list(pl_main or []) + list(pl_other or [])
    return {
        'status': status_text,
        'error': error_text,
        'issue_signature': _u138_issue_signature(status_text, error_text),
        'inv_rows': len(inv_rows),
        'pl_rows': len(pl_rows),
        'inv_signature': _u138_rows_signature(inv_rows),
        'pl_signature': _u138_rows_signature(pl_rows),
    }


def _u138_decide_detail_action(old_state: Dict[str, Any], new_state: Dict[str, Any]) -> str:
    old_state = _u138_normalize_state_entry(old_state)
    new_state = _u138_normalize_state_entry(new_state)
    if not nz_str(old_state.get('status', '')):
        return 'REWRITE_DETAIL'
    if (
        old_state.get('status', '') == new_state.get('status', '') and
        old_state.get('issue_signature', '') == new_state.get('issue_signature', '') and
        int(old_state.get('inv_rows', 0) or 0) == int(new_state.get('inv_rows', 0) or 0) and
        int(old_state.get('pl_rows', 0) or 0) == int(new_state.get('pl_rows', 0) or 0) and
        nz_str(old_state.get('inv_signature', '')) == nz_str(new_state.get('inv_signature', '')) and
        nz_str(old_state.get('pl_signature', '')) == nz_str(new_state.get('pl_signature', ''))
    ):
        return 'SKIP_ALL'
    if (
        int(old_state.get('inv_rows', 0) or 0) == int(new_state.get('inv_rows', 0) or 0) and
        int(old_state.get('pl_rows', 0) or 0) == int(new_state.get('pl_rows', 0) or 0) and
        nz_str(old_state.get('inv_signature', '')) == nz_str(new_state.get('inv_signature', '')) and
        nz_str(old_state.get('pl_signature', '')) == nz_str(new_state.get('pl_signature', ''))
    ):
        return 'UPDATE_SUBDETAIL_ONLY'
    return 'REWRITE_DETAIL'


def _u134_load_substate(output_path: str) -> Dict[str, Any]:
    path = _u134_substate_path(output_path)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, dict):
            out: Dict[str, Any] = {}
            for k, v in data.items():
                pk = _u104_canonical_path(k)
                if pk:
                    out[pk] = _u138_normalize_state_entry(v)
            return out
    except Exception:
        pass
    return {}


def _u134_save_substate(output_path: str, state_map: Dict[str, Any]):
    path = _u134_substate_path(output_path)
    tmp = path + '.tmp'
    payload = {}
    for k, v in (state_map or {}).items():
        pk = _u104_canonical_path(k)
        if pk:
            payload[pk] = _u138_normalize_state_entry(v)
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _u137_eval_precheck_with_fallback(scan_map: Dict[str, str], fingerprint_cache: Dict[str, str], substate_cache: Dict[str, Any],
                                      parse_cache: Dict[str, Any], repair_options) -> Dict[str, Any]:
    stats = {
        'scanned': 0,
        'loaded_substate': len(substate_cache or {}),
        'fp_match': 0,
        'fp_miss': 0,
        'fp_match_fingerprint': 0,
        'fp_match_parsecache': 0,
        'status_ok': 0,
        'status_not_ok': 0,
        'status_missing': 0,
        'skip_ok': 0,
        'skip_non_ok_stable': 0,
        'skip_all': 0,
        'candidate': 0,
    }
    entries = (parse_cache or {}).get('entries', {}) if isinstance(parse_cache, dict) else {}
    candidates: List[str] = []
    for raw_key, actual_path in sorted((scan_map or {}).items()):
        pathkey = _u104_canonical_path(raw_key) or _u104_canonical_path(actual_path)
        stats['scanned'] += 1
        old_state = _u138_normalize_state_entry((substate_cache or {}).get(pathkey, {}))
        old_status = nz_str(old_state.get('status', '')).upper()
        if not old_status:
            stats['status_missing'] += 1
        elif old_status == 'OK':
            stats['status_ok'] += 1
        else:
            stats['status_not_ok'] += 1

        fp_same = False
        old_fp = nz_str((fingerprint_cache or {}).get(pathkey, ''))
        cur_fp = _compute_file_fingerprint(actual_path)
        if old_fp and cur_fp and old_fp == cur_fp:
            fp_same = True
            stats['fp_match_fingerprint'] += 1
        else:
            entry = entries.get(pathkey, {}) if isinstance(entries, dict) else {}
            old_parse_fp = nz_str((entry or {}).get('parse_fp', ''))
            # V22: chi hash parse_fp khi da co old_parse_fp de so sanh.
            # Neu cache rong / bootstrap lan dau, tinh hash o precheck la vo ich vi file van la candidate.
            # Cat duoc do tre dau run: BENCH_INIT_PRECHECK khong doc/hash 720 file them mot lan.
            if old_parse_fp:
                cur_parse_fp = _u130_build_parse_fp(actual_path, repair_options)
                if cur_parse_fp and old_parse_fp == cur_parse_fp:
                    fp_same = True
                    stats['fp_match_parsecache'] += 1
        if fp_same:
            stats['fp_match'] += 1
        else:
            stats['fp_miss'] += 1

        if fp_same and old_status:
            stats['skip_all'] += 1
            if old_status == 'OK':
                stats['skip_ok'] += 1
            else:
                stats['skip_non_ok_stable'] += 1
            continue
        candidates.append(actual_path)
    stats['candidate'] = len(candidates)
    stats['candidates'] = candidates
    return stats


def _u137_refresh_fingerprint_cache_from_scan(scan_map: Dict[str, str], fingerprint_cache: Dict[str, str]):
    for raw_key, actual_path in (scan_map or {}).items():
        pathkey = _u104_canonical_path(raw_key) or _u104_canonical_path(actual_path)
        cur_fp = _compute_file_fingerprint(actual_path)
        if pathkey and cur_fp:
            fingerprint_cache[pathkey] = cur_fp


def _u135_build_final_substate(old_state_map: Dict[str, Any], parsed_payloads: Dict[str, Any], action_map: Dict[str, str]) -> Dict[str, Any]:
    final_state: Dict[str, Any] = {}
    for pk, meta in (old_state_map or {}).items():
        pathkey = _u104_canonical_path(pk)
        if pathkey:
            final_state[pathkey] = _u138_normalize_state_entry(meta)
    for pk, payload in (parsed_payloads or {}).items():
        pathkey = _u104_canonical_path(pk)
        if not pathkey:
            continue
        action = nz_str((action_map or {}).get(pathkey, 'REWRITE_DETAIL')).upper()
        if action == 'SKIP_ALL':
            continue
        final_state[pathkey] = _u138_normalize_state_entry((payload or {}).get('state_entry', _u138_state_from_payload(payload)))
    return final_state


def _u139_issue_snapshot_path(output_path: str) -> str:
    return output_path + '.issue_snapshot_u139.json'


def _u139_build_issue_snapshot(state_map: Dict[str, Any]) -> List[Dict[str, Any]]:
    issues: List[Dict[str, Any]] = []
    for k, v in (state_map or {}).items():
        pathkey = _u104_canonical_path(k)
        if not pathkey:
            continue
        meta = _u138_normalize_state_entry(v)
        status = nz_str(meta.get('status', '')).upper()
        if status not in ('WARNING', 'ERROR'):
            continue
        detail = nz_str(meta.get('error', '')) or nz_str(meta.get('issue_signature', '')) or status
        issues.append({
            'file_path': pathkey,
            'file_name': get_file_name_from_path(pathkey),
            'status': status,
            'detail': detail,
        })
    issues.sort(key=lambda x: (0 if x.get('status') == 'ERROR' else 1, normalize_simple_text(x.get('file_name', '')), normalize_simple_text(x.get('detail', ''))))
    return issues


def _u139_save_issue_snapshot(output_path: str, items: List[Dict[str, Any]]):
    path = _u139_issue_snapshot_path(output_path)
    tmp = path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(list(items or []), f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _u139_load_issue_snapshot(output_path: str) -> List[Dict[str, Any]]:
    path = _u139_issue_snapshot_path(output_path)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        out = []
        for rec in list(data or []):
            if not isinstance(rec, dict):
                continue
            file_path = _u104_canonical_path(rec.get('file_path', ''))
            status = nz_str(rec.get('status', '')).upper()
            detail = nz_str(rec.get('detail', ''))
            if not file_path or status not in ('WARNING', 'ERROR'):
                continue
            out.append({
                'file_path': file_path,
                'file_name': nz_str(rec.get('file_name', '')) or get_file_name_from_path(file_path),
                'status': status,
                'detail': detail,
            })
        return out
    except Exception:
        return []


def _u139_issue_items_from_snapshot(output_path: str) -> List[Tuple[str, str, str]]:
    out = []
    for rec in _u139_load_issue_snapshot(output_path):
        out.append((nz_str(rec.get('file_path', '')), nz_str(rec.get('status', '')), nz_str(rec.get('detail', ''))))
    return out


def load_issue_snapshot_items(output_path: str) -> List[Tuple[str, str, str]]:
    return _u139_issue_items_from_snapshot(output_path)


def _u132_make_error_arr(file_path: str, error_text: str) -> List[Any]:
    arr = [''] * len(SUB_DETAIL_HEADERS)
    try:
        arr[0] = get_file_name_from_path(file_path)
    except Exception:
        arr[0] = os.path.basename(file_path)
    arr[21] = 'ERROR'
    arr[22] = nz_str(error_text)[:2000]
    arr[23] = _u104_canonical_path(file_path)
    return arr


def _u132_safe_parse_payload(file_path: str) -> Tuple[List[Any], Any, str]:
    try:
        arr = process_one_sub_file(file_path)
        bundle4 = _v72_build_detail_bundle(file_path)
        return arr, bundle4, ''
    except Exception as exc:
        return _u132_make_error_arr(file_path, str(exc)), ([], [], [], []), nz_str(exc)


class Processor(Processor):
    """
    Runtime-refactored Processor.

    Notes:
    - Keeps existing business helpers and data-extraction logic intact.
    - Consolidates the latest control flow directly into Processor.run.
    - Avoids broad `except Exception: pass` in the active execution path.
    """

    def _safe_log(self, step: str, level: str, detail: str, file_path: str = "", file_name: str = "") -> None:
        try:
            self.logger.log(step, level.lower(), detail, file_name=file_name, file_path=file_path)
        except (AttributeError, TypeError):
            _logger_runtime.warning("Logger bridge failed at step=%s detail=%s", step, detail, exc_info=True)

    def _safe_collect_scanned_files(self, folder_path: str, output_path: str) -> set:
        try:
            return _u87_norm_path_set(_u99_norm_path(p) for p in collect_source_files(folder_path, output_path))
        except FileNotFoundError:
            _logger_runtime.error("Folder not found while scanning: %s", folder_path, exc_info=True)
            return set()
        except OSError:
            _logger_runtime.error("OS error while scanning folder: %s", folder_path, exc_info=True)
            return set()
        except Exception:
            _logger_runtime.error("Unexpected scan failure for folder: %s", folder_path, exc_info=True)
            return set()

    def _safe_current_supported(self, output_path: str) -> set:
        try:
            return _u99_collect_supported_paths_from_output(output_path)
        except FileNotFoundError:
            _logger_runtime.error("Output workbook not found while collecting supported paths: %s", output_path, exc_info=True)
            return set()
        except KeyError:
            _logger_runtime.error("Expected sheet missing while collecting supported paths: %s", output_path, exc_info=True)
            return set()
        except Exception:
            _logger_runtime.error("Unexpected supported-path collection failure: %s", output_path, exc_info=True)
            return set()

    def _safe_bootstrap_needed(self, output_path: str) -> bool:
        try:
            return _u101_need_bootstrap(output_path)
        except FileNotFoundError:
            _logger_runtime.error("Output workbook not found while checking bootstrap: %s", output_path, exc_info=True)
            return True
        except KeyError:
            _logger_runtime.error("Expected sheet/index missing while checking bootstrap: %s", output_path, exc_info=True)
            return True
        except Exception:
            _logger_runtime.error("Unexpected bootstrap check failure: %s", output_path, exc_info=True)
            return True

    def _safe_processed_paths_from_summary(self, summary) -> set:
        try:
            return _u103_extract_processed_paths_from_summary(summary)
        except AttributeError:
            _logger_runtime.error("Summary object does not expose changed_detail_files", exc_info=True)
            return set()
        except Exception:
            _logger_runtime.error("Unexpected processed-path extraction failure", exc_info=True)
            return set()

    def _safe_changed_paths(self, output_path: str, scanned_files: set, current_supported: set, summary, bootstrap_needed: bool) -> set:
        if bootstrap_needed:
            return set(current_supported)

        processed_paths = self._safe_processed_paths_from_summary(summary)
        changed_paths = _u87_norm_path_set(p for p in processed_paths if p in current_supported)

        if changed_paths:
            return changed_paths

        try:
            fallback = _u102_collect_supported_scanned_paths(output_path, scanned_files)
            return _u87_norm_path_set(p for p in fallback if p in current_supported)
        except FileNotFoundError:
            _logger_runtime.error("Output workbook not found while using fallback changed-path logic: %s", output_path, exc_info=True)
            return set()
        except KeyError:
            _logger_runtime.error("Expected sheet/header missing while using fallback changed-path logic: %s", output_path, exc_info=True)
            return set()
        except Exception:
            _logger_runtime.error("Unexpected fallback changed-path failure: %s", output_path, exc_info=True)
            return set()

    def _safe_load_source_cache(self, output_path: str) -> None:
        _U90_SOURCE_CACHE_CTX['output_path'] = output_path
        try:
            _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
        except FileNotFoundError:
            _logger_runtime.warning("Source cache file not found for %s; using empty cache.", output_path)
            _U90_SOURCE_CACHE_CTX['cache'] = _u90_empty_source_cache()
        except (ValueError, EOFError, OSError):
            _logger_runtime.error("Corrupted/unreadable source cache for %s; resetting cache.", output_path, exc_info=True)
            _U90_SOURCE_CACHE_CTX['cache'] = _u90_empty_source_cache()
        except Exception:
            _logger_runtime.error("Unexpected source-cache load failure for %s; resetting cache.", output_path, exc_info=True)
            _U90_SOURCE_CACHE_CTX['cache'] = _u90_empty_source_cache()

    def _safe_save_source_cache(self, output_path: str) -> None:
        try:
            cache = _u90_get_source_cache()
            _u90_save_source_cache(output_path, cache)
        except OSError:
            _logger_runtime.error("OS error while saving source cache: %s", output_path, exc_info=True)
        except Exception:
            _logger_runtime.error("Unexpected source-cache save failure: %s", output_path, exc_info=True)

    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options: Optional[RepairOptions] = None):
        repair_options = repair_options or RepairOptions()
        _u85_reset_bench()
        t_all_0 = time.perf_counter()

        cache_seed = _u90_empty_source_cache()
        self._safe_load_source_cache(output_path)
        cache = _u90_get_source_cache()
        if not isinstance(cache, dict):
            cache = cache_seed
            _U90_SOURCE_CACHE_CTX['cache'] = cache
        cache['stats'] = _u90_empty_source_cache()['stats']

        try:
            self._safe_log('FILE_UPSERT_ENTER_REFACTORED', 'INFO', f'folder={folder_path}')
            self._safe_log('SOURCE_PARSE_CACHE_ENTER_REFACTORED', 'INFO', output_path)

            _U84_WIDTH_CACHE_CTX['output_path'] = output_path
            try:
                _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)
            except (FileNotFoundError, ValueError, OSError):
                _logger_runtime.warning("Width cache unavailable/corrupt for %s; using empty cache.", output_path, exc_info=True)
                _U84_WIDTH_CACHE_CTX['cache'] = {}

            scanned_files = self._safe_collect_scanned_files(folder_path, output_path)

            t0 = time.perf_counter()
            try:
                summary = CORE_RUN_V6(self, folder_path, output_path, progress_callback=progress_callback, repair_options=repair_options)
            except FileNotFoundError:
                _logger_runtime.error("CORE_RUN_V6 missing file during processing for folder=%s", folder_path, exc_info=True)
                summary = {'changed_detail_files': []}
            except KeyError:
                _logger_runtime.error("CORE_RUN_V6 missing expected key/sheet for folder=%s", folder_path, exc_info=True)
                summary = {'changed_detail_files': []}
            except ValueError:
                _logger_runtime.error("CORE_RUN_V6 value error for folder=%s", folder_path, exc_info=True)
                summary = {'changed_detail_files': []}
            except Exception:
                _logger_runtime.error("Unexpected CORE_RUN_V6 failure for folder=%s", folder_path, exc_info=True)
                summary = {'changed_detail_files': []}
            core_seconds = max(0.0, time.perf_counter() - t0)

            current_supported = self._safe_current_supported(output_path)
            bootstrap_needed = self._safe_bootstrap_needed(output_path)
            changed_paths = self._safe_changed_paths(output_path, scanned_files, current_supported, summary, bootstrap_needed)

            t1 = time.perf_counter()
            try:
                upsert_stats = _u99_rebuild_upsert(
                    output_path,
                    current_supported,
                    changed_paths=changed_paths,
                    prep_label='DETAIL_PREP_U79'
                )
            except FileNotFoundError:
                _logger_runtime.error("Upsert failed because workbook/file was missing: %s", output_path, exc_info=True)
                upsert_stats = {'supported_files': len(current_supported), 'changed': len(changed_paths),
                                'inv_append': 0, 'inv_update_same': 0, 'inv_update_resize': 0,
                                'inv_index_repaired': 0, 'pl_append': 0, 'pl_update_same': 0,
                                'pl_update_resize': 0, 'pl_index_repaired': 0,
                                'cache_hit': 0, 'rebuilt_files': 0, 'cache_path': _u87_detail_cache_path(output_path)}
            except KeyError:
                _logger_runtime.error("Upsert failed because expected columns/sheets were missing in %s", output_path, exc_info=True)
                upsert_stats = {'supported_files': len(current_supported), 'changed': len(changed_paths),
                                'inv_append': 0, 'inv_update_same': 0, 'inv_update_resize': 0,
                                'inv_index_repaired': 0, 'pl_append': 0, 'pl_update_same': 0,
                                'pl_update_resize': 0, 'pl_index_repaired': 0,
                                'cache_hit': 0, 'rebuilt_files': 0, 'cache_path': _u87_detail_cache_path(output_path)}
            except Exception:
                _logger_runtime.error("Unexpected upsert failure for %s", output_path, exc_info=True)
                upsert_stats = {'supported_files': len(current_supported), 'changed': len(changed_paths),
                                'inv_append': 0, 'inv_update_same': 0, 'inv_update_resize': 0,
                                'inv_index_repaired': 0, 'pl_append': 0, 'pl_update_same': 0,
                                'pl_update_resize': 0, 'pl_index_repaired': 0,
                                'cache_hit': 0, 'rebuilt_files': 0, 'cache_path': _u87_detail_cache_path(output_path)}
            rebuild_seconds = max(0.0, time.perf_counter() - t1)

            total_seconds = max(0.0, time.perf_counter() - t_all_0)

            self._safe_save_source_cache(output_path)
            stats = {}
            try:
                cache = _u90_get_source_cache()
                stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
            except Exception:
                _logger_runtime.error("Unable to read source-cache stats after save: %s", output_path, exc_info=True)
                stats = {}

            bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
            bench_rows.append(make_log_row(
                output_path, 'SMART_REBUILD_STATS', 'INFO',
                f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | "
                f"changed={upsert_stats.get('changed',0)} | bootstrap={1 if bootstrap_needed else 0} | "
                f"inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | "
                f"inv_update_resize={upsert_stats.get('inv_update_resize',0)} | inv_index_repaired={upsert_stats.get('inv_index_repaired',0)} | "
                f"pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | "
                f"pl_update_resize={upsert_stats.get('pl_update_resize',0)} | pl_index_repaired={upsert_stats.get('pl_index_repaired',0)} | "
                f"cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | "
                f"cache={upsert_stats.get('cache_path','')}"
            ))
            bench_rows.append(make_log_row(
                output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | "
                f"bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | "
                f"detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | "
                f"cache={_u90_source_cache_path(output_path)}"
            ))
            try:
                _u85_append_bench_log_rows(output_path, bench_rows)
            except Exception:
                _logger_runtime.error("Unable to append benchmark rows to workbook log: %s", output_path, exc_info=True)

            return summary
        finally:
            _U84_WIDTH_CACHE_CTX['output_path'] = ''
            _U84_WIDTH_CACHE_CTX['cache'] = {}
            _U85_BENCH_CTX['enabled'] = False
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None



# =========================================================
# REFACTORED MODULAR PROCESSOR
# Muc tieu:
# - Tach Processor thanh cac method ro lop:
#     scan
#     core_run
#     resolve_changed_paths
#     upsert_inv_pl
#     persist_caches
# - Them file fingerprint cache de skip file khong doi
# - Giu logic business moi nhat dang chay on dinh
# - Khong monkey-patch runtime
# Luu y:
# - Cac helper legacy van duoc giu lai vi business parsing con phu thuoc
# - Runtime se di qua class Processor ben duoi
# =========================================================

import logging
import traceback
from dataclasses import dataclass
from typing import Optional, Set, Dict, Any


_RUNTIME_LOGGER = logging.getLogger("detaisublog.runtime")
if not _RUNTIME_LOGGER.handlers:
    _RUNTIME_LOGGER.setLevel(logging.INFO)
    _h = logging.FileHandler("detaisublog_runtime.log", encoding="utf-8")
    _h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
    _RUNTIME_LOGGER.addHandler(_h)


@dataclass
class FingerprintEntry:
    path: str
    fingerprint: str


def _rt_log_exception(msg: str, exc: Exception):
    try:
        _RUNTIME_LOGGER.error("%s | %s\n%s", msg, str(exc), traceback.format_exc())
    except Exception:
        pass


def _safe_norm_path_runtime(p: str) -> str:
    try:
        if not p:
            return ""
        return os.path.abspath(str(p)).replace("\\", "/").replace("//", "/").upper()
    except Exception:
        return _u99_norm_path(p)


def _fingerprint_cache_path(output_path: str) -> str:
    try:
        return str(Path(output_path).with_suffix(Path(output_path).suffix + ".fingerprintcache.json"))
    except Exception:
        return str(output_path) + ".fingerprintcache.json"


def _load_fingerprint_cache(output_path: str) -> Dict[str, str]:
    fp = _fingerprint_cache_path(output_path)
    try:
        if os.path.exists(fp):
            with open(fp, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    out = {}
                    for k, v in data.items():
                        nk = _safe_norm_path_runtime(k)
                        if nk and isinstance(v, str):
                            out[nk] = v
                    return out
    except (OSError, json.JSONDecodeError, ValueError) as exc:
        _rt_log_exception(f"load_fingerprint_cache_failed: {fp}", exc)
    return {}


def _save_fingerprint_cache(output_path: str, cache_map: Dict[str, str]):
    fp = _fingerprint_cache_path(output_path)
    try:
        data = {k: v for k, v in sorted((cache_map or {}).items()) if k and isinstance(v, str)}
        with open(fp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except OSError as exc:
        _rt_log_exception(f"save_fingerprint_cache_failed: {fp}", exc)


def _compute_file_fingerprint(path: str) -> str:
    p = _safe_norm_path_runtime(path)
    if not p:
        return ""
    try:
        st = os.stat(path)
        raw = f"{p}|{int(st.st_size)}|{int(st.st_mtime_ns)}"
        return hashlib.md5(raw.encode("utf-8", errors="ignore")).hexdigest()
    except FileNotFoundError as exc:
        _rt_log_exception(f"fingerprint_file_missing: {path}", exc)
        return ""
    except OSError as exc:
        _rt_log_exception(f"fingerprint_oserror: {path}", exc)
        return ""


class Processor(Processor):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._runtime_logger = _RUNTIME_LOGGER
        self._fingerprint_cache: Dict[str, str] = {}
        self._current_scanned_files: Set[str] = set()
        self._processed_paths_from_core: Set[str] = set()

    # -------------------------
    # scan
    # -------------------------
    def scan(self, folder_path: str, output_path: str) -> Set[str]:
        try:
            files = collect_source_files(folder_path, output_path)
            normed = _u87_norm_path_set(_safe_norm_path_runtime(p) for p in files)
            self._current_scanned_files = set(normed)
            return set(normed)
        except (FileNotFoundError, OSError, ValueError) as exc:
            _rt_log_exception(f"scan_failed: {folder_path}", exc)
            self._current_scanned_files = set()
            return set()
        except Exception as exc:
            _rt_log_exception(f"scan_unexpected_failed: {folder_path}", exc)
            self._current_scanned_files = set()
            return set()

    # -------------------------
    # core_run
    # -------------------------
    def core_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        try:
            summary = CORE_RUN_V6(
                self,
                folder_path,
                output_path,
                progress_callback=progress_callback,
                repair_options=repair_options
            )
            self._processed_paths_from_core = _u103_extract_processed_paths_from_summary(summary)
            return summary or {}
        except (FileNotFoundError, ValueError, KeyError) as exc:
            _rt_log_exception(f"core_run_failed: {folder_path}", exc)
            return {}
        except Exception as exc:
            _rt_log_exception(f"core_run_unexpected_failed: {folder_path}", exc)
            return {}

    # -------------------------
    # resolve_changed_paths
    # -------------------------
    def resolve_changed_paths(self, output_path: str, summary: Dict[str, Any], scanned_files: Set[str]) -> Set[str]:
        current_supported = set()
        try:
            current_supported = _u99_collect_supported_paths_from_output(output_path)
        except Exception as exc:
            _rt_log_exception("collect_supported_failed", exc)
            current_supported = set()

        try:
            bootstrap_needed = _u101_need_bootstrap(output_path)
        except Exception as exc:
            _rt_log_exception("bootstrap_check_failed", exc)
            bootstrap_needed = True

        if bootstrap_needed:
            return set(current_supported)

        processed_paths = set(self._processed_paths_from_core or _u103_extract_processed_paths_from_summary(summary))
        changed_by_core = _u87_norm_path_set(p for p in processed_paths if p in current_supported)

        if changed_by_core:
            return set(changed_by_core)

        try:
            return set(_u102_collect_supported_scanned_paths(output_path, scanned_files))
        except Exception as exc:
            _rt_log_exception("resolve_changed_paths_fallback_failed", exc)
            return set()

    # -------------------------
    # fingerprint cache helpers
    # -------------------------
    def load_fingerprint_cache(self, output_path: str):
        self._fingerprint_cache = _load_fingerprint_cache(output_path)

    def persist_caches(self, output_path: str):
        try:
            _save_fingerprint_cache(output_path, self._fingerprint_cache)
        except Exception as exc:
            _rt_log_exception("persist_fingerprint_cache_failed", exc)

        try:
            cache = _u90_get_source_cache()
            _u90_save_source_cache(output_path, cache)
        except Exception as exc:
            _rt_log_exception("persist_source_cache_failed", exc)

    def _filter_changed_by_fingerprint(self, changed_paths: Set[str]) -> Set[str]:
        final_changed = set()
        for p in changed_paths or set():
            fp = _compute_file_fingerprint(p)
            if not fp:
                final_changed.add(p)
                continue
            old_fp = self._fingerprint_cache.get(p, "")
            if old_fp != fp:
                final_changed.add(p)
                self._fingerprint_cache[p] = fp
        return final_changed

    # -------------------------
    # upsert_inv_pl
    # -------------------------
    def upsert_inv_pl(self, output_path: str, summary: Dict[str, Any], changed_paths: Set[str]):
        current_supported = set()
        try:
            current_supported = _u99_collect_supported_paths_from_output(output_path)
        except Exception as exc:
            _rt_log_exception("upsert_collect_supported_failed", exc)
            current_supported = set()

        changed_paths = self._filter_changed_by_fingerprint(set(changed_paths or set()))

        try:
            return _u99_rebuild_upsert(
                output_path,
                current_supported,
                changed_paths=changed_paths,
                prep_label='DETAIL_PREP_U79'
            )
        except (ValueError, KeyError, FileNotFoundError) as exc:
            _rt_log_exception("upsert_inv_pl_failed", exc)
            return {
                'supported_files': len(current_supported),
                'changed': 0,
                'inv_append': 0,
                'inv_update_same': 0,
                'inv_update_resize': 0,
                'inv_index_repaired': 0,
                'pl_append': 0,
                'pl_update_same': 0,
                'pl_update_resize': 0,
                'pl_index_repaired': 0,
                'cache_hit': 0,
                'rebuilt_files': 0,
                'errors': 1,
                'cache_path': _u87_detail_cache_path(output_path),
            }
        except Exception as exc:
            _rt_log_exception("upsert_inv_pl_unexpected_failed", exc)
            return {
                'supported_files': len(current_supported),
                'changed': 0,
                'inv_append': 0,
                'inv_update_same': 0,
                'inv_update_resize': 0,
                'inv_index_repaired': 0,
                'pl_append': 0,
                'pl_update_same': 0,
                'pl_update_resize': 0,
                'pl_index_repaired': 0,
                'cache_hit': 0,
                'rebuilt_files': 0,
                'errors': 1,
                'cache_path': _u87_detail_cache_path(output_path),
            }

    # -------------------------
    # run
    # -------------------------
    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        _u85_reset_bench()
        t_all_0 = time.perf_counter()
        _U90_SOURCE_CACHE_CTX['output_path'] = output_path
        _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
        cache = _u90_get_source_cache()
        cache['stats'] = _u90_empty_source_cache()['stats']

        try:
            self._runtime_logger.info("RUN_START folder=%s output=%s", folder_path, output_path)
            self.logger.log('FILE_UPSERT_ENTER_REFACTORED', 'info')
            self.logger.log('SOURCE_PARSE_CACHE_ENTER_REFACTORED', 'info')

            _U84_WIDTH_CACHE_CTX['output_path'] = output_path
            _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

            self.load_fingerprint_cache(output_path)

            scanned_files = self.scan(folder_path, output_path)

            t0 = time.perf_counter()
            summary = self.core_run(
                folder_path,
                output_path,
                progress_callback=progress_callback,
                repair_options=repair_options
            )
            core_seconds = max(0.0, time.perf_counter() - t0)

            changed_paths = self.resolve_changed_paths(output_path, summary, scanned_files)

            t1 = time.perf_counter()
            upsert_stats = self.upsert_inv_pl(output_path, summary, changed_paths)
            rebuild_seconds = max(0.0, time.perf_counter() - t1)

            self.persist_caches(output_path)

            total_seconds = max(0.0, time.perf_counter() - t_all_0)

            try:
                stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
            except Exception:
                stats = {}

            bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
            bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                            f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | changed={upsert_stats.get('changed',0)} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | inv_index_repaired={upsert_stats.get('inv_index_repaired',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | pl_index_repaired={upsert_stats.get('pl_index_repaired',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
            bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                            f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
            _u85_append_bench_log_rows(output_path, bench_rows)
            return summary

        except Exception as exc:
            _rt_log_exception("processor_run_fatal", exc)
            raise
        finally:
            _U84_WIDTH_CACHE_CTX['output_path'] = ''
            _U84_WIDTH_CACHE_CTX['cache'] = {}
            _U85_BENCH_CTX['enabled'] = False
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None



# =========================================================
# U100 TRUE INCREMENTAL PRODUCTION
# Nguyen tac:
# - Fingerprint la nguon chan ly de quyet dinh file nao thuc su thay doi
# - Khong dung changed_paths tu summary/core lam nguon quyet dinh chinh
# - Chi xu ly file thuoc folder dang scan va hien dang co Status in (OK, WARNING)
# - Chi commit fingerprint sau khi upsert thanh cong
# - Log ro so file skip theo fingerprint de kiem chung incremental
# =========================================================

class Processor(Processor):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._pending_fingerprint_updates: Dict[str, str] = {}
        self._last_fingerprint_skipped = 0
        self._last_fingerprint_changed = 0
        self._last_supported_scanned = 0

    # -------------------------
    # scan
    # -------------------------
    def scan(self, folder_path: str, output_path: str) -> Set[str]:
        try:
            files = collect_source_files(folder_path, output_path)
            normed = _u87_norm_path_set(_safe_norm_path_runtime(p) for p in files)
            self._current_scanned_files = set(normed)
            return set(normed)
        except (FileNotFoundError, OSError, ValueError) as exc:
            _rt_log_exception(f"u100_scan_failed: {folder_path}", exc)
            self._current_scanned_files = set()
            return set()
        except Exception as exc:
            _rt_log_exception(f"u100_scan_unexpected_failed: {folder_path}", exc)
            self._current_scanned_files = set()
            return set()

    # -------------------------
    # core_run
    # -------------------------
    def core_run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        try:
            summary = CORE_RUN_V6(
                self,
                folder_path,
                output_path,
                progress_callback=progress_callback,
                repair_options=repair_options
            )
            self._processed_paths_from_core = _u103_extract_processed_paths_from_summary(summary)
            return summary or {}
        except (FileNotFoundError, ValueError, KeyError) as exc:
            _rt_log_exception(f"u100_core_run_failed: {folder_path}", exc)
            return {}
        except Exception as exc:
            _rt_log_exception(f"u100_core_run_unexpected_failed: {folder_path}", exc)
            return {}

    # -------------------------
    # fingerprint cache helpers
    # -------------------------
    def load_fingerprint_cache(self, output_path: str):
        self._fingerprint_cache = _load_fingerprint_cache(output_path)
        self._pending_fingerprint_updates = {}
        self._last_fingerprint_skipped = 0
        self._last_fingerprint_changed = 0
        self._last_supported_scanned = 0

    def _collect_supported_scanned_paths(self, output_path: str, scanned_files: Set[str]) -> Set[str]:
        scanned_files = _u87_norm_path_set(_safe_norm_path_runtime(p) for p in (scanned_files or set()))
        current_supported = set()
        try:
            current_supported = _u99_collect_supported_paths_from_output(output_path)
        except Exception as exc:
            _rt_log_exception("u100_collect_supported_failed", exc)
            current_supported = set()
        supported_scanned = _u87_norm_path_set(p for p in current_supported if p in scanned_files)
        self._last_supported_scanned = len(supported_scanned)
        return set(supported_scanned)

    def _filter_changed_by_fingerprint(self, candidate_paths: Set[str]) -> Set[str]:
        changed = set()
        skipped = 0
        pending = {}

        for p in sorted(candidate_paths or set()):
            fp = _compute_file_fingerprint(p)
            if not fp:
                # khong fingerprint duoc -> cu xu ly de an toan
                changed.add(p)
                continue

            old_fp = self._fingerprint_cache.get(p, "")
            if old_fp == fp:
                skipped += 1
                continue

            changed.add(p)
            pending[p] = fp

        self._pending_fingerprint_updates = pending
        self._last_fingerprint_skipped = skipped
        self._last_fingerprint_changed = len(changed)
        return changed

    def persist_caches(self, output_path: str, success_changed_paths: Optional[Set[str]] = None):
        try:
            success_changed_paths = set(success_changed_paths or set())
            for p in success_changed_paths:
                new_fp = self._pending_fingerprint_updates.get(p)
                if new_fp:
                    self._fingerprint_cache[p] = new_fp
            _save_fingerprint_cache(output_path, self._fingerprint_cache)
        except Exception as exc:
            _rt_log_exception("u100_persist_fingerprint_cache_failed", exc)

        try:
            cache = _u90_get_source_cache()
            _u90_save_source_cache(output_path, cache)
        except Exception as exc:
            _rt_log_exception("u100_persist_source_cache_failed", exc)

    # -------------------------
    # resolve_changed_paths
    # -------------------------
    def resolve_changed_paths(self, output_path: str, summary: Dict[str, Any], scanned_files: Set[str]) -> Set[str]:
        try:
            bootstrap_needed = _u101_need_bootstrap(output_path)
        except Exception as exc:
            _rt_log_exception("u100_bootstrap_check_failed", exc)
            bootstrap_needed = True

        supported_scanned = self._collect_supported_scanned_paths(output_path, scanned_files)

        if bootstrap_needed:
            # bootstrap lan dau phai seed toan bo file co trong folder dang scan va co status OK/WARNING
            return set(supported_scanned)

        # TRUE INCREMENTAL:
        # - Khong lay changed_paths tu summary/core lam nguon chan ly nua
        # - Summary chi dung de giup debug, khong override fingerprint
        return set(self._filter_changed_by_fingerprint(supported_scanned))

    # -------------------------
    # upsert_inv_pl
    # -------------------------
    def upsert_inv_pl(self, output_path: str, changed_paths: Set[str]):
        current_supported = set()
        try:
            current_supported = _u99_collect_supported_paths_from_output(output_path)
        except Exception as exc:
            _rt_log_exception("u100_upsert_collect_supported_failed", exc)
            current_supported = set()

        try:
            return _u99_rebuild_upsert(
                output_path,
                current_supported,
                changed_paths=set(changed_paths or set()),
                prep_label='DETAIL_PREP_U79'
            )
        except (ValueError, KeyError, FileNotFoundError) as exc:
            _rt_log_exception("u100_upsert_inv_pl_failed", exc)
            return {
                'supported_files': len(current_supported),
                'changed': 0,
                'inv_append': 0,
                'inv_update_same': 0,
                'inv_update_resize': 0,
                'inv_index_repaired': 0,
                'pl_append': 0,
                'pl_update_same': 0,
                'pl_update_resize': 0,
                'pl_index_repaired': 0,
                'cache_hit': 0,
                'rebuilt_files': 0,
                'errors': 1,
                'cache_path': _u87_detail_cache_path(output_path),
            }
        except Exception as exc:
            _rt_log_exception("u100_upsert_inv_pl_unexpected_failed", exc)
            return {
                'supported_files': len(current_supported),
                'changed': 0,
                'inv_append': 0,
                'inv_update_same': 0,
                'inv_update_resize': 0,
                'inv_index_repaired': 0,
                'pl_append': 0,
                'pl_update_same': 0,
                'pl_update_resize': 0,
                'pl_index_repaired': 0,
                'cache_hit': 0,
                'rebuilt_files': 0,
                'errors': 1,
                'cache_path': _u87_detail_cache_path(output_path),
            }

    # -------------------------
    # run
    # -------------------------
    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        _u85_reset_bench()
        t_all_0 = time.perf_counter()
        _U90_SOURCE_CACHE_CTX['output_path'] = output_path
        _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
        cache = _u90_get_source_cache()
        cache['stats'] = _u90_empty_source_cache()['stats']

        success_changed_paths = set()

        try:
            self._runtime_logger.info("U100_RUN_START folder=%s output=%s", folder_path, output_path)
            self.logger.log('FILE_UPSERT_ENTER_U100_FINAL', 'info')
            self.logger.log('SOURCE_PARSE_CACHE_ENTER_U100_FINAL', 'info')

            _U84_WIDTH_CACHE_CTX['output_path'] = output_path
            _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

            self.load_fingerprint_cache(output_path)

            scanned_files = self.scan(folder_path, output_path)

            t0 = time.perf_counter()
            summary = self.core_run(
                folder_path,
                output_path,
                progress_callback=progress_callback,
                repair_options=repair_options
            )
            core_seconds = max(0.0, time.perf_counter() - t0)

            changed_paths = self.resolve_changed_paths(output_path, summary, scanned_files)

            t1 = time.perf_counter()
            upsert_stats = self.upsert_inv_pl(output_path, changed_paths)
            rebuild_seconds = max(0.0, time.perf_counter() - t1)

            if upsert_stats.get('errors', 0) == 0:
                success_changed_paths = set(changed_paths)

            self.persist_caches(output_path, success_changed_paths=success_changed_paths)

            total_seconds = max(0.0, time.perf_counter() - t_all_0)

            try:
                stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
            except Exception:
                stats = {}

            bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
            bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                            f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | supported_scanned={self._last_supported_scanned} | changed={upsert_stats.get('changed',0)} | fp_changed={self._last_fingerprint_changed} | fp_skipped={self._last_fingerprint_skipped} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | inv_index_repaired={upsert_stats.get('inv_index_repaired',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | pl_index_repaired={upsert_stats.get('pl_index_repaired',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
            bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                            f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
            _u85_append_bench_log_rows(output_path, bench_rows)
            return summary

        except Exception as exc:
            _rt_log_exception("u100_processor_run_fatal", exc)
            raise
        finally:
            _U84_WIDTH_CACHE_CTX['output_path'] = ''
            _U84_WIDTH_CACHE_CTX['cache'] = {}
            _U85_BENCH_CTX['enabled'] = False
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None



# =========================================================
# U101 FIX PATH NORMALIZATION - PRODUCTION
# Muc tieu:
# - Dam bao supported_scanned > 0 neu folder scan va SUB_DETAIL cung chua file do
# - Dung mot ham normalize duong dan duy nhat cho scan / summary / output / cache
# - Giu logic incremental cua U100
# =========================================================

def _u101_norm_path_strict(p) -> str:
    try:
        if p is None:
            return ""
        s = str(p).strip().strip('"').strip("'")
        if not s:
            return ""
        s = os.path.abspath(s)
        s = os.path.normpath(s)
        s = os.path.normcase(s)
        s = s.replace("\\", "/")
        while '//' in s:
            s = s.replace('//', '/')
        return s.upper()
    except Exception as exc:
        _rt_log_exception(f"u101_norm_path_failed: {p}", exc)
        return ""


# override runtime normalizer for all later calls in this file
_safe_norm_path_runtime = _u101_norm_path_strict


class Processor(Processor):

    def scan(self, folder_path: str, output_path: str) -> Set[str]:
        try:
            files = collect_source_files(folder_path, output_path)
            normed = set()
            for p in files:
                np = _u101_norm_path_strict(p)
                if np:
                    normed.add(np)
            self._current_scanned_files = normed
            return set(normed)
        except (FileNotFoundError, OSError, ValueError) as exc:
            _rt_log_exception(f"u101_scan_failed: {folder_path}", exc)
            self._current_scanned_files = set()
            return set()
        except Exception as exc:
            _rt_log_exception(f"u101_scan_unexpected_failed: {folder_path}", exc)
            self._current_scanned_files = set()
            return set()

    def _collect_supported_scanned_paths(self, output_path: str, scanned_files: Set[str]) -> Set[str]:
        scanned_norm = set()
        for p in scanned_files or set():
            np = _u101_norm_path_strict(p)
            if np:
                scanned_norm.add(np)

        current_supported = set()
        try:
            raw_supported = _u99_collect_supported_paths_from_output(output_path)
            for p in raw_supported:
                np = _u101_norm_path_strict(p)
                if np:
                    current_supported.add(np)
        except Exception as exc:
            _rt_log_exception("u101_collect_supported_failed", exc)
            current_supported = set()

        supported_scanned = {p for p in current_supported if p in scanned_norm}
        self._last_supported_scanned = len(supported_scanned)

        # debug samples when mismatch happens
        if scanned_norm and not supported_scanned:
            try:
                sample_supported = list(sorted(current_supported))[:3]
                sample_scanned = list(sorted(scanned_norm))[:3]
                self.logger.log('DEBUG_PATH_MATCH_U101', 'INFO',
                                f"supported_scanned=0 | sample_supported={sample_supported} | sample_scanned={sample_scanned}")
            except Exception:
                pass

        return supported_scanned

    def _filter_changed_by_fingerprint(self, candidate_paths: Set[str]) -> Set[str]:
        changed = set()
        skipped = 0
        pending = {}

        # normalize cache keys one more time defensively
        if self._fingerprint_cache:
            self._fingerprint_cache = {
                _u101_norm_path_strict(k): v
                for k, v in self._fingerprint_cache.items()
                if _u101_norm_path_strict(k) and isinstance(v, str)
            }

        for p in sorted(candidate_paths or set()):
            np = _u101_norm_path_strict(p)
            if not np:
                continue
            fp = _compute_file_fingerprint(np)
            if not fp:
                changed.add(np)
                continue

            old_fp = self._fingerprint_cache.get(np, "")
            if old_fp == fp:
                skipped += 1
                continue

            changed.add(np)
            pending[np] = fp

        self._pending_fingerprint_updates = pending
        self._last_fingerprint_skipped = skipped
        self._last_fingerprint_changed = len(changed)
        return changed

    def load_fingerprint_cache(self, output_path: str):
        raw = _load_fingerprint_cache(output_path)
        self._fingerprint_cache = {}
        for k, v in raw.items():
            nk = _u101_norm_path_strict(k)
            if nk and isinstance(v, str):
                self._fingerprint_cache[nk] = v
        self._pending_fingerprint_updates = {}
        self._last_fingerprint_skipped = 0
        self._last_fingerprint_changed = 0
        self._last_supported_scanned = 0

    def persist_caches(self, output_path: str, success_changed_paths: Optional[Set[str]] = None):
        try:
            success_changed_paths = set(success_changed_paths or set())
            for p in success_changed_paths:
                np = _u101_norm_path_strict(p)
                new_fp = self._pending_fingerprint_updates.get(np)
                if new_fp:
                    self._fingerprint_cache[np] = new_fp
            _save_fingerprint_cache(output_path, self._fingerprint_cache)
        except Exception as exc:
            _rt_log_exception("u101_persist_fingerprint_cache_failed", exc)

        try:
            cache = _u90_get_source_cache()
            _u90_save_source_cache(output_path, cache)
        except Exception as exc:
            _rt_log_exception("u101_persist_source_cache_failed", exc)

    def upsert_inv_pl(self, output_path: str, changed_paths: Set[str]):
        current_supported = set()
        try:
            raw_supported = _u99_collect_supported_paths_from_output(output_path)
            for p in raw_supported:
                np = _u101_norm_path_strict(p)
                if np:
                    current_supported.add(np)
        except Exception as exc:
            _rt_log_exception("u101_upsert_collect_supported_failed", exc)
            current_supported = set()

        try:
            return _u99_rebuild_upsert(
                output_path,
                current_supported,
                changed_paths={_u101_norm_path_strict(p) for p in (changed_paths or set()) if _u101_norm_path_strict(p)},
                prep_label='DETAIL_PREP_U79'
            )
        except (ValueError, KeyError, FileNotFoundError) as exc:
            _rt_log_exception("u101_upsert_inv_pl_failed", exc)
            return {
                'supported_files': len(current_supported),
                'changed': 0,
                'inv_append': 0,
                'inv_update_same': 0,
                'inv_update_resize': 0,
                'inv_index_repaired': 0,
                'pl_append': 0,
                'pl_update_same': 0,
                'pl_update_resize': 0,
                'pl_index_repaired': 0,
                'cache_hit': 0,
                'rebuilt_files': 0,
                'errors': 1,
                'cache_path': _u87_detail_cache_path(output_path),
            }
        except Exception as exc:
            _rt_log_exception("u101_upsert_inv_pl_unexpected_failed", exc)
            return {
                'supported_files': len(current_supported),
                'changed': 0,
                'inv_append': 0,
                'inv_update_same': 0,
                'inv_update_resize': 0,
                'inv_index_repaired': 0,
                'pl_append': 0,
                'pl_update_same': 0,
                'pl_update_resize': 0,
                'pl_index_repaired': 0,
                'cache_hit': 0,
                'rebuilt_files': 0,
                'errors': 1,
                'cache_path': _u87_detail_cache_path(output_path),
            }

    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        _u85_reset_bench()
        t_all_0 = time.perf_counter()
        _U90_SOURCE_CACHE_CTX['output_path'] = output_path
        _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
        cache = _u90_get_source_cache()
        cache['stats'] = _u90_empty_source_cache()['stats']

        success_changed_paths = set()

        try:
            self._runtime_logger.info("U101_RUN_START folder=%s output=%s", folder_path, output_path)
            self.logger.log('FILE_UPSERT_ENTER_U101_FINAL', 'info')
            self.logger.log('SOURCE_PARSE_CACHE_ENTER_U101_FINAL', 'info')

            _U84_WIDTH_CACHE_CTX['output_path'] = output_path
            _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

            self.load_fingerprint_cache(output_path)

            scanned_files = self.scan(folder_path, output_path)

            t0 = time.perf_counter()
            summary = self.core_run(
                folder_path,
                output_path,
                progress_callback=progress_callback,
                repair_options=repair_options
            )
            core_seconds = max(0.0, time.perf_counter() - t0)

            changed_paths = self.resolve_changed_paths(output_path, summary, scanned_files)

            t1 = time.perf_counter()
            upsert_stats = self.upsert_inv_pl(output_path, changed_paths)
            rebuild_seconds = max(0.0, time.perf_counter() - t1)

            if upsert_stats.get('errors', 0) == 0:
                success_changed_paths = set(changed_paths)

            self.persist_caches(output_path, success_changed_paths=success_changed_paths)

            total_seconds = max(0.0, time.perf_counter() - t_all_0)

            try:
                stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
            except Exception:
                stats = {}

            bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
            bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                            f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | supported_scanned={self._last_supported_scanned} | changed={upsert_stats.get('changed',0)} | fp_changed={self._last_fingerprint_changed} | fp_skipped={self._last_fingerprint_skipped} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | inv_index_repaired={upsert_stats.get('inv_index_repaired',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | pl_index_repaired={upsert_stats.get('pl_index_repaired',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
            bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                            f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
            _u85_append_bench_log_rows(output_path, bench_rows)
            return summary

        except Exception as exc:
            _rt_log_exception("u101_processor_run_fatal", exc)
            raise
        finally:
            _U84_WIDTH_CACHE_CTX['output_path'] = ''
            _U84_WIDTH_CACHE_CTX['cache'] = {}
            _U85_BENCH_CTX['enabled'] = False
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None



# =========================================================
# U102 TRUE INCREMENTAL REAL FIX
# Muc tieu:
# - Fix fingerprint cache triệt để
# - Debug đầy đủ mismatch
# - Đảm bảo skip thật
# - Giảm rebuild mạnh khi rerun cùng folder / file không đổi
# =========================================================

def _u102_fingerprint_cache_path(output_path: str) -> str:
    try:
        p = Path(output_path)
        return str(p.with_suffix(p.suffix + ".fingerprintcache_u102.json"))
    except Exception:
        return str(output_path) + ".fingerprintcache_u102.json"


def _u102_load_fingerprint_cache(output_path: str) -> Dict[str, str]:
    fp = _u102_fingerprint_cache_path(output_path)
    out: Dict[str, str] = {}
    try:
        if os.path.exists(fp):
            with open(fp, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                for k, v in data.items():
                    nk = _u101_norm_path_strict(k)
                    if nk and isinstance(v, str):
                        out[nk] = v
    except (OSError, json.JSONDecodeError, ValueError) as exc:
        _rt_log_exception(f"u102_load_fingerprint_cache_failed: {fp}", exc)
    return out


def _u102_save_fingerprint_cache(output_path: str, cache_map: Dict[str, str]):
    fp = _u102_fingerprint_cache_path(output_path)
    try:
        normed = {}
        for k, v in (cache_map or {}).items():
            nk = _u101_norm_path_strict(k)
            if nk and isinstance(v, str):
                normed[nk] = v
        with open(fp, "w", encoding="utf-8") as f:
            json.dump(dict(sorted(normed.items())), f, ensure_ascii=False, indent=2)
    except OSError as exc:
        _rt_log_exception(f"u102_save_fingerprint_cache_failed: {fp}", exc)


class Processor(Processor):

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._fp_cache_loaded = 0
        self._fp_cache_saved = 0
        self._fp_mismatch_logged = 0
        self._full_rebuild_warning = 0
        self._sample_fp_skipped = []
        self._sample_fp_changed = []

    def load_fingerprint_cache(self, output_path: str):
        raw = _u102_load_fingerprint_cache(output_path)
        self._fingerprint_cache = {}
        for k, v in raw.items():
            nk = _u101_norm_path_strict(k)
            if nk and isinstance(v, str):
                self._fingerprint_cache[nk] = v
        self._pending_fingerprint_updates = {}
        self._last_fingerprint_skipped = 0
        self._last_fingerprint_changed = 0
        self._last_supported_scanned = 0
        self._fp_cache_loaded = len(self._fingerprint_cache)
        self._fp_cache_saved = 0
        self._fp_mismatch_logged = 0
        self._full_rebuild_warning = 0
        self._sample_fp_skipped = []
        self._sample_fp_changed = []
        try:
            self.logger.log('FP_CACHE_STATUS_U102', 'INFO',
                            f"loaded={self._fp_cache_loaded} | path={_u102_fingerprint_cache_path(output_path)}")
        except Exception:
            pass

    def persist_caches(self, output_path: str, success_changed_paths: Optional[Set[str]] = None):
        try:
            success_changed_paths = set(success_changed_paths or set())
            for p in success_changed_paths:
                np = _u101_norm_path_strict(p)
                new_fp = self._pending_fingerprint_updates.get(np)
                if new_fp:
                    self._fingerprint_cache[np] = new_fp

            # normalize all keys once more before save
            self._fingerprint_cache = {
                _u101_norm_path_strict(k): v
                for k, v in self._fingerprint_cache.items()
                if _u101_norm_path_strict(k) and isinstance(v, str)
            }
            _u102_save_fingerprint_cache(output_path, self._fingerprint_cache)
            self._fp_cache_saved = len(self._fingerprint_cache)
            try:
                self.logger.log('FP_CACHE_STATUS_U102', 'INFO',
                                f"saved={self._fp_cache_saved} | path={_u102_fingerprint_cache_path(output_path)}")
            except Exception:
                pass
        except Exception as exc:
            _rt_log_exception("u102_persist_fingerprint_cache_failed", exc)

        try:
            cache = _u90_get_source_cache()
            _u90_save_source_cache(output_path, cache)
        except Exception as exc:
            _rt_log_exception("u102_persist_source_cache_failed", exc)

    def _collect_supported_scanned_paths(self, output_path: str, scanned_files: Set[str]) -> Set[str]:
        scanned_norm = set()
        for p in scanned_files or set():
            np = _u101_norm_path_strict(p)
            if np:
                scanned_norm.add(np)

        current_supported = set()
        try:
            raw_supported = _u99_collect_supported_paths_from_output(output_path)
            for p in raw_supported:
                np = _u101_norm_path_strict(p)
                if np:
                    current_supported.add(np)
        except Exception as exc:
            _rt_log_exception("u102_collect_supported_failed", exc)
            current_supported = set()

        supported_scanned = {p for p in current_supported if p in scanned_norm}
        self._last_supported_scanned = len(supported_scanned)

        if scanned_norm and not supported_scanned:
            try:
                sample_supported = list(sorted(current_supported))[:3]
                sample_scanned = list(sorted(scanned_norm))[:3]
                self.logger.log('DEBUG_PATH_MATCH_U102', 'INFO',
                                f"supported_scanned=0 | sample_supported={sample_supported} | sample_scanned={sample_scanned}")
            except Exception:
                pass

        return supported_scanned

    def _filter_changed_by_fingerprint(self, candidate_paths: Set[str]) -> Set[str]:
        changed = set()
        skipped = 0
        pending = {}
        self._sample_fp_skipped = []
        self._sample_fp_changed = []
        self._fp_mismatch_logged = 0

        # defensively normalize existing cache keys
        self._fingerprint_cache = {
            _u101_norm_path_strict(k): v
            for k, v in self._fingerprint_cache.items()
            if _u101_norm_path_strict(k) and isinstance(v, str)
        }

        for p in sorted(candidate_paths or set()):
            np = _u101_norm_path_strict(p)
            if not np:
                continue

            fp = _compute_file_fingerprint(np)
            if not fp:
                changed.add(np)
                pending[np] = ""
                if len(self._sample_fp_changed) < 3:
                    self._sample_fp_changed.append(np)
                continue

            old_fp = self._fingerprint_cache.get(np, "")
            if old_fp and old_fp == fp:
                skipped += 1
                if len(self._sample_fp_skipped) < 3:
                    self._sample_fp_skipped.append(np)
                continue

            changed.add(np)
            pending[np] = fp
            if len(self._sample_fp_changed) < 3:
                self._sample_fp_changed.append(np)

            if old_fp and old_fp != fp and self._fp_mismatch_logged < 5:
                self._fp_mismatch_logged += 1
                try:
                    self.logger.log(
                        'FP_MISMATCH_U102',
                        'INFO',
                        f"path={np} | old={old_fp[:12]} | new={fp[:12]}"
                    )
                except Exception:
                    pass

        self._pending_fingerprint_updates = pending
        self._last_fingerprint_skipped = skipped
        self._last_fingerprint_changed = len(changed)

        if candidate_paths and len(changed) == len(candidate_paths):
            self._full_rebuild_warning = 1
            try:
                self.logger.log(
                    'FULL_REBUILD_WARNING_U102',
                    'WARNING',
                    f"all_files_marked_changed={len(changed)} | candidate_paths={len(candidate_paths)}"
                )
            except Exception:
                pass
        else:
            self._full_rebuild_warning = 0

        return changed

    def resolve_changed_paths(self, output_path: str, summary: Dict[str, Any], scanned_files: Set[str]) -> Set[str]:
        try:
            bootstrap_needed = _u101_need_bootstrap(output_path)
        except Exception as exc:
            _rt_log_exception("u102_bootstrap_check_failed", exc)
            bootstrap_needed = True

        supported_scanned = self._collect_supported_scanned_paths(output_path, scanned_files)

        if bootstrap_needed:
            # bootstrap seed toàn bộ supported_scanned của folder hiện tại
            self._last_supported_scanned = len(supported_scanned)
            return set(supported_scanned)

        return set(self._filter_changed_by_fingerprint(supported_scanned))

    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        _u85_reset_bench()
        t_all_0 = time.perf_counter()
        _U90_SOURCE_CACHE_CTX['output_path'] = output_path
        _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
        cache = _u90_get_source_cache()
        cache['stats'] = _u90_empty_source_cache()['stats']

        success_changed_paths = set()

        try:
            self._runtime_logger.info("U102_RUN_START folder=%s output=%s", folder_path, output_path)
            self.logger.log('FILE_UPSERT_ENTER_U102_FINAL', 'info')
            self.logger.log('SOURCE_PARSE_CACHE_ENTER_U102_FINAL', 'info')

            _U84_WIDTH_CACHE_CTX['output_path'] = output_path
            _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

            self.load_fingerprint_cache(output_path)

            scanned_files = self.scan(folder_path, output_path)

            t0 = time.perf_counter()
            summary = self.core_run(
                folder_path,
                output_path,
                progress_callback=progress_callback,
                repair_options=repair_options
            )
            core_seconds = max(0.0, time.perf_counter() - t0)

            changed_paths = self.resolve_changed_paths(output_path, summary, scanned_files)

            t1 = time.perf_counter()
            upsert_stats = self.upsert_inv_pl(output_path, changed_paths)
            rebuild_seconds = max(0.0, time.perf_counter() - t1)

            if upsert_stats.get('errors', 0) == 0:
                success_changed_paths = set(changed_paths)

            self.persist_caches(output_path, success_changed_paths=success_changed_paths)

            total_seconds = max(0.0, time.perf_counter() - t_all_0)

            try:
                stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
            except Exception:
                stats = {}

            bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
            bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                            f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | supported_scanned={self._last_supported_scanned} | changed={upsert_stats.get('changed',0)} | fp_changed={self._last_fingerprint_changed} | fp_skipped={self._last_fingerprint_skipped} | fp_cache_loaded={self._fp_cache_loaded} | fp_cache_saved={self._fp_cache_saved} | full_rebuild_warning={self._full_rebuild_warning} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | inv_index_repaired={upsert_stats.get('inv_index_repaired',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | pl_index_repaired={upsert_stats.get('pl_index_repaired',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
            bench_rows.append(make_log_row(output_path, 'FP_SAMPLES_U102', 'INFO',
                            f"changed_sample={self._sample_fp_changed} | skipped_sample={self._sample_fp_skipped}"))
            bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                            f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
            _u85_append_bench_log_rows(output_path, bench_rows)
            return summary

        except Exception as exc:
            _rt_log_exception("u102_processor_run_fatal", exc)
            raise
        finally:
            _U84_WIDTH_CACHE_CTX['output_path'] = ''
            _U84_WIDTH_CACHE_CTX['cache'] = {}
            _U85_BENCH_CTX['enabled'] = False
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None



# =========================================================
# U103 CANONICAL PATH PRODUCTION
# Muc tieu:
# - Dua tat ca File Path ve 1 kieu duy nhat: D:/...
# - Dung cung 1 canonicalizer cho:
#     scan / cache / compare / ghi worksheet / hidden index
# - Sua goc re loi path mix "\" va "/" dang lam fingerprint + incremental lech
# =========================================================

def _u103_canonical_path(p) -> str:
    try:
        if p is None:
            return ""
        s = str(p).strip().strip('"').strip("'")
        if not s:
            return ""
        s = os.path.abspath(s)
        s = os.path.normpath(s)
        s = s.replace("\\", "/")
        while '//' in s:
            s = s.replace('//', '/')
        # Giữ đúng ổ đĩa nhưng thống nhất kiểu slash và upper toàn bộ để compare/cache ổn định
        return s.upper()
    except Exception as exc:
        _rt_log_exception(f"u103_canonical_path_failed: {p}", exc)
        return ""


# Canonicalizer duy nhat cho runtime
_safe_norm_path_runtime = _u103_canonical_path
_u101_norm_path_strict = _u103_canonical_path


def _u103_normalize_sheet_file_path_column(ws, logger=None) -> int:
    try:
        path_col = find_header_col(ws, "File Path")
        if path_col <= 0:
            return 0
        changed = 0
        for r in range(2, ws.max_row + 1):
            old = ws.cell(r, path_col).value
            if old is None or str(old).strip() == "":
                continue
            new = _u103_canonical_path(old)
            if new and str(old) != new:
                ws.cell(r, path_col).value = new
                changed += 1
        return changed
    except Exception as exc:
        _rt_log_exception(f"u103_normalize_sheet_file_path_column_failed:{getattr(ws, 'title', '')}", exc)
        return 0


def _u103_normalize_output_workbook_paths(output_path: str, logger=None) -> Dict[str, int]:
    wb = open_or_create_output_workbook(output_path)
    counts = {"SUB_DETAIL": 0, "LOG_SUB_DETAIL": 0, "INV": 0, "PL": 0, "INDEX": 0}
    try:
        for sh in ("SUB_DETAIL", "LOG_SUB_DETAIL", "INV", "PL"):
            if sh in wb.sheetnames:
                counts[sh] = _u103_normalize_sheet_file_path_column(wb[sh], logger=logger)

        if _U98_BLOCK_INDEX_SHEET in wb.sheetnames:
            ws_idx = wb[_U98_BLOCK_INDEX_SHEET]
            # cot 2 la FilePath
            changed = 0
            for r in range(2, ws_idx.max_row + 1):
                old = ws_idx.cell(r, 2).value
                if old is None or str(old).strip() == "":
                    continue
                new = _u103_canonical_path(old)
                if new and str(old) != new:
                    ws_idx.cell(r, 2).value = new
                    changed += 1
            counts["INDEX"] = changed

        safe_save_workbook_atomic(wb, output_path)
        return counts
    finally:
        wb.close()


class Processor(Processor):

    def scan(self, folder_path: str, output_path: str) -> Set[str]:
        try:
            files = collect_source_files(folder_path, output_path)
            normed = set()
            for p in files:
                np = _u103_canonical_path(p)
                if np:
                    normed.add(np)
            self._current_scanned_files = normed
            return set(normed)
        except (FileNotFoundError, OSError, ValueError) as exc:
            _rt_log_exception(f"u103_scan_failed: {folder_path}", exc)
            self._current_scanned_files = set()
            return set()
        except Exception as exc:
            _rt_log_exception(f"u103_scan_unexpected_failed: {folder_path}", exc)
            self._current_scanned_files = set()
            return set()

    def load_fingerprint_cache(self, output_path: str):
        raw = _u102_load_fingerprint_cache(output_path)
        self._fingerprint_cache = {}
        for k, v in raw.items():
            nk = _u103_canonical_path(k)
            if nk and isinstance(v, str):
                self._fingerprint_cache[nk] = v
        self._pending_fingerprint_updates = {}
        self._last_fingerprint_skipped = 0
        self._last_fingerprint_changed = 0
        self._last_supported_scanned = 0
        self._fp_cache_loaded = len(self._fingerprint_cache)
        self._fp_cache_saved = 0
        self._fp_mismatch_logged = 0
        self._full_rebuild_warning = 0
        self._sample_fp_skipped = []
        self._sample_fp_changed = []
        try:
            self.logger.log('FP_CACHE_STATUS_U103', 'INFO',
                            f"loaded={self._fp_cache_loaded} | path={_u102_fingerprint_cache_path(output_path)}")
        except Exception:
            pass

    def persist_caches(self, output_path: str, success_changed_paths: Optional[Set[str]] = None):
        try:
            success_changed_paths = set(success_changed_paths or set())
            for p in success_changed_paths:
                np = _u103_canonical_path(p)
                new_fp = self._pending_fingerprint_updates.get(np)
                if new_fp:
                    self._fingerprint_cache[np] = new_fp

            self._fingerprint_cache = {
                _u103_canonical_path(k): v
                for k, v in self._fingerprint_cache.items()
                if _u103_canonical_path(k) and isinstance(v, str)
            }
            _u102_save_fingerprint_cache(output_path, self._fingerprint_cache)
            self._fp_cache_saved = len(self._fingerprint_cache)
            try:
                self.logger.log('FP_CACHE_STATUS_U103', 'INFO',
                                f"saved={self._fp_cache_saved} | path={_u102_fingerprint_cache_path(output_path)}")
            except Exception:
                pass
        except Exception as exc:
            _rt_log_exception("u103_persist_fingerprint_cache_failed", exc)

        try:
            cache = _u90_get_source_cache()
            _u90_save_source_cache(output_path, cache)
        except Exception as exc:
            _rt_log_exception("u103_persist_source_cache_failed", exc)

    def _collect_supported_scanned_paths(self, output_path: str, scanned_files: Set[str]) -> Set[str]:
        scanned_norm = set()
        for p in scanned_files or set():
            np = _u103_canonical_path(p)
            if np:
                scanned_norm.add(np)

        current_supported = set()
        try:
            raw_supported = _u99_collect_supported_paths_from_output(output_path)
            for p in raw_supported:
                np = _u103_canonical_path(p)
                if np:
                    current_supported.add(np)
        except Exception as exc:
            _rt_log_exception("u103_collect_supported_failed", exc)
            current_supported = set()

        supported_scanned = {p for p in current_supported if p in scanned_norm}
        self._last_supported_scanned = len(supported_scanned)

        if scanned_norm and not supported_scanned:
            try:
                sample_supported = list(sorted(current_supported))[:3]
                sample_scanned = list(sorted(scanned_norm))[:3]
                self.logger.log('DEBUG_PATH_MATCH_U103', 'INFO',
                                f"supported_scanned=0 | sample_supported={sample_supported} | sample_scanned={sample_scanned}")
            except Exception:
                pass

        return supported_scanned

    def _filter_changed_by_fingerprint(self, candidate_paths: Set[str]) -> Set[str]:
        changed = set()
        skipped = 0
        pending = {}
        self._sample_fp_skipped = []
        self._sample_fp_changed = []
        self._fp_mismatch_logged = 0

        self._fingerprint_cache = {
            _u103_canonical_path(k): v
            for k, v in self._fingerprint_cache.items()
            if _u103_canonical_path(k) and isinstance(v, str)
        }

        for p in sorted(candidate_paths or set()):
            np = _u103_canonical_path(p)
            if not np:
                continue

            fp = _compute_file_fingerprint(np)
            if not fp:
                changed.add(np)
                pending[np] = ""
                if len(self._sample_fp_changed) < 3:
                    self._sample_fp_changed.append(np)
                continue

            old_fp = self._fingerprint_cache.get(np, "")
            if old_fp and old_fp == fp:
                skipped += 1
                if len(self._sample_fp_skipped) < 3:
                    self._sample_fp_skipped.append(np)
                continue

            changed.add(np)
            pending[np] = fp
            if len(self._sample_fp_changed) < 3:
                self._sample_fp_changed.append(np)

            if old_fp and old_fp != fp and self._fp_mismatch_logged < 5:
                self._fp_mismatch_logged += 1
                try:
                    self.logger.log(
                        'FP_MISMATCH_U103',
                        'INFO',
                        f"path={np} | old={old_fp[:12]} | new={fp[:12]}"
                    )
                except Exception:
                    pass

        self._pending_fingerprint_updates = pending
        self._last_fingerprint_skipped = skipped
        self._last_fingerprint_changed = len(changed)

        if candidate_paths and len(changed) == len(candidate_paths):
            self._full_rebuild_warning = 1
            try:
                self.logger.log(
                    'FULL_REBUILD_WARNING_U103',
                    'WARNING',
                    f"all_files_marked_changed={len(changed)} | candidate_paths={len(candidate_paths)}"
                )
            except Exception:
                pass
        else:
            self._full_rebuild_warning = 0

        return changed

    def upsert_inv_pl(self, output_path: str, changed_paths: Set[str]):
        current_supported = set()
        try:
            raw_supported = _u99_collect_supported_paths_from_output(output_path)
            for p in raw_supported:
                np = _u103_canonical_path(p)
                if np:
                    current_supported.add(np)
        except Exception as exc:
            _rt_log_exception("u103_upsert_collect_supported_failed", exc)
            current_supported = set()

        try:
            return _u99_rebuild_upsert(
                output_path,
                current_supported,
                changed_paths={_u103_canonical_path(p) for p in (changed_paths or set()) if _u103_canonical_path(p)},
                prep_label='DETAIL_PREP_U79'
            )
        except (ValueError, KeyError, FileNotFoundError) as exc:
            _rt_log_exception("u103_upsert_inv_pl_failed", exc)
            return {
                'supported_files': len(current_supported),
                'changed': 0,
                'inv_append': 0,
                'inv_update_same': 0,
                'inv_update_resize': 0,
                'inv_index_repaired': 0,
                'pl_append': 0,
                'pl_update_same': 0,
                'pl_update_resize': 0,
                'pl_index_repaired': 0,
                'cache_hit': 0,
                'rebuilt_files': 0,
                'errors': 1,
                'cache_path': _u87_detail_cache_path(output_path),
            }
        except Exception as exc:
            _rt_log_exception("u103_upsert_inv_pl_unexpected_failed", exc)
            return {
                'supported_files': len(current_supported),
                'changed': 0,
                'inv_append': 0,
                'inv_update_same': 0,
                'inv_update_resize': 0,
                'inv_index_repaired': 0,
                'pl_append': 0,
                'pl_update_same': 0,
                'pl_update_resize': 0,
                'pl_index_repaired': 0,
                'cache_hit': 0,
                'rebuilt_files': 0,
                'errors': 1,
                'cache_path': _u87_detail_cache_path(output_path),
            }

    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        _u85_reset_bench()
        t_all_0 = time.perf_counter()
        _U90_SOURCE_CACHE_CTX['output_path'] = output_path
        _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
        cache = _u90_get_source_cache()
        cache['stats'] = _u90_empty_source_cache()['stats']

        success_changed_paths = set()

        try:
            self._runtime_logger.info("U103_RUN_START folder=%s output=%s", folder_path, output_path)
            self.logger.log('FILE_UPSERT_ENTER_U103_FINAL', 'info')
            self.logger.log('SOURCE_PARSE_CACHE_ENTER_U103_FINAL', 'info')

            _U84_WIDTH_CACHE_CTX['output_path'] = output_path
            _U84_WIDTH_CACHE_CTX['cache'] = _u84_load_width_cache(output_path)

            # normalize workbook path columns first to eliminate mixed slash styles
            path_fix_counts = _u103_normalize_output_workbook_paths(output_path, logger=self.logger)

            self.load_fingerprint_cache(output_path)

            scanned_files = self.scan(folder_path, output_path)

            t0 = time.perf_counter()
            summary = self.core_run(
                folder_path,
                output_path,
                progress_callback=progress_callback,
                repair_options=repair_options
            )
            core_seconds = max(0.0, time.perf_counter() - t0)

            changed_paths = self.resolve_changed_paths(output_path, summary, scanned_files)

            t1 = time.perf_counter()
            upsert_stats = self.upsert_inv_pl(output_path, changed_paths)
            rebuild_seconds = max(0.0, time.perf_counter() - t1)

            if upsert_stats.get('errors', 0) == 0:
                success_changed_paths = set(changed_paths)

            self.persist_caches(output_path, success_changed_paths=success_changed_paths)

            total_seconds = max(0.0, time.perf_counter() - t_all_0)

            try:
                stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
            except Exception:
                stats = {}

            bench_rows = _u85_make_bench_rows(folder_path, output_path, core_seconds, rebuild_seconds, total_seconds)
            bench_rows.append(make_log_row(output_path, 'PATH_CANONICAL_STATS_U103', 'INFO',
                            f"SUB_DETAIL={path_fix_counts.get('SUB_DETAIL',0)} | LOG_SUB_DETAIL={path_fix_counts.get('LOG_SUB_DETAIL',0)} | INV={path_fix_counts.get('INV',0)} | PL={path_fix_counts.get('PL',0)} | INDEX={path_fix_counts.get('INDEX',0)}"))
            bench_rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                            f"supported_files={upsert_stats.get('supported_files',0)} | scanned={len(scanned_files)} | supported_scanned={self._last_supported_scanned} | changed={upsert_stats.get('changed',0)} | fp_changed={self._last_fingerprint_changed} | fp_skipped={self._last_fingerprint_skipped} | fp_cache_loaded={self._fp_cache_loaded} | fp_cache_saved={self._fp_cache_saved} | full_rebuild_warning={self._full_rebuild_warning} | inv_append={upsert_stats.get('inv_append',0)} | inv_update_same={upsert_stats.get('inv_update_same',0)} | inv_update_resize={upsert_stats.get('inv_update_resize',0)} | inv_index_repaired={upsert_stats.get('inv_index_repaired',0)} | pl_append={upsert_stats.get('pl_append',0)} | pl_update_same={upsert_stats.get('pl_update_same',0)} | pl_update_resize={upsert_stats.get('pl_update_resize',0)} | pl_index_repaired={upsert_stats.get('pl_index_repaired',0)} | cache_hit={upsert_stats.get('cache_hit',0)} | rebuilt={upsert_stats.get('rebuilt_files',0)} | cache={upsert_stats.get('cache_path','')}"))
            bench_rows.append(make_log_row(output_path, 'FP_SAMPLES_U103', 'INFO',
                            f"changed_sample={self._sample_fp_changed} | skipped_sample={self._sample_fp_skipped}"))
            bench_rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                            f"arr_hit={int(stats.get('arr_hit',0) or 0)} | arr_miss={int(stats.get('arr_miss',0) or 0)} | bundle_hit={int(stats.get('bundle_hit',0) or 0)} | bundle_miss={int(stats.get('bundle_miss',0) or 0)} | detail_hit={int(stats.get('detail_hit',0) or 0)} | detail_miss={int(stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
            _u85_append_bench_log_rows(output_path, bench_rows)
            return summary

        except Exception as exc:
            _rt_log_exception("u103_processor_run_fatal", exc)
            raise
        finally:
            _U84_WIDTH_CACHE_CTX['output_path'] = ''
            _U84_WIDTH_CACHE_CTX['cache'] = {}
            _U85_BENCH_CTX['enabled'] = False
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None



# =========================================================
# U104 TRUE DELTA ENGINE
# Muc tieu:
# - Giu nguyen logic doc / cot dau ra goc cua SUB_DETAIL / INV / PL
# - Khong rebuild toan bo INV / PL moi lan chay
# - File OK + khong doi fingerprint -> bo qua hoan toan
# - File WARNING/ERROR hoac file moi / file da doi -> chi xu ly dung file do
# - Thong nhat duong dan va trang thai giua SUB_DETAIL / INV / PL
# - Cho phep them 1 cot cuoi cung: PathKey (khoa canon duong dan)
# =========================================================

U104_PATHKEY_HEADER = "PathKey"
U104_EMPTY_FILL = PatternFill(fill_type=None)
U104_EMPTY_FONT = Font()


def _u104_canonical_path(p) -> str:
    return _u103_canonical_path(p)


def _u104_safe_equal(a, b) -> bool:
    if isinstance(a, datetime) or isinstance(b, datetime):
        return normalize_date_compare(a) == normalize_date_compare(b)
    return nz_str(a) == nz_str(b)


def _u104_group_desc_ranges(row_numbers: List[int]) -> List[Tuple[int, int]]:
    rows = sorted({int(r) for r in row_numbers if int(r) >= 2}, reverse=True)
    if not rows:
        return []
    ranges: List[Tuple[int, int]] = []
    start = rows[0]
    count = 1
    prev = rows[0]
    for r in rows[1:]:
        if r == prev - 1:
            prev = r
            count += 1
            continue
        ranges.append((prev, count))
        start = r
        prev = r
        count = 1
    ranges.append((prev, count))
    return ranges


def _u104_delete_rows_desc(ws, row_numbers: List[int]) -> int:
    deleted = 0
    for start_row, count in _u104_group_desc_ranges(row_numbers):
        ws.delete_rows(start_row, count)
        deleted += count
    return deleted


def _u104_apply_header_style(cell, fill_obj):
    cell.font = HEADER_FONT
    cell.fill = fill_obj


def _u104_ensure_pathkey_column(ws, base_headers: List[str], header_fill) -> int:
    # ensure base headers stay untouched in the original positions
    if nz_str(ws.cell(1, 1).value) == "":
        for idx, header in enumerate(base_headers, start=1):
            ws.cell(1, idx).value = sanitize_excel_text(header)
            _u104_apply_header_style(ws.cell(1, idx), header_fill)
    else:
        for idx, header in enumerate(base_headers, start=1):
            if nz_str(ws.cell(1, idx).value) == "":
                ws.cell(1, idx).value = sanitize_excel_text(header)
                _u104_apply_header_style(ws.cell(1, idx), header_fill)

    for c in range(1, ws.max_column + 1):
        if nz_str(ws.cell(1, c).value) == U104_PATHKEY_HEADER:
            return c

    pathkey_col = ws.max_column + 1
    ws.cell(1, pathkey_col).value = U104_PATHKEY_HEADER
    _u104_apply_header_style(ws.cell(1, pathkey_col), header_fill)
    return pathkey_col


def _u104_clear_row_style(ws, row_idx: int, col_count: int):
    for c in range(1, col_count + 1):
        cell = ws.cell(row_idx, c)
        cell.fill = U104_EMPTY_FILL
        cell.font = U104_EMPTY_FONT


def _u104_set_subdetail_row_style(ws, row_idx: int, status_text: str, col_count: int):
    _u104_clear_row_style(ws, row_idx, col_count)
    st = nz_str(status_text).upper()
    if st in STATUS_FILLS:
        fill = STATUS_FILLS[st]
        for c in range(1, col_count + 1):
            ws.cell(row_idx, c).fill = fill


def _u104_set_detail_row_style(ws, row_idx: int, row_type_text: str, col_count: int):
    _u104_clear_row_style(ws, row_idx, col_count)
    if nz_str(row_type_text).upper() == 'TOTAL':
        for c in range(1, col_count + 1):
            ws.cell(row_idx, c).fill = TOTAL_ROW_FILL
            ws.cell(row_idx, c).font = TOTAL_FONT


def _u104_write_subdetail_row(ws, row_idx: int, arr: List[Any], pathkey_col: int):
    base_len = len(SUB_DETAIL_HEADERS)
    arr = list(arr[:base_len]) + [""] * max(0, base_len - len(arr))
    if base_len >= 24:
        arr[23] = _u104_canonical_path(arr[23])
    pathkey = _u104_canonical_path(arr[23]) if base_len >= 24 else ""
    for c in range(1, base_len + 1):
        header = SUB_DETAIL_HEADERS[c - 1]
        val = arr[c - 1]
        if c in (15, 19):
            ws.cell(row_idx, c).value = coerce_excel_date_value(val)
            if ws.cell(row_idx, c).value not in (None, ""):
                ws.cell(row_idx, c).number_format = 'dd/mm/yyyy'
        else:
            ws.cell(row_idx, c).value = sanitize_excel_text(val)
    ws.cell(row_idx, pathkey_col).value = pathkey
    _u104_set_subdetail_row_style(ws, row_idx, arr[21] if len(arr) > 21 else '', pathkey_col)


def _u104_append_subdetail_row(ws, arr: List[Any], pathkey_col: int) -> int:
    row_idx = ws.max_row + 1 if ws.max_row >= 1 else 2
    if row_idx < 2:
        row_idx = 2
    _u104_write_subdetail_row(ws, row_idx, arr, pathkey_col)
    return row_idx


def _u104_prepare_inv_rows(rows: List[List[Any]]) -> List[List[Any]]:
    out: List[List[Any]] = []
    for arr in rows or []:
        row = list(arr[:len(INV_DETAIL_HEADERS)]) + [""] * max(0, len(INV_DETAIL_HEADERS) - len(arr))
        if len(row) >= len(INV_DETAIL_HEADERS):
            row[20] = _u104_canonical_path(row[20])
        out.append(row)
    return out


def _u104_prepare_pl_rows(rows: List[List[Any]]) -> List[List[Any]]:
    out: List[List[Any]] = []
    for arr in rows or []:
        row = list(arr[:len(PL_DETAIL_HEADERS)]) + [""] * max(0, len(PL_DETAIL_HEADERS) - len(arr))
        if len(row) >= len(PL_DETAIL_HEADERS):
            row[16] = _u104_canonical_path(row[16])
        out.append(row)
    return out


def _u104_append_inv_rows(ws, rows: List[List[Any]], pathkey_col: int) -> int:
    if not rows:
        return 0
    appended = 0
    col_count = pathkey_col
    for arr in rows:
        row_idx = ws.max_row + 1 if ws.max_row >= 1 else 2
        if row_idx < 2:
            row_idx = 2
        pathkey = _u104_canonical_path(arr[20]) if len(arr) >= len(INV_DETAIL_HEADERS) else ""
        for c in range(1, len(INV_DETAIL_HEADERS) + 1):
            header = INV_DETAIL_HEADERS[c - 1]
            val = arr[c - 1] if c - 1 < len(arr) else ""
            if c in (1,):
                ws.cell(row_idx, c).value = coerce_excel_date_value(val)
                if ws.cell(row_idx, c).value not in (None, ""):
                    ws.cell(row_idx, c).number_format = 'dd/mm/yyyy'
            else:
                ws.cell(row_idx, c).value = sanitize_excel_text(val)
        ws.cell(row_idx, pathkey_col).value = pathkey
        _u104_set_detail_row_style(ws, row_idx, arr[17] if len(arr) > 17 else '', col_count)
        appended += 1
    return appended


def _u104_append_pl_rows(ws, rows: List[List[Any]], pathkey_col: int) -> int:
    if not rows:
        return 0
    appended = 0
    col_count = pathkey_col
    for arr in rows:
        row_idx = ws.max_row + 1 if ws.max_row >= 1 else 2
        if row_idx < 2:
            row_idx = 2
        pathkey = _u104_canonical_path(arr[16]) if len(arr) >= len(PL_DETAIL_HEADERS) else ""
        for c in range(1, len(PL_DETAIL_HEADERS) + 1):
            header = PL_DETAIL_HEADERS[c - 1]
            val = arr[c - 1] if c - 1 < len(arr) else ""
            if c in (1,):
                ws.cell(row_idx, c).value = coerce_excel_date_value(val)
                if ws.cell(row_idx, c).value not in (None, ""):
                    ws.cell(row_idx, c).number_format = 'dd/mm/yyyy'
            else:
                ws.cell(row_idx, c).value = sanitize_excel_text(val)
        ws.cell(row_idx, pathkey_col).value = pathkey
        _u104_set_detail_row_style(ws, row_idx, arr[13] if len(arr) > 13 else '', col_count)
        appended += 1
    return appended


def _u104_collect_subdetail_state(ws, pathkey_col: int) -> Tuple[Dict[str, Dict[str, Any]], List[int], int]:
    state: Dict[str, Dict[str, Any]] = {}
    dup_rows: List[int] = []
    backfill = 0
    path_col = find_header_col(ws, 'File Path')
    status_col = find_header_col(ws, 'Status')
    sig_col = find_header_col(ws, 'File Signature')
    for r in range(2, ws.max_row + 1):
        pk = nz_str(ws.cell(r, pathkey_col).value)
        raw_path = ws.cell(r, path_col).value if path_col > 0 else ''
        if not pk:
            pk = _u104_canonical_path(raw_path)
            if pk:
                ws.cell(r, pathkey_col).value = pk
                backfill += 1
        if not pk:
            continue
        status = nz_str(ws.cell(r, status_col).value).upper() if status_col > 0 else ''
        sig = nz_str(ws.cell(r, sig_col).value) if sig_col > 0 else ''
        if pk in state:
            dup_rows.append(state[pk]['row_idx'])
        state[pk] = {
            'row_idx': r,
            'status': status,
            'signature': sig,
        }
    return state, sorted(set(dup_rows), reverse=True), backfill


def _u104_collect_detail_rownums_by_pathkey(ws, file_path_col: int, pathkey_col: int) -> Tuple[Dict[str, List[int]], int]:
    mapping: Dict[str, List[int]] = {}
    backfill = 0
    for r in range(2, ws.max_row + 1):
        pk = nz_str(ws.cell(r, pathkey_col).value)
        raw_path = ws.cell(r, file_path_col).value if file_path_col > 0 else ''
        if not pk:
            pk = _u104_canonical_path(raw_path)
            if pk:
                ws.cell(r, pathkey_col).value = pk
                backfill += 1
        if not pk:
            continue
        mapping.setdefault(pk, []).append(r)
    return mapping, backfill


def _u104_existing_row_equals_subdetail(ws, row_idx: int, arr: List[Any], pathkey_col: int) -> bool:
    base_len = len(SUB_DETAIL_HEADERS)
    arr = list(arr[:base_len]) + [""] * max(0, base_len - len(arr))
    if base_len >= 24:
        arr[23] = _u104_canonical_path(arr[23])
    for c in range(1, base_len + 1):
        old = ws.cell(row_idx, c).value
        new = arr[c - 1]
        if c in (15, 19):
            if normalize_date_compare(old) != normalize_date_compare(new):
                return False
        else:
            if nz_str(old) != nz_str(new):
                return False
    if nz_str(ws.cell(row_idx, pathkey_col).value) != _u104_canonical_path(arr[23]):
        return False
    return True


def _u104_make_log_rows_for_zero_work(cnt_total: int, cnt_skip_ok: int) -> List[List[Any]]:
    return [
        make_log_row('', 'START', 'INFO', f'Bat dau quet {cnt_total} file.'),
        make_log_row('', 'SKIP', 'INFO', f'Zero-work: bo qua {cnt_skip_ok} file OK da co san va khong doi fingerprint.'),
        make_log_row('', 'DONE', 'INFO', f'Tong={cnt_total}; OK=0; WARNING=0; ERROR=0; SkipPath={cnt_skip_ok}; SkipSignature=0; Repaired=0')
    ]


def _u104_build_runtime_bench_rows(folder_path: str, output_path: str, core_seconds: float, total_seconds: float, stats_map: Dict[str, Any], source_cache_stats: Dict[str, Any]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    rows.append(make_log_row(output_path, 'BENCH_CORE_WRITE', 'INFO', f'{core_seconds:.3f}s'))
    rows.append(make_log_row(output_path, 'BENCH_REBUILD_DETAIL', 'INFO', '0.000s'))
    rows.append(make_log_row(output_path, 'BENCH_TOTAL', 'INFO', f'{total_seconds:.3f}s'))
    rows.append(make_log_row(output_path, 'SMART_REBUILD_STATS', 'INFO',
                    f"scanned={int(stats_map.get('scanned',0) or 0)} | candidate={int(stats_map.get('candidate',0) or 0)} | ok_skip={int(stats_map.get('ok_skip',0) or 0)} | parsed={int(stats_map.get('parsed',0) or 0)} | repaired={int(stats_map.get('repaired',0) or 0)} | parse_reused={int(stats_map.get('parse_reused',0) or 0)} | parse_fresh={int(stats_map.get('parse_fresh',0) or 0)} | sub_insert={int(stats_map.get('sub_insert',0) or 0)} | sub_update={int(stats_map.get('sub_update',0) or 0)} | inv_deleted={int(stats_map.get('inv_deleted',0) or 0)} | inv_appended={int(stats_map.get('inv_appended',0) or 0)} | pl_deleted={int(stats_map.get('pl_deleted',0) or 0)} | pl_appended={int(stats_map.get('pl_appended',0) or 0)} | fp_changed={int(stats_map.get('fp_changed',0) or 0)} | fp_skipped={int(stats_map.get('fp_skipped',0) or 0)} | workbook_saved={int(stats_map.get('workbook_saved',0) or 0)}"))
    if source_cache_stats and ('lookup_total' in source_cache_stats or 'cache_hit_reuse' in source_cache_stats):
        rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                        f"lookup_total={int(source_cache_stats.get('lookup_total',0) or 0)} | cache_hit_reuse={int(source_cache_stats.get('cache_hit_reuse',0) or 0)} | cache_miss_parse={int(source_cache_stats.get('cache_miss_parse',0) or 0)} | cache_bypass={int(source_cache_stats.get('cache_bypass',0) or 0)} | parse_real_total={int(source_cache_stats.get('parse_real_total',0) or 0)} | parse_reused_total={int(source_cache_stats.get('parse_reused_total',0) or 0)} | cache_store_total={int(source_cache_stats.get('cache_store_total',0) or 0)} | cache={_u130_parse_cache_path(output_path)}"))
    else:
        rows.append(make_log_row(output_path, 'SOURCE_PARSE_CACHE_STATS', 'INFO',
                        f"arr_hit={int(source_cache_stats.get('arr_hit',0) or 0)} | arr_miss={int(source_cache_stats.get('arr_miss',0) or 0)} | bundle_hit={int(source_cache_stats.get('bundle_hit',0) or 0)} | bundle_miss={int(source_cache_stats.get('bundle_miss',0) or 0)} | detail_hit={int(source_cache_stats.get('detail_hit',0) or 0)} | detail_miss={int(source_cache_stats.get('detail_miss',0) or 0)} | cache={_u90_source_cache_path(output_path)}"))
    return rows


class Processor(Processor):

    def _u104_scan_map(self, folder_path: str, output_path: str) -> Dict[str, str]:
        out: Dict[str, str] = {}
        try:
            raw_files = collect_source_files(folder_path, output_path)
        except Exception:
            raw_files = []
        out_path = _u104_canonical_path(output_path)
        for fp in raw_files or []:
            np = _u104_canonical_path(fp)
            if not np or np == out_path:
                continue
            if np not in out:
                out[np] = fp
        self._current_scanned_files = set(out.keys())
        return out

    def load_fingerprint_cache(self, output_path: str):
        raw = _u102_load_fingerprint_cache(output_path)
        self._fingerprint_cache = {}
        for k, v in raw.items():
            nk = _u104_canonical_path(k)
            if nk and isinstance(v, str):
                self._fingerprint_cache[nk] = v
        self._pending_fingerprint_updates = {}
        self._last_fingerprint_skipped = 0
        self._last_fingerprint_changed = 0
        self._last_supported_scanned = 0
        self._fp_cache_loaded = len(self._fingerprint_cache)
        self._fp_cache_saved = 0
        self._sample_fp_skipped = []
        self._sample_fp_changed = []
        try:
            self.logger.log('FP_CACHE_STATUS_U104', 'INFO',
                            f"loaded={self._fp_cache_loaded} | path={_u102_fingerprint_cache_path(output_path)}")
        except Exception:
            pass

    def persist_caches(self, output_path: str, success_changed_paths: Optional[Set[str]] = None):
        try:
            success_changed_paths = set(success_changed_paths or set())
            for p in success_changed_paths:
                np = _u104_canonical_path(p)
                new_fp = self._pending_fingerprint_updates.get(np)
                if new_fp:
                    self._fingerprint_cache[np] = new_fp
            _u102_save_fingerprint_cache(output_path, self._fingerprint_cache)
            self._fp_cache_saved = len(self._fingerprint_cache)
            try:
                self.logger.log('FP_CACHE_STATUS_U104', 'INFO',
                                f"saved={self._fp_cache_saved} | path={_u102_fingerprint_cache_path(output_path)}")
            except Exception:
                pass
        except Exception as exc:
            _rt_log_exception('u104_persist_fingerprint_cache_failed', exc)
        try:
            cache = _u90_get_source_cache()
            _u90_save_source_cache(output_path, cache)
        except Exception as exc:
            _rt_log_exception('u104_persist_source_cache_failed', exc)

    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        repair_options = repair_options or RepairOptions()
        t_all_0 = time.perf_counter()
        _U90_SOURCE_CACHE_CTX['output_path'] = output_path
        _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
        cache = _u90_get_source_cache()
        cache['stats'] = _u90_empty_source_cache()['stats']

        self.load_fingerprint_cache(output_path)

        scan_map = self._u104_scan_map(folder_path, output_path)
        scanned_keys = set(scan_map.keys())
        self._last_supported_scanned = len(scanned_keys)

        wb_out = open_or_create_output_workbook(output_path)
        workbook_saved = 0
        success_changed_paths: Set[str] = set()
        stats_map: Dict[str, Any] = {
            'scanned': len(scanned_keys),
            'candidate': 0,
            'ok_skip': 0,
            'parsed': 0,
            'repaired': 0,
            'sub_insert': 0,
            'sub_update': 0,
            'inv_deleted': 0,
            'inv_appended': 0,
            'pl_deleted': 0,
            'pl_appended': 0,
            'fp_changed': 0,
            'fp_skipped': 0,
            'workbook_saved': 0,
            'parse_reused': 0,
            'parse_fresh': 0,
        }

        try:
            ws_out = ensure_sheet(wb_out, 'SUB_DETAIL')
            ws_log = ensure_sheet(wb_out, 'LOG_SUB_DETAIL')
            ws_inv = ensure_sheet(wb_out, 'INV')
            ws_pl = ensure_sheet(wb_out, 'PL')
            if 'Sheet' in wb_out.sheetnames and wb_out['Sheet'].max_row == 1 and wb_out['Sheet'].max_column == 1 and nz_str(wb_out['Sheet']['A1'].value) == '':
                try:
                    del wb_out['Sheet']
                except Exception:
                    pass

            init_sub_detail_header(ws_out)
            init_log_header(ws_log)
            init_inv_detail_header(ws_inv)
            init_pl_detail_header(ws_pl)

            sub_pathkey_col = _u104_ensure_pathkey_column(ws_out, SUB_DETAIL_HEADERS, HEADER_FILL_SUB)
            inv_pathkey_col = _u104_ensure_pathkey_column(ws_inv, INV_DETAIL_HEADERS, HEADER_FILL_SUB)
            pl_pathkey_col = _u104_ensure_pathkey_column(ws_pl, PL_DETAIL_HEADERS, HEADER_FILL_SUB)

            sub_state, sub_dup_rows, sub_backfill = _u104_collect_subdetail_state(ws_out, sub_pathkey_col)
            if sub_dup_rows:
                _u104_delete_rows_desc(ws_out, sub_dup_rows)
                sub_state, _, sub_backfill2 = _u104_collect_subdetail_state(ws_out, sub_pathkey_col)
                sub_backfill += sub_backfill2

            inv_rownums_by_key, inv_backfill = _u104_collect_detail_rownums_by_pathkey(ws_inv, 21, inv_pathkey_col)
            pl_rownums_by_key, pl_backfill = _u104_collect_detail_rownums_by_pathkey(ws_pl, 17, pl_pathkey_col)

            workbook_dirty = bool(sub_backfill or inv_backfill or pl_backfill or sub_dup_rows)

            cnt_total = len(scanned_keys)
            cnt_ok = 0
            cnt_warn = 0
            cnt_err = 0
            cnt_skip_path = 0
            cnt_skip_sig = 0
            cnt_repaired = 0
            issue_files: List[Tuple[str, str, str]] = []
            log_rows: List[List[Any]] = [make_log_row('', 'START', 'INFO', f'Bat dau quet {cnt_total} file.')]

            candidate_actual_paths: List[str] = []
            candidate_pending_fp: Dict[str, str] = {}
            self._sample_fp_skipped = []
            self._sample_fp_changed = []

            for pathkey, actual_path in sorted(scan_map.items()):
                fp_now = _compute_file_fingerprint(actual_path)
                old_fp = self._fingerprint_cache.get(pathkey, '')
                old_status = nz_str((sub_state.get(pathkey) or {}).get('status', '')).upper()

                if old_status == 'OK' and fp_now and old_fp and old_fp == fp_now:
                    cnt_skip_path += 1
                    stats_map['ok_skip'] += 1
                    self._last_fingerprint_skipped += 1
                    if len(self._sample_fp_skipped) < 3:
                        self._sample_fp_skipped.append(pathkey)
                    continue

                candidate_actual_paths.append(actual_path)
                candidate_pending_fp[pathkey] = fp_now or ''
                if len(self._sample_fp_changed) < 3:
                    self._sample_fp_changed.append(pathkey)

            stats_map['candidate'] = len(candidate_actual_paths)
            stats_map['fp_changed'] = len(candidate_actual_paths)
            stats_map['fp_skipped'] = cnt_skip_path
            self._last_fingerprint_changed = len(candidate_actual_paths)
            self._pending_fingerprint_updates = dict(candidate_pending_fp)

            parsed_payloads: Dict[str, Dict[str, Any]] = {}

            if candidate_actual_paths:
                def job(path: str):
                    arr, inv_rows, pl_rows = process_one_sub_file_bundle(path)
                    return path, arr, inv_rows, pl_rows

                futures = {}
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                    for fp in candidate_actual_paths:
                        futures[executor.submit(job, fp)] = fp

                    completed = 0
                    total_jobs = max(1, len(futures))
                    for future in as_completed(futures):
                        file_path, arr, inv_rows_one, pl_rows_one = future.result()
                        completed += 1

                        file_path_can = _u104_canonical_path(file_path)
                        arr = list(arr[:len(SUB_DETAIL_HEADERS)]) + [''] * max(0, len(SUB_DETAIL_HEADERS) - len(arr))
                        arr[23] = file_path_can

                        status_text = nz_str(arr[21]).upper()
                        err_text = nz_str(arr[22])

                        if status_text in ('WARNING', 'ERROR'):
                            issue_files.append((file_path_can, status_text, err_text))
                            log_rows.append(make_log_row(file_path_can, 'CHECK', status_text, err_text))
                            self.logger.log(f"{status_text}: {get_file_name_from_path(file_path_can)} -> {err_text}", 'warn' if status_text == 'WARNING' else 'error')

                            if repair_options.use_folder_truth or repair_options.use_manual_values:
                                repaired, detail = repair_excel_metadata(file_path, arr, repair_options)
                                if repaired:
                                    cnt_repaired += 1
                                    stats_map['repaired'] += 1
                                    log_rows.append(make_log_row(file_path_can, 'REPAIR', 'OK', detail))
                                    arr2, inv_rows_one, pl_rows_one = process_one_sub_file_bundle(file_path)
                                    arr = list(arr2[:len(SUB_DETAIL_HEADERS)]) + [''] * max(0, len(SUB_DETAIL_HEADERS) - len(arr2))
                                    arr[23] = file_path_can
                                    status_text = nz_str(arr[21]).upper()
                                    err_text = nz_str(arr[22])
                                else:
                                    log_rows.append(make_log_row(file_path_can, 'REPAIR', 'SKIP', detail))

                        if status_text == 'OK':
                            cnt_ok += 1
                        elif status_text == 'WARNING':
                            cnt_warn += 1
                        else:
                            cnt_err += 1

                        inv_rows_one = _u104_prepare_inv_rows(inv_rows_one)
                        pl_rows_one = _u104_prepare_pl_rows(pl_rows_one)
                        for row in inv_rows_one:
                            row[18] = status_text
                            row[20] = file_path_can
                        for row in pl_rows_one:
                            row[14] = status_text
                            row[16] = file_path_can

                        parsed_payloads[file_path_can] = {
                            'arr': arr,
                            'status': status_text,
                            'error': err_text,
                            'inv_rows': inv_rows_one,
                            'pl_rows': pl_rows_one,
                        }
                        success_changed_paths.add(file_path_can)
                        stats_map['parsed'] += 1

                        if progress_callback:
                            progress_callback(completed, total_jobs, file_path_can)

            # SUB_DETAIL targeted upsert
            for pathkey, payload in parsed_payloads.items():
    # --- KHỞI TẠO BIẾN ĐỂ TRÁNH LỖI VÀ TỐI ƯU TỐC ĐỘ ---
                rewrite_payloads = {} 

            for pathkey, payload in parsed_payloads.items():
                arr = payload['arr']
                existing = sub_state.get(pathkey)
                
                # Nếu file đã tồn tại và dữ liệu không đổi -> Bỏ qua hoàn toàn (Tối ưu tốc độ)
                if existing and _u104_existing_row_equals_subdetail(ws_out, existing['row_idx'], arr):
                    continue
                
                if existing:
                    # Trường hợp UPDATE: Cập nhật dòng tổng hợp
                    _u104_write_subdetail_row(ws_out, existing['row_idx'], arr, sub_pathkey_col)
                    stats_map['sub_update'] += 1
                    workbook_dirty = True
                    # Đánh dấu file này cần xóa dòng cũ và ghi lại Detail mới
                    rewrite_payloads[pathkey] = payload 
                else:
                    # Trường hợp INSERT: Thêm dòng tổng hợp mới
                    row_idx = _u104_append_subdetail_row(ws_out, arr, sub_pathkey_col)
                    sub_state[pathkey] = {'row_idx': row_idx, 'status': nz_str(arr[21]).upper()}
                    stats_map['sub_insert'] += 1
                    workbook_dirty = True
                    # Đánh dấu file mới này cần ghi Detail
                    rewrite_payloads[pathkey] = payload

            # Chỉ tác động vào những file thực sự thay đổi (giảm thiểu thao tác delete_rows chậm chạp)
            affected_keys = set(rewrite_payloads.keys())
            rewrite_keys = affected_keys
            inv_delete_rows: List[int] = []
            pl_delete_rows: List[int] = []
            for pathkey in affected_keys:
                inv_delete_rows.extend(inv_rownums_by_key.get(pathkey, []))
                pl_delete_rows.extend(pl_rownums_by_key.get(pathkey, []))

            if inv_delete_rows:
                stats_map['inv_deleted'] = _u104_delete_rows_desc(ws_inv, inv_delete_rows)
                workbook_dirty = True
            if pl_delete_rows:
                stats_map['pl_deleted'] = _u104_delete_rows_desc(ws_pl, pl_delete_rows)
                workbook_dirty = True

            inv_rows_to_append: List[List[Any]] = []
            pl_rows_to_append: List[List[Any]] = []
            for pathkey in sorted(rewrite_keys):
                payload = parsed_payloads[pathkey]
                if payload['status'] in ('OK', 'WARNING'):
                    inv_rows_to_append.extend(payload['inv_rows'])
                    pl_rows_to_append.extend(payload['pl_rows'])

            if inv_rows_to_append:
                stats_map['inv_appended'] = _u104_append_inv_rows(ws_inv, inv_rows_to_append, inv_pathkey_col)
                workbook_dirty = True
            if pl_rows_to_append:
                stats_map['pl_appended'] = _u104_append_pl_rows(ws_pl, pl_rows_to_append, pl_pathkey_col)
                workbook_dirty = True

            if issue_files:
                log_rows.append(make_log_row('', 'SUMMARY', 'INFO', 'TONG HOP FILE WARNING/ERROR'))
                for fp, st, detail in issue_files:
                    log_rows.append(make_log_row(fp, 'SUMMARY', st, detail))

            if cnt_skip_path > 0:
                log_rows.append(make_log_row('', 'SKIP', 'INFO', f'So luong bo qua do da co File Path OK va khong doi: {cnt_skip_path}'))
            log_rows.append(make_log_row('', 'DONE', 'INFO',
                                         f'Tong={cnt_total}; OK={cnt_ok}; WARNING={cnt_warn}; ERROR={cnt_err}; SkipPath={cnt_skip_path}; SkipSignature={cnt_skip_sig}; Repaired={cnt_repaired}'))

            core_seconds = max(0.0, time.perf_counter() - t_all_0)
            source_cache_stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
            stats_map['workbook_saved'] = 1 if workbook_dirty else 0
            bench_rows = _u104_build_runtime_bench_rows(folder_path, output_path, core_seconds, max(0.0, time.perf_counter() - t_all_0), stats_map, source_cache_stats)

            if workbook_dirty:
                append_log_rows(ws_log, log_rows + bench_rows)
                safe_save_workbook_atomic(wb_out, output_path)
                workbook_saved = 1
            else:
                # zero-work: khong ghi them log vao workbook de tranh save lai toan file
                self.logger.log(f'ZERO_WORK_U104: scanned={cnt_total} | skip_ok={cnt_skip_path} | candidate=0', 'info')

            stats_map['workbook_saved'] = workbook_saved

            self.persist_caches(output_path, success_changed_paths=success_changed_paths)
            try:
                final_substate = {pk: nz_str((meta or {}).get('status', '')).upper() for pk, meta in sub_state.items()}
                for pk, payload in parsed_payloads.items():
                    final_substate[_u104_canonical_path(pk)] = nz_str((payload or {}).get('status', '')).upper()
                _u134_save_substate(output_path, final_substate)
            except Exception as exc:
                try:
                    self.logger.log(f'SUBSTATE_SAVE_ERROR: {exc}', 'warn')
                except Exception:
                    pass

            return {
                'total': cnt_total,
                'ok': cnt_ok,
                'warning': cnt_warn,
                'error': cnt_err,
                'skip_path': cnt_skip_path,
                'skip_signature': cnt_skip_sig,
                'repaired': cnt_repaired,
                'issues': issue_files,
                'scanned_files': sorted(scan_map.values()),
            }
        finally:
            wb_out.close()
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None



# =========================================================
# U105 RESTORE S.* COLUMNS + TOTAL W HANDLING UPCHARGE
# - Giu delta engine cua U104
# - Khoi phuc INV/PL theo schema V73_INV_HEADERS / V76_PL_HEADERS
# - Dung lai helper goc:
#     _v73_apply_invoice_extras
#     _v76_apply_pl_sequence
#     _v74_seq_seed_from_rows
# - Dam bao cac cot:
#     INV: S. Invoice#, S.DateWInv#, S.MODEL, S.MODEL Y, S.COLOR,
#          S.WIDTH, S.Q'TY (Pairs), S.UNIT PRICE (USD), TOTAL W HANDLING UPCHARGE
#     PL : S. Invoice#, S.DateWInv#
# =========================================================

U105_INV_HEADERS = list(V73_INV_HEADERS) + [U104_PATHKEY_HEADER]
U105_PL_HEADERS = list(V76_PL_HEADERS) + [U104_PATHKEY_HEADER]


def _u105_snapshot_rows(ws, headers: List[str]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    last_row = _v74_last_used_row(ws)
    for r in range(2, last_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        if all(nz_str(v) == '' for v in vals):
            continue
        rows.append(vals)
    return rows


def _u105_merge_seed_map(base: Dict[Tuple[str, str], int], extra: Dict[Tuple[str, str], int]) -> Dict[Tuple[str, str], int]:
    out = dict(base or {})
    for k, v in (extra or {}).items():
        try:
            iv = int(v)
        except Exception:
            continue
        if iv > out.get(k, 0):
            out[k] = iv
    return out


def _u105_prepare_inv_compact(rows: List[List[Any]], status_text: str, file_path_can: str) -> List[List[Any]]:
    out: List[List[Any]] = []
    for arr in rows or []:
        row = list(arr[:21]) + [''] * max(0, 21 - len(arr))
        row[18] = status_text
        row[20] = file_path_can
        out.append(row)
    return out


def _u105_prepare_pl_compact(rows: List[List[Any]], status_text: str, file_path_can: str) -> List[List[Any]]:
    out: List[List[Any]] = []
    for arr in rows or []:
        row = list(arr[:17]) + [''] * max(0, 17 - len(arr))
        row[14] = status_text
        row[16] = file_path_can
        out.append(row)
    return out


def _u105_append_inv_rows(ws, rows: List[List[Any]], pathkey_col: int) -> int:
    if not rows:
        return 0
    rows2: List[List[Any]] = []
    for arr in rows:
        row = list(arr[:len(V73_INV_HEADERS)]) + [''] * max(0, len(V73_INV_HEADERS) - len(arr))
        pathkey = _u104_canonical_path(row[29]) if len(row) > 29 else ''
        row.append(pathkey)
        rows2.append(row)
    return _v74_append_rows_exact(ws, rows2, U105_INV_HEADERS, [1], 27)


def _u105_append_pl_rows(ws, rows: List[List[Any]], pathkey_col: int) -> int:
    if not rows:
        return 0
    rows2: List[List[Any]] = []
    for arr in rows:
        row = list(arr[:len(V76_PL_HEADERS)]) + [''] * max(0, len(V76_PL_HEADERS) - len(arr))
        pathkey = _u104_canonical_path(row[18]) if len(row) > 18 else ''
        row.append(pathkey)
        rows2.append(row)
    return _v74_append_rows_exact(ws, rows2, U105_PL_HEADERS, [1], 16)


class Processor(Processor):

    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        repair_options = repair_options or RepairOptions()
        t_all_0 = time.perf_counter()
        _U90_SOURCE_CACHE_CTX['output_path'] = output_path
        _U90_SOURCE_CACHE_CTX['cache'] = _u90_load_source_cache(output_path)
        cache = _u90_get_source_cache()
        cache['stats'] = _u90_empty_source_cache()['stats']

        self.load_fingerprint_cache(output_path)
        scan_map = self._u104_scan_map(folder_path, output_path)
        scanned_keys = set(scan_map.keys())
        self._last_supported_scanned = len(scanned_keys)

        wb_out = open_or_create_output_workbook(output_path)
        workbook_saved = 0
        success_changed_paths: Set[str] = set()
        stats_map: Dict[str, Any] = {
            'scanned': len(scanned_keys),
            'candidate': 0,
            'ok_skip': 0,
            'parsed': 0,
            'repaired': 0,
            'sub_insert': 0,
            'sub_update': 0,
            'inv_deleted': 0,
            'inv_appended': 0,
            'pl_deleted': 0,
            'pl_appended': 0,
            'fp_changed': 0,
            'fp_skipped': 0,
            'workbook_saved': 0,
        }

        try:
            ws_out = ensure_sheet(wb_out, 'SUB_DETAIL')
            ws_log = ensure_sheet(wb_out, 'LOG_SUB_DETAIL')
            ws_inv = ensure_sheet(wb_out, 'INV')
            ws_pl = ensure_sheet(wb_out, 'PL')
            if 'Sheet' in wb_out.sheetnames and wb_out['Sheet'].max_row == 1 and wb_out['Sheet'].max_column == 1 and nz_str(wb_out['Sheet']['A1'].value) == '':
                try:
                    del wb_out['Sheet']
                except Exception:
                    pass

            init_sub_detail_header(ws_out)
            init_log_header(ws_log)
            _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
            _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

            sub_pathkey_col = _u104_ensure_pathkey_column(ws_out, SUB_DETAIL_HEADERS, HEADER_FILL_SUB)
            inv_pathkey_col = _u104_ensure_pathkey_column(ws_inv, V73_INV_HEADERS, HEADER_FILL_SUB)
            pl_pathkey_col = _u104_ensure_pathkey_column(ws_pl, V76_PL_HEADERS, HEADER_FILL_SUB)

            sub_state, sub_dup_rows, sub_backfill = _u104_collect_subdetail_state(ws_out, sub_pathkey_col)
            if sub_dup_rows:
                _u104_delete_rows_desc(ws_out, sub_dup_rows)
                sub_state, _, sub_backfill2 = _u104_collect_subdetail_state(ws_out, sub_pathkey_col)
                sub_backfill += sub_backfill2

            inv_rownums_by_key, inv_backfill = _u104_collect_detail_rownums_by_pathkey(ws_inv, 30, inv_pathkey_col)
            pl_rownums_by_key, pl_backfill = _u104_collect_detail_rownums_by_pathkey(ws_pl, 19, pl_pathkey_col)

            workbook_dirty = bool(sub_backfill or inv_backfill or pl_backfill or sub_dup_rows)

            cnt_total = len(scanned_keys)
            cnt_ok = 0
            cnt_warn = 0
            cnt_err = 0
            cnt_skip_path = 0
            cnt_skip_sig = 0
            cnt_repaired = 0
            issue_files: List[Tuple[str, str, str]] = []
            log_rows: List[List[Any]] = [make_log_row('', 'START', 'INFO', f'Bat dau quet {cnt_total} file.')]

            candidate_actual_paths: List[str] = []
            candidate_pending_fp: Dict[str, str] = {}
            self._sample_fp_skipped = []
            self._sample_fp_changed = []

            for pathkey, actual_path in sorted(scan_map.items()):
                fp_now = _compute_file_fingerprint(actual_path)
                old_fp = self._fingerprint_cache.get(pathkey, '')
                old_status = nz_str((sub_state.get(pathkey) or {}).get('status', '')).upper()
                if old_status == 'OK' and fp_now and old_fp and old_fp == fp_now:
                    cnt_skip_path += 1
                    stats_map['ok_skip'] += 1
                    self._last_fingerprint_skipped += 1
                    if len(self._sample_fp_skipped) < 3:
                        self._sample_fp_skipped.append(pathkey)
                    continue
                candidate_actual_paths.append(actual_path)
                candidate_pending_fp[pathkey] = fp_now or ''
                if len(self._sample_fp_changed) < 3:
                    self._sample_fp_changed.append(pathkey)

            stats_map['candidate'] = len(candidate_actual_paths)
            stats_map['fp_changed'] = len(candidate_actual_paths)
            stats_map['fp_skipped'] = cnt_skip_path
            self._last_fingerprint_changed = len(candidate_actual_paths)
            self._pending_fingerprint_updates = dict(candidate_pending_fp)

            parsed_payloads: Dict[str, Dict[str, Any]] = {}

            if candidate_actual_paths:
                def job(path: str):
                    arr, inv_rows, pl_rows = process_one_sub_file_bundle(path)
                    return path, arr, inv_rows, pl_rows

                futures = {}
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                    for fp in candidate_actual_paths:
                        futures[executor.submit(job, fp)] = fp

                    completed = 0
                    total_jobs = max(1, len(futures))
                    for future in as_completed(futures):
                        file_path, arr, inv_rows_one, pl_rows_one = future.result()
                        completed += 1

                        file_path_can = _u104_canonical_path(file_path)
                        arr = list(arr[:len(SUB_DETAIL_HEADERS)]) + [''] * max(0, len(SUB_DETAIL_HEADERS) - len(arr))
                        arr[23] = file_path_can
                        status_text = nz_str(arr[21]).upper()
                        err_text = nz_str(arr[22])

                        if status_text in ('WARNING', 'ERROR'):
                            issue_files.append((file_path_can, status_text, err_text))
                            log_rows.append(make_log_row(file_path_can, 'CHECK', status_text, err_text))
                            self.logger.log(f"{status_text}: {get_file_name_from_path(file_path_can)} -> {err_text}", 'warn' if status_text == 'WARNING' else 'error')
                            if repair_options.use_folder_truth or repair_options.use_manual_values:
                                repaired, detail = repair_excel_metadata(file_path, arr, repair_options)
                                if repaired:
                                    cnt_repaired += 1
                                    stats_map['repaired'] += 1
                                    log_rows.append(make_log_row(file_path_can, 'REPAIR', 'OK', detail))
                                    arr2, inv_rows_one, pl_rows_one = process_one_sub_file_bundle(file_path)
                                    arr = list(arr2[:len(SUB_DETAIL_HEADERS)]) + [''] * max(0, len(SUB_DETAIL_HEADERS) - len(arr2))
                                    arr[23] = file_path_can
                                    status_text = nz_str(arr[21]).upper()
                                    err_text = nz_str(arr[22])
                                else:
                                    log_rows.append(make_log_row(file_path_can, 'REPAIR', 'SKIP', detail))

                        if status_text == 'OK':
                            cnt_ok += 1
                        elif status_text == 'WARNING':
                            cnt_warn += 1
                        else:
                            cnt_err += 1

                        parsed_payloads[file_path_can] = {
                            'arr': arr,
                            'status': status_text,
                            'error': err_text,
                            'inv_rows_compact': _u105_prepare_inv_compact(inv_rows_one, status_text, file_path_can),
                            'pl_rows_compact': _u105_prepare_pl_compact(pl_rows_one, status_text, file_path_can),
                        }
                        success_changed_paths.add(file_path_can)
                        stats_map['parsed'] += 1
                        if progress_callback:
                            progress_callback(completed, total_jobs, file_path_can)

            for pathkey, payload in parsed_payloads.items():
                arr = payload['arr']
                existing = sub_state.get(pathkey)
                if existing and _u104_existing_row_equals_subdetail(ws_out, existing['row_idx'], arr, sub_pathkey_col):
                    continue
                if existing:
                    _u104_write_subdetail_row(ws_out, existing['row_idx'], arr, sub_pathkey_col)
                    stats_map['sub_update'] += 1
                    workbook_dirty = True
                else:
                    row_idx = _u104_append_subdetail_row(ws_out, arr, sub_pathkey_col)
                    sub_state[pathkey] = {'row_idx': row_idx, 'status': nz_str(arr[21]).upper(), 'signature': nz_str(arr[24])}
                    stats_map['sub_insert'] += 1
                    workbook_dirty = True

            affected_keys = set(parsed_payloads.keys())
            inv_delete_rows: List[int] = []
            pl_delete_rows: List[int] = []
            for pathkey in affected_keys:
                inv_delete_rows.extend(inv_rownums_by_key.get(pathkey, []))
                pl_delete_rows.extend(pl_rownums_by_key.get(pathkey, []))

            if inv_delete_rows:
                stats_map['inv_deleted'] = _u104_delete_rows_desc(ws_inv, inv_delete_rows)
                workbook_dirty = True
            if pl_delete_rows:
                stats_map['pl_deleted'] = _u104_delete_rows_desc(ws_pl, pl_delete_rows)
                workbook_dirty = True

            inv_existing_rows = _u105_snapshot_rows(ws_inv, V73_INV_HEADERS)
            pl_existing_rows = _u105_snapshot_rows(ws_pl, V76_PL_HEADERS)
            inv_seed = _v74_seq_seed_from_rows(inv_existing_rows, V73_INV_HEADERS, 'S. Invoice#')
            pl_seed = _v74_seq_seed_from_rows(pl_existing_rows, V76_PL_HEADERS, 'S. Invoice#')

            inv_rows_to_append: List[List[Any]] = []
            pl_rows_to_append: List[List[Any]] = []
            for pathkey in sorted(affected_keys):
                payload = parsed_payloads[pathkey]
                if payload['status'] in ('OK', 'WARNING'):
                    inv_expanded = _v73_apply_invoice_extras(payload['inv_rows_compact'], existing_seeds=inv_seed)
                    inv_seed = _u105_merge_seed_map(inv_seed, _v74_seq_seed_from_rows(inv_expanded, V73_INV_HEADERS, 'S. Invoice#'))
                    pl_expanded = _v76_apply_pl_sequence(payload['pl_rows_compact'], existing_seeds=pl_seed)
                    pl_seed = _u105_merge_seed_map(pl_seed, _v74_seq_seed_from_rows(pl_expanded, V76_PL_HEADERS, 'S. Invoice#'))
                    inv_rows_to_append.extend(inv_expanded)
                    pl_rows_to_append.extend(pl_expanded)

            if inv_rows_to_append:
                stats_map['inv_appended'] = _u105_append_inv_rows(ws_inv, inv_rows_to_append, inv_pathkey_col)
                workbook_dirty = True
            if pl_rows_to_append:
                stats_map['pl_appended'] = _u105_append_pl_rows(ws_pl, pl_rows_to_append, pl_pathkey_col)
                workbook_dirty = True

            if issue_files:
                log_rows.append(make_log_row('', 'SUMMARY', 'INFO', 'TONG HOP FILE WARNING/ERROR'))
                for fp, st, detail in issue_files:
                    log_rows.append(make_log_row(fp, 'SUMMARY', st, detail))
            if cnt_skip_path > 0:
                log_rows.append(make_log_row('', 'SKIP', 'INFO', f'So luong bo qua do da co File Path OK va khong doi: {cnt_skip_path}'))
            log_rows.append(make_log_row('', 'DONE', 'INFO',
                                         f'Tong={cnt_total}; OK={cnt_ok}; WARNING={cnt_warn}; ERROR={cnt_err}; SkipPath={cnt_skip_path}; SkipSignature={cnt_skip_sig}; Repaired={cnt_repaired}'))

            core_seconds = max(0.0, time.perf_counter() - t_all_0)
            source_cache_stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
            stats_map['workbook_saved'] = 1 if workbook_dirty else 0
            bench_rows = _u104_build_runtime_bench_rows(folder_path, output_path, core_seconds, max(0.0, time.perf_counter() - t_all_0), stats_map, source_cache_stats)

            if workbook_dirty:
                append_log_rows(ws_log, log_rows + bench_rows)
                safe_save_workbook_atomic(wb_out, output_path)
                workbook_saved = 1
            else:
                self.logger.log(f'ZERO_WORK_U105: scanned={cnt_total} | skip_ok={cnt_skip_path} | candidate=0', 'info')

            stats_map['workbook_saved'] = workbook_saved
            self.persist_caches(output_path, success_changed_paths=success_changed_paths)

            return {
                'total': cnt_total,
                'ok': cnt_ok,
                'warning': cnt_warn,
                'error': cnt_err,
                'skip_path': cnt_skip_path,
                'skip_signature': cnt_skip_sig,
                'repaired': cnt_repaired,
                'issues': issue_files,
                'scanned_files': sorted(scan_map.values()),
            }
        finally:
            wb_out.close()
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None


# =========================================================
# U106 CACHE RESET + ONE-TIME FULL RESYNC
# Muc tieu:
# - Xoa tac dong cua stale core_bundle/source cache
# - Chay resync 1 lan cho workbook chua co marker U106
# - Sau lan do quay lai delta theo fingerprint + status
# =========================================================

U106_ENGINE_VERSION = 'U106'
U106_META_SHEET = '_ENGINE_META'


def _u106_get_engine_version(wb) -> str:
    try:
        if U106_META_SHEET not in wb.sheetnames:
            return ''
        ws = wb[U106_META_SHEET]
        return nz_str(ws['B1'].value)
    except Exception:
        return ''


def _u106_set_engine_version(wb, version_text: str):
    try:
        ws = ensure_sheet(wb, U106_META_SHEET)
        ws.sheet_state = 'hidden'
        ws['A1'] = 'DELTA_ENGINE_VERSION'
        ws['B1'] = version_text
    except Exception:
        pass


class Processor(Processor):

    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        repair_options = repair_options or RepairOptions()
        t_all_0 = time.perf_counter()
        _U90_SOURCE_CACHE_CTX['output_path'] = output_path
        # Reset source cache de tranh dung stale core_bundle cua cac ban patch truoc
        _U90_SOURCE_CACHE_CTX['cache'] = _u90_empty_source_cache()
        cache = _u90_get_source_cache()
        cache['stats'] = _u90_empty_source_cache()['stats']

        self.load_fingerprint_cache(output_path)
        scan_map = self._u104_scan_map(folder_path, output_path)
        scanned_keys = set(scan_map.keys())
        self._last_supported_scanned = len(scanned_keys)

        wb_out = open_or_create_output_workbook(output_path)
        workbook_saved = 0
        success_changed_paths: Set[str] = set()
        stats_map: Dict[str, Any] = {
            'scanned': len(scanned_keys),
            'candidate': 0,
            'ok_skip': 0,
            'parsed': 0,
            'repaired': 0,
            'sub_insert': 0,
            'sub_update': 0,
            'inv_deleted': 0,
            'inv_appended': 0,
            'pl_deleted': 0,
            'pl_appended': 0,
            'fp_changed': 0,
            'fp_skipped': 0,
            'workbook_saved': 0,
        }

        try:
            force_full_resync = (_u106_get_engine_version(wb_out) != U106_ENGINE_VERSION)

            ws_out = ensure_sheet(wb_out, 'SUB_DETAIL')
            ws_log = ensure_sheet(wb_out, 'LOG_SUB_DETAIL')
            ws_inv = ensure_sheet(wb_out, 'INV')
            ws_pl = ensure_sheet(wb_out, 'PL')
            if 'Sheet' in wb_out.sheetnames and wb_out['Sheet'].max_row == 1 and wb_out['Sheet'].max_column == 1 and nz_str(wb_out['Sheet']['A1'].value) == '':
                try:
                    del wb_out['Sheet']
                except Exception:
                    pass

            init_sub_detail_header(ws_out)
            init_log_header(ws_log)
            _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
            _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

            sub_pathkey_col = _u104_ensure_pathkey_column(ws_out, SUB_DETAIL_HEADERS, HEADER_FILL_SUB)
            inv_pathkey_col = _u104_ensure_pathkey_column(ws_inv, V73_INV_HEADERS, HEADER_FILL_SUB)
            pl_pathkey_col = _u104_ensure_pathkey_column(ws_pl, V76_PL_HEADERS, HEADER_FILL_SUB)

            sub_state, sub_dup_rows, sub_backfill = _u104_collect_subdetail_state(ws_out, sub_pathkey_col)
            if sub_dup_rows:
                _u104_delete_rows_desc(ws_out, sub_dup_rows)
                sub_state, _, sub_backfill2 = _u104_collect_subdetail_state(ws_out, sub_pathkey_col)
                sub_backfill += sub_backfill2

            inv_rownums_by_key, inv_backfill = _u104_collect_detail_rownums_by_pathkey(ws_inv, 30, inv_pathkey_col)
            pl_rownums_by_key, pl_backfill = _u104_collect_detail_rownums_by_pathkey(ws_pl, 19, pl_pathkey_col)

            workbook_dirty = bool(sub_backfill or inv_backfill or pl_backfill or sub_dup_rows or force_full_resync)

            cnt_total = len(scanned_keys)
            cnt_ok = 0
            cnt_warn = 0
            cnt_err = 0
            cnt_skip_path = 0
            cnt_skip_sig = 0
            cnt_repaired = 0
            issue_files: List[Tuple[str, str, str]] = []
            log_rows: List[List[Any]] = [make_log_row('', 'START', 'INFO', f'Bat dau quet {cnt_total} file.')]
            if force_full_resync:
                log_rows.append(make_log_row('', 'RESYNC', 'INFO', 'Force full resync 1 lan de lam sach source cache cu va dong bo lai INV/PL.'))

            candidate_actual_paths: List[str] = []
            candidate_pending_fp: Dict[str, str] = {}
            self._sample_fp_skipped = []
            self._sample_fp_changed = []

            for pathkey, actual_path in sorted(scan_map.items()):
                fp_now = _compute_file_fingerprint(actual_path)
                old_fp = self._fingerprint_cache.get(pathkey, '')
                old_status = nz_str((sub_state.get(pathkey) or {}).get('status', '')).upper()
                if (not force_full_resync) and old_status == 'OK' and fp_now and old_fp and old_fp == fp_now:
                    cnt_skip_path += 1
                    stats_map['ok_skip'] += 1
                    self._last_fingerprint_skipped += 1
                    if len(self._sample_fp_skipped) < 3:
                        self._sample_fp_skipped.append(pathkey)
                    continue
                candidate_actual_paths.append(actual_path)
                candidate_pending_fp[pathkey] = fp_now or ''
                if len(self._sample_fp_changed) < 3:
                    self._sample_fp_changed.append(pathkey)

            stats_map['candidate'] = len(candidate_actual_paths)
            stats_map['fp_changed'] = len(candidate_actual_paths)
            stats_map['fp_skipped'] = cnt_skip_path
            self._last_fingerprint_changed = len(candidate_actual_paths)
            self._pending_fingerprint_updates = dict(candidate_pending_fp)

            parsed_payloads: Dict[str, Dict[str, Any]] = {}

            if candidate_actual_paths:
                def job(path: str):
                    arr, inv_rows, pl_rows = process_one_sub_file_bundle(path)
                    return path, arr, inv_rows, pl_rows

                futures = {}
                with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
                    for fp in candidate_actual_paths:
                        futures[executor.submit(job, fp)] = fp

                    completed = 0
                    total_jobs = max(1, len(futures))
                    for future in as_completed(futures):
                        file_path, arr, inv_rows_one, pl_rows_one = future.result()
                        completed += 1

                        file_path_can = _u104_canonical_path(file_path)
                        arr = list(arr[:len(SUB_DETAIL_HEADERS)]) + [''] * max(0, len(SUB_DETAIL_HEADERS) - len(arr))
                        arr[23] = file_path_can
                        status_text = nz_str(arr[21]).upper()
                        err_text = nz_str(arr[22])

                        if status_text in ('WARNING', 'ERROR'):
                            issue_files.append((file_path_can, status_text, err_text))
                            log_rows.append(make_log_row(file_path_can, 'CHECK', status_text, err_text))
                            self.logger.log(f"{status_text}: {get_file_name_from_path(file_path_can)} -> {err_text}", 'warn' if status_text == 'WARNING' else 'error')
                            if repair_options.use_folder_truth or repair_options.use_manual_values:
                                repaired, detail = repair_excel_metadata(file_path, arr, repair_options)
                                if repaired:
                                    cnt_repaired += 1
                                    stats_map['repaired'] += 1
                                    log_rows.append(make_log_row(file_path_can, 'REPAIR', 'OK', detail))
                                    arr2, inv_rows_one, pl_rows_one = process_one_sub_file_bundle(file_path)
                                    arr = list(arr2[:len(SUB_DETAIL_HEADERS)]) + [''] * max(0, len(SUB_DETAIL_HEADERS) - len(arr2))
                                    arr[23] = file_path_can
                                    status_text = nz_str(arr[21]).upper()
                                    err_text = nz_str(arr[22])
                                else:
                                    log_rows.append(make_log_row(file_path_can, 'REPAIR', 'SKIP', detail))

                        if status_text == 'OK':
                            cnt_ok += 1
                        elif status_text == 'WARNING':
                            cnt_warn += 1
                        else:
                            cnt_err += 1

                        parsed_payloads[file_path_can] = {
                            'arr': arr,
                            'status': status_text,
                            'error': err_text,
                            'inv_rows_compact': _u105_prepare_inv_compact(inv_rows_one, status_text, file_path_can),
                            'pl_rows_compact': _u105_prepare_pl_compact(pl_rows_one, status_text, file_path_can),
                        }
                        success_changed_paths.add(file_path_can)
                        stats_map['parsed'] += 1
                        if progress_callback:
                            progress_callback(completed, total_jobs, file_path_can)

            for pathkey, payload in parsed_payloads.items():
                arr = payload['arr']
                existing = sub_state.get(pathkey)
                if existing and _u104_existing_row_equals_subdetail(ws_out, existing['row_idx'], arr, sub_pathkey_col):
                    pass
                elif existing:
                    _u104_write_subdetail_row(ws_out, existing['row_idx'], arr, sub_pathkey_col)
                    stats_map['sub_update'] += 1
                    workbook_dirty = True
                else:
                    row_idx = _u104_append_subdetail_row(ws_out, arr, sub_pathkey_col)
                    sub_state[pathkey] = {'row_idx': row_idx, 'status': nz_str(arr[21]).upper(), 'signature': nz_str(arr[24])}
                    stats_map['sub_insert'] += 1
                    workbook_dirty = True

            affected_keys = set(parsed_payloads.keys())
            inv_delete_rows: List[int] = []
            pl_delete_rows: List[int] = []
            for pathkey in affected_keys:
                inv_delete_rows.extend(inv_rownums_by_key.get(pathkey, []))
                pl_delete_rows.extend(pl_rownums_by_key.get(pathkey, []))

            if inv_delete_rows:
                stats_map['inv_deleted'] = _u104_delete_rows_desc(ws_inv, inv_delete_rows)
                workbook_dirty = True
            if pl_delete_rows:
                stats_map['pl_deleted'] = _u104_delete_rows_desc(ws_pl, pl_delete_rows)
                workbook_dirty = True

            inv_existing_rows = _u105_snapshot_rows(ws_inv, V73_INV_HEADERS)
            pl_existing_rows = _u105_snapshot_rows(ws_pl, V76_PL_HEADERS)
            inv_seed = _v74_seq_seed_from_rows(inv_existing_rows, V73_INV_HEADERS, 'S. Invoice#')
            pl_seed = _v74_seq_seed_from_rows(pl_existing_rows, V76_PL_HEADERS, 'S. Invoice#')

            inv_rows_to_append: List[List[Any]] = []
            pl_rows_to_append: List[List[Any]] = []
            for pathkey in sorted(affected_keys):
                payload = parsed_payloads[pathkey]
                if payload['status'] in ('OK', 'WARNING'):
                    inv_expanded = _v73_apply_invoice_extras(payload['inv_rows_compact'], existing_seeds=inv_seed)
                    inv_seed = _u105_merge_seed_map(inv_seed, _v74_seq_seed_from_rows(inv_expanded, V73_INV_HEADERS, 'S. Invoice#'))
                    pl_expanded = _v76_apply_pl_sequence(payload['pl_rows_compact'], existing_seeds=pl_seed)
                    pl_seed = _u105_merge_seed_map(pl_seed, _v74_seq_seed_from_rows(pl_expanded, V76_PL_HEADERS, 'S. Invoice#'))
                    inv_rows_to_append.extend(inv_expanded)
                    pl_rows_to_append.extend(pl_expanded)

            if inv_rows_to_append:
                stats_map['inv_appended'] = _u105_append_inv_rows(ws_inv, inv_rows_to_append, inv_pathkey_col)
                workbook_dirty = True
            if pl_rows_to_append:
                stats_map['pl_appended'] = _u105_append_pl_rows(ws_pl, pl_rows_to_append, pl_pathkey_col)
                workbook_dirty = True

            if issue_files:
                log_rows.append(make_log_row('', 'SUMMARY', 'INFO', 'TONG HOP FILE WARNING/ERROR'))
                for fp, st, detail in issue_files:
                    log_rows.append(make_log_row(fp, 'SUMMARY', st, detail))
            if cnt_skip_path > 0:
                log_rows.append(make_log_row('', 'SKIP', 'INFO', f'So luong bo qua do da co File Path OK va khong doi: {cnt_skip_path}'))
            log_rows.append(make_log_row('', 'DONE', 'INFO',
                                         f'Tong={cnt_total}; OK={cnt_ok}; WARNING={cnt_warn}; ERROR={cnt_err}; SkipPath={cnt_skip_path}; SkipSignature={cnt_skip_sig}; Repaired={cnt_repaired}'))

            _u106_set_engine_version(wb_out, U106_ENGINE_VERSION)
            workbook_dirty = True if (workbook_dirty or force_full_resync) else workbook_dirty

            core_seconds = max(0.0, time.perf_counter() - t_all_0)
            source_cache_stats = cache.get('stats', {}) if isinstance(cache, dict) else {}
            stats_map['workbook_saved'] = 1 if workbook_dirty else 0
            bench_rows = _u104_build_runtime_bench_rows(folder_path, output_path, core_seconds, max(0.0, time.perf_counter() - t_all_0), stats_map, source_cache_stats)

            if workbook_dirty:
                append_log_rows(ws_log, log_rows + bench_rows)
                safe_save_workbook_atomic(wb_out, output_path)
                workbook_saved = 1
            else:
                self.logger.log(f'ZERO_WORK_U106: scanned={cnt_total} | skip_ok={cnt_skip_path} | candidate=0', 'info')

            stats_map['workbook_saved'] = workbook_saved
            self.persist_caches(output_path, success_changed_paths=success_changed_paths)

            return {
                'total': cnt_total,
                'ok': cnt_ok,
                'warning': cnt_warn,
                'error': cnt_err,
                'skip_path': cnt_skip_path,
                'skip_signature': cnt_skip_sig,
                'repaired': cnt_repaired,
                'issues': issue_files,
                'scanned_files': sorted(scan_map.values()),
            }
        finally:
            wb_out.close()
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None


# =========================================================
# U107 USE _v72_build_detail_bundle DIRECTLY
# Muc tieu:
# - Bo hoan toan nhanh parse wrapper process_one_sub_file_bundle dang gay sai INV/PL
# - Dung lai bundle logic goc da tung merge dung:
#     _v72_build_detail_bundle
#     _v73_apply_invoice_extras
#     _v76_apply_pl_sequence
# - Standard SUB/VSF PL va SA cung di chung 1 luong bundle 4 phan
# =========================================================

U107_ENGINE_VERSION = 'U107'


def _u107_expand_bundle_rows(bundle4, inv_seed, pl_seed):
    inv_main, pl_main, inv_other, pl_other = bundle4
    inv_main = list(inv_main or [])
    pl_main = list(pl_main or [])
    inv_other = list(inv_other or [])
    pl_other = list(pl_other or [])

    inv_main_ext = _v73_apply_invoice_extras(inv_main, existing_seeds=inv_seed)
    inv_seed = _u105_merge_seed_map(inv_seed, _v74_seq_seed_from_rows(inv_main_ext, V73_INV_HEADERS, 'S. Invoice#'))
    inv_other_ext = _v73_apply_invoice_extras(inv_other, existing_seeds=inv_seed)
    inv_seed = _u105_merge_seed_map(inv_seed, _v74_seq_seed_from_rows(inv_other_ext, V73_INV_HEADERS, 'S. Invoice#'))

    pl_main_ext = _v76_apply_pl_sequence(pl_main, existing_seeds=pl_seed)
    pl_seed = _u105_merge_seed_map(pl_seed, _v74_seq_seed_from_rows(pl_main_ext, V76_PL_HEADERS, 'S. Invoice#'))
    pl_other_ext = _v76_apply_pl_sequence(pl_other, existing_seeds=pl_seed)
    pl_seed = _u105_merge_seed_map(pl_seed, _v74_seq_seed_from_rows(pl_other_ext, V76_PL_HEADERS, 'S. Invoice#'))

    return inv_main_ext + inv_other_ext, pl_main_ext + pl_other_ext, inv_seed, pl_seed


def _u136_build_delete_plan(rownums: List[int]) -> List[Tuple[int, int]]:
    nums = sorted({int(x) for x in (rownums or []) if int(x) > 1})
    if not nums:
        return []
    blocks: List[Tuple[int, int]] = []
    start = prev = nums[0]
    for n in nums[1:]:
        if n == prev + 1:
            prev = n
            continue
        blocks.append((start, prev - start + 1))
        start = prev = n
    blocks.append((start, prev - start + 1))
    blocks.sort(key=lambda x: x[0], reverse=True)
    return blocks


def _u136_execute_delete_plan(ws, plan: List[Tuple[int, int]]) -> int:
    deleted = 0
    for start, count in sorted(plan or [], key=lambda x: x[0], reverse=True):
        if start > 1 and count > 0:
            ws.delete_rows(start, count)
            deleted += count
    return deleted


def _u136_parse_candidate_paths(candidate_paths: List[str], parse_cache: Dict[str, Any], repair_options, progress_callback=None,
                                progress_start: int = 0, progress_total: int = 1, force_full_resync: bool = False,
                                logger=None) -> Dict[str, Any]:
    parsed_payloads: Dict[str, Dict[str, Any]] = {}
    issue_files: List[Tuple[str, str, str]] = []
    log_rows: List[List[Any]] = []
    success_changed_paths: Set[str] = set()
    parse_fp_map: Dict[str, str] = {}
    late_deleted_paths: Set[str] = set()
    parse_cache_dirty = False
    cnt_ok = cnt_warn = cnt_err = cnt_repaired = 0
    stats = {'parsed': 0, 'parse_reused': 0, 'parse_fresh': 0, 'repaired': 0, 'late_deleted': 0}
    completed = 0
    total_jobs = max(1, int(progress_total or 1))

    def _accept_payload(file_path: str, arr: List[Any], bundle4: Any, source_tag: str):
        nonlocal cnt_ok, cnt_warn, cnt_err
        file_path_can = _u104_canonical_path(file_path)
        arr = list(arr[:len(SUB_DETAIL_HEADERS)]) + [''] * max(0, len(SUB_DETAIL_HEADERS) - len(arr))
        arr[23] = file_path_can
        status_text = nz_str(arr[21]).upper()
        err_text = nz_str(arr[22])
        if status_text in ('WARNING', 'ERROR'):
            issue_files.append((file_path_can, status_text, err_text))
            log_rows.append(make_log_row(file_path_can, 'CHECK', status_text, err_text))
            if logger is not None:
                logger.log(f"{status_text}: {get_file_name_from_path(file_path_can)} -> {err_text} [{source_tag}]", 'warn' if status_text == 'WARNING' else 'error')
        if status_text == 'OK':
            cnt_ok += 1
        elif status_text == 'WARNING':
            cnt_warn += 1
        else:
            cnt_err += 1
        parsed_payloads[file_path_can] = {'arr': arr, 'status': status_text, 'error': err_text, 'bundle4': bundle4, 'state_entry': _u138_state_from_payload({'arr': arr, 'status': status_text, 'error': err_text, 'bundle4': bundle4})}
        success_changed_paths.add(file_path_can)
        stats['parsed'] += 1
        return file_path_can, arr, status_text, err_text, bundle4

    fresh_paths: List[str] = []
    for file_path in (candidate_paths or []):
        file_path_can = _u104_canonical_path(file_path)
        if not os.path.exists(file_path):
            late_deleted_paths.add(file_path_can)
            stats['late_deleted'] += 1
            log_rows.append(make_log_row(file_path_can, 'CACHE_GATE', 'INFO', 'MISSING_ON_DISK_DROP'))
            if logger is not None:
                logger.log(f'DELETED_PATH_LATE_DROP: {file_path_can}', 'info')
            completed += 1
            if progress_callback:
                progress_callback(progress_start + completed, total_jobs, file_path_can)
            continue
        parse_fp = _u130_build_parse_fp(file_path, repair_options)
        parse_fp_map[file_path_can] = parse_fp
        hit, entry, reason = _u130_lookup_parse_cache(parse_cache, file_path, parse_fp, force_full_resync)
        if hit and entry is not None:
            _accept_payload(file_path, _u130_deepcopy_payload(entry.get('arr', [])), _u130_deepcopy_payload(entry.get('bundle4')), 'CACHE')
            stats['parse_reused'] += 1
            completed += 1
            if progress_callback:
                progress_callback(progress_start + completed, total_jobs, file_path_can)
        else:
            log_rows.append(make_log_row(file_path_can, 'CACHE_GATE', 'INFO', reason))
            fresh_paths.append(file_path)

    if fresh_paths:
        def job(path: str):
            arr, bundle4, exc_text = _u132_safe_parse_payload(path)
            return path, arr, bundle4, exc_text
        futures = {}
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            for fp in fresh_paths:
                futures[executor.submit(job, fp)] = fp
            for future in as_completed(futures):
                file_path, arr, bundle4, exc_text = future.result()
                file_path_can = _u104_canonical_path(file_path)
                arr = list(arr[:len(SUB_DETAIL_HEADERS)]) + [''] * max(0, len(SUB_DETAIL_HEADERS) - len(arr))
                arr[23] = file_path_can
                status_text = nz_str(arr[21]).upper()
                err_text = nz_str(arr[22]) or nz_str(exc_text)
                if exc_text:
                    log_rows.append(make_log_row(file_path_can, 'PARSE_SAFE', 'ERROR', err_text))
                    if logger is not None:
                        logger.log(f"ERROR: {get_file_name_from_path(file_path_can)} -> {err_text} [SAFE]", 'error')
                if status_text in ('WARNING', 'ERROR') and (repair_options.use_folder_truth or repair_options.use_manual_values) and not exc_text:
                    repaired, detail = repair_excel_metadata(file_path, arr, repair_options)
                    if repaired:
                        cnt_repaired += 1
                        stats['repaired'] += 1
                        log_rows.append(make_log_row(file_path_can, 'REPAIR', 'OK', detail))
                        arr, bundle4, exc_text2 = _u132_safe_parse_payload(file_path)
                        arr = list(arr[:len(SUB_DETAIL_HEADERS)]) + [''] * max(0, len(SUB_DETAIL_HEADERS) - len(arr))
                        arr[23] = file_path_can
                        status_text = nz_str(arr[21]).upper()
                        err_text = nz_str(arr[22]) or nz_str(exc_text2)
                        if exc_text2:
                            log_rows.append(make_log_row(file_path_can, 'PARSE_SAFE', 'ERROR', err_text))
                            if logger is not None:
                                logger.log(f"ERROR: {get_file_name_from_path(file_path_can)} -> {err_text} [SAFE-REPAIR]", 'error')
                    else:
                        log_rows.append(make_log_row(file_path_can, 'REPAIR', 'SKIP', detail))
                file_path_can, arr, status_text, err_text, bundle4 = _accept_payload(file_path, arr, bundle4, 'FRESH' if not exc_text else 'SAFE')
                _u130_store_parse_cache(parse_cache, file_path_can, parse_fp_map.get(file_path_can, ''), arr, bundle4)
                parse_cache_dirty = True
                _u130_stats_add(parse_cache, 'parse_real_total')
                stats['parse_fresh'] += 1
                completed += 1
                if progress_callback:
                    progress_callback(progress_start + completed, total_jobs, file_path_can)

    return {
        'parsed_payloads': parsed_payloads,
        'issue_files': issue_files,
        'log_rows': log_rows,
        'success_changed_paths': success_changed_paths,
        'parse_cache_dirty': parse_cache_dirty,
        'cnt_ok': cnt_ok,
        'cnt_warn': cnt_warn,
        'cnt_err': cnt_err,
        'cnt_repaired': cnt_repaired,
        'stats': stats,
        'late_deleted_paths': late_deleted_paths,
    }




# =========================================================
# V23 SIDE-CAR PRUNE + V24 DIAGNOSTIC/FOLDER-AUDIT GATES
# - Không đụng parser / business logic.
# - Giảm open/save khi deleted path chỉ còn trong sidecar cache.
# - Thêm sub-index vào detailindex để lần sau phân biệt stale sidecar với workbook row thật.
# - Cho phép hoãn Folder sheet metadata-only change để tránh mở/save workbook chỉ vì folder audit phụ.
# =========================================================

V23_DEFER_FOLDER_AUDIT_METADATA_ONLY = True


def _v23_get_detail_index_rows(detail_index_cache: Dict[str, Any], section: str) -> Dict[str, List[int]]:
    try:
        sec = (detail_index_cache or {}).get(section, {}) if isinstance(detail_index_cache, dict) else {}
        rows = (sec or {}).get('rows', {}) if isinstance(sec, dict) else {}
        return rows if isinstance(rows, dict) else {}
    except Exception:
        return {}


def _v23_split_deleted_paths_by_index(deleted_paths: List[str],
                                      detail_index_cache: Dict[str, Any],
                                      substate_cache: Dict[str, Any]) -> Tuple[List[str], List[str], Dict[str, int]]:
    """
    Chỉ prune sidecar-only khi detailindex đã có section 'sub'.
    Nếu chưa có sub-index thì giữ hành vi cũ để an toàn.
    """
    deleted = sorted({_u104_canonical_path(p) for p in (deleted_paths or []) if _u104_canonical_path(p)})
    has_sub_index = isinstance(detail_index_cache, dict) and isinstance((detail_index_cache or {}).get('sub'), dict)
    if not deleted or not has_sub_index:
        return deleted, [], {'sidecar_pruned': 0, 'workbook_deleted': len(deleted), 'has_sub_index': 1 if has_sub_index else 0}

    sub_rows = _v23_get_detail_index_rows(detail_index_cache, 'sub')
    inv_rows = _v23_get_detail_index_rows(detail_index_cache, 'inv')
    pl_rows = _v23_get_detail_index_rows(detail_index_cache, 'pl')

    workbook_deleted: List[str] = []
    sidecar_only: List[str] = []

    for pk in deleted:
        has_rows = bool(sub_rows.get(pk) or inv_rows.get(pk) or pl_rows.get(pk))
        # Nếu không có row trong index schema mới thì xem là stale sidecar-only.
        # Đây là lý do phải có sub-index; nếu thiếu sub-index thì không prune ở nhánh này.
        if has_rows:
            workbook_deleted.append(pk)
        else:
            sidecar_only.append(pk)

    return sorted(workbook_deleted), sorted(sidecar_only), {
        'sidecar_pruned': len(sidecar_only),
        'workbook_deleted': len(workbook_deleted),
        'has_sub_index': 1,
    }


def _v23_prune_sidecar_keys(parse_cache: Dict[str, Any],
                            fingerprint_cache: Dict[str, str],
                            substate_cache: Dict[str, Any],
                            keys: List[str]) -> int:
    dropped = 0
    keyset = {_u104_canonical_path(k) for k in (keys or []) if _u104_canonical_path(k)}
    if not keyset:
        return 0
    try:
        entries = parse_cache.get('entries') if isinstance(parse_cache, dict) else None
        if isinstance(entries, dict):
            for pk in list(keyset):
                if entries.pop(pk, None) is not None:
                    dropped += 1
    except Exception:
        pass
    try:
        if isinstance(fingerprint_cache, dict):
            for pk in list(keyset):
                if fingerprint_cache.pop(pk, None) is not None:
                    dropped += 1
    except Exception:
        pass
    try:
        if isinstance(substate_cache, dict):
            for pk in list(keyset):
                if substate_cache.pop(pk, None) is not None:
                    dropped += 1
    except Exception:
        pass
    return dropped


def _v23_folder_audit_should_defer(folder_audit_stats: Dict[str, Any],
                                   folder_delta_prebuilt: Any,
                                   core_needs_workbook: bool) -> bool:
    if not V23_DEFER_FOLDER_AUDIT_METADATA_ONLY:
        return False
    if core_needs_workbook:
        return False
    stats = folder_audit_stats or {}
    try:
        changed = int(stats.get('new_or_changed_files', 0) or 0)
        deleted = int(stats.get('deleted_files', 0) or 0)
        affected = int(stats.get('affected_folders', 0) or 0)
    except Exception:
        changed, deleted, affected = 0, 0, 0

    # Không hoãn deleted/new-folder vì Folder sheet cần loại/thêm dòng.
    # Chỉ hoãn changed metadata khi không có tín hiệu Cus/workbook dirty rõ ràng.
    if changed <= 0 or deleted > 0 or affected <= 0:
        return False

    delta = folder_delta_prebuilt if isinstance(folder_delta_prebuilt, dict) else {}
    if bool(delta.get('cus_dirty', False)) or bool(delta.get('workbook_dirty', False)):
        return False
    # Nếu extension nói rõ Folder sheet dirty thì vẫn cho phép hoãn metadata-only
    # khi không có cus_dirty/core_dirty; Folder sheet phụ sẽ được cập nhật ở lần core-open kế tiếp.
    return True


def _v23_build_detail_index_payload(sub_state: Dict[str, Any], ws_out, inv_map: Dict[str, List[int]], pl_map: Dict[str, List[int]], ws_inv, ws_pl) -> Dict[str, Any]:
    payload = _u131_build_detail_index_payload(inv_map, pl_map, ws_inv, ws_pl)
    sub_rows: Dict[str, List[int]] = {}
    try:
        for pk, meta in (sub_state or {}).items():
            pathkey = _u104_canonical_path(pk)
            row_idx = int((meta or {}).get('row_idx', 0) or 0) if isinstance(meta, dict) else 0
            if pathkey and row_idx >= 2:
                sub_rows[pathkey] = [row_idx]
    except Exception:
        sub_rows = {}
    payload['schema'] = 2
    payload['sub'] = {'rows': sub_rows, 'max_row': int(getattr(ws_out, 'max_row', 0) or 0)}
    return payload



# =========================================================
# V24 DIAGNOSTIC + FOLDER-AUDIT DELTA GATE
# - Không thay parser / business rule.
# - Thêm log quyết định mở workbook để phân biệt real mutation vs gate miss.
# - Folder audit không được tự ép mở/save workbook nếu chỉ là Folder sheet metadata phụ.
# - Khi Folder audit bị defer, KHÔNG save cache folder audit để tránh cache nói đã đồng bộ
#   trong khi sheet Folder trong workbook chưa được ghi.
# =========================================================

V24_DEFER_FOLDER_SHEET_ONLY_WHEN_CORE_CLEAN = True
V24_SAVE_FOLDER_CACHE_WHEN_DEFERRED = False


def _v24_int(value: Any, default: int = 0) -> int:
    try:
        if value is None:
            return default
        return int(value)
    except Exception:
        return default


def _v24_bool(value: Any) -> bool:
    try:
        if isinstance(value, str):
            return value.strip().lower() in ('1', 'true', 'yes', 'y', 'ok')
        return bool(value)
    except Exception:
        return False


def _v24_folder_audit_gate(folder_audit_stats: Dict[str, Any],
                           folder_delta_prebuilt: Any,
                           core_needs_workbook: bool) -> Dict[str, Any]:
    """
    Phân loại folder audit trước khi mở workbook.

    Nguyên tắc an toàn:
    - Nếu core đã cần workbook: có thể apply folder audit piggyback, nhưng không phải lý do riêng để mở workbook.
    - Nếu core không cần workbook:
        + cus_dirty: mở workbook vì Cus no có thể ảnh hưởng SUB_DETAIL/INV/PL.
        + folder sheet only / metadata only: defer, không mở/save workbook.
    - Khi defer: không save folder audit cache, để lần core-open kế tiếp vẫn còn delta và có thể apply thật.
    """
    stats = folder_audit_stats or {}
    delta = folder_delta_prebuilt if isinstance(folder_delta_prebuilt, dict) else {}

    changed = _v24_int(stats.get('new_or_changed_files', 0))
    deleted = _v24_int(stats.get('deleted_files', 0))
    affected = _v24_int(stats.get('affected_folders', 0))
    total_scanned = _v24_int(stats.get('total_scanned', 0))
    cache_dirty = _v24_bool(stats.get('cache_dirty', False))

    cus_dirty = _v24_bool(delta.get('cus_dirty', False))
    folder_sheet_dirty = _v24_bool(delta.get('folder_sheet_dirty', False))
    raw_workbook_dirty = _v24_bool(delta.get('workbook_dirty', False))

    folder_data_changed = bool(changed > 0 or deleted > 0 or affected > 0 or folder_sheet_dirty)
    has_any_delta = bool(folder_data_changed or cus_dirty or raw_workbook_dirty or cache_dirty)

    gate = {
        'total_scanned': total_scanned,
        'changed': changed,
        'deleted': deleted,
        'affected': affected,
        'cache_dirty': 1 if cache_dirty else 0,
        'cus_dirty': 1 if cus_dirty else 0,
        'folder_sheet_dirty': 1 if folder_sheet_dirty else 0,
        'raw_workbook_dirty': 1 if raw_workbook_dirty else 0,
        'folder_data_changed': 1 if folder_data_changed else 0,
        'core_needs_workbook': 1 if core_needs_workbook else 0,
        'needs_workbook': 0,
        'apply_required': 0,
        'deferred': 0,
        'save_cache_allowed': 1,
        'reason': 'no_folder_delta',
    }

    if not has_any_delta:
        return gate

    if core_needs_workbook:
        gate.update({
            'needs_workbook': 0,          # workbook đã mở vì core, folder không phải lý do riêng
            'apply_required': 1 if (folder_data_changed or cus_dirty or raw_workbook_dirty) else 0,
            'deferred': 0,
            'save_cache_allowed': 1,
            'reason': 'piggyback_core_open',
        })
        return gate

    if cus_dirty:
        gate.update({
            'needs_workbook': 1,
            'apply_required': 1,
            'deferred': 0,
            'save_cache_allowed': 1,
            'reason': 'cus_dirty_requires_workbook',
        })
        return gate

    if V24_DEFER_FOLDER_SHEET_ONLY_WHEN_CORE_CLEAN and folder_data_changed:
        gate.update({
            'needs_workbook': 0,
            'apply_required': 0,
            'deferred': 1,
            'save_cache_allowed': 1 if V24_SAVE_FOLDER_CACHE_WHEN_DEFERRED else 0,
            'reason': 'folder_sheet_only_deferred',
        })
        return gate

    # Fallback an toàn nếu extension báo workbook_dirty nhưng không phân loại được cus/folder rõ ràng.
    if raw_workbook_dirty:
        gate.update({
            'needs_workbook': 1,
            'apply_required': 1,
            'deferred': 0,
            'save_cache_allowed': 1,
            'reason': 'raw_workbook_dirty_unclassified',
        })
        return gate

    # Chỉ cache sidecar thay đổi, không mở workbook.
    gate.update({
        'needs_workbook': 0,
        'apply_required': 0,
        'deferred': 0,
        'save_cache_allowed': 1,
        'reason': 'cache_only',
    })
    return gate


def _v24_open_decision(core_needs_workbook: bool,
                       folder_audit_needs_workbook: bool,
                       pre_force_full_resync: bool,
                       deleted_paths: List[str],
                       pre_action_subdetail_only: int,
                       pre_action_rewrite_detail: int,
                       pre_action_skip_all: int,
                       precheck_stats: Dict[str, Any],
                       folder_gate: Dict[str, Any]) -> Dict[str, Any]:
    reasons: List[str] = []
    if pre_force_full_resync:
        reasons.append('bootstrap_or_force_full_resync')
    if deleted_paths:
        reasons.append('deleted_cleanup')
    if pre_action_rewrite_detail:
        if _v24_int((precheck_stats or {}).get('status_missing', 0)) > 0:
            reasons.append('new_file_or_missing_state')
        else:
            reasons.append('detail_material_change')
    if pre_action_subdetail_only:
        reasons.append('subdetail_only_change')
    if folder_audit_needs_workbook:
        reasons.append('folder_audit_' + nz_str((folder_gate or {}).get('reason', 'delta')))
    if not reasons:
        reasons.append('no_workbook_mutation')
    return {
        'open_required': 1 if (core_needs_workbook or folder_audit_needs_workbook) else 0,
        'reason': '+'.join(reasons),
        'core_needs_workbook': 1 if core_needs_workbook else 0,
        'folder_needs_workbook': 1 if folder_audit_needs_workbook else 0,
        'rewrite_detail': int(pre_action_rewrite_detail or 0),
        'subdetail_only': int(pre_action_subdetail_only or 0),
        'skip_all': int(pre_action_skip_all or 0),
        'deleted_paths': len(deleted_paths or []),
    }


def _v24_fmt_map(d: Dict[str, Any], keys: Optional[List[str]] = None) -> str:
    try:
        if keys is None:
            keys = list((d or {}).keys())
        return ' | '.join(f"{k}={nz_str((d or {}).get(k, ''))}" for k in keys)
    except Exception:
        return nz_str(d)
class Processor(Processor):

    def run(self, folder_path: str, output_path: str, progress_callback=None, repair_options=None):
        repair_options = repair_options or RepairOptions()
        t_all_0 = time.perf_counter()
        _u133_stage_progress(progress_callback, 1, 6, '1/6 Load cache sidecar...')
        t_init_cache_0 = time.perf_counter()
        parse_cache = _u130_load_parse_cache(output_path)
        parse_cache['stats'] = _u130_empty_parse_cache()['stats']
        parse_cache_dirty = False
        detail_index_cache = _u131_load_detail_index(output_path)
        substate_cache = _u134_load_substate(output_path)
        init_cache_seconds = max(0.0, time.perf_counter() - t_init_cache_0)

        self.load_fingerprint_cache(output_path)
        _u133_stage_progress(progress_callback, 2, 6, '2/6 Scan danh sach file nguon...')
        t_init_scan_0 = time.perf_counter()
        scan_map = self._u104_scan_map(folder_path, output_path)
        scan_missing_paths: Set[str] = set()
        for _pk, _actual in list((scan_map or {}).items()):
            try:
                if not os.path.exists(_actual):
                    scan_missing_paths.add(_pk)
                    scan_map.pop(_pk, None)
            except Exception:
                scan_missing_paths.add(_pk)
                scan_map.pop(_pk, None)
        scanned_keys = set(scan_map.keys())
        self._last_supported_scanned = len(scanned_keys)
        cached_pathkeys = set(k for k in (substate_cache or {}).keys() if nz_str(k)) | set(k for k in (self._fingerprint_cache or {}).keys() if nz_str(k))
        try:
            inv_rows_cache = (((detail_index_cache or {}).get('inv', {}) or {}).get('rows', {}) or {}) if isinstance(detail_index_cache, dict) else {}
            pl_rows_cache = (((detail_index_cache or {}).get('pl', {}) or {}).get('rows', {}) or {}) if isinstance(detail_index_cache, dict) else {}
            sub_rows_cache = (((detail_index_cache or {}).get('sub', {}) or {}).get('rows', {}) or {}) if isinstance(detail_index_cache, dict) else {}
            if isinstance(inv_rows_cache, dict):
                cached_pathkeys.update(k for k in inv_rows_cache.keys() if nz_str(k))
            if isinstance(pl_rows_cache, dict):
                cached_pathkeys.update(k for k in pl_rows_cache.keys() if nz_str(k))
            if isinstance(sub_rows_cache, dict):
                cached_pathkeys.update(k for k in sub_rows_cache.keys() if nz_str(k))
        except Exception:
            pass
        deleted_paths = sorted((set(pk for pk in cached_pathkeys if pk and pk not in scanned_keys) | set(scan_missing_paths)))
        # V23: neu detailindex da co sub-index, tach path chi con trong sidecar de prune truoc open workbook.
        deleted_paths, deleted_sidecar_only, deleted_split_stats = _v23_split_deleted_paths_by_index(deleted_paths, detail_index_cache, substate_cache)
        if deleted_sidecar_only:
            dropped_sidecar = _v23_prune_sidecar_keys(parse_cache, self._fingerprint_cache, substate_cache, deleted_sidecar_only)
            try:
                self.logger.log(
                    f"V23_DELETED_SIDECAR_ONLY_PRUNED: keys={len(deleted_sidecar_only)} | dropped={dropped_sidecar} | workbook_deleted={len(deleted_paths)}",
                    'info'
                )
            except Exception:
                pass
        init_scan_seconds = max(0.0, time.perf_counter() - t_init_scan_0)

        _u133_stage_progress(progress_callback, 3, 6, '3/6 Precheck fingerprint/status...')
        t_init_pre_0 = time.perf_counter()
        precheck_stats = _u137_eval_precheck_with_fallback(scan_map, self._fingerprint_cache, substate_cache, parse_cache, repair_options)
        precheck_candidates: List[str] = list(precheck_stats.get('candidates', []) or [])
        deleted_path_set = set(deleted_paths)
        deleted_candidates = [pk for pk in precheck_candidates if pk in deleted_path_set]
        if deleted_candidates:
            precheck_candidates = [pk for pk in precheck_candidates if pk not in deleted_path_set]
            precheck_stats['candidate'] = len(precheck_candidates)
        precheck_skipped = int(precheck_stats.get('skip_all', 0) or 0)
        precheck_skip_ok = int(precheck_stats.get('skip_ok', 0) or 0)
        precheck_skip_non_ok = int(precheck_stats.get('skip_non_ok_stable', 0) or 0)
        init_precheck_seconds = max(0.0, time.perf_counter() - t_init_pre_0)
        self.logger.log(
            'PRECHECK_SUBSTATE_STATS: '
            f"loaded={precheck_stats.get('loaded_substate', 0)} | "
            f"fp_match={precheck_stats.get('fp_match', 0)} | "
            f"fp_match_fingerprint={precheck_stats.get('fp_match_fingerprint', 0)} | "
            f"fp_match_parsecache={precheck_stats.get('fp_match_parsecache', 0)} | "
            f"fp_miss={precheck_stats.get('fp_miss', 0)} | "
            f"status_ok={precheck_stats.get('status_ok', 0)} | "
            f"status_not_ok={precheck_stats.get('status_not_ok', 0)} | "
            f"status_missing={precheck_stats.get('status_missing', 0)} | "
            f"candidate={precheck_stats.get('candidate', 0)} | "
            f"skip_ok={precheck_skip_ok} | "
            f"skip_non_ok_stable={precheck_skip_non_ok} | "
            f"skip_all={precheck_skipped}",
            'info'
        )
        if deleted_paths:
            self.logger.log(f'DELETED_PATH_PENDING: {len(deleted_paths)}', 'info')
        if scan_missing_paths:
            self.logger.log(f'DELETED_PATH_SCAN_DROP: {len(scan_missing_paths)}', 'info')
        if deleted_candidates:
            self.logger.log(f'DELETED_PATH_EXCLUDED_FROM_CANDIDATE: {len(deleted_candidates)}', 'info')

        cnt_total = len(scanned_keys)
        cnt_ok = 0
        cnt_warn = 0
        cnt_err = 0
        cnt_skip_path = precheck_skipped
        cnt_skip_sig = 0
        cnt_repaired = 0
        issue_files: List[Tuple[str, str, str]] = []
        log_rows: List[List[Any]] = [make_log_row('', 'START', 'INFO', f'Bat dau quet {cnt_total} file.')]

        if os.path.exists(output_path) and substate_cache and not precheck_candidates and not deleted_paths:
            if deleted_sidecar_only:
                try:
                    self.persist_caches(output_path, success_changed_paths=set())
                    _u130_prune_parse_cache(parse_cache, scanned_keys)
                    _u130_save_parse_cache(output_path, parse_cache)
                    _u134_save_substate(output_path, substate_cache)
                except Exception as exc:
                    try:
                        self.logger.log(f'V23_SIDECAR_PRUNE_SAVE_ERROR: {exc}', 'warn')
                    except Exception:
                        pass
            total_real_seconds = max(0.0, time.perf_counter() - t_all_0)
            zero_work_issue_snapshot = _u139_build_issue_snapshot(substate_cache)
            stable_issue_items = [
                (nz_str(x.get('file_path', '')), nz_str(x.get('status', '')), nz_str(x.get('detail', '')))
                for x in zero_work_issue_snapshot
            ]
            cnt_warning = sum(1 for _, st, _ in stable_issue_items if nz_str(st).upper() == 'WARNING')
            cnt_error = sum(1 for _, st, _ in stable_issue_items if nz_str(st).upper() == 'ERROR')
            cnt_ok = max(0, len(scanned_keys) - cnt_warning - cnt_error)
            self.logger.log(f'PRECHECK_ZERO_WORK_SKIP_OPEN_WB: scanned={len(scanned_keys)} | skip_all={precheck_skipped} | stable_issue={len(stable_issue_items)}', 'info')
            try:
                ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_INIT_CACHE: {init_cache_seconds:.3f}s")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_INIT_SCAN_MAP: {init_scan_seconds:.3f}s | scanned={len(scanned_keys)}")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_INIT_PRECHECK: {init_precheck_seconds:.3f}s | loaded={precheck_stats.get('loaded_substate', 0)} | candidate=0 | skip_ok={precheck_skip_ok} | skip_non_ok_stable={precheck_skip_non_ok} | skip_all={precheck_skipped}")
                _u130_append_runtime_bench_log(output_path, f"{ts} | ZERO_WORK_ISSUES_REHYDRATED: {len(stable_issue_items)} issue(s)")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_OPEN_DECISION: open_required=0 | reason=precheck_zero_work | core_needs_workbook=0 | folder_needs_workbook=0 | rewrite_detail=0 | subdetail_only=0 | deleted_paths=0 | skip_all={precheck_skipped}")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_FOLDER_DELTA_GATE: reason=not_built_precheck_zero_work | needs_workbook=0 | apply_required=0 | deferred=0 | save_cache_allowed=1")
                _u130_append_runtime_bench_log(output_path, f"{ts} | PRECHECK_ZERO_WORK_SKIP_OPEN_WB: {total_real_seconds:.3f}s")
            except Exception:
                pass
            _u133_stage_progress(progress_callback, 6, 6, '6/6 Zero-work, bo qua mo workbook')
            return {
                'total': len(scanned_keys),
                'ok': cnt_ok,
                'warning': cnt_warning,
                'error': cnt_error,
                'skip_path': precheck_skipped,
                'skip_signature': 0,
                'repaired': 0,
                'issues': stable_issue_items,
                'scanned_files': sorted(scan_map.values()),
            }

        _u133_stage_progress(progress_callback, 4, 6, '4/6 Parse candidate truoc mo workbook...')
        t_parse_before_open_0 = time.perf_counter()
        pre_force_full_resync = not (os.path.exists(output_path) and substate_cache)
        preparse_candidate_paths = list(scan_map.values()) if pre_force_full_resync else list(precheck_candidates)
        preparse_candidate_count = len(preparse_candidate_paths)
        preparse_result = _u136_parse_candidate_paths(
            preparse_candidate_paths,
            parse_cache,
            repair_options,
            progress_callback=progress_callback,
            progress_start=0,
            progress_total=max(1, preparse_candidate_count),
            force_full_resync=pre_force_full_resync,
            logger=getattr(self, 'logger', None),
        )
        late_deleted_paths = set(preparse_result.get('late_deleted_paths', set()) or set())
        if late_deleted_paths:
            deleted_paths = sorted(set(deleted_paths) | late_deleted_paths)
            self.logger.log(f'DELETED_PATH_LATE_DROP: {len(late_deleted_paths)}', 'info')
        parse_before_open_seconds = max(0.0, time.perf_counter() - t_parse_before_open_0)
        parse_cache_dirty = parse_cache_dirty or bool(preparse_result.get('parse_cache_dirty'))
        parsed_payloads: Dict[str, Dict[str, Any]] = dict(preparse_result.get('parsed_payloads', {}) or {})
        issue_files.extend(preparse_result.get('issue_files', []) or [])
        log_rows.extend(preparse_result.get('log_rows', []) or [])
        cnt_ok += int(preparse_result.get('cnt_ok', 0) or 0)
        cnt_warn += int(preparse_result.get('cnt_warn', 0) or 0)
        cnt_err += int(preparse_result.get('cnt_err', 0) or 0)
        cnt_repaired += int(preparse_result.get('cnt_repaired', 0) or 0)
        success_changed_paths: Set[str] = set(preparse_result.get('success_changed_paths', set()) or set())

        # =========================================================
        # V21 MATERIAL-CHANGE GATE + FOLDER-DELTA GATE
        # Muc tieu:
        # - Tach ro "co candidate can parse" voi "co thay doi that su can mo/ghi workbook".
        # - Neu candidate parse xong nhung state detail khong doi va Folder audit khong doi,
        #   thi dung truoc open workbook de tranh OPEN_WB_POST_PARSE + SAVE_FINAL.
        # - Khong dung parser/business logic; chi dung state_entry da co san de quyet dinh runtime.
        # =========================================================
        pre_action_map: Dict[str, str] = {}
        pre_subdetail_payloads: Dict[str, Dict[str, Any]] = {}
        pre_rewrite_payloads: Dict[str, Dict[str, Any]] = {}
        pre_action_skip_all = 0
        pre_action_subdetail_only = 0
        pre_action_rewrite_detail = 0
        for pathkey, payload in parsed_payloads.items():
            old_state_entry = _u138_normalize_state_entry((substate_cache or {}).get(pathkey, {}))
            new_state_entry = _u138_normalize_state_entry((payload or {}).get('state_entry', _u138_state_from_payload(payload)))
            payload['state_entry'] = new_state_entry
            action = _u138_decide_detail_action(old_state_entry, new_state_entry)
            pre_action_map[pathkey] = action
            if action == 'SKIP_ALL':
                pre_action_skip_all += 1
            elif action == 'UPDATE_SUBDETAIL_ONLY':
                pre_action_subdetail_only += 1
                pre_subdetail_payloads[pathkey] = payload
            else:
                pre_action_rewrite_detail += 1
                pre_subdetail_payloads[pathkey] = payload
                pre_rewrite_payloads[pathkey] = payload

        folder_delta_prebuilt = None
        folder_audit_build_seconds = 0.0
        folder_audit_apply_seconds = 0.0
        folder_audit_cache_dirty = False
        folder_audit_cache_payload = None
        folder_audit_stats = {}
        folder_audit_apply = {}
        folder_audit_needs_workbook = False
        folder_audit_deferred_metadata = False
        save_folder_audit_cache_func = None
        apply_delta_inplace_func = None
        folder_audit_apply_required = False
        folder_audit_gate: Dict[str, Any] = {}
        folder_audit_cache_save_allowed = True
        core_needs_workbook = bool(pre_subdetail_payloads or deleted_paths)
        open_decision: Dict[str, Any] = {}
        try:
            from folder_audit_ext import build_delta_state, apply_delta_inplace as _fad_apply_delta_inplace, save_cache as _fad_save_cache
            save_folder_audit_cache_func = _fad_save_cache
            apply_delta_inplace_func = _fad_apply_delta_inplace
            t_faudit_gate_0 = time.perf_counter()
            folder_delta_prebuilt = build_delta_state(folder_path, output_path, logger=getattr(self, 'logger', None))
            folder_audit_build_seconds = max(0.0, time.perf_counter() - t_faudit_gate_0)
            folder_audit_stats = dict((folder_delta_prebuilt or {}).get('stats', {}) or {})
            folder_audit_cache_dirty = bool(folder_audit_stats.get('cache_dirty', False))
            folder_audit_cache_payload = (folder_delta_prebuilt or {}).get('cache')
            # V24:
            # - Phân loại folder audit thành: piggyback core-open / cus_dirty / folder_sheet_only_deferred / cache_only.
            # - Folder sheet only không được tự ép mở workbook khi core không có thay đổi.
            # - Khi defer thì không save folder audit cache để tránh cache và workbook lệch nhau.
            folder_audit_gate = _v24_folder_audit_gate(
                folder_audit_stats,
                folder_delta_prebuilt,
                core_needs_workbook=core_needs_workbook,
            )
            folder_audit_needs_workbook = bool(folder_audit_gate.get('needs_workbook', 0))
            folder_audit_apply_required = bool(folder_audit_gate.get('apply_required', 0))
            folder_audit_deferred_metadata = bool(folder_audit_gate.get('deferred', 0))
            folder_audit_cache_save_allowed = bool(folder_audit_gate.get('save_cache_allowed', 1))
            try:
                self.logger.log('V24_FOLDER_DELTA_GATE: ' + _v24_fmt_map(folder_audit_gate, [
                    'reason', 'core_needs_workbook', 'needs_workbook', 'apply_required', 'deferred',
                    'save_cache_allowed', 'changed', 'deleted', 'affected', 'cus_dirty',
                    'folder_sheet_dirty', 'raw_workbook_dirty', 'cache_dirty'
                ]), 'info')
            except Exception:
                pass
        except Exception as exc:
            try:
                self.logger.log(f'FOLDER_AUDIT_EXT_ERROR_BUILD_GATE: {exc}', 'warn')
            except Exception:
                pass

        core_needs_workbook = bool(pre_subdetail_payloads or deleted_paths)
        open_decision = _v24_open_decision(
            core_needs_workbook=core_needs_workbook,
            folder_audit_needs_workbook=folder_audit_needs_workbook,
            pre_force_full_resync=pre_force_full_resync,
            deleted_paths=deleted_paths,
            pre_action_subdetail_only=pre_action_subdetail_only,
            pre_action_rewrite_detail=pre_action_rewrite_detail,
            pre_action_skip_all=pre_action_skip_all,
            precheck_stats=precheck_stats,
            folder_gate=folder_audit_gate,
        )
        try:
            self.logger.log('V24_OPEN_DECISION: ' + _v24_fmt_map(open_decision, [
                'open_required', 'reason', 'core_needs_workbook', 'folder_needs_workbook',
                'rewrite_detail', 'subdetail_only', 'deleted_paths', 'skip_all'
            ]), 'info')
        except Exception:
            pass
        if (not core_needs_workbook) and (not folder_audit_needs_workbook):
            # Candidate co the da parse do fingerprint/cache miss, nhung output logic khong doi.
            # Luu lai cache/fingerprint/issue snapshot, khong mo workbook va khong save workbook.
            try:
                _u137_refresh_fingerprint_cache_from_scan(scan_map, self._fingerprint_cache)
                self.persist_caches(output_path, success_changed_paths=success_changed_paths)
            except Exception as exc:
                try:
                    self.logger.log(f'V21_CACHE_PERSIST_SKIP_OPEN_ERROR: {exc}', 'warn')
                except Exception:
                    pass
            try:
                _u130_prune_parse_cache(parse_cache, scanned_keys)
                _u130_save_parse_cache(output_path, parse_cache)
            except Exception as exc:
                try:
                    self.logger.log(f'PARSE_CACHE_SAVE_ERROR: {exc}', 'warn')
                except Exception:
                    pass
            if folder_audit_cache_dirty and folder_audit_cache_payload and save_folder_audit_cache_func and folder_audit_cache_save_allowed:
                try:
                    save_folder_audit_cache_func(output_path, folder_audit_cache_payload)
                except Exception as exc:
                    try:
                        self.logger.log(f'FOLDER_AUDIT_CACHE_SAVE_ERROR: {exc}', 'warn')
                    except Exception:
                        pass
            try:
                stable_issue_items = [
                    (nz_str(x.get('file_path', '')), nz_str(x.get('status', '')), nz_str(x.get('detail', '')))
                    for x in _u139_build_issue_snapshot(substate_cache)
                ]
                _u139_save_issue_snapshot(output_path, [
                    {'file_path': fp, 'file_name': get_file_name_from_path(fp), 'status': st, 'detail': detail}
                    for fp, st, detail in stable_issue_items
                ])
            except Exception:
                stable_issue_items = _u139_issue_items_from_snapshot(output_path)

            cnt_warning = sum(1 for _, st, _ in stable_issue_items if nz_str(st).upper() == 'WARNING')
            cnt_error = sum(1 for _, st, _ in stable_issue_items if nz_str(st).upper() == 'ERROR')
            cnt_ok = max(0, len(scanned_keys) - cnt_warning - cnt_error)
            total_real_seconds = max(0.0, time.perf_counter() - t_all_0)
            try:
                ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_INIT_CACHE: {init_cache_seconds:.3f}s")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_INIT_SCAN_MAP: {init_scan_seconds:.3f}s | scanned={len(scanned_keys)}")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_INIT_PRECHECK: {init_precheck_seconds:.3f}s | loaded={precheck_stats.get('loaded_substate', 0)} | candidate={len(precheck_candidates)} | skip_ok={precheck_skip_ok} | skip_non_ok_stable={precheck_skip_non_ok} | skip_all={precheck_skipped}")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_PARSE_BEFORE_OPEN_WB: {parse_before_open_seconds:.3f}s | candidate={preparse_candidate_count}")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_MATERIAL_GATE: skip_open=1 | core_needs_workbook=0 | folder_needs_workbook=0 | action_skip_all={pre_action_skip_all} | subdetail_only={pre_action_subdetail_only} | rewrite_detail={pre_action_rewrite_detail} | folder_changed={folder_audit_stats.get('new_or_changed_files', 0)} | folder_deleted={folder_audit_stats.get('deleted_files', 0)} | folder_affected={folder_audit_stats.get('affected_folders', 0)} | folder_deferred={1 if folder_audit_deferred_metadata else 0} | sidecar_deleted_pruned={len(deleted_sidecar_only)}")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_OPEN_DECISION: {_v24_fmt_map(open_decision, ['open_required', 'reason', 'core_needs_workbook', 'folder_needs_workbook', 'rewrite_detail', 'subdetail_only', 'deleted_paths', 'skip_all'])}")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_FOLDER_DELTA_GATE: {_v24_fmt_map(folder_audit_gate, ['reason', 'needs_workbook', 'apply_required', 'deferred', 'save_cache_allowed', 'changed', 'deleted', 'affected', 'cus_dirty', 'folder_sheet_dirty', 'raw_workbook_dirty', 'cache_dirty'])}")
                _u130_append_runtime_bench_log(output_path, f"{ts} | V21_MATERIAL_GATE_SKIP_OPEN_WB: {total_real_seconds:.3f}s | parsed={len(parsed_payloads)} | issues={len(stable_issue_items)}")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_SAVE_FINAL: 0.000s | workbook_saved=0")
                _u130_append_runtime_bench_log(output_path, f"{ts} | BENCH_TOTAL_REAL: {total_real_seconds:.3f}s | core={total_real_seconds:.3f}s | folder_build={folder_audit_build_seconds:.3f}s | folder_apply=0.000s | save=0.000s | workbook_saved=0")
            except Exception:
                pass
            try:
                self.logger.log(f'V21_MATERIAL_GATE_SKIP_OPEN_WB: parsed={len(parsed_payloads)} | issues={len(stable_issue_items)} | total={total_real_seconds:.3f}s', 'info')
            except Exception:
                pass
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None
            _u133_stage_progress(progress_callback, 6, 6, '6/6 V24 skip open/save vi khong co material change')
            return {
                'total': len(scanned_keys),
                'ok': cnt_ok,
                'warning': cnt_warning,
                'error': cnt_error,
                'skip_path': precheck_skipped,
                'skip_signature': 0,
                'repaired': cnt_repaired,
                'issues': stable_issue_items,
                'scanned_files': sorted(scan_map.values()),
            }

        _u133_stage_progress(progress_callback, 5, 6, '5/6 Mo workbook va ghi thay doi...')
        t_init_open_0 = time.perf_counter()
        wb_out = open_or_create_output_workbook(output_path)
        open_wb_post_parse_seconds = max(0.0, time.perf_counter() - t_init_open_0)
        init_open_wb_seconds = open_wb_post_parse_seconds
        workbook_saved = 0
        stats_map: Dict[str, Any] = {
            'scanned': len(scanned_keys),
            'candidate': preparse_candidate_count,
            'ok_skip': precheck_skipped,
            'parsed': int((preparse_result.get('stats', {}) or {}).get('parsed', 0) or 0),
            'repaired': int((preparse_result.get('stats', {}) or {}).get('repaired', 0) or 0),
            'sub_insert': 0,
            'sub_update': 0,
            'inv_deleted': 0,
            'inv_appended': 0,
            'pl_deleted': 0,
            'pl_appended': 0,
            'fp_changed': preparse_candidate_count,
            'fp_skipped': precheck_skipped,
            'workbook_saved': 0,
            'parse_reused': int((preparse_result.get('stats', {}) or {}).get('parse_reused', 0) or 0),
            'parse_fresh': int((preparse_result.get('stats', {}) or {}).get('parse_fresh', 0) or 0),
            'action_skip_all': 0,
            'action_subdetail_only': 0,
            'action_rewrite_detail': 0,
        }

        stats_map['action_skip_all'] = pre_action_skip_all
        stats_map['action_subdetail_only'] = pre_action_subdetail_only
        stats_map['action_rewrite_detail'] = pre_action_rewrite_detail

        try:
            force_full_resync = (_u106_get_engine_version(wb_out) != U107_ENGINE_VERSION)

            ws_out = ensure_sheet(wb_out, 'SUB_DETAIL')
            ws_log = ensure_sheet(wb_out, 'LOG_SUB_DETAIL')
            ws_inv = ensure_sheet(wb_out, 'INV')
            ws_pl = ensure_sheet(wb_out, 'PL')
            if 'Sheet' in wb_out.sheetnames and wb_out['Sheet'].max_row == 1 and wb_out['Sheet'].max_column == 1 and nz_str(wb_out['Sheet']['A1'].value) == '':
                try:
                    del wb_out['Sheet']
                except Exception:
                    pass

            init_sub_detail_header(ws_out)
            init_log_header(ws_log)
            _v74_ensure_headers(ws_inv, V73_INV_HEADERS)
            _v74_ensure_headers(ws_pl, V76_PL_HEADERS)

            sub_pathkey_col = _u104_ensure_pathkey_column(ws_out, SUB_DETAIL_HEADERS, HEADER_FILL_SUB)
            inv_pathkey_col = _u104_ensure_pathkey_column(ws_inv, V73_INV_HEADERS, HEADER_FILL_SUB)
            pl_pathkey_col = _u104_ensure_pathkey_column(ws_pl, V76_PL_HEADERS, HEADER_FILL_SUB)

            sub_state, sub_dup_rows, sub_backfill = _u104_collect_subdetail_state(ws_out, sub_pathkey_col)
            if sub_dup_rows:
                _u104_delete_rows_desc(ws_out, sub_dup_rows)
                sub_state, _, sub_backfill2 = _u104_collect_subdetail_state(ws_out, sub_pathkey_col)
                sub_backfill += sub_backfill2

            inv_rownums_by_key, inv_backfill = {}, 0
            pl_rownums_by_key, pl_backfill = {}, 0

            workbook_dirty = bool(sub_backfill or inv_backfill or pl_backfill or sub_dup_rows or force_full_resync)
            if force_full_resync:
                log_rows.append(make_log_row('', 'RESYNC', 'INFO', 'Force full resync 1 lan bang _v72_build_detail_bundle de khoi phuc INV/PL dung logic goc.'))
                missing_paths = [actual_path for _, actual_path in sorted(scan_map.items()) if _u104_canonical_path(actual_path) not in parsed_payloads]
                if missing_paths:
                    extra_result = _u136_parse_candidate_paths(
                        missing_paths,
                        parse_cache,
                        repair_options,
                        progress_callback=None,
                        progress_start=0,
                        progress_total=max(1, len(missing_paths)),
                        force_full_resync=True,
                        logger=getattr(self, 'logger', None),
                    )
                    parse_cache_dirty = parse_cache_dirty or bool(extra_result.get('parse_cache_dirty'))
                    parsed_payloads.update(extra_result.get('parsed_payloads', {}) or {})
                    issue_files.extend(extra_result.get('issue_files', []) or [])
                    log_rows.extend(extra_result.get('log_rows', []) or [])
                    cnt_ok += int(extra_result.get('cnt_ok', 0) or 0)
                    cnt_warn += int(extra_result.get('cnt_warn', 0) or 0)
                    cnt_err += int(extra_result.get('cnt_err', 0) or 0)
                    cnt_repaired += int(extra_result.get('cnt_repaired', 0) or 0)
                    success_changed_paths.update(extra_result.get('success_changed_paths', set()) or set())
                    stats_map['parsed'] += int((extra_result.get('stats', {}) or {}).get('parsed', 0) or 0)
                    stats_map['parse_reused'] += int((extra_result.get('stats', {}) or {}).get('parse_reused', 0) or 0)
                    stats_map['parse_fresh'] += int((extra_result.get('stats', {}) or {}).get('parse_fresh', 0) or 0)
                    stats_map['repaired'] += int((extra_result.get('stats', {}) or {}).get('repaired', 0) or 0)
                    stats_map['candidate'] = len(parsed_payloads)

            # V21: dung lai action_map da tinh truoc khi mo workbook.
            # Neu force_full_resync parse them missing_paths sau khi da mo workbook thi tinh bo sung cho cac path do.
            action_map: Dict[str, str] = dict(pre_action_map)
            subdetail_payloads: Dict[str, Dict[str, Any]] = dict(pre_subdetail_payloads)
            rewrite_payloads: Dict[str, Dict[str, Any]] = dict(pre_rewrite_payloads)
            for pathkey, payload in parsed_payloads.items():
                if pathkey in action_map:
                    continue
                old_state_entry = _u138_normalize_state_entry((substate_cache or {}).get(pathkey, {}))
                new_state_entry = _u138_normalize_state_entry((payload or {}).get('state_entry', _u138_state_from_payload(payload)))
                payload['state_entry'] = new_state_entry
                action = _u138_decide_detail_action(old_state_entry, new_state_entry)
                action_map[pathkey] = action
                if action == 'SKIP_ALL':
                    stats_map['action_skip_all'] += 1
                elif action == 'UPDATE_SUBDETAIL_ONLY':
                    stats_map['action_subdetail_only'] += 1
                    subdetail_payloads[pathkey] = payload
                else:
                    stats_map['action_rewrite_detail'] += 1
                    subdetail_payloads[pathkey] = payload
                    rewrite_payloads[pathkey] = payload

            t_sub_upsert_0 = time.perf_counter()
            for pathkey, payload in subdetail_payloads.items():
                arr = payload['arr']
                existing = sub_state.get(pathkey)
                if existing and _u104_existing_row_equals_subdetail(ws_out, existing['row_idx'], arr, sub_pathkey_col):
                    pass
                elif existing:
                    _u104_write_subdetail_row(ws_out, existing['row_idx'], arr, sub_pathkey_col)
                    stats_map['sub_update'] += 1
                    workbook_dirty = True
                else:
                    row_idx = _u104_append_subdetail_row(ws_out, arr, sub_pathkey_col)
                    sub_state[pathkey] = {'row_idx': row_idx, 'status': nz_str(arr[21]).upper(), 'signature': nz_str(arr[24])}
                    stats_map['sub_insert'] += 1
                    workbook_dirty = True
            if deleted_paths:
                sub_delete_rows = sorted((sub_state.get(pk, {}) or {}).get('row_idx') for pk in deleted_paths if isinstance(sub_state.get(pk, {}), dict) and (sub_state.get(pk, {}) or {}).get('row_idx'))
                sub_delete_rows = [int(r) for r in sub_delete_rows if int(r) >= 2]
                if sub_delete_rows:
                    _u104_delete_rows_desc(ws_out, sub_delete_rows)
                    workbook_dirty = True
                    stats_map['sub_deleted'] = len(sub_delete_rows)
                    sub_state, _, _ = _u104_collect_subdetail_state(ws_out, sub_pathkey_col)
            sub_upsert_seconds = max(0.0, time.perf_counter() - t_sub_upsert_0)

            # V21 material gate: chi rewrite detail khi action that su yeu cau rewrite
            rewrite_keys = set(rewrite_payloads.keys())
            cleanup_keys = set(deleted_paths)
            delete_plan_keys = rewrite_keys | cleanup_keys
            t_detail_index_0 = time.perf_counter()
            inv_delete_rows, inv_index_hit = _u131_collect_delete_rows_for_keys(ws_inv, (detail_index_cache or {}).get('inv', {}), delete_plan_keys, inv_pathkey_col)
            pl_delete_rows, pl_index_hit = _u131_collect_delete_rows_for_keys(ws_pl, (detail_index_cache or {}).get('pl', {}), delete_plan_keys, pl_pathkey_col)
            if not inv_index_hit:
                inv_rownums_by_key, inv_backfill = _u104_collect_detail_rownums_by_pathkey(ws_inv, 30, inv_pathkey_col)
                inv_delete_rows = []
                for pathkey in delete_plan_keys:
                    inv_delete_rows.extend(inv_rownums_by_key.get(pathkey, []))
            if not pl_index_hit:
                pl_rownums_by_key, pl_backfill = _u104_collect_detail_rownums_by_pathkey(ws_pl, 19, pl_pathkey_col)
                pl_delete_rows = []
                for pathkey in delete_plan_keys:
                    pl_delete_rows.extend(pl_rownums_by_key.get(pathkey, []))
            detail_index_seconds = max(0.0, time.perf_counter() - t_detail_index_0)

            t_delete_plan_0 = time.perf_counter()
            inv_delete_plan = _u136_build_delete_plan(inv_delete_rows)
            pl_delete_plan = _u136_build_delete_plan(pl_delete_rows)
            delete_plan_seconds = max(0.0, time.perf_counter() - t_delete_plan_0)

            t_detail_delete_0 = time.perf_counter()
            if inv_delete_plan:
                stats_map['inv_deleted'] = _u136_execute_delete_plan(ws_inv, inv_delete_plan)
                workbook_dirty = True
            if pl_delete_plan:
                stats_map['pl_deleted'] = _u136_execute_delete_plan(ws_pl, pl_delete_plan)
                workbook_dirty = True
            detail_delete_seconds = max(0.0, time.perf_counter() - t_detail_delete_0)

            t_detail_seed_0 = time.perf_counter()
            inv_existing_rows = _u105_snapshot_rows(ws_inv, V73_INV_HEADERS)
            pl_existing_rows = _u105_snapshot_rows(ws_pl, V76_PL_HEADERS)
            inv_seed = _v74_seq_seed_from_rows(inv_existing_rows, V73_INV_HEADERS, 'S. Invoice#')
            pl_seed = _v74_seq_seed_from_rows(pl_existing_rows, V76_PL_HEADERS, 'S. Invoice#')
            detail_seed_seconds = max(0.0, time.perf_counter() - t_detail_seed_0)

            inv_rows_to_append: List[List[Any]] = []
            pl_rows_to_append: List[List[Any]] = []
            t_detail_expand_0 = time.perf_counter()
            for pathkey in sorted(rewrite_keys):
                payload = parsed_payloads[pathkey]
                if payload['status'] in ('OK', 'WARNING'):
                    inv_ext, pl_ext, inv_seed, pl_seed = _u107_expand_bundle_rows(payload['bundle4'], inv_seed, pl_seed)
                    inv_rows_to_append.extend(inv_ext)
                    pl_rows_to_append.extend(pl_ext)
            detail_expand_seconds = max(0.0, time.perf_counter() - t_detail_expand_0)

            t_detail_append_0 = time.perf_counter()
            if inv_rows_to_append:
                stats_map['inv_appended'] = _u105_append_inv_rows(ws_inv, inv_rows_to_append, inv_pathkey_col)
                workbook_dirty = True
            if pl_rows_to_append:
                stats_map['pl_appended'] = _u105_append_pl_rows(ws_pl, pl_rows_to_append, pl_pathkey_col)
                workbook_dirty = True
            detail_append_seconds = max(0.0, time.perf_counter() - t_detail_append_0)

            if issue_files:
                log_rows.append(make_log_row('', 'SUMMARY', 'INFO', 'TONG HOP FILE WARNING/ERROR'))
                for fp, st, detail in issue_files:
                    log_rows.append(make_log_row(fp, 'SUMMARY', st, detail))
            if cnt_skip_path > 0:
                log_rows.append(make_log_row('', 'SKIP', 'INFO', f'So luong bo qua do da co File Path OK va khong doi: {cnt_skip_path}'))
            log_rows.append(make_log_row('', 'DONE', 'INFO',
                                         f'Tong={cnt_total}; OK={cnt_ok}; WARNING={cnt_warn}; ERROR={cnt_err}; SkipPath={cnt_skip_path}; SkipSignature={cnt_skip_sig}; Repaired={cnt_repaired}'))

            _u106_set_engine_version(wb_out, U107_ENGINE_VERSION)
            workbook_dirty = True if (workbook_dirty or force_full_resync) else workbook_dirty

            core_seconds = max(0.0, time.perf_counter() - t_all_0)
            save_seconds = 0.0

            # V21: Folder audit da build truoc open de lam folder-delta gate.
            # Den day chi apply vao workbook neu gate bao co delta that.
            try:
                if folder_delta_prebuilt is not None and apply_delta_inplace_func is not None and folder_audit_apply_required:
                    t_faudit_apply_0 = time.perf_counter()
                    folder_audit_apply = apply_delta_inplace_func(wb_out, folder_delta_prebuilt, logger=getattr(self, 'logger', None))
                    folder_audit_apply_seconds = max(0.0, time.perf_counter() - t_faudit_apply_0)
                    if folder_audit_apply.get('workbook_dirty'):
                        workbook_dirty = True
                elif folder_delta_prebuilt is not None:
                    folder_audit_apply = {
                        'folder_sheet_rewritten': 0,
                        'cus_updates': {'SUB_DETAIL': 0, 'INV': 0, 'PL': 0},
                        'workbook_dirty': False,
                    }
            except Exception as exc:
                try:
                    self.logger.log(f'FOLDER_AUDIT_EXT_ERROR_APPLY: {exc}', 'warn')
                except Exception:
                    pass

            stats_map['workbook_saved'] = 1 if workbook_dirty else 0
            source_cache_stats = _u130_get_stats(parse_cache)
            bench_rows = _u104_build_runtime_bench_rows(
                folder_path,
                output_path,
                core_seconds,
                max(0.0, time.perf_counter() - t_all_0),
                stats_map,
                source_cache_stats
            )
            bench_rows.append(make_log_row('', 'BENCH_CACHE_GATE', 'INFO',
                                           f"lookup_total={source_cache_stats.get('lookup_total', 0)} | cache_hit_reuse={source_cache_stats.get('cache_hit_reuse', 0)} | cache_miss_parse={source_cache_stats.get('cache_miss_parse', 0)} | cache_bypass={source_cache_stats.get('cache_bypass', 0)} | parse_real_total={source_cache_stats.get('parse_real_total', 0)} | parse_reused_total={source_cache_stats.get('parse_reused_total', 0)} | cache_store_total={source_cache_stats.get('cache_store_total', 0)}"))
            bench_rows.append(make_log_row('', 'BENCH_INIT_CACHE', 'INFO', f'{init_cache_seconds:.3f}s'))
            bench_rows.append(make_log_row('', 'BENCH_INIT_SCAN_MAP', 'INFO', f'{init_scan_seconds:.3f}s | scanned={len(scanned_keys)}'))
            bench_rows.append(make_log_row('', 'BENCH_INIT_PRECHECK', 'INFO', f"{init_precheck_seconds:.3f}s | loaded={precheck_stats.get('loaded_substate', 0)} | candidate={len(precheck_candidates)} | skip_ok={precheck_skip_ok} | skip_non_ok_stable={precheck_skip_non_ok} | skip_all={precheck_skipped}"))
            bench_rows.append(make_log_row('', 'PRECHECK_SUBSTATE_STATS', 'INFO',
                                           f"fp_match={precheck_stats.get('fp_match', 0)} | fp_match_fingerprint={precheck_stats.get('fp_match_fingerprint', 0)} | fp_match_parsecache={precheck_stats.get('fp_match_parsecache', 0)} | fp_miss={precheck_stats.get('fp_miss', 0)} | status_ok={precheck_stats.get('status_ok', 0)} | status_not_ok={precheck_stats.get('status_not_ok', 0)} | status_missing={precheck_stats.get('status_missing', 0)} | skip_non_ok_stable={precheck_skip_non_ok}"))
            bench_rows.append(make_log_row('', 'BENCH_PARSE_BEFORE_OPEN_WB', 'INFO', f'{parse_before_open_seconds:.3f}s | candidate={preparse_candidate_count}'))
            bench_rows.append(make_log_row('', 'BENCH_OPEN_WB_POST_PARSE', 'INFO', f'{open_wb_post_parse_seconds:.3f}s'))
            bench_rows.append(make_log_row('', 'BENCH_DELETE_PLAN', 'INFO', f'{delete_plan_seconds:.3f}s | inv_blocks={len(inv_delete_plan)} | pl_blocks={len(pl_delete_plan)}'))
            bench_rows.append(make_log_row('', 'BENCH_FOLDER_AUDIT_BUILD', 'INFO',
                                           f"{folder_audit_build_seconds:.3f}s | scanned={folder_audit_stats.get('total_scanned', 0)} | unchanged={folder_audit_stats.get('unchanged_files', 0)} | changed={folder_audit_stats.get('new_or_changed_files', 0)} | deleted={folder_audit_stats.get('deleted_files', 0)} | affected_folders={folder_audit_stats.get('affected_folders', 0)}"))
            bench_rows.append(make_log_row('', 'BENCH_FOLDER_AUDIT_APPLY', 'INFO',
                                           f"{folder_audit_apply_seconds:.3f}s | folder_sheet_rewritten={folder_audit_apply.get('folder_sheet_rewritten', 0)} | cus_updates={folder_audit_apply.get('cus_updates', {})} | workbook_dirty={1 if folder_audit_apply.get('workbook_dirty') else 0}"))
            bench_rows.append(make_log_row('', 'BENCH_DETAIL_INDEX', 'INFO',
                                           f"{detail_index_seconds:.3f}s | inv_index_hit={1 if inv_index_hit else 0} | pl_index_hit={1 if pl_index_hit else 0} | delete_plan_keys={len(delete_plan_keys)}"))
            bench_rows.append(make_log_row('', 'BENCH_SUBDETAIL_UPSERT', 'INFO',
                                           f"{sub_upsert_seconds:.3f}s | sub_update={stats_map.get('sub_update', 0)} | sub_insert={stats_map.get('sub_insert', 0)} | sub_deleted={stats_map.get('sub_deleted', 0)} | parsed={stats_map.get('parsed', 0)} | deleted_paths={len(deleted_paths)}"))
            bench_rows.append(make_log_row('', 'BENCH_ACTION_SPLIT', 'INFO',
                                           f"skip_all={stats_map.get('action_skip_all', 0)} | subdetail_only={stats_map.get('action_subdetail_only', 0)} | rewrite_detail={stats_map.get('action_rewrite_detail', 0)}"))
            bench_rows.append(make_log_row('', 'BENCH_OPEN_DECISION', 'INFO',
                                           _v24_fmt_map(open_decision, ['open_required', 'reason', 'core_needs_workbook', 'folder_needs_workbook', 'rewrite_detail', 'subdetail_only', 'deleted_paths', 'skip_all'])))
            bench_rows.append(make_log_row('', 'BENCH_FOLDER_DELTA_GATE', 'INFO',
                                           _v24_fmt_map(folder_audit_gate, ['reason', 'needs_workbook', 'apply_required', 'deferred', 'save_cache_allowed', 'changed', 'deleted', 'affected', 'cus_dirty', 'folder_sheet_dirty', 'raw_workbook_dirty', 'cache_dirty'])))
            bench_rows.append(make_log_row('', 'BENCH_V24_GATES', 'INFO',
                                           f"sidecar_deleted_pruned={len(deleted_sidecar_only)} | folder_deferred={1 if folder_audit_deferred_metadata else 0} | workbook_deleted_paths={len(deleted_paths)} | folder_apply_required={1 if folder_audit_apply_required else 0}"))
            bench_rows.append(make_log_row('', 'BENCH_DETAIL_DELETE', 'INFO',
                                           f"{detail_delete_seconds:.3f}s | inv_deleted={stats_map.get('inv_deleted', 0)} | pl_deleted={stats_map.get('pl_deleted', 0)}"))
            bench_rows.append(make_log_row('', 'BENCH_DETAIL_SEED', 'INFO',
                                           f"{detail_seed_seconds:.3f}s | inv_rows_snapshot={len(inv_existing_rows)} | pl_rows_snapshot={len(pl_existing_rows)}"))
            bench_rows.append(make_log_row('', 'BENCH_DETAIL_EXPAND', 'INFO',
                                           f"{detail_expand_seconds:.3f}s | inv_rows_to_append={len(inv_rows_to_append)} | pl_rows_to_append={len(pl_rows_to_append)} | rewrite_keys={len(rewrite_keys)}"))
            bench_rows.append(make_log_row('', 'BENCH_DETAIL_APPEND', 'INFO',
                                           f"{detail_append_seconds:.3f}s | inv_appended={stats_map.get('inv_appended', 0)} | pl_appended={stats_map.get('pl_appended', 0)}"))

            if workbook_dirty:
                append_log_rows(ws_log, log_rows + bench_rows)
                t_save_0 = time.perf_counter()
                safe_save_workbook_atomic(wb_out, output_path)
                save_seconds = max(0.0, time.perf_counter() - t_save_0)
                workbook_saved = 1
            else:
                self.logger.log(f'ZERO_WORK_U107: scanned={cnt_total} | skip_ok={cnt_skip_path} | candidate=0', 'info')

            stats_map['workbook_saved'] = workbook_saved
            try:
                if isinstance(parse_cache, dict):
                    entries = parse_cache.get('entries')
                    if isinstance(entries, dict):
                        for pk in list(deleted_paths):
                            entries.pop(pk, None)
                for pk in list(deleted_paths):
                    self._fingerprint_cache.pop(pk, None)
            except Exception:
                pass
            _u137_refresh_fingerprint_cache_from_scan(scan_map, self._fingerprint_cache)
            self.persist_caches(output_path, success_changed_paths=success_changed_paths)
            try:
                _u130_prune_parse_cache(parse_cache, scanned_keys)
                _u130_save_parse_cache(output_path, parse_cache)
            except Exception as exc:
                try:
                    self.logger.log(f'PARSE_CACHE_SAVE_ERROR: {exc}', 'warn')
                except Exception:
                    pass
            try:
                inv_index_map, _ = _u104_collect_detail_rownums_by_pathkey(ws_inv, 30, inv_pathkey_col)
                pl_index_map, _ = _u104_collect_detail_rownums_by_pathkey(ws_pl, 19, pl_pathkey_col)
                sub_index_state, _, _ = _u104_collect_subdetail_state(ws_out, sub_pathkey_col)
                _u131_save_detail_index(output_path, _v23_build_detail_index_payload(sub_index_state, ws_out, inv_index_map, pl_index_map, ws_inv, ws_pl))
            except Exception as exc:
                try:
                    self.logger.log(f'DETAIL_INDEX_SAVE_ERROR: {exc}', 'warn')
                except Exception:
                    pass
            if folder_audit_cache_dirty and folder_audit_cache_payload and save_folder_audit_cache_func and (folder_audit_cache_save_allowed or folder_audit_apply_required):
                try:
                    save_folder_audit_cache_func(output_path, folder_audit_cache_payload)
                except Exception as exc:
                    try:
                        self.logger.log(f'FOLDER_AUDIT_CACHE_SAVE_ERROR: {exc}', 'warn')
                    except Exception:
                        pass

            try:
                total_real_seconds = max(0.0, time.perf_counter() - t_all_0)
                save_line = f'BENCH_SAVE_FINAL: {save_seconds:.3f}s | workbook_saved={workbook_saved}'
                total_line = f'BENCH_TOTAL_REAL: {total_real_seconds:.3f}s | core={core_seconds:.3f}s | folder_build={folder_audit_build_seconds:.3f}s | folder_apply={folder_audit_apply_seconds:.3f}s | save={save_seconds:.3f}s | workbook_saved={workbook_saved}'
                self.logger.log(save_line, 'info')
                self.logger.log(total_line, 'info')
                _ts_v24 = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                _u130_append_runtime_bench_log(output_path, f"{_ts_v24} | BENCH_OPEN_DECISION: {_v24_fmt_map(open_decision, ['open_required', 'reason', 'core_needs_workbook', 'folder_needs_workbook', 'rewrite_detail', 'subdetail_only', 'deleted_paths', 'skip_all'])}")
                _u130_append_runtime_bench_log(output_path, f"{_ts_v24} | BENCH_FOLDER_DELTA_GATE: {_v24_fmt_map(folder_audit_gate, ['reason', 'needs_workbook', 'apply_required', 'deferred', 'save_cache_allowed', 'changed', 'deleted', 'affected', 'cus_dirty', 'folder_sheet_dirty', 'raw_workbook_dirty', 'cache_dirty'])}")
                _u130_append_runtime_bench_log(output_path, f"{_ts_v24} | {save_line}")
                _u130_append_runtime_bench_log(output_path, f"{_ts_v24} | {total_line}")
            except Exception:
                pass

            final_substate = {}
            try:
                pruned_substate_cache = {k: v for k, v in (substate_cache or {}).items() if k not in deleted_paths}
                final_substate = _u135_build_final_substate(pruned_substate_cache, parsed_payloads, action_map)
                _u134_save_substate(output_path, final_substate)
            except Exception as exc:
                try:
                    self.logger.log(f'SUBSTATE_SAVE_ERROR: {exc}', 'warn')
                except Exception:
                    pass

            try:
                hydrated_issue_snapshot = _u139_build_issue_snapshot(final_substate or substate_cache)
                try:
                    _u139_save_issue_snapshot(output_path, hydrated_issue_snapshot)
                except Exception as exc:
                    try:
                        self.logger.log(f'ISSUE_SNAPSHOT_SAVE_ERROR: {exc}', 'warn')
                    except Exception:
                        pass
                issue_files = [
                    (nz_str(x.get('file_path', '')), nz_str(x.get('status', '')), nz_str(x.get('detail', '')))
                    for x in hydrated_issue_snapshot
                ]
                cnt_warn = sum(1 for _, st, _ in issue_files if nz_str(st).upper() == 'WARNING')
                cnt_err = sum(1 for _, st, _ in issue_files if nz_str(st).upper() == 'ERROR')
                cnt_ok = max(0, cnt_total - cnt_warn - cnt_err)
            except Exception as exc:
                try:
                    self.logger.log(f'ISSUE_HYDRATE_FROM_SUBSTATE_ERROR: {exc}', 'warn')
                except Exception:
                    pass

            return {
                'total': cnt_total,
                'ok': cnt_ok,
                'warning': cnt_warn,
                'error': cnt_err,
                'skip_path': cnt_skip_path,
                'skip_signature': cnt_skip_sig,
                'repaired': cnt_repaired,
                'issues': issue_files,
                'scanned_files': sorted(scan_map.values()),
            }

        finally:
            wb_out.close()
            _U90_SOURCE_CACHE_CTX['output_path'] = ''
            _U90_SOURCE_CACHE_CTX['cache'] = None


# =========================================================
# U108 SAFE SA-INVOICE HEADER HANDLING
# - Khong de 1 file SA loi header lam crash ca run
# - Neu khong tim thay header SA-INVOICE hop le:
#     + log warning runtime
#     + bo qua SA-INV Others cho file do
#     + thu fallback standard parser neu co the
# =========================================================

def _u108_build_detail_bundle_safe(file_path: str) -> Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]:
    wb_data = load_workbook_data(file_path)

    if _v6_is_sa_family(file_path):
        inv_sheet_name = "INVOICE" if "INVOICE" in wb_data.sheets else _u77_sheet_by_prefix(wb_data, ["INV"])
        pl_sheet_name = _u77_sheet_by_prefix(wb_data, ["PL"])
        if not pl_sheet_name and "Sheet1" in wb_data.sheets:
            pl_sheet_name = "Sheet1"
    else:
        inv_sheet_name = _u77_sheet_by_prefix(wb_data, ["INV", "INVOICE"])
        pl_sheet_name = _u77_sheet_by_prefix(wb_data, ["PL"])

    report_meta = _v6_collect_meta(wb_data, inv_sheet_name, pl_sheet_name)
    arr = process_one_sub_file(file_path)
    status_text = nz_str(arr[21]).upper()

    inv_rows: List[List[Any]] = []
    pl_rows: List[List[Any]] = []
    inv_other_rows: List[List[Any]] = []
    pl_other_rows: List[List[Any]] = []

    if _v6_is_sa_family(file_path):
        try:
            inv_other_rows = _u79_extract_sa_inv(file_path, wb_data, report_meta, status_text)
        except Exception as exc:
            _rt_log_exception(f'u108_sa_inv_skip::{file_path}', exc)
            try:
                inv_rows = _v6_extract_sub_vsf_inv(file_path, wb_data, report_meta, status_text)
            except Exception as exc2:
                _rt_log_exception(f'u108_sa_inv_fallback_fail::{file_path}', exc2)
                inv_rows = []
        try:
            pl_other_rows = _u78_extract_sa_pl(file_path, wb_data, report_meta, status_text)
        except Exception as exc:
            _rt_log_exception(f'u108_sa_pl_skip::{file_path}', exc)
            try:
                pl_rows = _v6_extract_sub_vsf_pl(file_path, wb_data, report_meta, status_text)
            except Exception as exc2:
                _rt_log_exception(f'u108_sa_pl_fallback_fail::{file_path}', exc2)
                pl_rows = []
    else:
        inv_rows = _v6_extract_sub_vsf_inv(file_path, wb_data, report_meta, status_text)
        pl_rows = _v6_extract_sub_vsf_pl(file_path, wb_data, report_meta, status_text)

    return inv_rows, pl_rows, inv_other_rows, pl_other_rows


# Override final bundle entry used by U107 run
_v72_build_detail_bundle = _u108_build_detail_bundle_safe

if __name__ == '__main__':
    main()
