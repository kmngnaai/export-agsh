from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Iterable

import openpyxl


SHEETS_TO_CHECK = ("SUB_DETAIL", "INV", "PL", "Folder")
REQUIRED_SHEETS = ("SUB_DETAIL", "INV")
TRACKED_STATUSES = ("OK", "WARNING", "ERROR")


def normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_header(value: Any) -> str:
    return normalize_text(value).casefold()


def normalize_status(value: Any) -> str:
    return normalize_text(value).upper()


def row_has_data(row: Iterable[Any]) -> bool:
    return any(normalize_text(value) for value in row)


def inspect_sheet(ws) -> tuple[int, dict[str, int] | None]:
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return 0, None

    header_map = {
        normalize_header(value): index
        for index, value in enumerate(header)
        if normalize_header(value)
    }
    status_index = header_map.get("status")
    status_counts = {status: 0 for status in TRACKED_STATUSES} if status_index is not None else None
    data_rows = 0

    for row in rows:
        if not row_has_data(row):
            continue
        data_rows += 1
        if status_counts is None or status_index >= len(row):
            continue
        status = normalize_status(row[status_index])
        if status in status_counts:
            status_counts[status] += 1

    return data_rows, status_counts


def main() -> int:
    if len(sys.argv) != 2:
        print("FAIL: usage: python tests/check_output_baseline.py <output.xlsx>")
        return 1

    workbook_path = Path(sys.argv[1])
    if not workbook_path.is_file():
        print(f"FAIL: output Excel file does not exist: {workbook_path}")
        return 1

    try:
        workbook = openpyxl.load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        print(f"FAIL: cannot open output Excel file: {workbook_path}")
        print(f"FAIL: {exc}")
        return 1

    try:
        missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in workbook.sheetnames]
        if missing:
            print(f"FAIL: missing required sheet(s): {', '.join(missing)}")
            return 1

        for sheet_name in SHEETS_TO_CHECK:
            if sheet_name not in workbook.sheetnames:
                print(f"SKIP: sheet not found: {sheet_name}")
                continue

            data_rows, status_counts = inspect_sheet(workbook[sheet_name])
            if status_counts is None:
                print(f"SHEET: {sheet_name} -> rows={data_rows}")
            else:
                print(
                    f"SHEET: {sheet_name} -> rows={data_rows} "
                    f"OK={status_counts['OK']} "
                    f"WARNING={status_counts['WARNING']} "
                    f"ERROR={status_counts['ERROR']}"
                )

        print("PASS: output baseline workbook is readable and required sheets are present")
        return 0
    finally:
        workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
