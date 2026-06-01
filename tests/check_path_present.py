from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import openpyxl


SHEETS_TO_CHECK = ("SUB_DETAIL", "INV", "PL", "Folder")


def normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def main() -> int:
    if len(sys.argv) != 3:
        print(
            "FAIL: usage: python tests/check_path_present.py "
            "<output.xlsx> <text-that-must-be-present>"
        )
        return 1

    workbook_path = Path(sys.argv[1])
    search_text = normalize_text(sys.argv[2])

    if not workbook_path.is_file():
        print(f"FAIL: output Excel file does not exist: {workbook_path}")
        return 1

    if not search_text:
        print("FAIL: search text must not be empty")
        return 1

    try:
        workbook = openpyxl.load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        print(f"FAIL: cannot open output Excel file: {workbook_path}")
        print(f"FAIL: {exc}")
        return 1

    needle = search_text.casefold()
    found = False

    try:
        for sheet_name in SHEETS_TO_CHECK:
            if sheet_name not in workbook.sheetnames:
                print(f"SKIP: sheet not found: {sheet_name}")
                continue

            ws = workbook[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    value_text = normalize_text(cell.value)
                    if value_text and needle in value_text.casefold():
                        print(
                            f"PASS: found path text -> "
                            f"sheet={sheet_name} "
                            f"row={cell.row} "
                            f"column={cell.column} "
                            f"value={value_text!r}"
                        )
                        found = True

        if found:
            return 0

        print(
            "FAIL: path text was not found in "
            f"{', '.join(SHEETS_TO_CHECK)} -> {search_text!r}"
        )
        return 1
    finally:
        workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
