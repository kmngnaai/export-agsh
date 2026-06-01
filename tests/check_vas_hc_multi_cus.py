from __future__ import annotations

import sys
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils.datetime import from_excel


SHEET_NAME = "INV"
TARGET_INVOICE = "INV000000491337"
TARGET_DATE = "11/05/2026"
REQUIRED_COLUMNS = (
    "Invoice No",
    "Date",
    "Row Type",
    "PO No",
    "Cus no",
    "Cus no.-",
    "Vas",
    "HC",
)
EXPECTED = {
    "308518255900": (Decimal("285.01"), Decimal("160.00")),
    "308518722800": (Decimal("219.30"), Decimal("160.00")),
    "308519134360": (Decimal("502.64"), Decimal("480.00")),
}
EXPECTED_TOTAL = (Decimal("1006.95"), Decimal("800.00"))


def normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_invoice(value: Any) -> str:
    return normalize_text(value).upper()


def normalize_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            return from_excel(value).strftime("%d/%m/%Y")
        except Exception:
            return normalize_text(value)

    text = normalize_text(value)
    if not text:
        return ""

    for fmt in (
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return text


def normalize_cus_no(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = normalize_text(value)
    if text.endswith(".0"):
        try:
            return str(int(Decimal(text)))
        except (InvalidOperation, ValueError):
            pass
    return text


def to_decimal(value: Any) -> Decimal:
    if value is None or normalize_text(value) == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    text = normalize_text(value).replace(",", "")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Cannot parse number: {value!r}") from exc


def format_decimal(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01')):.2f}"


def fail(message: str) -> int:
    print(f"FAIL: {message}", file=sys.stderr)
    return 1


def main() -> int:
    if len(sys.argv) != 2:
        return fail("usage: python tests/check_vas_hc_multi_cus.py <output.xlsx>")

    output_path = Path(sys.argv[1]).expanduser().resolve()
    if not output_path.is_file():
        return fail(f"output workbook not found: {output_path}")

    try:
        wb = openpyxl.load_workbook(output_path, read_only=True, data_only=True)
    except Exception as exc:
        return fail(f"cannot open workbook: {exc}")

    try:
        if SHEET_NAME not in wb.sheetnames:
            return fail(f"missing sheet: {SHEET_NAME}")

        ws = wb[SHEET_NAME]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return fail(f"sheet {SHEET_NAME} is empty")

        header_map = {normalize_text(value): index for index, value in enumerate(header)}
        missing_columns = [name for name in REQUIRED_COLUMNS if name not in header_map]
        if missing_columns:
            return fail(f"missing required columns: {', '.join(missing_columns)}")

        matched: dict[str, tuple[Decimal, Decimal]] = {}
        duplicates: set[str] = set()
        current_cus_no = ""

        for row in rows:
            invoice_no = normalize_invoice(row[header_map["Invoice No"]])
            invoice_date = normalize_date(row[header_map["Date"]])
            row_type = normalize_text(row[header_map["Row Type"]]).upper()
            po_no = normalize_text(row[header_map["PO No"]]).upper()
            if invoice_no != TARGET_INVOICE or invoice_date != TARGET_DATE:
                continue

            cus_no = normalize_cus_no(row[header_map["Cus no"]])
            cus_slot = normalize_cus_no(row[header_map["Cus no.-"]])
            row_cus_no = cus_no or cus_slot

            if (row_type == "DETAIL" or row_type.startswith("HANDLING")) and row_cus_no:
                current_cus_no = row_cus_no

            is_total = row_type == "TOTAL" or po_no.startswith("TOTAL")
            if not is_total:
                continue

            effective_cus_no = row_cus_no or current_cus_no
            if not effective_cus_no:
                return fail("TOTAL row has no Cus no / Cus no.- and no previous block Cus no")

            if effective_cus_no in matched:
                duplicates.add(effective_cus_no)
                continue

            try:
                vas = to_decimal(row[header_map["Vas"]])
                hc = to_decimal(row[header_map["HC"]])
            except ValueError as exc:
                return fail(f"{effective_cus_no}: {exc}")
            matched[effective_cus_no] = (vas, hc)

        if duplicates:
            return fail(f"duplicate TOTAL rows for effective Cus no: {', '.join(sorted(duplicates))}")

        expected_keys = set(EXPECTED)
        actual_keys = set(matched)
        missing_keys = sorted(expected_keys - actual_keys)
        extra_keys = sorted(actual_keys - expected_keys)
        if missing_keys:
            return fail(f"missing expected Cus no: {', '.join(missing_keys)}")
        if extra_keys:
            return fail(f"unexpected Cus no in TOTAL rows: {', '.join(extra_keys)}")

        for cus_no, (expected_vas, expected_hc) in EXPECTED.items():
            actual_vas, actual_hc = matched[cus_no]
            if actual_vas != expected_vas or actual_hc != expected_hc:
                return fail(
                    f"{cus_no} expected Vas={format_decimal(expected_vas)} "
                    f"HC={format_decimal(expected_hc)}, actual "
                    f"Vas={format_decimal(actual_vas)} HC={format_decimal(actual_hc)}"
                )
            print(f"PASS: {cus_no} -> Vas={format_decimal(actual_vas)} HC={format_decimal(actual_hc)}")

        total_vas = sum((values[0] for values in matched.values()), Decimal("0"))
        total_hc = sum((values[1] for values in matched.values()), Decimal("0"))
        if (total_vas, total_hc) != EXPECTED_TOTAL:
            return fail(
                f"invoice total expected Vas={format_decimal(EXPECTED_TOTAL[0])} "
                f"HC={format_decimal(EXPECTED_TOTAL[1])}, actual "
                f"Vas={format_decimal(total_vas)} HC={format_decimal(total_hc)}"
            )

        print(f"PASS: invoice total -> Vas={format_decimal(total_vas)} HC={format_decimal(total_hc)}")
        return 0
    finally:
        wb.close()


if __name__ == "__main__":
    raise SystemExit(main())
