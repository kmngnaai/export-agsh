from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import openpyxl


SHEETS_TO_CHECK = ("SUB_DETAIL", "INV", "PL", "Folder")
MAX_SAMPLE_POSITIONS = 10


def normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def main() -> int:
    if len(sys.argv) != 3:
        print(
            "FAIL: usage: python tests/check_path_occurrences.py "
            "<output.xlsx> <text-to-find>"
        )
        return 1

    workbook_path = Path(sys.argv[1])
    search_text = normalize_text(sys.argv[2])

    if not workbook_path.is_file():
        print(f"FAIL: output Excel file does not exist: {workbook_path}")
        return 1

    if not search_text:
        print("FAIL: search text must not be empty")
        return 1

    try:
        workbook = openpyxl.load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        print(f"FAIL: cannot open output Excel file: {workbook_path}")
        print(f"FAIL: {exc}")
        return 1

    needle = search_text.casefold()
    total_occurrences = 0

    try:
        for sheet_name in SHEETS_TO_CHECK:
            if sheet_name not in workbook.sheetnames:
                print(f"SKIP: sheet not found: {sheet_name}")
                continue

            sheet_occurrences = 0
            sample_positions: list[tuple[int, int, str]] = []
            ws = workbook[sheet_name]

            for row in ws.iter_rows():
                for cell in row:
                    value_text = normalize_text(cell.value)
                    if not value_text or needle not in value_text.casefold():
                        continue

                    sheet_occurrences += 1
                    if len(sample_positions) < MAX_SAMPLE_POSITIONS:
                        sample_positions.append((cell.row, cell.column, value_text))

            total_occurrences += sheet_occurrences
            print(f"SHEET: {sheet_name} -> occurrences={sheet_occurrences}")
            for row_number, column_number, value_text in sample_positions:
                print(
                    f"  POSITION: row={row_number} "
                    f"column={column_number} "
                    f"value={value_text!r}"
                )

        print(f"TOTAL: occurrences={total_occurrences}")
        if total_occurrences <= 0:
            print(f"FAIL: path text was not found -> {search_text!r}")
            return 1

        print(f"PASS: path text was found -> {search_text!r}")
        return 0
    finally:
        workbook.close()


if __name__ == "__main__":
    raise SystemExit(main())
