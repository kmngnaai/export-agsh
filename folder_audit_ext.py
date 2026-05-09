from __future__ import annotations

import json
import os
import re
import tempfile
from datetime import datetime
from typing import Any, Dict, Iterable, List, Sequence, Set, Tuple

import openpyxl

CACHE_VERSION = 4
TOKHAI_PREFIX = "TOKHAIHQ7X_QDTQ_"
CUS_NO_HEADER = "Cus no"
CUS_NO_FIRST_HEADER = "Cus no.1"
CUS_NO_SLOT_HEADER = "Cus no.-"
FOLDER_SHEET_NAME = "Folder"
TARGET_SHEETS = ("SUB_DETAIL", "INV", "PL")

FOLDER_HEADERS = [
    "Name",
    "Extension",
    "Date accessed",
    "Date modified",
    "Date created",
    "Folder Path",
    "Folder name",
    "Date folder",
    "File no",
    "Check",
]

FILE_TYPE_LABELS: Dict[str, str] = {
    "1": "Tờ khai",
    "2": "hợp đồng",
    "3": "Sub",
    "4": "BL",
    "5": "HĐ",
    "6": "PXK",
}


def _nz(v: Any) -> str:
    if v is None:
        return ""
    try:
        return str(v).strip()
    except Exception:
        return ""


def _uc(v: Any) -> str:
    return _nz(v).upper()


def _norm_path(path: str) -> str:
    try:
        return os.path.abspath(path).replace("/", "\\").rstrip("\\ ").lower()
    except Exception:
        return _nz(path).replace("/", "\\").rstrip("\\ ").lower()


def _display_folder_path(path: str) -> str:
    return os.path.abspath(path).rstrip("\\/") + os.sep


def _fmt_ts(ts: float) -> str:
    try:
        return datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return ""


def _parse_date_folder(folder_name: str) -> str:
    token = _nz(folder_name).split("-", 1)[0].strip()
    if re.fullmatch(r"\d{8}", token):
        try:
            return datetime.strptime(token, "%Y%m%d").strftime("%d/%m/%Y")
        except Exception:
            return ""
    return ""


def _log(logger: Any, msg: str, level: str = "info") -> None:
    try:
        if logger is None:
            return
        if hasattr(logger, "log"):
            logger.log(msg, level)
            return
        if callable(logger):
            logger(msg, level)
    except Exception:
        pass


def _cache_path_for_output(output_path: str) -> str:
    return output_path + ".folderauditcache.json"


def _atomic_write_json(path: str, payload: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    fd, temp_path = tempfile.mkstemp(prefix="tmp_folderaudit_", suffix=".json", dir=os.path.dirname(path) or ".")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        os.replace(temp_path, path)
    finally:
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception:
            pass


def _empty_cache(root_folder: str) -> Dict[str, Any]:
    return {
        "version": CACHE_VERSION,
        "root_folder": _norm_path(root_folder),
        "files": {},
        "folders": {},
    }


def load_cache(root_folder: str, output_path: str, logger: Any = None) -> Dict[str, Any]:
    cache_path = _cache_path_for_output(output_path)
    if not os.path.exists(cache_path):
        return _empty_cache(root_folder)
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            payload = json.load(f)
        if int(payload.get("version", 0)) != CACHE_VERSION:
            return _empty_cache(root_folder)
        if _norm_path(payload.get("root_folder", "")) != _norm_path(root_folder):
            return _empty_cache(root_folder)
        payload.setdefault("files", {})
        payload.setdefault("folders", {})
        return payload
    except Exception as exc:
        _log(logger, f"FOLDER_AUDIT_CACHE_READ_ERROR::{exc}", "warn")
        return _empty_cache(root_folder)


def save_cache(output_path: str, cache: Dict[str, Any]) -> None:
    _atomic_write_json(_cache_path_for_output(output_path), cache)


def extract_cus_no_from_name(file_name: str) -> str:
    m = re.search(r"TOKHAIHQ7X_QDTQ_(\d+)", _uc(file_name))
    return m.group(1) if m else ""


def _is_sub_family(name_upper: str) -> bool:
    return (
        "SUB" in name_upper
        or "VSF" in name_upper
        or "SA2" in name_upper
        or name_upper.startswith("SA")
        or name_upper.startswith("COPY OF SA")
    )


def classify_file_no(file_name: str) -> str:
    name_upper = _uc(file_name)
    stem_upper = _uc(os.path.splitext(file_name)[0])
    prefix4 = stem_upper[:4]
    prefix3 = stem_upper[:3]
    if TOKHAI_PREFIX in name_upper:
        return "1"
    if "PXK" in prefix4:
        return "6"
    if ("HĐ" in prefix4) or ("HD" in prefix4):
        return "5"
    if "AG" in prefix3:
        return "2"
    if stem_upper.endswith("BL") or bool(re.search(r"(^|[^A-Z0-9])BL$", stem_upper)):
        return "4"
    if _is_sub_family(name_upper):
        return "3"
    return ""


def _join_missing_labels(labels: Sequence[str]) -> str:
    labels = [x for x in labels if _nz(x)]
    if not labels:
        return ""
    if len(labels) == 1:
        return labels[0]
    if len(labels) == 2:
        return f"{labels[0]} và {labels[1]}"
    return ", ".join(labels[:-1]) + f" và {labels[-1]}"


def build_folder_check(file_nos: Iterable[str]) -> str:
    existing = {str(x).strip() for x in file_nos if _nz(x)}
    missing = sorted({"1", "2", "3", "4", "5", "6"} - existing, key=int)
    if not missing:
        return "ĐỦ 6"
    return f"Thiếu {_join_missing_labels([FILE_TYPE_LABELS.get(x, x) for x in missing])}"


def scan_current_metadata(root_folder: str, output_path: str, logger: Any = None) -> Dict[str, Dict[str, Any]]:
    root_folder = os.path.abspath(root_folder)
    output_norm = _norm_path(output_path)
    cache_norm = _norm_path(_cache_path_for_output(output_path))
    current: Dict[str, Dict[str, Any]] = {}
    stack = [root_folder]
    scanned = 0
    while stack:
        folder = stack.pop()
        try:
            with os.scandir(folder) as it:
                entries = sorted(list(it), key=lambda e: e.name.lower())
        except Exception as exc:
            _log(logger, f"FOLDER_AUDIT_SCAN_SKIP::{folder}::{exc}", "warn")
            continue
        for entry in reversed(entries):
            if entry.is_dir(follow_symlinks=False):
                stack.append(entry.path)
        for entry in entries:
            if not entry.is_file(follow_symlinks=False):
                continue
            if entry.name.startswith("~$"):
                continue
            file_path = os.path.abspath(entry.path)
            norm = _norm_path(file_path)
            if norm in (output_norm, cache_norm):
                continue
            try:
                st = entry.stat(follow_symlinks=False)
            except Exception:
                continue
            scanned += 1
            current[norm] = {
                "file_path": file_path,
                "folder_key": _norm_path(os.path.dirname(file_path)),
                "file_name": entry.name,
                "size": int(getattr(st, "st_size", 0) or 0),
                "mtime_ns": int(getattr(st, "st_mtime_ns", 0) or 0),
                "atime": float(getattr(st, "st_atime", 0.0) or 0.0),
                "mtime": float(getattr(st, "st_mtime", 0.0) or 0.0),
                "ctime": float(getattr(st, "st_ctime", 0.0) or 0.0),
            }
    _log(logger, f"FOLDER_AUDIT_SCAN_DONE::files={scanned}", "info")
    return current


def _excel_int_or_text(v: str) -> Any:
    t = _nz(v)
    if t.isdigit():
        try:
            return int(t)
        except Exception:
            return t
    return t


def _build_folder_row(meta: Dict[str, Any], file_no: str, check_text: str = "") -> List[Any]:
    file_path = meta["file_path"]
    folder_path = os.path.dirname(file_path)
    folder_name = os.path.basename(folder_path)
    return [
        meta.get("file_name", ""),
        os.path.splitext(meta.get("file_name", ""))[1],
        _fmt_ts(meta.get("atime", 0.0)),
        _fmt_ts(meta.get("mtime", 0.0)),
        _fmt_ts(meta.get("ctime", 0.0)),
        _display_folder_path(folder_path),
        folder_name,
        _parse_date_folder(folder_name),
        _excel_int_or_text(file_no),
        check_text,
    ]


def _new_record(meta: Dict[str, Any]) -> Dict[str, Any]:
    file_name = meta["file_name"]
    file_no = classify_file_no(file_name)
    cus_no = extract_cus_no_from_name(file_name)
    return {
        "file_path": meta["file_path"],
        "folder_key": meta["folder_key"],
        "file_name": file_name,
        "size": meta["size"],
        "mtime_ns": meta["mtime_ns"],
        "file_no": file_no,
        "cus_no": cus_no,
        "folder_cus_no": "",
        "folder_row": _build_folder_row(meta, file_no, ""),
    }


def _folder_summary(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    file_nos = [str(rec.get("file_no", "")) for rec in records if _nz(rec.get("file_no", ""))]
    check_text = build_folder_check(file_nos)
    cus_list: List[str] = []
    seen: Set[str] = set()
    for rec in sorted(records, key=lambda x: _uc(x.get("file_name", ""))):
        cus_no = _nz(rec.get("cus_no", ""))
        if cus_no and cus_no not in seen:
            cus_list.append(cus_no)
            seen.add(cus_no)
    return {
        "cus_no": " | ".join(cus_list),
        "cus_list": cus_list,
        "check": check_text,
        "file_nos": sorted(set(file_nos), key=lambda x: int(x)),
    }


def _apply_summary_to_records(records: List[Dict[str, Any]], summary: Dict[str, Any]) -> None:
    check_text = _nz(summary.get("check", ""))
    folder_cus_no = _nz(summary.get("cus_no", ""))
    for rec in records:
        row = list(rec.get("folder_row", []) or [])
        while len(row) < len(FOLDER_HEADERS):
            row.append("")
        row[9] = check_text
        rec["folder_row"] = row[:len(FOLDER_HEADERS)]
        rec["folder_cus_no"] = folder_cus_no


def build_folder_rows_from_cache(files_state: Dict[str, Dict[str, Any]]) -> List[List[Any]]:
    ordered = sorted(
        files_state.values(),
        key=lambda rec: (
            _nz((rec.get("folder_row") or ["", "", "", "", "", "", "", "", "", ""])[5]).lower(),
            _uc((rec.get("folder_row") or [""])[0]),
        ),
    )
    rows: List[List[Any]] = []
    for rec in ordered:
        row = list(rec.get("folder_row", []) or [])
        while len(row) < len(FOLDER_HEADERS):
            row.append("")
        rows.append(row[:len(FOLDER_HEADERS)])
    return rows


def build_delta_state(root_folder: str, output_path: str, logger: Any = None) -> Dict[str, Any]:
    old_cache = load_cache(root_folder, output_path, logger=logger)
    old_files: Dict[str, Dict[str, Any]] = dict(old_cache.get("files", {}) or {})
    old_folders: Dict[str, Dict[str, Any]] = dict(old_cache.get("folders", {}) or {})
    current_meta = scan_current_metadata(root_folder, output_path, logger=logger)

    current_keys = set(current_meta.keys())
    old_keys = set(old_files.keys())
    deleted_keys = old_keys - current_keys

    files_state: Dict[str, Dict[str, Any]] = {}
    affected_folders: Set[str] = set()
    unchanged_files = 0
    new_or_changed_files = 0

    for key, meta in current_meta.items():
        old = old_files.get(key)
        if old and int(old.get("size", 0) or 0) == int(meta["size"]) and int(old.get("mtime_ns", 0) or 0) == int(meta["mtime_ns"]):
            files_state[key] = old
            unchanged_files += 1
        else:
            files_state[key] = _new_record(meta)
            new_or_changed_files += 1
            affected_folders.add(meta["folder_key"])

    for key in deleted_keys:
        old = old_files.get(key) or {}
        fk = _nz(old.get("folder_key", ""))
        if fk:
            affected_folders.add(fk)

    folder_state = dict(old_folders)
    by_folder: Dict[str, List[Dict[str, Any]]] = {}
    for rec in files_state.values():
        by_folder.setdefault(_nz(rec.get("folder_key", "")), []).append(rec)

    cache_dirty = bool(new_or_changed_files or deleted_keys)
    real_affected: Set[str] = set()

    for fk in affected_folders:
        records = by_folder.get(fk, [])
        old_summary = folder_state.get(fk, {})
        if not records:
            if fk in folder_state:
                del folder_state[fk]
                cache_dirty = True
                real_affected.add(fk)
            continue
        new_summary = _folder_summary(records)
        if old_summary != new_summary:
            cache_dirty = True
            real_affected.add(fk)
        folder_state[fk] = new_summary
        _apply_summary_to_records(records, new_summary)

    return {
        "cache": {
            "version": CACHE_VERSION,
            "root_folder": _norm_path(root_folder),
            "files": files_state,
            "folders": folder_state,
        },
        "desired_folder_rows": build_folder_rows_from_cache(files_state),
        "affected_folders": real_affected if real_affected else affected_folders,
        "stats": {
            "total_scanned": len(current_meta),
            "unchanged_files": unchanged_files,
            "new_or_changed_files": new_or_changed_files,
            "deleted_files": len(deleted_keys),
            "affected_folders": len(real_affected if real_affected else affected_folders),
            "cache_dirty": cache_dirty,
        },
    }


def _ensure_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)


def _header_map(ws) -> Dict[str, int]:
    result: Dict[str, int] = {}
    for c in range(1, max(1, ws.max_column or 1) + 1):
        name = _nz(ws.cell(1, c).value)
        if name:
            result[name] = c
    return result


def _ensure_header_at_end(ws, header_name: str) -> Tuple[int, bool]:
    hdr = _header_map(ws)
    if header_name in hdr:
        return hdr[header_name], False
    col = max(1, (ws.max_column or 0) + 1)
    ws.cell(1, col).value = header_name
    return col, True


def _ensure_header_after(ws, after_header: str, header_name: str) -> Tuple[int, bool]:
    hdr = _header_map(ws)
    if header_name in hdr:
        return hdr[header_name], False
    after_col = hdr.get(after_header, 0)
    if after_col <= 0:
        return _ensure_header_at_end(ws, header_name)
    ws.insert_cols(after_col + 1, 1)
    ws.cell(1, after_col + 1).value = header_name
    return after_col + 1, True


def _read_rows(ws, num_cols: int) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for r in range(2, (ws.max_row or 0) + 1):
        arr = []
        has_data = False
        for c in range(1, num_cols + 1):
            v = ws.cell(r, c).value
            arr.append("" if v is None else v)
            if _nz(v):
                has_data = True
        if has_data:
            rows.append(arr)
    return rows


def _sync_folder_sheet_inplace(ws, desired_rows: List[List[Any]]) -> bool:
    current_headers = [_nz(ws.cell(1, c).value) for c in range(1, len(FOLDER_HEADERS) + 1)]
    current_rows = _read_rows(ws, len(FOLDER_HEADERS))
    desired = [["" if v is None else v for v in row[:len(FOLDER_HEADERS)]] for row in desired_rows]
    if current_headers == FOLDER_HEADERS and current_rows == desired:
        return False
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for c, h in enumerate(FOLDER_HEADERS, start=1):
        ws.cell(1, c).value = h
    row_idx = 2
    for arr in desired:
        for c, v in enumerate(arr, start=1):
            cell = ws.cell(row_idx, c)
            cell.value = v
            if c == 9 and isinstance(v, int):
                cell.number_format = "0"
        row_idx += 1
    return True


def _extract_s_invoice_seq(s_invoice_value: Any) -> int:
    text = _nz(s_invoice_value)
    m = re.search(r'-(\d+)$', text)
    if not m:
        return 0
    try:
        return int(m.group(1))
    except Exception:
        return 0


def _build_cus_no_first(summary: Dict[str, Any]) -> str:
    cus_list = [x for x in (summary.get("cus_list") or []) if _nz(x)]
    if not cus_list:
        joined = _nz(summary.get("cus_no", ""))
        if joined:
            cus_list = [_nz(x) for x in joined.split("|") if _nz(x)]
    return _nz(cus_list[0]) if cus_list else ""


def _build_cus_no_slot(summary: Dict[str, Any], s_invoice_value: Any) -> str:
    cus_list = [x for x in (summary.get('cus_list') or []) if _nz(x)]
    if not cus_list:
        joined = _nz(summary.get('cus_no', ''))
        if joined:
            cus_list = [_nz(x) for x in joined.split('|') if _nz(x)]
    if not cus_list:
        return ''
    seq = _extract_s_invoice_seq(s_invoice_value)
    if seq <= 0:
        return ''
    idx = (seq - 1) // 50
    if idx < 0 or idx >= len(cus_list):
        return ''
    return _nz(cus_list[idx])


def _normalize_invoice_no_for_sequence(v: Any) -> str:
    text = _uc(v)
    return "".join(ch for ch in text if ch.isalnum() or ch == ".")


def _try_datetime_for_sequence(v: Any):
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return openpyxl.utils.datetime.from_excel(float(v))
        except Exception:
            return None
    text = _nz(v)
    if not text:
        return None
    text = text.replace(".", "/").replace("-", "/")
    text = re.sub(r"\s+00:00:00$", "", text).strip()
    for fmt in (
        "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d",
        "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%m/%d/%Y",
        "%d/%m/%y", "%m/%d/%y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            pass
    return None


def _normalize_date_for_sequence(v: Any) -> str:
    dt = _try_datetime_for_sequence(v)
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d")
    return _nz(v)


def _date_serial_text_for_sdatewinv(v: Any) -> str:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return str(int(round(float(v))))
        except Exception:
            return ""
    dt = _try_datetime_for_sequence(v)
    if isinstance(dt, datetime):
        try:
            return str(int(round(openpyxl.utils.datetime.to_excel(dt))))
        except Exception:
            return ""
    return ""


def _make_sdatewinv_by_current_date(date_value: Any, s_invoice_value: str) -> str:
    s_invoice_value = _nz(s_invoice_value)
    if not s_invoice_value:
        return ""
    serial_text = _date_serial_text_for_sdatewinv(date_value)
    return f"{serial_text}{s_invoice_value}" if serial_text else s_invoice_value


def _is_non_sequence_row_type(row_type_value: Any) -> bool:
    text = _uc(row_type_value).replace(":", " ")
    text = re.sub(r"\s+", " ", text).strip()
    if not text:
        return True
    if "TOTAL" in text:
        return True
    if "HANDLING" in text or "UPCHARGE" in text or "DEDUCT" in text:
        return True
    return text != "DETAIL"


def _sync_s_invoice_by_cus_no_inplace(ws) -> Tuple[bool, int]:
    """
    Re-number S. Invoice# after Cus no.- is available.

    Old key: (Invoice No, Date)
    New key: (Invoice No, Date, Cus no.-)

    HANDLING CHARGE / UPCHARGE / DEDUCT / TOTAL rows are forced blank.
    S.DateWInv# is recalculated from the same Date + new S. Invoice#.
    """
    hdr = _header_map(ws)
    date_col = hdr.get("Date", 0)
    inv_col = hdr.get("Invoice No", 0)
    s_inv_col = hdr.get("S. Invoice#", 0)
    s_date_col = hdr.get("S.DateWInv#", 0)
    row_type_col = hdr.get("Row Type", 0)
    cus_slot_col = hdr.get(CUS_NO_SLOT_HEADER, 0)

    if min(date_col, inv_col, s_inv_col, s_date_col, row_type_col, cus_slot_col) <= 0:
        return False, 0

    dirty = False
    updated = 0
    counters: Dict[Tuple[str, str, str], int] = {}

    for r in range(2, (ws.max_row or 0) + 1):
        row_type_value = ws.cell(r, row_type_col).value
        if _is_non_sequence_row_type(row_type_value):
            desired_s_inv = ""
            desired_s_date = ""
        else:
            invoice_no = _normalize_invoice_no_for_sequence(ws.cell(r, inv_col).value)
            date_key = _normalize_date_for_sequence(ws.cell(r, date_col).value)
            cus_slot = _nz(ws.cell(r, cus_slot_col).value)
            key = (invoice_no, date_key, cus_slot)
            counters[key] = counters.get(key, 0) + 1
            desired_s_inv = f"{invoice_no}-{counters[key]}" if invoice_no else f"-{counters[key]}"
            desired_s_date = _make_sdatewinv_by_current_date(ws.cell(r, date_col).value, desired_s_inv)

        old_s_inv = _nz(ws.cell(r, s_inv_col).value)
        if old_s_inv != desired_s_inv:
            ws.cell(r, s_inv_col).value = desired_s_inv
            dirty = True
            updated += 1

        old_s_date = _nz(ws.cell(r, s_date_col).value)
        if old_s_date != desired_s_date:
            ws.cell(r, s_date_col).value = desired_s_date
            dirty = True
            updated += 1

    return dirty, updated


def _sync_cus_no_sheet_inplace(ws, affected_folders: Set[str], folder_state: Dict[str, Dict[str, Any]]) -> Tuple[bool, int]:
    hdr = _header_map(ws)
    file_path_col = hdr.get("File Path", 0)
    if file_path_col <= 0:
        return False, 0
    cus_col, cus_added = _ensure_header_at_end(ws, CUS_NO_HEADER)
    cus_first_col, first_added = _ensure_header_after(ws, CUS_NO_HEADER, CUS_NO_FIRST_HEADER)
    cus_slot_col, slot_added = _ensure_header_after(ws, CUS_NO_FIRST_HEADER, CUS_NO_SLOT_HEADER)
    hdr = _header_map(ws)
    file_path_col = hdr.get("File Path", file_path_col)
    s_invoice_col = hdr.get("S. Invoice#", 0)
    dirty = cus_added or first_added or slot_added
    updated = 0
    if not affected_folders:
        return dirty, updated
    for r in range(2, (ws.max_row or 0) + 1):
        file_path = _nz(ws.cell(r, file_path_col).value)
        folder_key = _norm_path(os.path.dirname(file_path)) if file_path else ""
        if folder_key not in affected_folders:
            continue
        summary = folder_state.get(folder_key) or {}
        desired_cus = _nz(summary.get("cus_no", ""))
        s_invoice_value = ws.cell(r, s_invoice_col).value if s_invoice_col > 0 else ""
        desired_first = _build_cus_no_first(summary) if _nz(s_invoice_value) else ""
        desired_slot = _build_cus_no_slot(summary, s_invoice_value) if s_invoice_col > 0 else ""

        old_cus = _nz(ws.cell(r, cus_col).value)
        if old_cus != desired_cus:
            cell = ws.cell(r, cus_col)
            cell.value = _excel_int_or_text(desired_cus)
            if desired_cus.isdigit():
                cell.number_format = "0"
            updated += 1
            dirty = True

        old_first = _nz(ws.cell(r, cus_first_col).value)
        if old_first != desired_first:
            cell = ws.cell(r, cus_first_col)
            cell.value = _excel_int_or_text(desired_first)
            if desired_first.isdigit():
                cell.number_format = "0"
            updated += 1
            dirty = True

        old_slot = _nz(ws.cell(r, cus_slot_col).value)
        if old_slot != desired_slot:
            cell = ws.cell(r, cus_slot_col)
            cell.value = _excel_int_or_text(desired_slot)
            if desired_slot.isdigit():
                cell.number_format = "0"
            updated += 1
            dirty = True
    return dirty, updated


def apply_delta_inplace(wb, delta: Dict[str, Any], logger: Any = None) -> Dict[str, Any]:
    desired_folder_rows: List[List[Any]] = list(delta.get("desired_folder_rows", []) or [])
    affected_folders: Set[str] = set(delta.get("affected_folders", []) or [])
    cache = delta.get("cache", {}) or {}
    folder_state = cache.get("folders", {}) or {}

    workbook_dirty = False
    folder_sheet_rewritten = 0
    cus_updates = {"SUB_DETAIL": 0, "INV": 0, "PL": 0}
    sequence_updates = {"INV": 0, "PL": 0}

    ws_folder = _ensure_sheet(wb, FOLDER_SHEET_NAME)
    if _sync_folder_sheet_inplace(ws_folder, desired_folder_rows):
        workbook_dirty = True
        folder_sheet_rewritten = 1

    for sheet_name in TARGET_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        dirty, updated = _sync_cus_no_sheet_inplace(wb[sheet_name], affected_folders, folder_state)
        cus_updates[sheet_name] = updated
        if dirty:
            workbook_dirty = True

    # S. Invoice# must be recalculated AFTER Cus no.- is available.
    # Only INV/PL have Date + Invoice No + S. Invoice# + S.DateWInv#.
    for sheet_name in ("INV", "PL"):
        if sheet_name not in wb.sheetnames:
            continue
        dirty, updated = _sync_s_invoice_by_cus_no_inplace(wb[sheet_name])
        sequence_updates[sheet_name] = updated
        if dirty:
            workbook_dirty = True

    result = {
        "workbook_dirty": workbook_dirty,
        "folder_sheet_rewritten": folder_sheet_rewritten,
        "cus_updates": cus_updates,
        "sequence_updates": sequence_updates,
        "affected_folders": len(affected_folders),
        "cache_dirty": bool((delta.get("stats", {}) or {}).get("cache_dirty", False)),
    }
    _log(logger, f"FOLDER_AUDIT_APPLY_INPLACE::{result}", "info")
    return result


def apply_folder_audit_extension(root_folder: str, output_path: str, logger: Any = None) -> Dict[str, Any]:
    delta = build_delta_state(root_folder, output_path, logger=logger)
    if not os.path.exists(output_path):
        raise FileNotFoundError(output_path)
    wb = openpyxl.load_workbook(output_path)
    try:
        apply_result = apply_delta_inplace(wb, delta, logger=logger)
        if apply_result["workbook_dirty"]:
            wb.save(output_path)
    finally:
        wb.close()
    if (delta.get("stats", {}) or {}).get("cache_dirty"):
        save_cache(output_path, delta["cache"])
    return {
        **(delta.get("stats", {}) or {}),
        **apply_result,
        "cache_path": _cache_path_for_output(output_path),
    }


def main() -> None:
    import sys
    if len(sys.argv) < 3:
        print("Usage: python folder_audit_ext.py <root_folder> <output.xlsx>")
        raise SystemExit(1)
    print(apply_folder_audit_extension(sys.argv[1], sys.argv[2]))


if __name__ == "__main__":
    main()
